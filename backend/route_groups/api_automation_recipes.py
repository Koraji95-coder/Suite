from __future__ import annotations

import copy
import json
import mimetypes
import os
import re
import shutil
import tempfile
import time
import uuid
from datetime import datetime
from functools import wraps
from pathlib import Path, PurePosixPath
from typing import Any, Callable, Dict, List, Optional, Tuple

from flask import Blueprint, jsonify, request, send_file
from flask_limiter import Limiter
from openpyxl import Workbook
from werkzeug.utils import safe_join, secure_filename

from backend.runtime_paths import (
    is_absolute_path_value,
    resolve_runtime_directory,
    resolve_runtime_path,
)

MAX_DRAWINGS = 75
MAX_RULES = 100
MAX_OPERATIONS = 8000
STATE_DIRNAME = "suite_automation_recipe_api"


def create_automation_recipe_blueprint(
    *,
    limiter: Limiter,
    logger: Any,
    require_supabase_user: Callable,
    is_valid_api_key: Callable[[Optional[str]], bool],
    send_autocad_dotnet_command: Optional[Callable[[str, Dict[str, Any]], Dict[str, Any]]] = None,
    send_autocad_acade_command: Optional[Callable[[str, Dict[str, Any]], Dict[str, Any]]] = None,
) -> Blueprint:
    bp = Blueprint("automation_recipe_api", __name__, url_prefix="/api")

    def _state_dir() -> str:
        path = os.path.join(tempfile.gettempdir(), STATE_DIRNAME)
        os.makedirs(path, exist_ok=True)
        return path

    def _state_path(filename: str) -> str:
        return os.path.join(_state_dir(), filename)

    def _load_state_index(filename: str) -> Dict[str, Dict[str, Any]]:
        path = _state_path(filename)
        if not os.path.exists(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as handle:
                payload = json.load(handle)
            if not isinstance(payload, dict):
                return {}
            return {
                str(key): value
                for key, value in payload.items()
                if isinstance(value, dict)
            }
        except Exception:
            logger.warning("Automation recipe API state file could not be loaded: %s", path)
            return {}

    def _persist_state_index(filename: str, payload: Dict[str, Dict[str, Any]]) -> None:
        path = _state_path(filename)
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2)

    generated_reports: Dict[str, Dict[str, Any]] = _load_state_index("reports.json")
    generated_runs: Dict[str, Dict[str, Any]] = _load_state_index("runs.json")

    def require_user_or_api_key(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            provided_key = request.headers.get("X-API-Key")
            if provided_key and is_valid_api_key(provided_key):
                return f(*args, **kwargs)
            wrapped = require_supabase_user(f)
            return wrapped(*args, **kwargs)

        return decorated

    def _request_error_response(*, message: str, request_id: str, status_code: int):
        return jsonify({"success": False, "error": message, "requestId": request_id}), status_code

    def _normalize_text(value: Any) -> str:
        return str(value or "").strip()

    def _normalize_nullable_text(value: Any) -> str | None:
        normalized = _normalize_text(value)
        return normalized or None

    def _normalize_string_array(value: Any) -> List[str]:
        if not isinstance(value, list):
            return []
        output: List[str] = []
        seen: set[str] = set()
        for entry in value:
            normalized = _normalize_text(entry)
            if not normalized:
                continue
            key = normalized.lower()
            if key in seen:
                continue
            seen.add(key)
            output.append(normalized)
        return output

    def _normalize_step_source(value: Any) -> str:
        normalized = _normalize_text(value).lower()
        if normalized == "autowire":
            return "autowire"
        if normalized == "cad-utils":
            return "cad-utils"
        return "autodraft"

    def _normalize_binding_kinds(value: Any) -> List[str]:
        allowed = {
            "title-block",
            "drawing-row",
            "deliverable-row",
            "drawing-content",
            "terminal-wiring",
            "schedule-row",
            "note-only",
        }
        return [entry for entry in _normalize_string_array(value) if entry in allowed]

    def _normalize_drawing_key(value: Any) -> str:
        return re.sub(r"[^A-Z0-9]+", "", _normalize_text(value).upper())

    def _resolve_existing_directory(path_value: Any) -> Path | None:
        normalized = _normalize_text(path_value)
        if not normalized or not is_absolute_path_value(normalized):
            return None
        candidate = resolve_runtime_directory(normalized)
        if candidate is None:
            return None
        try:
            return candidate.resolve()
        except Exception:
            return None

    def _project_root_realpath(project_root: Path) -> str:
        return os.path.realpath(str(project_root))

    def _ensure_under_project_root(project_root: Path, candidate_path: str, *, field_name: str) -> Path:
        project_root_real = _project_root_realpath(project_root)
        candidate_real = os.path.realpath(candidate_path)
        try:
            if os.path.commonpath([project_root_real, candidate_real]) != project_root_real:
                raise ValueError
        except ValueError as exc:
            raise ValueError(f"{field_name} resolves outside the project root.") from exc
        return Path(candidate_real)

    def _resolve_project_scoped_file(
        project_root: Path | None,
        path_value: Any,
        *,
        field_name: str,
        allowed_suffixes: Tuple[str, ...],
    ) -> Path:
        normalized = _normalize_text(path_value)
        if not normalized:
            raise ValueError(f"{field_name} is required.")
        if project_root is None:
            raise ValueError(f"{field_name} requires a valid projectRootPath.")
        project_root_real = _project_root_realpath(project_root)
        if is_absolute_path_value(normalized):
            resolved_candidate = resolve_runtime_path(normalized)
            if resolved_candidate is None:
                raise ValueError(f"{field_name} was not found.")
            candidate = _ensure_under_project_root(
                project_root,
                str(resolved_candidate),
                field_name=field_name,
            )
        else:
            candidate_text = safe_join(project_root_real, normalized.replace("\\", "/"))
            if candidate_text is None:
                raise ValueError(f"{field_name} resolves outside the project root.")
            candidate = _ensure_under_project_root(
                project_root,
                candidate_text,
                field_name=field_name,
            )
        if candidate.suffix.lower() not in allowed_suffixes:
            raise ValueError(
                f"{field_name} must use one of the following extensions: {', '.join(allowed_suffixes)}."
            )
        if not candidate.exists() or not candidate.is_file():
            raise ValueError(f"{field_name} was not found: {candidate}")
        return candidate

    def _workspace_destination_path(workspace_root: str, relative_path: str, fallback_name: str) -> str:
        normalized_relative = _normalize_text(relative_path) or _normalize_text(fallback_name) or "drawing.dwg"
        normalized_relative = normalized_relative.replace("\\", "/")
        candidate_parts = [
            part for part in PurePosixPath(normalized_relative).parts if part not in {"", ".", "/"}
        ]
        if any(part == ".." for part in candidate_parts):
            raise ValueError(f"Drawing path '{relative_path}' resolves outside the workspace.")
        destination = safe_join(os.path.join(workspace_root, "drawings"), *candidate_parts)
        if destination is None:
            raise ValueError(f"Drawing path '{relative_path}' resolves outside the workspace.")
        return destination

    def _resolve_workspace_source_path(drawing: Dict[str, str]) -> str:
        drawing_path = _normalize_text(drawing.get("path"))
        if not drawing_path:
            raise ValueError("Selected drawing path is missing.")

        source_root = _resolve_existing_directory(drawing.get("sourceRootPath"))
        if source_root is not None:
            resolved_source = _ensure_under_project_root(
                source_root,
                drawing_path,
                field_name="selectedDrawingPaths",
            )
        else:
            resolved_runtime_path = resolve_runtime_path(drawing_path)
            if resolved_runtime_path is None:
                raise ValueError("Selected drawing is not available.")
            resolved_source = Path(os.path.realpath(str(resolved_runtime_path)))

        resolved_text = str(resolved_source)
        if not os.path.isfile(resolved_text):
            raise ValueError(f"Selected drawing '{resolved_text}' does not exist.")
        return resolved_text

    def _request_id() -> str:
        raw = (
            request.headers.get("X-Request-ID")
            or request.headers.get("X-Request-Id")
            or ""
        )
        if not raw:
            payload = request.get_json(silent=True) or {}
            if isinstance(payload, dict):
                raw = str(payload.get("requestId") or payload.get("request_id") or "")
        normalized = re.sub(r"[^A-Za-z0-9._:-]+", "-", str(raw or "").strip())
        return normalized[:128] if normalized else f"automation-recipe-{int(time.time() * 1000)}"

    def _call_bridge(action: str, payload: Dict[str, Any], request_id: str) -> Dict[str, Any]:
        if send_autocad_dotnet_command is None:
            raise RuntimeError("AutoCAD .NET bridge is not configured.")
        response = send_autocad_dotnet_command(
            action,
            {
                **payload,
                "requestId": request_id,
            },
        )
        if not isinstance(response, dict):
            raise RuntimeError("Malformed response from AutoCAD .NET bridge.")
        if not response.get("ok"):
            raise RuntimeError(
                str(response.get("error") or response.get("message") or "Unknown bridge error.")
            )
        result_payload = response.get("result")
        if not isinstance(result_payload, dict):
            raise RuntimeError("Invalid .NET bridge result payload.")
        return result_payload

    def _call_acade_host(action: str, payload: Dict[str, Any], request_id: str) -> Dict[str, Any]:
        if send_autocad_acade_command is None:
            raise RuntimeError("AutoCAD in-process ACADE host is not configured.")
        response = send_autocad_acade_command(
            action,
            {
                **payload,
                "requestId": request_id,
            },
        )
        if not isinstance(response, dict):
            raise RuntimeError("Malformed response from the AutoCAD in-process ACADE host.")
        if not response.get("ok"):
            raise RuntimeError(
                str(response.get("error") or response.get("message") or "Unknown ACADE host error.")
            )
        result_payload = response.get("result")
        if not isinstance(result_payload, dict):
            raise RuntimeError("Invalid in-process ACADE host result payload.")
        return result_payload

    def _candidate_plugin_roots() -> List[str]:
        candidates: List[str] = []
        env_roots = [
            os.environ.get("ProgramFiles"),
            os.environ.get("ProgramFiles(x86)"),
            os.environ.get("APPDATA"),
            os.environ.get("ProgramData"),
            os.environ.get("ALLUSERSPROFILE"),
            os.path.join(os.environ.get("USERPROFILE", ""), "AppData", "Roaming"),
        ]
        for root in env_roots:
            normalized_root = _normalize_text(root)
            if not normalized_root:
                continue
            if normalized_root.endswith("Autodesk\\ApplicationPlugins"):
                candidate = os.path.join(normalized_root, "SuiteCadAuthoring.bundle")
            else:
                candidate = os.path.join(normalized_root, "Autodesk", "ApplicationPlugins", "SuiteCadAuthoring.bundle")
            candidate = os.path.abspath(candidate)
            if candidate not in candidates:
                candidates.append(candidate)
        return candidates

    def _detect_plugin_status() -> Dict[str, Any]:
        for candidate in _candidate_plugin_roots():
            dll_path = os.path.join(candidate, "Contents", "Win64", "SuiteCadAuthoring.dll")
            deps_path = os.path.join(candidate, "Contents", "Win64", "SuiteCadAuthoring.deps.json")
            runtime_config_path = os.path.join(
                candidate,
                "Contents",
                "Win64",
                "SuiteCadAuthoring.runtimeconfig.json",
            )
            package_path = os.path.join(candidate, "PackageContents.xml")
            if all(os.path.exists(path) for path in [dll_path, deps_path, runtime_config_path, package_path]):
                return {
                    "ok": True,
                    "bundleRoot": candidate,
                    "dllPath": dll_path,
                    "errors": [],
                }

        return {
            "ok": False,
            "bundleRoot": _candidate_plugin_roots()[0] if _candidate_plugin_roots() else "",
            "dllPath": "",
            "errors": ["Suite CAD authoring plugin bundle was not found on this workstation."],
        }

    def _find_acade_support(work_package: Dict[str, Any]) -> Dict[str, Any]:
        project_root = _resolve_existing_directory(work_package.get("projectRootPath"))
        raw_project_file = _normalize_text(work_package.get("acadeProjectFilePath"))
        if not project_root and not _normalize_text(work_package.get("projectRootPath")):
            return {
                "projectFile": None,
                "supportFiles": [],
                "warnings": [
                    "Project root path is missing, so ACADE support files could not be checked."
                ],
            }
        if not project_root:
            return {
                "projectFile": None,
                "supportFiles": [],
                "warnings": [
                    f"Project root '{_normalize_text(work_package.get('projectRootPath'))}' was not found for ACADE support checks."
                ],
            }

        project_file = None
        support_files: List[str] = []
        warnings: List[str] = []
        if raw_project_file:
            try:
                resolved_project_file = _resolve_project_scoped_file(
                    project_root,
                    raw_project_file,
                    field_name="acadeProjectFilePath",
                    allowed_suffixes=(".wdp",),
                )
            except ValueError:
                warnings.append(
                    "ACADE project file path is invalid or outside the project root."
                )
            else:
                project_file = str(resolved_project_file)
                support_files.append(project_file)
        else:
            warnings.append(
                "ACADE support auto-discovery is disabled. Provide acadeProjectFilePath in the work package to enable project-file checks."
            )
        return {
            "projectFile": project_file,
            "supportFiles": support_files[:32],
            "warnings": warnings,
        }

    def _resolve_project_drawings(work_package: Dict[str, Any]) -> List[Dict[str, str]]:
        raw_selected = work_package.get("selectedDrawingPaths")
        if not isinstance(raw_selected, list) or len(raw_selected) == 0:
            raise ValueError("selectedDrawingPaths must contain at least one drawing path.")
        if len(raw_selected) > MAX_DRAWINGS:
            raise ValueError(f"Too many selected drawings. Maximum is {MAX_DRAWINGS}.")

        drawing_root_path = _resolve_existing_directory(work_package.get("drawingRootPath"))
        drawing_root = _project_root_realpath(drawing_root_path) if drawing_root_path else ""
        resolved: List[Dict[str, str]] = []
        seen_paths: set[str] = set()

        for entry in raw_selected:
            raw_path = _normalize_text(entry)
            if not raw_path:
                continue
            if is_absolute_path_value(raw_path):
                absolute_candidate = os.path.abspath(raw_path)
                if drawing_root_path is not None:
                    absolute_resolved = _ensure_under_project_root(
                        drawing_root_path,
                        absolute_candidate,
                        field_name="selectedDrawingPaths",
                    )
                    relative_path = os.path.relpath(str(absolute_resolved), drawing_root)
                else:
                    absolute_resolved = Path(os.path.realpath(absolute_candidate))
                    relative_path = os.path.basename(str(absolute_resolved)) or raw_path
            else:
                if not drawing_root_path or not drawing_root:
                    raise ValueError(
                        "drawingRootPath is required when selectedDrawingPaths are relative."
                    )
                candidate_text = safe_join(drawing_root, raw_path.replace("\\", "/"))
                if candidate_text is None:
                    raise ValueError(
                        f"Drawing path '{raw_path}' resolves outside the drawing root."
                    )
                absolute_resolved = _ensure_under_project_root(
                    drawing_root_path,
                    candidate_text,
                    field_name="selectedDrawingPaths",
                )
                relative_path = raw_path

            absolute_path = str(absolute_resolved)
            normalized_key = absolute_path.lower()
            if normalized_key in seen_paths:
                continue
            seen_paths.add(normalized_key)
            resolved.append(
                {
                    "path": absolute_path,
                    "relativePath": relative_path,
                    "drawingName": os.path.basename(absolute_path) or absolute_path,
                    "drawingNumber": Path(relative_path).stem,
                    "drawingKey": _normalize_drawing_key(Path(relative_path).stem),
                    "sourceRootPath": drawing_root,
                    "exists": "true" if os.path.isfile(absolute_path) else "",
                }
            )

        if not resolved:
            raise ValueError("No valid project drawings were resolved from the work package.")
        return resolved

    def _build_drawing_lookup(drawings: List[Dict[str, str]]) -> Dict[str, List[Dict[str, str]]]:
        lookup: Dict[str, List[Dict[str, str]]] = {}
        for drawing in drawings:
            keys = {
                _normalize_drawing_key(drawing.get("path")),
                _normalize_drawing_key(drawing.get("relativePath")),
                _normalize_drawing_key(drawing.get("drawingNumber")),
                _normalize_drawing_key(drawing.get("drawingName")),
            }
            for key in keys:
                if not key:
                    continue
                lookup.setdefault(key, []).append(drawing)
        return lookup

    def _resolve_schedule_rows(
        drawings: List[Dict[str, str]],
        strip_rows: List[Dict[str, Any]],
        connection_rows: List[Dict[str, Any]],
    ) -> List[str]:
        blockers: List[str] = []
        lookup = _build_drawing_lookup(drawings)
        all_rows = [
            ("TerminalStrips", row)
            for row in strip_rows
            if isinstance(row, dict)
        ] + [
            ("TerminalConnections", row)
            for row in connection_rows
            if isinstance(row, dict)
        ]
        for sheet_name, row in all_rows:
            row_id = _normalize_text(row.get("id")) or _normalize_text(row.get("rowId")) or "row"
            drawing_ref = _normalize_text(row.get("drawingPath") or row.get("drawingNumber"))
            matches = lookup.get(_normalize_drawing_key(drawing_ref), [])
            if len(matches) != 1:
                blockers.append(
                    f"{sheet_name} row '{row_id}' could not be resolved to exactly one selected issue-set drawing."
                )
        return blockers

    def _extract_enabled_steps(recipe: Dict[str, Any]) -> List[Dict[str, Any]]:
        raw_steps = recipe.get("steps")
        if not isinstance(raw_steps, list):
            return []
        steps: List[Dict[str, Any]] = []
        for raw_step in raw_steps:
            if not isinstance(raw_step, dict):
                continue
            if raw_step.get("enabled") is False:
                continue
            steps.append(raw_step)
        return steps

    def _normalize_report_artifact(
        *,
        artifact_id: str,
        label: str,
        kind: str,
        download_url: str | None,
        path: str | None,
        description: str | None,
    ) -> Dict[str, Any]:
        return {
            "id": artifact_id,
            "label": label,
            "kind": kind,
            "downloadUrl": download_url,
            "path": path,
            "description": description,
        }

    def _register_generated_report(
        report_path: str,
        *,
        filename: str | None = None,
        mimetype: str | None = None,
    ) -> Tuple[str, Dict[str, Any]]:
        report_id = uuid.uuid4().hex
        detected_mimetype = mimetype or mimetypes.guess_type(report_path)[0] or "application/octet-stream"
        report_entry = {
            "path": report_path,
            "filename": filename or os.path.basename(report_path),
            "mimetype": detected_mimetype,
        }
        generated_reports[report_id] = report_entry
        _persist_state_index("reports.json", generated_reports)
        return report_id, report_entry

    def _build_step_summary(
        *,
        source: str,
        label: str,
        enabled: bool,
        ready: bool,
        actionable: bool,
        planned_item_count: int,
        approved_item_count: int,
        warning_count: int,
        binding_kinds: List[str],
        summary: str,
        request_id: str | None = None,
        report_id: str | None = None,
    ) -> Dict[str, Any]:
        return {
            "id": f"{source}-step",
            "source": source,
            "label": label,
            "enabled": enabled,
            "ready": ready,
            "actionable": actionable,
            "plannedItemCount": max(0, planned_item_count),
            "approvedItemCount": max(0, approved_item_count),
            "warningCount": max(0, warning_count),
            "bindingKinds": binding_kinds,
            "summary": summary,
            "requestId": request_id,
            "reportId": report_id,
        }

    def _build_preflight(payload: Dict[str, Any], *, enforce_plugin: bool) -> Dict[str, Any]:
        work_package = payload.get("workPackage")
        recipe = payload.get("recipe")
        step_payloads = payload.get("stepPayloads") or {}
        if not isinstance(work_package, dict):
            raise ValueError("workPackage is required.")
        if not isinstance(recipe, dict):
            raise ValueError("recipe is required.")

        drawings = _resolve_project_drawings(work_package)
        enabled_steps = _extract_enabled_steps(recipe)
        if not enabled_steps:
            raise ValueError("At least one enabled recipe step is required.")

        issues: List[Dict[str, Any]] = []
        blockers: List[str] = []
        warnings: List[str] = []

        issue_set_id = _normalize_text(work_package.get("issueSetId"))
        if not issue_set_id:
            blockers.append("Offline automation recipes must be anchored to a selected issue set.")
            issues.append(
                {
                    "id": "missing:issue-set",
                    "severity": "blocker",
                    "label": "Issue set required",
                    "detail": "Offline automation recipes must be anchored to a selected issue set.",
                    "drawingPath": None,
                }
            )

        plugin_status = _detect_plugin_status()
        if enforce_plugin and not plugin_status["ok"]:
            blockers.extend(plugin_status["errors"])
        elif not plugin_status["ok"]:
            warnings.extend(plugin_status["errors"])

        acade_support = _find_acade_support(work_package)
        warnings.extend(acade_support["warnings"])

        for drawing in drawings:
            drawing_path = drawing["path"]
            if not drawing.get("exists"):
                blockers.append(f"Selected drawing '{drawing_path}' does not exist.")
                issues.append(
                    {
                        "id": f"missing:{drawing_path}",
                        "severity": "blocker",
                        "label": "Drawing missing",
                        "detail": f"Selected drawing '{drawing_path}' does not exist.",
                        "drawingPath": drawing_path,
                    }
                )

        title_block_status = _normalize_text(work_package.get("titleBlockSnapshotStatus")).lower()
        title_block_warning_count = int(work_package.get("titleBlockWarningCount") or 0)
        if title_block_status in {"stale", "warning", "needs-review"} or title_block_warning_count > 0:
            warnings.append(
                "Project title block / metadata snapshot has warnings. Review the package metadata before final issue."
            )

        autowire_payload = step_payloads.get("autowire") if isinstance(step_payloads, dict) else None
        if any(_normalize_step_source(step.get("source")) == "autowire" for step in enabled_steps):
            if not isinstance(autowire_payload, dict):
                blockers.append("Enabled wiring recipe step is missing its step payload.")
            else:
                schedule_snapshot_id = _normalize_text(autowire_payload.get("scheduleSnapshotId"))
                strip_rows = autowire_payload.get("stripRows")
                connection_rows = autowire_payload.get("connectionRows")
                if not schedule_snapshot_id:
                    blockers.append("Enabled wiring recipe step requires a terminal schedule snapshot id.")
                if not isinstance(strip_rows, list) or len(strip_rows) == 0:
                    blockers.append("Enabled wiring recipe step requires TerminalStrips rows.")
                else:
                    blockers.extend(
                        _resolve_schedule_rows(
                            drawings,
                            [row for row in strip_rows if isinstance(row, dict)],
                            [
                                row
                                for row in (connection_rows if isinstance(connection_rows, list) else [])
                                if isinstance(row, dict)
                            ],
                        )
                    )

        cad_utils_payload = step_payloads.get("cadUtils") if isinstance(step_payloads, dict) else None
        if any(_normalize_step_source(step.get("source")) == "cad-utils" for step in enabled_steps):
            if not isinstance(cad_utils_payload, dict):
                blockers.append("Enabled CAD utilities recipe step is missing its step payload.")
            else:
                raw_rules = cad_utils_payload.get("rules")
                if not isinstance(raw_rules, list) or len(raw_rules) == 0:
                    blockers.append("Enabled CAD utilities recipe step requires at least one rule.")
                else:
                    valid_rules = [
                        row
                        for row in raw_rules
                        if isinstance(row, dict) and _normalize_text(row.get("find"))
                    ]
                    if len(valid_rules) == 0:
                        blockers.append("Enabled CAD utilities recipe step does not contain any valid find/replace rules.")
                    elif len(valid_rules) > MAX_RULES:
                        blockers.append(f"Too many CAD utility rules. Maximum is {MAX_RULES}.")

        autodraft_payload = step_payloads.get("autodraft") if isinstance(step_payloads, dict) else None
        if any(_normalize_step_source(step.get("source")) == "autodraft" for step in enabled_steps):
            if not isinstance(autodraft_payload, dict):
                blockers.append("Enabled Bluebeam markup recipe step is missing its step payload.")
            else:
                raw_snapshots = autodraft_payload.get("markupSnapshots")
                if not isinstance(raw_snapshots, list) or len(raw_snapshots) == 0:
                    blockers.append(
                        "Enabled Bluebeam markup recipe step requires at least one published markup snapshot."
                    )
                else:
                    _operations, autodraft_warnings, autodraft_blockers = _normalize_autodraft_snapshots(
                        autodraft_payload,
                        drawings,
                    )
                    warnings.extend(autodraft_warnings)
                    blockers.extend(autodraft_blockers)

        warnings = _normalize_string_array(warnings)
        blockers = _normalize_string_array(blockers)
        return {
            "drawings": drawings,
            "enabledSteps": enabled_steps,
            "pluginStatus": plugin_status,
            "acadeSupport": acade_support,
            "issues": issues,
            "warnings": warnings,
            "blockers": blockers,
            "simulateOnCopy": recipe.get("simulateOnCopy") is not False,
        }

    def _normalize_terminal_operation(
        operation: Dict[str, Any],
        *,
        selected_ids: set[str],
    ) -> Dict[str, Any]:
        operation_id = _normalize_text(operation.get("operationId")) or uuid.uuid4().hex
        operation_type = _normalize_text(operation.get("operationType")) or "label-upsert"
        _source = _normalize_text(operation.get("source")) or "strip"
        drawing_path = _normalize_nullable_text(operation.get("drawingPath"))
        managed_value = _normalize_nullable_text(
            operation.get("routeKey") or operation.get("stripKey") or operation_id
        )
        managed_kind = "route" if "route" in operation_type else "strip"
        warnings = []
        operation_warning = _normalize_nullable_text(operation.get("warning"))
        if operation_warning:
            warnings.append(operation_warning)
        return {
            "id": operation_id,
            "source": "autowire",
            "operationType": operation_type,
            "drawingPath": drawing_path,
            "drawingName": _normalize_nullable_text(operation.get("drawingName")),
            "relativePath": _normalize_nullable_text(operation.get("relativePath")),
            "managedKey": {
                "source": "autowire",
                "entityKind": managed_kind,
                "value": managed_value or operation_id,
                "drawingPath": drawing_path,
            },
            "before": _normalize_nullable_text(operation.get("before")),
            "after": _normalize_nullable_text(operation.get("after")),
            "detail": _normalize_text(operation.get("detail")),
            "warnings": warnings,
            "artifactRefs": [],
            "approved": operation_type != "unresolved" and operation_id in selected_ids,
            "nativePayload": operation,
        }

    def _normalize_cad_match(
        match: Dict[str, Any],
        drawings: List[Dict[str, str]],
        index: int,
    ) -> Dict[str, Any]:
        drawing_lookup = {entry["path"].lower(): entry for entry in drawings}
        drawing_path = _normalize_text(match.get("drawingPath"))
        drawing_meta = drawing_lookup.get(drawing_path.lower()) if drawing_path else None
        drawing_name = _normalize_text(match.get("drawingName")) or (drawing_meta or {}).get("drawingName", "")
        relative_path = _normalize_text(match.get("relativePath")) or (drawing_meta or {}).get("relativePath", "")
        group_key = _normalize_text(match.get("groupKey")) or drawing_path or drawing_name or f"drawing-{index}"
        match_key = _normalize_text(match.get("matchKey")) or "::".join(
            [
                group_key,
                _normalize_text(match.get("handle")),
                _normalize_text(match.get("attributeTag")),
                _normalize_text(match.get("ruleId")),
                _normalize_text(match.get("before")),
                _normalize_text(match.get("after")),
                str(index),
            ]
        )
        normalized = {
            **match,
            "drawingPath": drawing_path or None,
            "drawingName": drawing_name or _normalize_text(match.get("file")) or "Drawing",
            "relativePath": relative_path or None,
            "groupKey": group_key,
            "matchKey": match_key,
            "file": _normalize_text(match.get("file")) or drawing_name or "Drawing",
        }
        return normalized

    def _normalize_cad_operation(
        match: Dict[str, Any],
        *,
        selected_keys: set[str],
    ) -> Dict[str, Any]:
        match_key = _normalize_text(match.get("matchKey")) or uuid.uuid4().hex
        drawing_path = _normalize_nullable_text(match.get("drawingPath"))
        entity_type = _normalize_text(match.get("entityType")) or "drawing-content"
        return {
            "id": match_key,
            "source": "cad-utils",
            "operationType": "replace",
            "drawingPath": drawing_path,
            "drawingName": _normalize_nullable_text(match.get("drawingName")),
            "relativePath": _normalize_nullable_text(match.get("relativePath")),
            "managedKey": {
                "source": "cad-utils",
                "entityKind": entity_type,
                "value": match_key,
                "drawingPath": drawing_path,
            },
            "before": _normalize_nullable_text(match.get("before")),
            "after": _normalize_nullable_text(match.get("after")),
            "detail": f"{entity_type} replacement via {_normalize_text(match.get('ruleId'))}",
            "warnings": [],
            "artifactRefs": [],
            "approved": match_key in selected_keys,
            "nativePayload": match,
        }

    def _resolve_work_package_drawing(
        drawing_ref: str,
        drawings: List[Dict[str, str]],
    ) -> Tuple[Optional[Dict[str, str]], str | None]:
        normalized_ref = _normalize_text(drawing_ref)
        if not normalized_ref:
            return None, "Drawing binding is missing."
        lookup = _build_drawing_lookup(drawings)
        matches = lookup.get(_normalize_drawing_key(normalized_ref), [])
        if len(matches) == 1:
            return matches[0], None
        if len(matches) == 0:
            return None, f"Drawing binding '{drawing_ref}' did not match any selected issue-set drawing."
        return None, f"Drawing binding '{drawing_ref}' matched more than one selected issue-set drawing."

    def _extract_markup_snapshot_preview_operations(snapshot: Dict[str, Any]) -> List[Dict[str, Any]]:
        compare_payload = snapshot.get("comparePayload")
        if not isinstance(compare_payload, dict):
            return []
        raw_operations = compare_payload.get("preview_operations")
        if not isinstance(raw_operations, list):
            raw_operations = compare_payload.get("previewOperations")
        if not isinstance(raw_operations, list):
            return []
        return [entry for entry in raw_operations if isinstance(entry, dict)]

    def _normalize_autodraft_operation(
        operation: Dict[str, Any],
        *,
        snapshot: Dict[str, Any],
        drawing: Dict[str, str],
        selected_ids: set[str],
    ) -> Dict[str, Any]:
        operation_id = _normalize_text(operation.get("id")) or uuid.uuid4().hex
        operation_type = _normalize_text(operation.get("operationType")) or "preview"
        managed_key = operation.get("managedKey") if isinstance(operation.get("managedKey"), dict) else None
        native_payload = (
            copy.deepcopy(operation.get("nativePayload"))
            if isinstance(operation.get("nativePayload"), dict)
            else {}
        )
        target_handle_refs = _normalize_string_array(operation.get("targetHandleRefs"))
        native_payload["markupSnapshotId"] = _normalize_text(snapshot.get("id"))
        native_payload["drawingPath"] = drawing["path"]
        native_payload["drawingName"] = drawing["drawingName"]
        native_payload["relativePath"] = drawing["relativePath"]
        native_payload["operationId"] = operation_id
        native_payload["operationType"] = operation_type
        native_payload["before"] = _normalize_nullable_text(operation.get("before"))
        native_payload["after"] = _normalize_nullable_text(operation.get("after"))
        native_payload["detail"] = _normalize_text(operation.get("detail")) or operation_type
        native_payload["targetHandleRefs"] = target_handle_refs
        if isinstance(managed_key, dict):
            native_payload["managedKey"] = {
                "source": "autodraft",
                "entityKind": _normalize_text((managed_key or {}).get("entityKind")) or "markup",
                "value": _normalize_text((managed_key or {}).get("value")) or operation_id,
                "drawingPath": drawing["path"],
            }

        warnings = _normalize_string_array(operation.get("warnings"))
        approved = (
            operation.get("approved") is not False
            and operation_id in selected_ids
        )
        return {
            "id": operation_id,
            "source": "autodraft",
            "operationType": operation_type,
            "drawingPath": drawing["path"],
            "drawingName": drawing["drawingName"],
            "relativePath": drawing["relativePath"],
            "managedKey": {
                "source": "autodraft",
                "entityKind": _normalize_text((managed_key or {}).get("entityKind")) or "markup",
                "value": _normalize_text((managed_key or {}).get("value")) or operation_id,
                "drawingPath": drawing["path"],
            }
            if isinstance(managed_key, dict)
            else {
                "source": "autodraft",
                "entityKind": "markup",
                "value": operation_id,
                "drawingPath": drawing["path"],
            },
            "targetHandleRefs": target_handle_refs,
            "before": _normalize_nullable_text(operation.get("before")),
            "after": _normalize_nullable_text(operation.get("after")),
            "detail": _normalize_text(operation.get("detail")) or operation_type,
            "warnings": warnings,
            "artifactRefs": _normalize_string_array(operation.get("artifactRefs")),
            "approved": approved,
            "nativePayload": native_payload,
        }

    def _normalize_autodraft_snapshots(
        autodraft_payload: Dict[str, Any],
        drawings: List[Dict[str, str]],
    ) -> Tuple[List[Dict[str, Any]], List[str], List[str]]:
        raw_snapshots = autodraft_payload.get("markupSnapshots")
        if not isinstance(raw_snapshots, list):
            raw_snapshots = []
        selected_ids = set(_normalize_string_array(autodraft_payload.get("selectedOperationIds")))
        blockers: List[str] = []
        warnings: List[str] = []
        operations: List[Dict[str, Any]] = []

        for snapshot in raw_snapshots:
            if not isinstance(snapshot, dict):
                continue
            snapshot_id = _normalize_text(snapshot.get("id")) or "markup-snapshot"
            drawing_ref = _normalize_text(snapshot.get("drawingPath"))
            drawing, drawing_error = _resolve_work_package_drawing(drawing_ref, drawings)
            if drawing is None:
                blockers.append(
                    f"Markup snapshot '{snapshot_id}' could not be resolved: {drawing_error or 'drawing binding failed.'}"
                )
                continue
            preview_operations = _extract_markup_snapshot_preview_operations(snapshot)
            if not preview_operations:
                blockers.append(
                    f"Markup snapshot '{snapshot_id}' does not contain any preview operations."
                )
                continue
            warnings.extend(_normalize_string_array(snapshot.get("warnings")))
            snapshot_selected_ids = set(
                _normalize_string_array(snapshot.get("selectedOperationIds"))
            )
            effective_selected_ids = selected_ids or snapshot_selected_ids
            if not effective_selected_ids:
                blockers.append(
                    f"Markup snapshot '{snapshot_id}' does not contain any selected operation ids."
                )
                continue
            for preview_operation in preview_operations:
                operation = _normalize_autodraft_operation(
                    preview_operation,
                    snapshot=snapshot,
                    drawing=drawing,
                    selected_ids=effective_selected_ids,
                )
                operations.append(operation)
                warnings.extend(operation["warnings"])

        return operations, _normalize_string_array(warnings), _normalize_string_array(blockers)

    def _copy_drawings_to_workspace(drawings: List[Dict[str, str]]) -> Tuple[str, Dict[str, str]]:
        workspace_root = tempfile.mkdtemp(prefix="suite-automation-workspace-")
        path_map: Dict[str, str] = {}
        for drawing in drawings:
            source_path = _resolve_workspace_source_path(drawing)
            relative_path = drawing["relativePath"] or drawing["drawingName"]
            destination_path = _workspace_destination_path(
                workspace_root,
                relative_path,
                drawing.get("drawingName") or "",
            )
            os.makedirs(os.path.dirname(destination_path), exist_ok=True)
            shutil.copy2(source_path, destination_path)
            path_map[source_path.lower()] = destination_path
        return workspace_root, path_map

    def _rewrite_operation_for_workspace(
        operation: Dict[str, Any],
        path_map: Dict[str, str],
    ) -> Dict[str, Any]:
        rewritten = copy.deepcopy(operation)
        native_payload = rewritten.get("nativePayload")
        if isinstance(native_payload, dict):
            drawing_path = _normalize_text(native_payload.get("drawingPath"))
            if drawing_path and drawing_path.lower() in path_map:
                rewritten["drawingPath"] = path_map[drawing_path.lower()]
                native_payload["drawingPath"] = path_map[drawing_path.lower()]
        drawing_path = _normalize_text(rewritten.get("drawingPath"))
        if drawing_path and drawing_path.lower() in path_map:
            rewritten["drawingPath"] = path_map[drawing_path.lower()]
        managed_key = rewritten.get("managedKey")
        if isinstance(managed_key, dict):
            managed_path = _normalize_text(managed_key.get("drawingPath"))
            if managed_path and managed_path.lower() in path_map:
                managed_key["drawingPath"] = path_map[managed_path.lower()]
        return rewritten

    def _export_combined_report(
        *,
        work_package: Dict[str, Any],
        recipe: Dict[str, Any],
        operations: List[Dict[str, Any]],
        warnings: List[str],
        artifacts: List[Dict[str, Any]],
        changed_drawing_count: int,
        changed_item_count: int,
        workspace_root: str | None,
    ) -> Tuple[str, str]:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = tempfile.mkdtemp(prefix="suite_automation_recipe_")
        workbook_path = os.path.join(out_dir, f"suite_automation_recipe_{timestamp}.xlsx")
        manifest_path = os.path.join(out_dir, f"suite_automation_recipe_{timestamp}.json")

        workbook = Workbook()
        operations_ws = workbook.active
        operations_ws.title = "Operations"
        drawings_ws = workbook.create_sheet("Summary")
        warnings_ws = workbook.create_sheet("Warnings")
        artifacts_ws = workbook.create_sheet("Artifacts")

        operations_ws.append(
            [
                "Source",
                "Drawing",
                "Relative Path",
                "Operation",
                "Before",
                "After",
                "Approved",
                "Warnings",
                "Detail",
            ]
        )
        for operation in operations:
            operations_ws.append(
                [
                    _normalize_text(operation.get("source")),
                    _normalize_text(operation.get("drawingName") or operation.get("drawingPath")),
                    _normalize_text(operation.get("relativePath")),
                    _normalize_text(operation.get("operationType")),
                    _normalize_text(operation.get("before")),
                    _normalize_text(operation.get("after")),
                    "yes" if operation.get("approved") else "no",
                    " | ".join(_normalize_string_array(operation.get("warnings"))),
                    _normalize_text(operation.get("detail")),
                ]
            )

        drawings_ws.append(["Metric", "Value"])
        drawings_ws.append(["Project", _normalize_text(work_package.get("projectId"))])
        drawings_ws.append(["Issue Set", _normalize_text(work_package.get("issueSetId"))])
        drawings_ws.append(["Recipe", _normalize_text(recipe.get("name"))])
        drawings_ws.append(["Changed drawings", changed_drawing_count])
        drawings_ws.append(["Changed items", changed_item_count])
        drawings_ws.append(["Simulate on copy", "yes" if recipe.get("simulateOnCopy") is not False else "no"])
        drawings_ws.append(["Workspace", workspace_root or "source drawings"])

        warnings_ws.append(["Warning"])
        for warning in warnings or ["No warnings."]:
            warnings_ws.append([warning])

        artifacts_ws.append(["Kind", "Label", "Download URL", "Path", "Description"])
        for artifact in artifacts:
            artifacts_ws.append(
                [
                    _normalize_text(artifact.get("kind")),
                    _normalize_text(artifact.get("label")),
                    _normalize_text(artifact.get("downloadUrl")),
                    _normalize_text(artifact.get("path")),
                    _normalize_text(artifact.get("description")),
                ]
            )

        workbook.save(workbook_path)
        with open(manifest_path, "w", encoding="utf-8") as manifest_file:
            json.dump(
                {
                    "projectId": work_package.get("projectId"),
                    "issueSetId": work_package.get("issueSetId"),
                    "recipeId": recipe.get("id"),
                    "recipeName": recipe.get("name"),
                    "changedDrawingCount": changed_drawing_count,
                    "changedItemCount": changed_item_count,
                    "warnings": warnings,
                    "workspaceRoot": workspace_root,
                    "operations": operations,
                    "artifacts": artifacts,
                },
                manifest_file,
                indent=2,
            )
        return workbook_path, manifest_path

    @bp.route("/cad/preflight/project-scope", methods=["POST"])
    @require_user_or_api_key
    @limiter.limit("30 per hour")
    def api_cad_preflight_project_scope():
        request_id = _request_id()
        try:
            payload = request.get_json(silent=True) or {}
            if not isinstance(payload, dict):
                payload = {}
            preflight = _build_preflight(payload, enforce_plugin=False)
            work_package = payload.get("workPackage") or {}
            recipe = payload.get("recipe") or {}
            return jsonify(
                {
                    "success": True,
                    "requestId": request_id,
                    "workPackageId": work_package.get("id"),
                    "recipeSnapshotId": recipe.get("id"),
                    "ok": len(preflight["blockers"]) == 0,
                    "simulateOnCopy": preflight["simulateOnCopy"],
                    "drawingCount": len(work_package.get("selectedDrawingPaths") or []),
                    "resolvedDrawingCount": len(preflight["drawings"]),
                    "pluginReady": preflight["pluginStatus"]["ok"],
                    "acadeContextFound": bool(preflight["acadeSupport"]["projectFile"]),
                    "issues": preflight["issues"],
                    "warnings": preflight["warnings"],
                    "blockers": preflight["blockers"],
                    "message": "CAD preflight completed.",
                }
            )
        except ValueError:
            return _request_error_response(
                message="Invalid CAD preflight request.",
                request_id=request_id,
                status_code=400,
            )
        except Exception:
            logger.exception("CAD preflight failed")
            return _request_error_response(
                message="CAD preflight failed.",
                request_id=request_id,
                status_code=500,
            )

    @bp.route("/automation-recipes/preview", methods=["POST"])
    @require_user_or_api_key
    @limiter.limit("20 per hour")
    def api_automation_recipe_preview():
        request_id = _request_id()
        try:
            payload = request.get_json(silent=True) or {}
            if not isinstance(payload, dict):
                payload = {}
            preflight = _build_preflight(payload, enforce_plugin=False)
            work_package = payload.get("workPackage") or {}
            recipe = payload.get("recipe") or {}
            step_payloads = payload.get("stepPayloads") or {}
            drawings = preflight["drawings"]
            operations: List[Dict[str, Any]] = []
            warnings = list(preflight["warnings"])
            blockers = list(preflight["blockers"])
            step_summaries: List[Dict[str, Any]] = []

            for step in preflight["enabledSteps"]:
                source = _normalize_step_source(step.get("source"))
                if source == "autowire":
                    autowire_payload = step_payloads.get("autowire") if isinstance(step_payloads, dict) else None
                    if not isinstance(autowire_payload, dict):
                        step_summaries.append(
                            _build_step_summary(
                                source=source,
                                label="Wiring authoring",
                                enabled=True,
                                ready=False,
                                actionable=False,
                                planned_item_count=0,
                                approved_item_count=0,
                                warning_count=1,
                                binding_kinds=["terminal-wiring", "schedule-row"],
                                summary="Wiring step payload is missing.",
                            )
                        )
                        continue
                    host_result = _call_acade_host(
                        "suite_terminal_authoring_project_preview",
                        {
                            "projectId": _normalize_text(work_package.get("projectId")),
                            "issueSetId": _normalize_text(work_package.get("issueSetId")),
                            "scheduleSnapshotId": _normalize_text(autowire_payload.get("scheduleSnapshotId")),
                            "drawings": drawings,
                            "stripRows": [
                                row for row in autowire_payload.get("stripRows", []) if isinstance(row, dict)
                            ],
                            "connectionRows": [
                                row for row in autowire_payload.get("connectionRows", []) if isinstance(row, dict)
                            ],
                        },
                        request_id,
                    )
                    if not host_result.get("success", False):
                        raise RuntimeError(
                            str(host_result.get("message") or "Wiring preview failed.")
                        )
                    selected_ids = set(_normalize_string_array(autowire_payload.get("selectedOperationIds")))
                    raw_operations = (host_result.get("data") or {}).get("operations") or []
                    normalized_operations = [
                        _normalize_terminal_operation(entry, selected_ids=selected_ids)
                        for entry in raw_operations
                        if isinstance(entry, dict)
                    ]
                    operations.extend(normalized_operations)
                    warnings.extend(_normalize_string_array(host_result.get("warnings")))
                    step_summaries.append(
                        _build_step_summary(
                            source=source,
                            label="Wiring authoring",
                            enabled=True,
                            ready=True,
                            actionable=True,
                            planned_item_count=len(normalized_operations),
                            approved_item_count=sum(1 for entry in normalized_operations if entry["approved"]),
                            warning_count=len(_normalize_string_array(host_result.get("warnings"))),
                            binding_kinds=["terminal-wiring", "schedule-row"],
                            summary=host_result.get("message") or "Wiring preview ready.",
                            request_id=(host_result.get("meta") or {}).get("requestId"),
                        )
                    )
                elif source == "cad-utils":
                    cad_payload = step_payloads.get("cadUtils") if isinstance(step_payloads, dict) else None
                    if not isinstance(cad_payload, dict):
                        step_summaries.append(
                            _build_step_summary(
                                source=source,
                                label="CAD utilities",
                                enabled=True,
                                ready=False,
                                actionable=False,
                                planned_item_count=0,
                                approved_item_count=0,
                                warning_count=1,
                                binding_kinds=["drawing-content"],
                                summary="CAD utilities step payload is missing.",
                            )
                        )
                        continue
                    host_result = _call_acade_host(
                        "suite_batch_find_replace_project_preview",
                        {
                            "rules": [
                                row for row in cad_payload.get("rules", []) if isinstance(row, dict)
                            ],
                            "drawings": drawings,
                            "blockNameHint": _normalize_text(cad_payload.get("blockNameHint")),
                        },
                        request_id,
                    )
                    if not host_result.get("success", False):
                        raise RuntimeError(
                            str(host_result.get("message") or "CAD utilities preview failed.")
                        )
                    selected_keys = set(_normalize_string_array(cad_payload.get("selectedPreviewKeys")))
                    raw_matches = (host_result.get("data") or {}).get("matches") or []
                    normalized_matches = [
                        _normalize_cad_match(entry, drawings, index)
                        for index, entry in enumerate(raw_matches)
                        if isinstance(entry, dict)
                    ]
                    normalized_operations = [
                        _normalize_cad_operation(entry, selected_keys=selected_keys)
                        for entry in normalized_matches
                    ]
                    operations.extend(normalized_operations)
                    warnings.extend(_normalize_string_array(host_result.get("warnings")))
                    step_summaries.append(
                        _build_step_summary(
                            source=source,
                            label="CAD utilities",
                            enabled=True,
                            ready=True,
                            actionable=True,
                            planned_item_count=len(normalized_operations),
                            approved_item_count=sum(1 for entry in normalized_operations if entry["approved"]),
                            warning_count=len(_normalize_string_array(host_result.get("warnings"))),
                            binding_kinds=["drawing-content"],
                            summary=host_result.get("message") or "CAD utilities preview ready.",
                            request_id=(host_result.get("meta") or {}).get("requestId"),
                        )
                    )
                else:
                    autodraft_payload = step_payloads.get("autodraft") if isinstance(step_payloads, dict) else None
                    _queue_items = autodraft_payload.get("queueItems") if isinstance(autodraft_payload, dict) else []
                    if not isinstance(autodraft_payload, dict):
                        step_summaries.append(
                            _build_step_summary(
                                source=source,
                                label="Bluebeam markup authoring",
                                enabled=True,
                                ready=False,
                                actionable=False,
                                planned_item_count=0,
                                approved_item_count=0,
                                warning_count=1,
                                binding_kinds=["title-block", "drawing-row", "deliverable-row", "note-only"],
                                summary="Bluebeam markup step payload is missing.",
                            )
                        )
                        continue
                    autodraft_operations, autodraft_warnings, autodraft_blockers = _normalize_autodraft_snapshots(
                        autodraft_payload,
                        drawings,
                    )
                    operations.extend(autodraft_operations)
                    warnings.extend(autodraft_warnings)
                    blockers.extend(autodraft_blockers)
                    planned_item_count = len(autodraft_operations)
                    approved_item_count = sum(
                        1 for entry in autodraft_operations if bool(entry.get("approved"))
                    )
                    step_summaries.append(
                        _build_step_summary(
                            source=source,
                            label="Bluebeam markup authoring",
                            enabled=True,
                            ready=planned_item_count > 0,
                            actionable=approved_item_count > 0,
                            planned_item_count=planned_item_count,
                            approved_item_count=approved_item_count,
                            warning_count=len(autodraft_warnings) + len(autodraft_blockers),
                            binding_kinds=["title-block", "drawing-row", "deliverable-row", "note-only"],
                            summary=(
                                f"Bluebeam markup snapshot preview resolved {planned_item_count} operation(s) "
                                f"with {approved_item_count} approved for apply."
                            ),
                            request_id=_normalize_nullable_text((autodraft_payload or {}).get("requestId")),
                        )
                    )

            return jsonify(
                {
                    "success": True,
                    "requestId": request_id,
                    "workPackageId": work_package.get("id"),
                    "recipeSnapshotId": recipe.get("id"),
                    "steps": step_summaries,
                    "operations": operations,
                    "warnings": _normalize_string_array(warnings),
                    "blockers": _normalize_string_array(blockers),
                    "message": f"Recipe preview built {len(operations)} CAD operation(s) across {len(drawings)} scoped drawing(s).",
                }
            )
        except ValueError:
            return _request_error_response(
                message="Invalid automation recipe preview request.",
                request_id=request_id,
                status_code=400,
            )
        except Exception:
            logger.exception("Automation recipe preview failed")
            return _request_error_response(
                message="Automation recipe preview failed.",
                request_id=request_id,
                status_code=500,
            )

    @bp.route("/automation-recipes/apply", methods=["POST"])
    @require_user_or_api_key
    @limiter.limit("15 per hour")
    def api_automation_recipe_apply():
        request_id = _request_id()
        try:
            payload = request.get_json(silent=True) or {}
            if not isinstance(payload, dict):
                payload = {}
            preflight = _build_preflight(payload, enforce_plugin=True)
            if preflight["blockers"]:
                return (
                    jsonify(
                        {
                            "success": False,
                            "error": "Recipe apply is blocked by preflight findings.",
                            "requestId": request_id,
                            "blockers": preflight["blockers"],
                        }
                    ),
                    400,
                )

            work_package = payload.get("workPackage") or {}
            recipe = payload.get("recipe") or {}
            step_payloads = payload.get("stepPayloads") or {}
            raw_operations = payload.get("operations")
            if not isinstance(raw_operations, list) or len(raw_operations) == 0:
                raise ValueError("operations must contain at least one approved preview row.")
            operations = [
                entry for entry in raw_operations if isinstance(entry, dict) and entry.get("approved") is not False
            ]
            if not operations:
                raise ValueError("No approved recipe operations were supplied for apply.")
            if len(operations) > MAX_OPERATIONS:
                raise ValueError(f"Too many recipe operations. Maximum is {MAX_OPERATIONS}.")

            drawings = preflight["drawings"]
            simulate_on_copy = preflight["simulateOnCopy"]
            workspace_root = None
            path_map: Dict[str, str] = {}
            if simulate_on_copy:
                workspace_root, path_map = _copy_drawings_to_workspace(drawings)
            else:
                path_map = {drawing["path"].lower(): drawing["path"] for drawing in drawings}

            rewritten_operations = [
                _rewrite_operation_for_workspace(entry, path_map)
                for entry in operations
            ]

            warnings = list(preflight["warnings"])
            drawing_results: List[Dict[str, Any]] = []
            combined_operations: List[Dict[str, Any]] = []
            changed_drawing_keys: set[str] = set()
            changed_item_count = 0

            autodraft_ops = [
                entry.get("nativePayload")
                for entry in rewritten_operations
                if _normalize_step_source(entry.get("source")) == "autodraft"
                and isinstance(entry.get("nativePayload"), dict)
            ]
            if autodraft_ops:
                bridge_result = _call_acade_host(
                    "suite_markup_authoring_project_apply",
                    {
                        "projectId": _normalize_text(work_package.get("projectId")),
                        "issueSetId": _normalize_text(work_package.get("issueSetId")),
                        "projectRootPath": _normalize_text(work_package.get("projectRootPath")),
                        "operations": autodraft_ops,
                    },
                    request_id,
                )
                if not bridge_result.get("success", False):
                    raise RuntimeError(str(bridge_result.get("message") or "Bluebeam markup apply failed."))
                data = bridge_result.get("data") or {}
                warnings.extend(_normalize_string_array(bridge_result.get("warnings")))
                changed_item_count += int(data.get("changedItemCount") or 0)
                for drawing in data.get("drawings") or []:
                    if isinstance(drawing, dict):
                        drawing_results.append({"source": "autodraft", **drawing})
                        drawing_path = _normalize_text(drawing.get("drawingPath"))
                        if drawing_path:
                            changed_drawing_keys.add(drawing_path.lower())
                for change in data.get("changes") or []:
                    if isinstance(change, dict):
                        combined_operations.append({"source": "autodraft", **change})

            autowire_ops = [
                entry.get("nativePayload")
                for entry in rewritten_operations
                if _normalize_step_source(entry.get("source")) == "autowire" and isinstance(entry.get("nativePayload"), dict)
            ]
            if autowire_ops:
                autowire_payload = step_payloads.get("autowire") if isinstance(step_payloads, dict) else {}
                bridge_result = _call_acade_host(
                    "suite_terminal_authoring_project_apply",
                    {
                        "projectId": _normalize_text(work_package.get("projectId")),
                        "issueSetId": _normalize_text(work_package.get("issueSetId")),
                        "scheduleSnapshotId": _normalize_text((autowire_payload or {}).get("scheduleSnapshotId")),
                        "projectRootPath": _normalize_text(work_package.get("projectRootPath")),
                        "operations": autowire_ops,
                    },
                    request_id,
                )
                if not bridge_result.get("success", False):
                    raise RuntimeError(str(bridge_result.get("message") or "Wiring apply failed."))
                data = bridge_result.get("data") or {}
                warnings.extend(_normalize_string_array(bridge_result.get("warnings")))
                changed_item_count += int(data.get("terminalStripUpdateCount") or 0) + int(data.get("managedRouteUpsertCount") or 0)
                for drawing in data.get("drawings") or []:
                    if isinstance(drawing, dict):
                        drawing_results.append({"source": "autowire", **drawing})
                        drawing_path = _normalize_text(drawing.get("drawingPath"))
                        if drawing_path:
                            changed_drawing_keys.add(drawing_path.lower())
                for change in data.get("changes") or []:
                    if isinstance(change, dict):
                        combined_operations.append({"source": "autowire", **change})

            cad_utils_matches = [
                entry.get("nativePayload")
                for entry in rewritten_operations
                if _normalize_step_source(entry.get("source")) == "cad-utils" and isinstance(entry.get("nativePayload"), dict)
            ]
            if cad_utils_matches:
                cad_payload = step_payloads.get("cadUtils") if isinstance(step_payloads, dict) else {}
                bridge_result = _call_acade_host(
                    "suite_batch_find_replace_project_apply",
                    {
                        "matches": cad_utils_matches,
                        "blockNameHint": _normalize_text((cad_payload or {}).get("blockNameHint")),
                    },
                    request_id,
                )
                if not bridge_result.get("success", False):
                    raise RuntimeError(str(bridge_result.get("message") or "CAD utilities apply failed."))
                data = bridge_result.get("data") or {}
                warnings.extend(_normalize_string_array(bridge_result.get("warnings")))
                changed_item_count += int(data.get("changedItemCount") or data.get("updated") or 0)
                for drawing in data.get("drawings") or []:
                    if isinstance(drawing, dict):
                        drawing_results.append({"source": "cad-utils", **drawing})
                        drawing_path = _normalize_text(drawing.get("drawingPath"))
                        if drawing_path:
                            changed_drawing_keys.add(drawing_path.lower())
                for change in data.get("changes") or []:
                    if isinstance(change, dict):
                        combined_operations.append({"source": "cad-utils", **change})

            run_id = uuid.uuid4().hex
            artifacts: List[Dict[str, Any]] = []
            workbook_path, manifest_path = _export_combined_report(
                work_package=work_package,
                recipe=recipe,
                operations=rewritten_operations,
                warnings=_normalize_string_array(warnings),
                artifacts=[],
                changed_drawing_count=len(changed_drawing_keys),
                changed_item_count=changed_item_count,
                workspace_root=workspace_root,
            )
            excel_report_id, excel_report_entry = _register_generated_report(
                workbook_path,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            manifest_report_id, manifest_report_entry = _register_generated_report(
                manifest_path,
                mimetype="application/json",
            )
            artifacts.extend(
                [
                    _normalize_report_artifact(
                        artifact_id=excel_report_id,
                        label="Recipe audit workbook",
                        kind="excel-report",
                        download_url=f"/api/cad/reports/{excel_report_id}",
                        path=excel_report_entry["path"],
                        description="Combined offline automation workbook.",
                    ),
                    _normalize_report_artifact(
                        artifact_id=manifest_report_id,
                        label="Recipe manifest",
                        kind="json-manifest",
                        download_url=f"/api/cad/reports/{manifest_report_id}",
                        path=manifest_report_entry["path"],
                        description="Combined offline automation manifest.",
                    ),
                ]
            )
            if workspace_root:
                artifacts.append(
                    _normalize_report_artifact(
                        artifact_id=f"workspace:{run_id}",
                        label="Simulation workspace",
                        kind="workspace",
                        download_url=None,
                        path=workspace_root,
                        description="Copied DWG workspace used for simulate-on-copy apply.",
                    )
                )

            generated_runs[run_id] = {
                "id": run_id,
                "requestId": request_id,
                "projectId": _normalize_text(work_package.get("projectId")),
                "issueSetId": _normalize_nullable_text(work_package.get("issueSetId")),
                "workPackageId": _normalize_nullable_text(work_package.get("id")),
                "recipeId": _normalize_nullable_text(recipe.get("id")),
                "simulateOnCopy": simulate_on_copy,
                "warnings": _normalize_string_array(warnings),
                "artifacts": artifacts,
                "workspaceRoot": workspace_root,
                "changedDrawingPaths": list(changed_drawing_keys),
                "operations": rewritten_operations,
            }
            _persist_state_index("runs.json", generated_runs)

            return jsonify(
                {
                    "success": True,
                    "requestId": request_id,
                    "runId": run_id,
                    "changedDrawingCount": len(changed_drawing_keys),
                    "changedItemCount": changed_item_count,
                    "reportId": excel_report_id,
                    "reportFilename": excel_report_entry["filename"],
                    "downloadUrl": f"/api/cad/reports/{excel_report_id}",
                    "warnings": _normalize_string_array(warnings),
                    "artifacts": artifacts,
                    "operations": rewritten_operations,
                    "message": f"Recipe apply completed across {len(changed_drawing_keys)} drawing(s) with {changed_item_count} changed item(s).",
                }
            )
        except ValueError:
            return _request_error_response(
                message="Invalid automation recipe apply request.",
                request_id=request_id,
                status_code=400,
            )
        except Exception:
            logger.exception("Automation recipe apply failed")
            return _request_error_response(
                message="Automation recipe apply failed.",
                request_id=request_id,
                status_code=500,
            )

    @bp.route("/automation-recipes/verify", methods=["POST"])
    @require_user_or_api_key
    @limiter.limit("20 per hour")
    def api_automation_recipe_verify():
        request_id = _request_id()
        try:
            payload = request.get_json(silent=True) or {}
            if not isinstance(payload, dict):
                payload = {}
            run_id = _normalize_text(payload.get("runId"))
            if not run_id:
                raise ValueError("runId is required.")
            run_record = generated_runs.get(run_id)
            if not run_record:
                return jsonify({"success": False, "error": "Run not found.", "requestId": request_id}), 404

            warnings = _normalize_string_array(run_record.get("warnings"))
            verified = True
            for drawing_path in run_record.get("changedDrawingPaths") or []:
                if not os.path.exists(drawing_path):
                    warnings.append(f"Changed drawing '{drawing_path}' is no longer available for verification.")
                    verified = False
            for artifact in run_record.get("artifacts") or []:
                artifact_path = _normalize_text((artifact or {}).get("path"))
                artifact_kind = _normalize_text((artifact or {}).get("kind"))
                if artifact_kind == "workspace":
                    continue
                if artifact_path and not os.path.exists(artifact_path):
                    warnings.append(f"Artifact '{artifact_path}' is no longer available.")
                    verified = False

            verification_dir = tempfile.mkdtemp(prefix="suite_automation_verify_")
            verification_filename = secure_filename(f"automation_verify_{run_id}.json") or (
                f"automation_verify_{uuid.uuid4().hex}.json"
            )
            verification_path = safe_join(verification_dir, verification_filename)
            if verification_path is None:
                raise ValueError("Verification artifact path could not be created.")
            with open(verification_path, "w", encoding="utf-8") as verification_file:
                json.dump(
                    {
                        "runId": run_id,
                        "verified": verified,
                        "warnings": warnings,
                        "changedDrawingPaths": run_record.get("changedDrawingPaths") or [],
                    },
                    verification_file,
                    indent=2,
                )
            verification_id, verification_entry = _register_generated_report(
                verification_path,
                mimetype="application/json",
            )
            verification_artifact = _normalize_report_artifact(
                artifact_id=verification_id,
                label="Verification manifest",
                kind="verification",
                download_url=f"/api/cad/reports/{verification_id}",
                path=verification_entry["path"],
                description="Offline verification result for the recipe run.",
            )
            run_record["artifacts"] = [*(run_record.get("artifacts") or []), verification_artifact]
            run_record["warnings"] = _normalize_string_array(warnings)
            _persist_state_index("runs.json", generated_runs)

            return jsonify(
                {
                    "success": True,
                    "requestId": request_id,
                    "runId": run_id,
                    "verified": verified,
                    "warnings": _normalize_string_array(warnings),
                    "artifacts": run_record["artifacts"],
                    "message": "Recipe verification completed." if verified else "Recipe verification completed with warnings.",
                }
            )
        except ValueError:
            return _request_error_response(
                message="Invalid automation recipe verify request.",
                request_id=request_id,
                status_code=400,
            )
        except Exception:
            logger.exception("Automation recipe verify failed")
            return _request_error_response(
                message="Automation recipe verify failed.",
                request_id=request_id,
                status_code=500,
            )

    @bp.route("/acade/reconcile/project-scope", methods=["POST"])
    @require_user_or_api_key
    @limiter.limit("20 per hour")
    def api_acade_reconcile_project_scope():
        request_id = _request_id()
        try:
            payload = request.get_json(silent=True) or {}
            if not isinstance(payload, dict):
                payload = {}
            work_package = payload.get("workPackage")
            if not isinstance(work_package, dict):
                raise ValueError("workPackage is required.")
            drawings = _resolve_project_drawings(work_package)
            preflight = _build_preflight(payload, enforce_plugin=False)
            acade_support = preflight["acadeSupport"]
            blockers = list(preflight["blockers"])
            warnings = list(preflight["warnings"])

            project_file = _normalize_text(acade_support.get("projectFile"))
            project_root = _resolve_existing_directory(work_package.get("projectRootPath"))
            if project_file and project_root is not None:
                try:
                    resolved_project_file = _resolve_project_scoped_file(
                        project_root,
                        project_file,
                        field_name="acadeProjectFilePath",
                        allowed_suffixes=(".wdp",),
                    )
                    with open(
                        str(resolved_project_file),
                        "r",
                        encoding="utf-8",
                        errors="ignore",
                    ) as handle:
                        content = handle.read().upper()
                    missing_drawings = [
                        drawing["relativePath"]
                        for drawing in drawings
                        if _normalize_drawing_key(drawing.get("drawingNumber")) not in content
                    ]
                    if missing_drawings:
                        warnings.append(
                            f"ACADE project file does not reference {len(missing_drawings)} selected drawing(s): {', '.join(missing_drawings[:8])}"
                        )
                except Exception:
                    warnings.append("ACADE project file was found, but Suite could not inspect its contents.")

            return jsonify(
                {
                    "success": True,
                    "requestId": request_id,
                    "drawingCount": len(drawings),
                    "acadeProjectFilePath": acade_support.get("projectFile"),
                    "acadeSupportFiles": acade_support.get("supportFiles") or [],
                    "blockers": _normalize_string_array(blockers),
                    "warnings": _normalize_string_array(warnings),
                    "message": "ACADE reconcile completed.",
                }
            )
        except ValueError:
            return _request_error_response(
                message="Invalid ACADE reconcile request.",
                request_id=request_id,
                status_code=400,
            )
        except Exception:
            logger.exception("ACADE reconcile failed")
            return _request_error_response(
                message="ACADE reconcile failed.",
                request_id=request_id,
                status_code=500,
            )

    @bp.route("/cad/reports/<report_id>", methods=["GET"])
    @require_user_or_api_key
    @limiter.limit("40 per hour")
    def api_cad_report_download(report_id: str):
        report = generated_reports.get(report_id)
        if not report:
            return jsonify({"success": False, "error": "Report not found."}), 404
        report_path = report.get("path") or ""
        if not report_path or not os.path.exists(report_path):
            generated_reports.pop(report_id, None)
            return jsonify({"success": False, "error": "Report is no longer available."}), 404

        return send_file(
            report_path,
            as_attachment=True,
            download_name=report.get("filename") or os.path.basename(report_path),
            mimetype=report.get("mimetype") or "application/octet-stream",
        )

    return bp
