# file: batch_find_replace_gui.py
import sys
import os
import getpass
import re
import time
import pythoncom
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from PIL import ImageTk, Image
import win32com.client
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import logging
import threading
import queue
import smtplib
from email.utils import formataddr
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage

def resource_path(rel_path: str) -> str: # Resource Loader for the Logo.
    """
    Return an absolute path to a resource, works for dev and for PyInstaller.
    When frozen by PyInstaller, files are unpacked into a temp folder
    exposed as sys._MEIPASS.
    """
    base_path = getattr(sys, "_MEIPASS", os.path.abspath("."))
    return os.path.join(base_path, rel_path)

# Email configuration (replace with your credentials)
SENDER_EMAIL = "Hyphaeos@gmail.com"   # Replace with your Gmail address
APP_PASSWORD = "ibhb mbcg wluz byri"      # Replace with your 16-character Gmail App Password
RECEIVER_EMAIL = "dustin.ward@root3power.com"

# Constants for configuration
COLOR_BG = '#121212'
COLOR_FG = '#1E90FF'
COLOR_ENTRY_BG = '#2C2C2C'
COLOR_ENTRY_FG = '#FFFFFF'
COLOR_BUTTON = '#000080'
FONT_HEADER = ("Arial", 14, "bold")
FONT_LABEL = ("Arial", 10, "bold")
FONT_ENTRY = ("Arial", 11)
MAX_REPLACEMENTS = 10
ENTRY_WIDTH = 40
PROGRESS_LENGTH = 850
LOGO_FILE = "root3_logo.png"
VERSION = "1.2"   # bumped

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class AutoCADReplacerGUI:
    """GUI application for batch find and replace in AutoCAD DWG files."""

    def __init__(self, root):
        """Initialize the GUI components."""
        self.root = root
        self.root.title("Batch Find & Replace")
        WIDTH, HEIGHT = 1440, 840
        self.root.geometry(f"{WIDTH}x{HEIGHT}")  # default client size
        self.root.minsize(1200, 720) 
            # Optional: center the window on the current display (why: avoids top-left spawn)
        try:
            self.root.update_idletasks()
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
            x = (screen_w - WIDTH) // 2
            y = (screen_h - HEIGHT) // 2
            self.root.geometry(f"{WIDTH}x{HEIGHT}+{x}+{y}")
        except Exception:
            pass  # safe fallback if window manager info isn't available yet
        self.root.configure(bg=COLOR_BG)
        self.root.resizable(True, True)

        # Load logo and set as window icon
        self.logo = None
        try:
            self.logo = ImageTk.PhotoImage(Image.open(resource_path(LOGO_FILE)).resize((80, 80)))
            self.root.iconphoto(True, self.logo)
        except Exception as e:
            logger.error(f"Could not load {LOGO_FILE}: {e}")
            messagebox.showwarning("Logo Load Error", f"Could not load logo: {e}. Continuing without logo.")

        # Header
        header_frame = tk.Frame(root, bg=COLOR_BG)
        header_frame.grid(row=0, column=0, pady=(5, 5), padx=10, sticky='nsew', columnspan=2)
        root.grid_rowconfigure(0, weight=0)
        root.grid_columnconfigure(0, weight=1)

        if self.logo:
            tk.Label(header_frame, image=self.logo, bg=COLOR_BG).pack(side=tk.LEFT, padx=(0, 10))

        header_text_frame = tk.Frame(header_frame, bg=COLOR_BG)
        header_text_frame.pack(side=tk.LEFT, expand=True)
        tk.Label(header_text_frame, text="Root3Power LLC", fg=COLOR_FG, bg=COLOR_BG, font=FONT_HEADER).pack(pady=(0, 5))
        tk.Label(header_text_frame, text="Batch Find & Replace", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack()

        # Layout weights
        root.grid_rowconfigure(1, weight=0)
        root.grid_rowconfigure(2, weight=0)
        root.grid_rowconfigure(3, weight=0)
        root.grid_rowconfigure(4, weight=0)
        root.grid_rowconfigure(5, weight=0)
        root.grid_rowconfigure(6, weight=1)  # log grows
        root.grid_rowconfigure(7, weight=0)
        root.grid_columnconfigure(0, weight=1)

        # Project Selection
        project_frame = tk.Frame(root, bg=COLOR_BG, padx=10, pady=5)
        project_frame.grid(row=1, column=0, sticky='nsew')
        tk.Label(project_frame, text="Project Selection", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(anchor='w')
        self.mode_var = tk.StringVar(value="directory")
        mode_frame = tk.Frame(project_frame, bg=COLOR_BG)
        mode_frame.pack(pady=5, fill='x')

        dir_radio = tk.Radiobutton(mode_frame, text="Entire Project Directory (Recursive)", variable=self.mode_var, value="directory",
                                   command=self.update_mode, takefocus=False, bg=COLOR_BG, fg=COLOR_ENTRY_FG,
                                   selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                                   indicatoron=False, font=FONT_LABEL)
        dir_radio.pack(side=tk.LEFT, padx=5)
        self.create_tooltip(dir_radio, "Process all DWG files in the selected folder. Use 'Include Subdirectories' to recurse.")

        file_radio = tk.Radiobutton(mode_frame, text="Select Specific Files", variable=self.mode_var, value="files",
                                    command=self.update_mode, takefocus=False, bg=COLOR_BG, fg=COLOR_ENTRY_FG,
                                    selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                                    indicatoron=False, font=FONT_LABEL)
        file_radio.pack(side=tk.LEFT, padx=5)
        self.create_tooltip(file_radio, "Pick individual DWG files (multi-select).")

        # Path + Browse
        path_frame = tk.Frame(project_frame, bg=COLOR_BG)
        path_frame.pack(pady=5, fill='x')
        self.path_entry = tk.Entry(path_frame, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white',
                                   font=FONT_ENTRY)
        self.path_entry.pack(side=tk.LEFT, padx=5, fill='x', expand=True)
        self.create_tooltip(self.path_entry, "Folder path or semicolon-separated DWG files.")

        browse_btn = tk.Button(path_frame, text="Browse", command=self.browse_path, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                               font=FONT_ENTRY, padx=10, pady=5, takefocus=False, relief="flat")
        browse_btn.pack(side=tk.LEFT, padx=5)
        self.create_tooltip(browse_btn, "Browse to select a directory or DWG files.")

        # Rules
        rules_frame = tk.Frame(root, bg=COLOR_BG, padx=10, pady=5)
        rules_frame.grid(row=2, column=0, sticky='nsew')
        tk.Label(rules_frame, text="Text Replacement Rules", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(anchor='w')
        labels_frame = tk.Frame(rules_frame, bg=COLOR_BG); labels_frame.pack(fill='x')
        tk.Label(labels_frame, text="Find", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).pack(side=tk.LEFT, padx=(0, 135))
        tk.Label(labels_frame, text="Replace With", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL).pack(side=tk.LEFT, padx=(110, 40))
        self.replacements_frame = tk.Frame(rules_frame, bg=COLOR_BG); self.replacements_frame.pack(fill='x')
        self.replacements = []; self.add_replacement_pair()
        add_rule_btn = tk.Button(rules_frame, text="Add Rule", command=self.add_replacement_pair, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                                 font=FONT_ENTRY, padx=10, pady=5, takefocus=False, relief="flat")
        add_rule_btn.pack(pady=5)
        self.create_tooltip(add_rule_btn, "Add another find/replace rule (max 10).")

        # Options
        options_frame = tk.Frame(root, bg=COLOR_BG, padx=10, pady=5)
        options_frame.grid(row=3, column=0, sticky='nsew')
        tk.Label(options_frame, text="Options", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(anchor='w')
        self.preview_var = tk.BooleanVar(value=False)
        preview_check = tk.Checkbutton(options_frame, text="Preview Mode (No Changes Saved)", variable=self.preview_var,
                                       fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON,
                                       activeforeground=COLOR_ENTRY_FG, font=FONT_ENTRY, takefocus=False, indicatoron=False)
        preview_check.pack(side=tk.LEFT, padx=5)
        self.create_tooltip(preview_check, "Dry run: show what would change without saving.")

        self.recursive_var = tk.BooleanVar(value=False)
        recursive_check = tk.Checkbutton(options_frame, text="Include Subdirectories", variable=self.recursive_var,
                                         fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON,
                                         activeforeground=COLOR_ENTRY_FG, font=FONT_ENTRY, takefocus=False, indicatoron=False)
        recursive_check.pack(side=tk.LEFT, padx=5)
        self.create_tooltip(recursive_check, "Search subfolders when a directory is selected.")

        # --- Title Block UI ---
        tb_frame = tk.Frame(root, bg=COLOR_BG, padx=10, pady=6)
        tb_frame.grid(row=4, column=0, sticky='nsew')

        tk.Label(tb_frame, text="Title Block", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=0, column=0, columnspan=16, sticky='w')

        self.tb_enable = tk.BooleanVar(value=False)
        tb_enable_chk = tk.Checkbutton(tb_frame, text="Enable", variable=self.tb_enable,
                                       fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON,
                                       activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                                       font=FONT_ENTRY, takefocus=False, indicatoron=False)
        tb_enable_chk.grid(row=0, column=16, sticky='e', padx=6)
        self.create_tooltip(tb_enable_chk, "Apply Title Block values to matching block attributes.")

        # Rev section selector (now Rev1..Rev5)
        tk.Label(tb_frame, text="Rev Section", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=1, column=0, sticky='e')
        self.tb_rev_choice = tk.StringVar(value="Rev1")
        rev_combo = ttk.Combobox(tb_frame, textvariable=self.tb_rev_choice, state="readonly", width=8,
                                 values=["Rev1","Rev2","Rev3","Rev4","Rev5"])
        rev_combo.grid(row=1, column=1, padx=4, sticky='w')
        self.create_tooltip(rev_combo, "Choose which revision slot (1–5) to write.")

        tk.Label(tb_frame, text="REV", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=1, column=2, sticky='e')
        self.tb_rev_val = tk.StringVar()
        rev_entry = tk.Entry(tb_frame, textvariable=self.tb_rev_val, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                             insertbackground='white', width=8)
        rev_entry.grid(row=1, column=3, padx=4, sticky='w')
        self.create_tooltip(rev_entry, "Revision code/letter for the selected slot (e.g., A, B, C).")

        tk.Label(tb_frame, text="DESC", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=1, column=4, sticky='e')
        self.tb_desc_val = tk.StringVar()
        desc_entry = tk.Entry(tb_frame, textvariable=self.tb_desc_val, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                              insertbackground='white', width=36)
        desc_entry.grid(row=1, column=5, padx=4, sticky='w')
        self.create_tooltip(desc_entry, "Revision description for the selected slot.")

        tk.Label(tb_frame, text="BY", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=1, column=6, sticky='e')
        self.tb_by_val = tk.StringVar()
        by_entry = tk.Entry(tb_frame, textvariable=self.tb_by_val, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                            insertbackground='white', width=10)
        by_entry.grid(row=1, column=7, padx=4, sticky='w')
        self.create_tooltip(by_entry, "Initials/name of drafter for the revision slot.")

        tk.Label(tb_frame, text="CHK", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=1, column=8, sticky='e')
        self.tb_chk_val = tk.StringVar()
        chk_entry = tk.Entry(tb_frame, textvariable=self.tb_chk_val, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                             insertbackground='white', width=10)
        chk_entry.grid(row=1, column=9, padx=4, sticky='w')
        self.create_tooltip(chk_entry, "Checker initials/name. Writes CHKn and CHKBYn.")

        tk.Label(tb_frame, text="DATE", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=1, column=10, sticky='e')
        self.tb_date_val = tk.StringVar()  # e.g., 09/26/25
        date_entry = tk.Entry(tb_frame, textvariable=self.tb_date_val, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                              insertbackground='white', width=12)
        date_entry.grid(row=1, column=11, padx=4, sticky='w')
        self.create_tooltip(date_entry, "Revision date for the selected slot (e.g., 09/26/25).")

        # Base title block fields (no number)
        row2 = 2
        tk.Label(tb_frame, text="DWNBY", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=row2, column=0, sticky='e', pady=(6,0))
        self.tb_dwnby = tk.StringVar()
        dwnby_entry = tk.Entry(tb_frame, textvariable=self.tb_dwnby, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                               insertbackground='white', width=12)
        dwnby_entry.grid(row=row2, column=1, padx=4, pady=(6,0), sticky='w')
        self.create_tooltip(dwnby_entry, "Drawn by; writes DWNBY.")

        tk.Label(tb_frame, text="DWNDATE", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=row2, column=2, sticky='e', pady=(6,0))
        self.tb_dwndate = tk.StringVar()
        dwndate_entry = tk.Entry(tb_frame, textvariable=self.tb_dwndate, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                                 insertbackground='white', width=12)
        dwndate_entry.grid(row=row2, column=3, padx=4, pady=(6,0), sticky='w')
        self.create_tooltip(dwndate_entry, "Drawn date; writes DWNDATE.")

        tk.Label(tb_frame, text="CHKBY", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=row2, column=4, sticky='e', pady=(6,0))
        self.tb_chkby = tk.StringVar()
        chkby_entry = tk.Entry(tb_frame, textvariable=self.tb_chkby, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                               insertbackground='white', width=12)
        chkby_entry.grid(row=row2, column=5, padx=4, pady=(6,0), sticky='w')
        self.create_tooltip(chkby_entry, "Checked by; writes CHKBY (base).")

        tk.Label(tb_frame, text="CHKDATE", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=row2, column=6, sticky='e', pady=(6,0))
        self.tb_chkdate = tk.StringVar()
        chkdate_entry = tk.Entry(tb_frame, textvariable=self.tb_chkdate, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                                 insertbackground='white', width=12)
        chkdate_entry.grid(row=row2, column=7, padx=4, pady=(6,0), sticky='w')
        self.create_tooltip(chkdate_entry, "Checked date; writes CHKDATE (base).")

        tk.Label(tb_frame, text="ENGR", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=row2, column=8, sticky='e', pady=(6,0))
        self.tb_engr = tk.StringVar()
        engr_entry = tk.Entry(tb_frame, textvariable=self.tb_engr, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                              insertbackground='white', width=12)
        engr_entry.grid(row=row2, column=9, padx=4, pady=(6,0), sticky='w')
        self.create_tooltip(engr_entry, "Engineer; writes ENGR.")

        tk.Label(tb_frame, text="ENGRDATE", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=row2, column=10, sticky='e', pady=(6,0))
        self.tb_engrdate = tk.StringVar()
        engrdate_entry = tk.Entry(tb_frame, textvariable=self.tb_engrdate, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                                  insertbackground='white', width=12)
        engrdate_entry.grid(row=row2, column=11, padx=4, pady=(6,0), sticky='w')
        self.create_tooltip(engrdate_entry, "Engineer date; writes ENGRDATE.")

        # Overall Revision (REV tag)
        tk.Label(tb_frame, text="Revision (REV)", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=row2, column=12, sticky='e', pady=(6,0))
        self.tb_overall_rev = tk.StringVar()
        overall_rev_entry = tk.Entry(tb_frame, textvariable=self.tb_overall_rev, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
                                     insertbackground='white', width=8)
        overall_rev_entry.grid(row=row2, column=13, padx=4, pady=(6,0), sticky='w')
        self.create_tooltip(overall_rev_entry, "Overall REV tag (non-numbered).")

        # Shift controls
        row3 = 3
        self.tb_shift_down = tk.BooleanVar(value=False)
        self.tb_shift_up = tk.BooleanVar(value=False)
        shift_down_chk = tk.Checkbutton(tb_frame, text="Shift Revisions Down (1→2→…→5)", variable=self.tb_shift_down,
                       fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON,
                       activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                       font=FONT_ENTRY, takefocus=False, indicatoron=False)
        shift_down_chk.grid(row=row3, column=0, columnspan=4, sticky='w', pady=(10,0))
        self.create_tooltip(shift_down_chk, "Copy Rev1→Rev2, Rev2→Rev3, … Rev5 unchanged.")

        shift_up_chk = tk.Checkbutton(tb_frame, text="Shift Revisions Up (5→4→…→1)", variable=self.tb_shift_up,
                       fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON,
                       activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                       font=FONT_ENTRY, takefocus=False, indicatoron=False)
        shift_up_chk.grid(row=row3, column=4, columnspan=4, sticky='w', pady=(10,0))
        self.create_tooltip(shift_up_chk, "Copy Rev5→Rev4, Rev4→Rev3, … Rev1 overwritten.")

        # ----- Clear Rev Lines queue -----
        clear_frame = tk.Frame(tb_frame, bg=COLOR_BG)
        clear_frame.grid(row=4, column=0, columnspan=12, sticky='w', pady=(10, 0))

        tk.Label(clear_frame, text="Clear Rev Lines:", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=0, column=0, sticky='w')

        self.tb_clear_choice = tk.StringVar(value="Rev")
        tb_clear_cb = ttk.Combobox(clear_frame, textvariable=self.tb_clear_choice, state="readonly", width=8,
                                   values=["Rev","Rev2","Rev3","Rev4","Rev5"])
        tb_clear_cb.grid(row=0, column=1, padx=6, sticky='w')
        self.create_tooltip(tb_clear_cb, "Pick a revision row to blank across all blocks.")

        add_clear_btn = tk.Button(clear_frame, text="Add Clear-Rev", command=self._ui_add_clear_rev,
                                  bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG, relief="flat")
        add_clear_btn.grid(row=0, column=2, padx=6)
        self.create_tooltip(add_clear_btn, "Queue the selected Rev row to be cleared.")

        remove_clear_btn = tk.Button(clear_frame, text="Remove Selected", command=self._ui_remove_selected_clear_rev,
                                     bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG, relief="flat")
        remove_clear_btn.grid(row=0, column=3, padx=6)
        self.create_tooltip(remove_clear_btn, "Remove selected entries from the clear queue.")

        self.tb_clear_list = tk.Listbox(clear_frame, height=3, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG)
        self.tb_clear_list.grid(row=1, column=0, columnspan=6, sticky='we', pady=(6,0))
        clear_frame.grid_columnconfigure(5, weight=1)
        self.create_tooltip(self.tb_clear_list, "Queued Rev rows to blank.")

        # ----- Stamp controls (freeze/unfreeze ISSUE layers) -----
        stamp_frame = tk.Frame(tb_frame, bg=COLOR_BG)
        stamp_frame.grid(row=5, column=0, columnspan=12, sticky='w', pady=(10, 0))

        tk.Label(stamp_frame, text="Stamp", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=0, column=0, sticky='w')

        tk.Label(stamp_frame, text="Issue:", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=FONT_LABEL)\
            .grid(row=0, column=1, padx=(10,2), sticky='e')

        self.stamp_issue = tk.StringVar(value="(leave as-is)")
        stamp_combo = ttk.Combobox(stamp_frame, textvariable=self.stamp_issue, state="readonly", width=20,
                    values=["(leave as-is)", "APPROVAL", "PRELIM", "CONSTRUCTION", "BID", "AS-BUILT", "REFERENCE"])
        stamp_combo.grid(row=0, column=2, padx=6, sticky='w')
        self.create_tooltip(stamp_combo, "Freeze all ISSUE-* layers, then thaw the selected one.")

        self.stamp_apply = tk.BooleanVar(value=True)
        stamp_chk = tk.Checkbutton(stamp_frame, text="Apply in run", variable=self.stamp_apply,
                    fg=COLOR_ENTRY_FG, bg=COLOR_BG, selectcolor=COLOR_BUTTON,
                    activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                    font=FONT_ENTRY, takefocus=False, indicatoron=False)
        stamp_chk.grid(row=0, column=3, padx=10, sticky='w')
        self.create_tooltip(stamp_chk, "Enable stamping; in Preview it logs without saving.")

        # Run + Progress
        run_frame = tk.Frame(root, bg=COLOR_BG, padx=10, pady=5)
        run_frame.grid(row=5, column=0, sticky='nsew')
        self.run_btn = tk.Button(run_frame, text="Run", command=self.start_process, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                                 font=FONT_ENTRY, padx=10, pady=5, takefocus=False, relief="flat")
        self.run_btn.pack(side=tk.LEFT, padx=5)
        self.create_tooltip(self.run_btn, "Start processing DWGs.")

        self.cancel_btn = tk.Button(run_frame, text="Cancel", command=self.cancel_process, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                                    font=FONT_ENTRY, padx=10, pady=5, state=tk.DISABLED, takefocus=False, relief="flat")
        self.cancel_btn.pack(side=tk.LEFT, padx=5)
        self.create_tooltip(self.cancel_btn, "Stop the current run after the active file.")

        self.progress = ttk.Progressbar(run_frame, length=PROGRESS_LENGTH, mode='determinate', style='black.Horizontal.TProgressbar')
        self.progress.pack(pady=5, fill='x')
        self.create_tooltip(self.progress, "Overall progress across selected files.")

        # Log
        self.status_log = scrolledtext.ScrolledText(root, height=4, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, state='disabled')
        self.status_log.grid(row=6, column=0, pady=5, padx=10, sticky='nsew')
        self.create_tooltip(self.status_log, "Live status, warnings, and errors.")

        # Bottom
        bottom_frame = tk.Frame(root, bg=COLOR_BG)
        bottom_frame.grid(row=7, column=0, pady=(0, 5), sticky='se')
        suggest_btn = tk.Button(bottom_frame, text="Suggest", command=self.open_suggestion_window, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                  font=FONT_ENTRY, relief="flat")
        suggest_btn.pack(side="left", padx=5)
        self.create_tooltip(suggest_btn, "Send a feature request or improvement idea.")

        bug_btn = tk.Button(bottom_frame, text="Report Bug", command=self.open_bug_report_window, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                  font=FONT_ENTRY, relief="flat")
        bug_btn.pack(side="left", padx=5)
        self.create_tooltip(bug_btn, "Report an issue you encountered.")

        tk.Label(bottom_frame, text=f"Version {VERSION}", fg=COLOR_ENTRY_FG, bg=COLOR_BG, font=("Arial", 8)).pack(side="right")

        self.log_message("Ready")

        # Style
        style = ttk.Style()
        style.configure('black.Horizontal.TProgressbar', background='blue', troughcolor='gray',
                        bordercolor='white', lightcolor='white', darkcolor='white')

        # Threading state
        self.process_thread = None
        self.cancel_event = threading.Event()
        self.queue = queue.Queue()

    # ---------- UI helpers ----------
    def apply_stamp_layers(self, doc, preview):
        """
        Freeze/unfreeze ISSUE-* layers based on the dropdown.
        We globally freeze all ISSUE-* layers, then thaw the chosen one.
        If '(leave as-is)' is selected, we do nothing.
        """
        if not self.stamp_apply.get():
            return False
        choice = (self.stamp_issue.get() or "(leave as-is)").strip().upper()
        if choice == "(LEAVE AS-IS)":
            return False

        layer_map = {
            "APPROVAL":     "ISSUE-APPROVAL",
            "AS-BUILT":     "ISSUE-AS-BUILT",
            "BID":          "ISSUE-BID",
            "CONSTRUCTION": "ISSUE-CONSTRUCTION",
            "PRELIM":       "ISSUE-PRELIM",
            "REFERENCE":    "ISSUE-REFERENCE",
        }
        target_layer = layer_map.get(choice)
        if not target_layer:
            return False

        present_layers = []
        try:
            for layer in doc.Layers:
                try:
                    name = layer.Name
                except Exception:
                    continue
                if name.upper().startswith("ISSUE-"):
                    present_layers.append(name)
        except Exception as e:
            self.log_message(f"Stamp: failed to enumerate layers: {e}", "error")
            return False

        if not present_layers:
            self.log_message("Stamp: no ISSUE-* layers found; skipping.")
            return False

        changed = False
        for lname in present_layers:
            try:
                lyr = doc.Layers.Item(lname)
                desired_freeze = (lname.upper() != target_layer.upper())
                if bool(lyr.Freeze) != desired_freeze:
                    if not preview:
                        lyr.Freeze = desired_freeze
                    self.log_message(f"Stamp: {'Froze' if desired_freeze else 'Thawed'} layer {lname}")
                    changed = True
            except Exception as e:
                self.log_message(f"Stamp: could not update layer {lname}: {e}", "error")

        try:
            lyr = doc.Layers.Item(target_layer)
            if lyr.LayerOn is False:
                if not preview:
                    lyr.LayerOn = True
        except Exception:
            pass

        return changed

    def _ui_add_clear_rev(self):
        choice = (self.tb_clear_choice.get() or "Rev").strip()
        if choice not in ("Rev", "Rev2", "Rev3", "Rev4", "Rev5"):
            return
        if choice not in self.tb_clear_list.get(0, tk.END):
            self.tb_clear_list.insert(tk.END, choice)

    def _ui_remove_selected_clear_rev(self):
        sel = list(self.tb_clear_list.curselection())
        for i in reversed(sel):
            self.tb_clear_list.delete(i)

    def create_tooltip(self, widget, text):
        def enter(event):
            # Why: simple, lightweight hover help that doesn’t steal focus
            self.tooltip = tk.Toplevel(widget)
            self.tooltip.wm_overrideredirect(True)
            self.tooltip.wm_geometry(f"+{event.x_root + 20}+{event.y_root + 20}")
            label = tk.Label(self.tooltip, text=text, bg='yellow', fg='black', relief='solid', borderwidth=1, padx=5, pady=3)
            label.pack()
        def leave(event):
            if hasattr(self, 'tooltip'):
                self.tooltip.destroy()
        widget.bind("<Enter>", enter); widget.bind("<Leave>", leave)

    def log_message(self, message, level='info'):
        self.status_log.config(state='normal')
        self.status_log.insert(tk.END, f"{datetime.now().strftime('%H:%M:%S')} - {message}\n")
        self.status_log.see(tk.END)
        self.status_log.config(state='disabled')
        (logger.error if level == 'error' else logger.info)(message)

    def update_mode(self):
        self.path_entry.delete(0, tk.END)

    def browse_path(self):
        if self.mode_var.get() == "directory":
            path = filedialog.askdirectory()
            if path:
                self.path_entry.delete(0, tk.END); self.path_entry.insert(0, path)
        else:
            paths = filedialog.askopenfilenames(filetypes=[("DWG Files", "*.dwg")])
            if paths:
                self.path_entry.delete(0, tk.END); self.path_entry.insert(0, ";".join(paths))

    def add_replacement_pair(self):
        if len(self.replacements) >= MAX_REPLACEMENTS:
            self.log_message("Error: Maximum of 10 find-replace pairs allowed.", 'error')
            return
        frame = tk.Frame(self.replacements_frame, bg=COLOR_BG); frame.pack(pady=2, fill='x')

        find_entry = tk.Entry(frame, width=ENTRY_WIDTH, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', font=FONT_ENTRY)
        find_entry.pack(side=tk.LEFT, padx=5, fill='x', expand=True)
        self.create_tooltip(find_entry, "Text/pattern to find. Enable 'Advanced' for regex.")
        tk.Label(frame, text="->", fg=COLOR_ENTRY_FG, bg=COLOR_BG).pack(side=tk.LEFT)
        replace_entry = tk.Entry(frame, width=ENTRY_WIDTH, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, insertbackground='white', font=FONT_ENTRY)
        replace_entry.pack(side=tk.LEFT, padx=5, fill='x', expand=True)
        self.create_tooltip(replace_entry, "Replacement text.")

        case_var = tk.BooleanVar()
        case_check = tk.Checkbutton(frame, text="Ignore Case", variable=case_var, fg=COLOR_ENTRY_FG, bg=COLOR_BG,
                                    selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                                    font=FONT_ENTRY, takefocus=False, indicatoron=False)
        case_check.pack(side=tk.LEFT, padx=2)
        self.create_tooltip(case_check, "Case-insensitive matching.")

        regex_var = tk.BooleanVar()
        regex_check = tk.Checkbutton(frame, text="Use Advanced Patterns", variable=regex_var, fg=COLOR_ENTRY_FG, bg=COLOR_BG,
                                     selectcolor=COLOR_BUTTON, activebackground=COLOR_BUTTON, activeforeground=COLOR_ENTRY_FG,
                                     font=FONT_ENTRY, takefocus=False, indicatoron=False)
        regex_check.pack(side=tk.LEFT, padx=2)
        self.create_tooltip(regex_check, "Enable regular expressions (advanced).")

        remove_btn = tk.Button(frame, text="Remove", command=lambda f=frame: self.remove_pair(f), bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                               font=FONT_ENTRY, padx=10, pady=5, takefocus=False, relief="flat")
        remove_btn.pack(side=tk.LEFT)
        self.create_tooltip(remove_btn, "Delete this rule.")
        self.replacements.append((find_entry, replace_entry, case_var, regex_var))

    def remove_pair(self, frame):
        frame.destroy()
        self.replacements = [pair for pair in self.replacements if pair[0].winfo_exists()]

    def validate_replacements(self, replacements):
        valid = []
        for find_entry, replace_entry, case_var, regex_var in replacements:
            find = find_entry.get().strip()
            replace = replace_entry.get().strip()
            if not find:
                continue
            if regex_var.get():
                try:
                    re.compile(find)
                except re.error as e:
                    self.log_message(f"Invalid regex '{find}': {e}", 'error')
                    return None
            valid.append((find_entry, replace_entry, case_var, regex_var))
        return valid

    # ---------- AutoCAD helpers ----------
    def get_autocad_app(self):
        try:
            acad = win32com.client.Dispatch("AutoCAD.Application")
            acad.Visible = True
            return acad
        except Exception as e:
            self.log_message(f"Error connecting to AutoCAD: {e}", 'error')
            return None

    def process_entity_text(self, entity, entity_type, original_text, replacements, preview, changes_list, file_path):
        """Apply find/replace to a single text-bearing entity."""
        new_text = original_text
        for find_entry, replace_entry, case_var, regex_var in replacements:
            find = find_entry.get().strip()
            replace = replace_entry.get().strip()
            if not find:
                continue
            if regex_var.get():
                flags = re.IGNORECASE if case_var.get() else 0
                new_text = re.sub(find, replace, new_text, flags=flags)
            elif case_var.get():
                lower_text = new_text.lower()
                new_text = lower_text.replace(find.lower(), replace)
                if new_text != original_text.lower():
                    new_text = original_text.replace(original_text.lower().replace(find.lower(), replace), new_text)
            else:
                new_text = new_text.replace(find, replace)

        if new_text != original_text:
            if not preview:
                try:
                    if entity_type == "AcDbText":
                        entity.TextString = new_text
                    elif entity_type == "AcDbMText":
                        try:
                            entity.Text = new_text
                        except Exception:
                            entity.Contents = new_text
                    elif entity_type == "AcDbMLeader" and hasattr(entity, "MTextContent"):
                        entity.MTextContent = new_text
                    elif entity_type.startswith("AcDbDimension"):
                        entity.TextOverride = new_text
                except Exception as e:
                    self.log_message(f"Warning: failed to write text on {file_path}: {e}", "error")
            changes_list.append({
                "File": file_path,
                "EntityType": entity_type,
                "OriginalText": original_text,
                "NewText": new_text,
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "User": getpass.getuser()
            })
            return True
        return False

    def _read_attr_map(self, entity):
        """Return {TAG: value} for a BlockReference (uppercased tags)."""
        out = {}
        if getattr(entity, "HasAttributes", False):
            for attr in entity.GetAttributes():
                tag = getattr(attr, "TagString", "")
                if tag:
                    out[tag.upper().strip()] = attr.TextString
        return out

    def _write_attr_if_exists(self, entity, tag, value):
        """Write attribute TAG=value on BlockReference if TAG exists. Returns True if written."""
        wrote = False
        if getattr(entity, "HasAttributes", False):
            for attr in entity.GetAttributes():
                if getattr(attr, "TagString", "").upper().strip() == tag.upper():
                    if attr.TextString != value:
                        attr.TextString = value
                    wrote = True
        return wrote

    def _clear_rev_line_on_blockref(self, entity, n, preview, changes_list, file_path):
        """
        Blank the Rev-n line on this BlockReference (REVn, DESCn, BYn, CHKn/CHKBYn, DATEn).
        Only blanks tags that exist; does not create new ones.
        """
        changed = False
        try:
            amap = self._read_attr_map(entity)
            if not amap:
                return False

            bases = ["REV", "DESC", "BY", "DATE", "CHK", "CHKBY"]
            for base in bases:
                tag = f"{base}{n}"
                if tag in amap:
                    if not preview:
                        self._write_attr_if_exists(entity, tag, "")
                    changes_list.append({
                        "File": file_path,
                        "EntityType": f"ClearRev {n}",
                        "OriginalText": amap.get(tag, ""),
                        "NewText": "",
                        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "User": getpass.getuser()
                    })
                    changed = True
        except Exception as e:
            self.log_message(f"Clear Rev error: {e}", "error")
        return changed

    def _shift_revisions_on_blockref(self, entity, direction, preview, changes_list, file_path):
        """
        Shift revisions up/down on this BlockReference.
        direction: 'down' => 1->2->3->4->5 ; 'up' => 5->4->3->2->1
        Handles: REVn, DESCn, BYn, CHKn, CHKBYn, DATEn for n=1..5
        """
        if direction not in ("down", "up"):
            return False
        changed = False
        try:
            amap = self._read_attr_map(entity)
            if not amap:
                return False

            bases = ["REV", "DESC", "BY", "DATE", "CHK", "CHKBY"]
            indices = list(range(1, 6))
            if direction == "down":
                src_order = indices[:-1]      # 1..4 move to 2..5
                dest_shift = +1
            else:  # up
                src_order = indices[:0:-1]    # 5..2 move to 4..1
                dest_shift = -1

            planned = {}
            for n in src_order:
                dest = n + dest_shift
                for base in bases:
                    src_tag  = f"{base}{n}"
                    dest_tag = f"{base}{dest}"
                    if src_tag in amap and dest_tag in amap:
                        planned[dest_tag] = amap[src_tag]

            if planned:
                for tag, val in planned.items():
                    try:
                        if not preview:
                            self._write_attr_if_exists(entity, tag, val)
                        changes_list.append({
                            "File": file_path,
                            "EntityType": f"ShiftRevisions {direction}",
                            "OriginalText": amap.get(tag, ""),
                            "NewText": val,
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
                        changed = True
                    except Exception as e:
                        self.log_message(f"Shift write failed for {tag}: {e}", "error")
        except Exception as e:
            self.log_message(f"Shift error: {e}", "error")
        return changed

    def process_block_attributes(self, entity, replacements, preview, changes_list, file_path):
        """Process attributes on a block reference (AcDbBlockReference)."""
        changed = False
        try:
            if entity.HasAttributes:
                attributes = entity.GetAttributes()
                for attr in attributes:
                    original_text = attr.TextString
                    new_text = original_text
                    for find_entry, replace_entry, case_var, regex_var in replacements:
                        find = find_entry.get().strip()
                        replace = replace_entry.get().strip()
                        if not find:
                            continue
                        if regex_var.get():
                            flags = re.IGNORECASE if case_var.get() else 0
                            new_text = re.sub(find, replace, new_text, flags=flags)
                        elif case_var.get():
                            lower_text = new_text.lower()
                            new_text = lower_text.replace(find.lower(), replace)
                            if new_text != original_text.lower():
                                new_text = original_text.replace(original_text.lower().replace(find.lower(), replace), new_text)
                        else:
                            new_text = new_text.replace(find, replace)

                    if new_text != original_text:
                        if not preview:
                            try:
                                attr.TextString = new_text
                            except Exception as e:
                                self.log_message(f"Warning: failed to write attribute on {file_path}: {e}", "error")
                                continue
                        changes_list.append({
                            "File": file_path,
                            "EntityType": "AcDbBlockReference Attribute",
                            "OriginalText": original_text,
                            "NewText": new_text,
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
                        changed = True
        except Exception as e:
            self.log_message(f"Error accessing attributes: {e}", "error")
        return changed

    def apply_title_block_to_blockref(self, entity, preview, changes_list, file_path):
        """
        If Title Block helper is enabled, set selected Rev section (REVn/DESCn/BYn/CHK(n|BYn)/DATEn)
        and base tags (DWNBY, DWNDATE, CHKBY, CHKDATE, ENGR, ENGRDATE) on this block reference.
        """
        if not self.tb_enable.get():
            return False
        if not getattr(entity, "HasAttributes", False):
            return False

        choice = (self.tb_rev_choice.get() or "Rev1").strip().lower()
        try:
            n = int(choice.replace("rev", ""))
            n = max(1, min(5, n))
        except Exception:
            n = 1

        targets = {}
        def set_target(tag, value):
            if value is not None:
                v = str(value).strip()
                if v != "":
                    targets[tag] = v

        set_target(f"REV{n}",   self.tb_rev_val.get())
        set_target(f"DESC{n}",  self.tb_desc_val.get())
        set_target(f"BY{n}",    self.tb_by_val.get())
        chk_val = (self.tb_chk_val.get() or "").strip()
        if chk_val:
            set_target(f"CHK{n}",   chk_val)
            set_target(f"CHKBY{n}", chk_val)
        set_target(f"DATE{n}",  self.tb_date_val.get())

        set_target("DWNBY",    self.tb_dwnby.get())
        set_target("DWNDATE",  self.tb_dwndate.get())
        set_target("CHKBY",    self.tb_chkby.get())
        set_target("CHKDATE",  self.tb_chkdate.get())
        set_target("ENGR",     self.tb_engr.get())
        set_target("ENGRDATE", self.tb_engrdate.get())
        set_target("REV",      self.tb_overall_rev.get())

        if not targets:
            return False

        changed = False
        try:
            for attr in entity.GetAttributes():
                tag = getattr(attr, "TagString", "").upper().strip()
                if not tag:
                    continue
                if tag in targets:
                    new_val = targets[tag]
                    old_val = attr.TextString
                    if new_val != old_val:
                        if not preview:
                            try:
                                attr.TextString = new_val
                            except Exception as e:
                                self.log_message(f"Title Block write failed for {tag}: {e}", "error")
                                continue
                        changes_list.append({
                            "File": file_path,
                            "EntityType": "TitleBlock Attribute",
                            "OriginalText": old_val,
                            "NewText": new_val,
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
                        changed = True
                elif (tag.startswith("CHK") and tag.endswith(str(n))) and chk_val:
                    old_val = attr.TextString
                    if old_val != chk_val:
                        if not preview:
                            try:
                                attr.TextString = chk_val
                            except Exception as e:
                                self.log_message(f"Title Block write failed for {tag}: {e}", "error")
                                continue
                        changes_list.append({
                            "File": file_path,
                            "EntityType": "TitleBlock Attribute",
                            "OriginalText": old_val,
                            "NewText": chk_val,
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
                        changed = True
        except Exception as e:
            self.log_message(f"Title Block error: {e}", "error")

        return changed

    def process_table(self, entity, replacements, preview, changes_list, file_path):
        changed = False
        for row in range(entity.Rows):
            for col in range(entity.Columns):
                original_text = entity.GetCellValue(row, col)
                if isinstance(original_text, str):
                    new_text = original_text
                    for find_entry, replace_entry, case_var, regex_var in replacements:
                        find = find_entry.get().strip()
                        replace = replace_entry.get().strip()
                        if not find:
                            continue
                        if regex_var.get():
                            flags = re.IGNORECASE if case_var.get() else 0
                            new_text = re.sub(find, replace, new_text, flags=flags)
                        elif case_var.get():
                            lower_text = new_text.lower()
                            new_text = lower_text.replace(find.lower(), replace)
                            if new_text != original_text.lower():
                                new_text = original_text.replace(original_text.lower().replace(find.lower(), replace), new_text)
                        else:
                            new_text = new_text.replace(find, replace)
                    if new_text != original_text:
                        if not preview:
                            try:
                                entity.SetCellValue(row, col, new_text)
                            except Exception as e:
                                self.log_message(f"Warning: failed to write table cell on {file_path}: {e}", "error")
                        changes_list.append({
                            "File": file_path,
                            "EntityType": "AcDbTable",
                            "OriginalText": original_text,
                            "NewText": new_text,
                            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "User": getpass.getuser()
                        })
                        changed = True
        return changed

    # --- recursive block processor: clear -> shift -> replacements -> title block; skips XREFs ---
    def process_block_reference(self, entity, replacements, preview, changes_list, file_path):
        """Process a block reference: clear revs, shifts, attributes, nested non-xref blocks, and Title Block helper."""
        changed = False

        # 0) Clear requested Rev lines first
        clear_items = list(self.tb_clear_list.get(0, tk.END))
        for item in clear_items:
            choice = item.strip().lower()
            n = 1 if choice == "rev" else int(choice.replace("rev", ""))
            if self._clear_rev_line_on_blockref(entity, n, preview, changes_list, file_path):
                changed = True

        # 1) Shift revisions
        shift_dir = None
        if self.tb_shift_down.get():
            shift_dir = "down"
        elif self.tb_shift_up.get():
            shift_dir = "up"
        if shift_dir:
            try:
                if self._shift_revisions_on_blockref(entity, shift_dir, preview, changes_list, file_path):
                    changed = True
            except Exception as e:
                self.log_message(f"Error shifting revisions: {e}", "error")

        # 2) Attribute find/replace
        if self.process_block_attributes(entity, replacements, preview, changes_list, file_path):
            changed = True

        # 3) Apply Title Block explicit values
        if self.apply_title_block_to_blockref(entity, preview, changes_list, file_path):
            changed = True

        # 4) Nested blocks (skip XREFs)
        try:
            if getattr(entity, "IsXRef", False):
                self.log_message(f"Skipping xref block: {getattr(entity, 'EffectiveName', 'Unknown')}")
                return changed

            block_name = entity.EffectiveName
            block_def = entity.Document.Blocks(block_name)

            for ent in block_def:
                ent_type = ent.EntityName
                original_text = None

                if ent_type == "AcDbText":
                    original_text = ent.TextString
                elif ent_type == "AcDbMText":
                    try:
                        original_text = ent.Text
                    except Exception:
                        try:
                            original_text = ent.Contents
                        except Exception:
                            original_text = ""
                elif ent_type.startswith("AcDbDimension"):
                    original_text = ent.TextOverride or ""
                elif ent_type == "AcDbMLeader" and hasattr(ent, "MTextContent"):
                    original_text = ent.MTextContent

                if original_text is not None:
                    if self.process_entity_text(ent, ent_type, original_text, replacements, preview, changes_list, file_path):
                        changed = True
                    continue

                if ent_type == "AcDbBlockReference":
                    if self.process_block_reference(ent, replacements, preview, changes_list, file_path):
                        changed = True
                    continue

                if ent_type == "AcDbTable":
                    if self.process_table(ent, replacements, preview, changes_list, file_path):
                        changed = True
                    continue

        except Exception as e:
            self.log_message(f"Error processing nested block {getattr(entity, 'EffectiveName', 'Unknown')}: {e}", "error")

        return changed

    # ---------- Main DWG processing ----------
    def process_dwg_file(self, acad, file_path, replacements, log, error_log, preview, total_files, processed_files):
        """Process a single DWG file with attribute support (title block + stamp), skipping xrefs."""
        self.log_message(f"Processing: {os.path.basename(file_path)}")
        try:
            # Retry loop for busy server
            for attempt in range(3):
                try:
                    doc = acad.Documents.Open(file_path)
                    pythoncom.PumpWaitingMessages()
                    time.sleep(1)
                    try:
                        self.apply_stamp_layers(doc, preview)
                    except Exception as e:
                        self.log_message(f"Stamp apply failed: {e}", "error")
                    break
                except Exception as e:
                    if "application is busy" in str(e).lower() and attempt < 2:
                        self.log_message(f"AutoCAD busy, retrying open ({attempt+1}/3)...")
                        time.sleep(2)
                        pythoncom.PumpWaitingMessages()
                        continue
                    else:
                        raise

            changed = False
            changes_list = []

            # Process Model + Paper
            spaces = [doc.ModelSpace, doc.PaperSpace]

            for space in spaces:
                for entity in space:
                    entity_type = entity.EntityName
                    original_text = None

                    if entity_type == "AcDbText":
                        original_text = entity.TextString
                    elif entity_type == "AcDbMText":
                        try:
                            original_text = entity.Text
                        except Exception:
                            try:
                                original_text = entity.Contents
                            except Exception:
                                original_text = ""
                    elif entity_type.startswith("AcDbDimension"):
                        original_text = entity.TextOverride or ""
                    elif entity_type == "AcDbMLeader" and hasattr(entity, "MTextContent"):
                        original_text = entity.MTextContent

                    if original_text is not None:
                        if self.process_entity_text(entity, entity_type, original_text, replacements, preview, changes_list, file_path):
                            changed = True
                        continue

                    if entity_type == "AcDbBlockReference":
                        if self.process_block_reference(entity, replacements, preview, changes_list, file_path):
                            changed = True
                        continue

                    if entity_type == "AcDbTable":
                        if self.process_table(entity, replacements, preview, changes_list, file_path):
                            changed = True
                        continue

            if changed and not preview:
                doc.Save()
                time.sleep(1); pythoncom.PumpWaitingMessages()

            doc.Close()
            time.sleep(1); pythoncom.PumpWaitingMessages()

            log.extend(changes_list)
            self.log_message(f"Processed: {os.path.basename(file_path)} - {processed_files + 1}/{total_files}")
            return True

        except Exception as e:
            error_log.append({
                "File": file_path,
                "Error": str(e),
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            })
            self.log_message(f"Error processing {file_path}: {str(e)}", 'error')
            return False

    # ---------- Runner / threading ----------
    def start_process(self):
        if self.process_thread and self.process_thread.is_alive():
            return
        self.cancel_event.clear()
        self.run_btn.config(state=tk.DISABLED)
        self.cancel_btn.config(state=tk.NORMAL)
        self.progress['value'] = 0
        self.process_thread = threading.Thread(target=self.run_process_thread)
        self.process_thread.start()
        self.root.after(100, self.check_queue)

    def cancel_process(self):
        self.cancel_event.set()
        self.log_message("Cancelling process...")

    def check_queue(self):
        try:
            while True:
                msg_type, msg = self.queue.get_nowait()
                if msg_type == 'status':
                    self.log_message(msg)
                elif msg_type == 'error':
                    self.log_message(msg, 'error')
                elif msg_type == 'progress':
                    self.progress['value'] = msg
                elif msg_type == 'done':
                    self.run_btn.config(state=tk.NORMAL)
                    self.cancel_btn.config(state=tk.DISABLED)
                    self.handle_completion(*msg)
        except queue.Empty:
            pass
        if self.process_thread and self.process_thread.is_alive():
            self.root.after(100, self.check_queue)
        else:
            self.run_btn.config(state=tk.NORMAL)
            self.cancel_btn.config(state=tk.DISABLED)

    def run_process_thread(self):
        """Run the process, continue through errors, and print a summary."""
        path = self.path_entry.get()
        if not path:
            self.log_message("No path selected.", 'error'); return

        replacements = self.validate_replacements(self.replacements)
        if replacements is None:
            return

        replacements = self.validate_replacements(self.replacements)
        if replacements is None:
            return

        try:
            clear_count = self.tb_clear_list.size()
        except Exception:
            try:
                clear_count = len(self.tb_clear_list.get(0, tk.END))
            except Exception:
                clear_count = 0

        doing_titleblock_ops = (
            self.tb_enable.get() or
            self.tb_shift_down.get() or
            self.tb_shift_up.get() or
            clear_count > 0
        )

        stamp_choice = (self.stamp_issue.get() or "").strip().upper()
        stamp_selected = self.stamp_apply.get() and stamp_choice != "(LEAVE AS-IS)"

        doing_any_ops = doing_titleblock_ops or stamp_selected

        if not replacements and not doing_any_ops:
            self.log_message(
                "No valid replacements entered and no Title Block or Stamp operations selected.",
                'error'
            )
            return

        if not replacements and doing_any_ops:
            if stamp_selected and not doing_titleblock_ops:
                self.log_message("Proceeding with Stamp operation only (no text or title block changes).")
            else:
                self.log_message("Proceeding with Title Block / Stamp operations only (no text replacements).")

        preview = self.preview_var.get()

        # Build file list
        if self.mode_var.get() == "directory":
            if not os.path.isdir(path):
                self.log_message("Invalid directory.", 'error'); return
            dwg_files = []
            if self.recursive_var.get():
                for root_dir, _, files in os.walk(path):
                    dwg_files.extend(os.path.join(root_dir, f) for f in files if f.lower().endswith('.dwg'))
            else:
                dwg_files = [os.path.join(path, f) for f in os.listdir(path) if f.lower().endswith('.dwg')]
            report_dir = path
        else:
            dwg_files = path.split(";")
            report_dir = os.path.dirname(dwg_files[0]) if dwg_files else ""

        if not dwg_files:
            self.log_message("No DWG files found.", 'error'); return

        acad = self.get_autocad_app()
        if not acad: return

        log = []; error_log = []
        total_files = len(dwg_files)
        success_count = 0; fail_count = 0

        for i, file_path in enumerate(dwg_files):
            if self.cancel_event.is_set():
                self.log_message("Process cancelled."); break
            ok = self.process_dwg_file(acad, file_path, replacements, log, error_log, preview, total_files, i)
            if ok:
                success_count += 1
            else:
                fail_count += 1
                self.log_message(f"Skipped {file_path} due to error, continuing...")

        # Summary
        summary_msg = f"Process completed. {success_count} succeeded, {fail_count} failed."
        if fail_count > 0:
            summary_msg += " See error log for details."
        self.log_message(summary_msg)

        # Report
        if log or error_log:
            report_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialdir=report_dir,
                                                       title="Save Report As", filetypes=[("Excel Files", "*.xlsx")])
            if report_path:
                try:
                    df_changes = pd.DataFrame(log) if log else pd.DataFrame()
                    df_errors = pd.DataFrame(error_log) if error_log else pd.DataFrame()
                    wb = Workbook()

                    # Changes sheet
                    if not df_changes.empty:
                        ws_changes = wb.active; ws_changes.title = "Changes"
                        headers = list(df_changes.columns)
                        for col_num, header in enumerate(headers, 1):
                            cell = ws_changes.cell(row=1, column=col_num, value=header); cell.font = Font(bold=True)
                        for row_num, row in enumerate(df_changes.itertuples(index=False), 2):
                            for col_num, value in enumerate(row, 1):
                                cell = ws_changes.cell(row=row_num, column=col_num, value=value)
                                if headers[col_num-1] == "File":
                                    hyperlink_value = str(value).replace("\\", "/")
                                    cell.hyperlink = f"file:///{hyperlink_value}"
                    else:
                        ws_changes = wb.active; ws_changes.title = "Changes"
                        ws_changes.cell(row=1, column=1, value="No changes recorded")

                    # Errors sheet
                    if not df_errors.empty:
                        ws_errors = wb.create_sheet("Errors")
                        headers = list(df_errors.columns)
                        for col_num, header in enumerate(headers, 1):
                            cell = ws_errors.cell(row=1, column=col_num, value=header); cell.font = Font(bold=True)
                        for row_num, row in enumerate(df_errors.itertuples(index=False), 2):
                            for col_num, value in enumerate(row, 1):
                                ws_errors.cell(row=row_num, column=col_num, value=value)

                    # Summary sheet
                    ws_summary = wb.create_sheet("Summary")
                    ws_summary["A1"].font = Font(bold=True); ws_summary["A1"] = "Timestamp"
                    ws_summary["B1"].font = Font(bold=True); ws_summary["B1"] = "Total"
                    ws_summary["C1"].font = Font(bold=True); ws_summary["C1"] = "Succeeded"
                    ws_summary["D1"].font = Font(bold=True); ws_summary["D1"] = "Failed"
                    ws_summary["A2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ws_summary["B2"] = total_files
                    ws_summary["C2"] = success_count
                    ws_summary["D2"] = fail_count

                    wb.save(report_path)
                    self.log_message(f"Report saved to {report_path}")
                except Exception as e:
                    self.log_message(f"Error saving report: {e}", "error")
            else:
                self.log_message("Report save cancelled.")
        else:
            self.log_message("No changes or errors detected.")

        # ---------- Suggest/Bug windows ----------
    def open_suggestion_window(self):
        suggestion_win = tk.Toplevel(self.root); suggestion_win.title("Make a Suggestion")
        suggestion_win.configure(bg=COLOR_BG); suggestion_win.resizable(False, False)

        tk.Label(suggestion_win, text="Enter your suggestion:", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=5)
        suggestion_text = scrolledtext.ScrolledText(suggestion_win, wrap="word", bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, height=5)
        suggestion_text.pack(padx=10, pady=5, fill="both")

        tk.Label(suggestion_win, text="Your name (optional):", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=(8, 2))
        user_name_entry = tk.Entry(suggestion_win, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, insertbackground='white')
        user_name_entry.pack(padx=10, pady=(0, 8), fill="x")

        tk.Label(suggestion_win, text="Your email (optional):", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=(2, 2))
        user_email_entry = tk.Entry(suggestion_win, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, insertbackground='white')
        user_email_entry.pack(padx=10, pady=(0, 8), fill="x")

        def submit_suggestion():
            suggestion = suggestion_text.get("1.0", "end").strip()
            user_name = user_name_entry.get().strip()
            user_email = user_email_entry.get().strip()
            if suggestion:
                send_email_single("Suggestion", suggestion, user_name=user_name or None, user_email=user_email or None)
                self.log_message("Suggestion submitted successfully.")
            else:
                self.log_message("Suggestion submission cancelled: No text entered.", 'error')
            suggestion_win.destroy()

        tk.Button(suggestion_win, text="Submit", command=submit_suggestion, bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG,
                font=FONT_ENTRY, relief="flat").pack(pady=5)
        self.log_message("Suggestion window opened.")

    def open_bug_report_window(self):
        bug_win = tk.Toplevel(self.root); bug_win.title("Report a Bug")
        bug_win.configure(bg=COLOR_BG); bug_win.resizable(False, False)

        tk.Label(bug_win, text="Describe the error:", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=5)
        bug_text = scrolledtext.ScrolledText(
            bug_win, wrap="word", bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG,
            font=FONT_ENTRY, height=5
        )
        bug_text.pack(padx=10, pady=5, fill="both")

        tk.Label(bug_win, text="Your name (optional):", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=(5, 2))
        user_name_entry = tk.Entry(bug_win, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, insertbackground='white')
        user_name_entry.pack(padx=10, pady=(0, 8), fill="x")

        tk.Label(bug_win, text="Your email (optional):", fg=COLOR_FG, bg=COLOR_BG, font=FONT_LABEL).pack(pady=(2, 2))
        user_email_entry = tk.Entry(bug_win, bg=COLOR_ENTRY_BG, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, insertbackground='white')
        user_email_entry.pack(padx=10, pady=(0, 8), fill="x")

        def submit_bug_report():
            bug_report = bug_text.get("1.0", "end").strip()
            user_name = user_name_entry.get().strip()
            user_email = user_email_entry.get().strip()
            if bug_report:
                send_email_single("Bug Report", bug_report, user_name=user_name or None, user_email=user_email or None)
                self.log_message("Bug report submitted successfully.")
            else:
                self.log_message("Bug report submission cancelled: No text entered.", 'error')
            bug_win.destroy()

        tk.Button(
            bug_win, text="Submit", command=submit_bug_report,
            bg=COLOR_BUTTON, fg=COLOR_ENTRY_FG, font=FONT_ENTRY, relief="flat"
        ).pack(pady=5)

        self.log_message("Bug report window opened.")

# ---------- Simplified / Privacy-first email sender ----------
def _escape_html(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def send_email_single(subject: str, body: str, user_name: str | None = None, user_email: str | None = None):
    """
    Send a single HTML email to RECEIVER_EMAIL (internal owner/dev).
    """
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        submitted_by_display = user_name or "(not provided)"
        submitted_email_display = user_email or "(not provided)"

        html = f"""
        <div style="background:#121212;color:#FFFFFF;font-family:Arial,sans-serif;padding:20px;">
          <div style="text-align:center;margin-bottom:15px;">
            <img src="cid:root3logo" alt="Root3Power Logo" style="width:80px;height:80px;"/>
            <h2 style="color:#1E90FF;margin:10px 0 0;">Root3Power LLC</h2>
            <p style="margin:0;font-size:14px;color:#4682B4;">{_escape_html(subject)}</p>
          </div>
          <div style="background:#1E1E1E;padding:15px;border-radius:8px;">
            <p><strong>Submitted By (name):</strong> {_escape_html(submitted_by_display)}</p>
            <p><strong>Submitted By (email):</strong> {_escape_html(submitted_email_display)}</p>
            <p><strong>Submitted At:</strong> {timestamp}</p>
            <p><strong>Message:</strong></p>
            <div style="background:#2C2C2C;padding:10px;border-radius:6px;color:#FFFFFF;white-space:pre-wrap;">
              {_escape_html(body)}
            </div>
          </div>

          <div style="margin-top:12px;color:#BBBBBB;font-size:12px;">
            App Version: {VERSION}
          </div>
        </div>
        """

        msg = MIMEMultipart("related")
        msg["Subject"] = subject
        msg["From"] = formataddr(("Batch Find & Replace", SENDER_EMAIL))
        msg["To"] = RECEIVER_EMAIL
        if user_email:
            # set Reply-To so the owner can reply directly to the submitter
            msg["Reply-To"] = user_email

        alt = MIMEMultipart("alternative")
        msg.attach(alt)
        alt.attach(MIMEText(html, "html"))

        # Embed logo when available
        try:
            with open(resource_path(LOGO_FILE), "rb") as f:
                logo = MIMEImage(f.read())
                logo.add_header("Content-ID", "<root3logo>")
                logo.add_header("Content-Disposition", "inline", filename=LOGO_FILE)
                msg.attach(logo)
        except Exception:
            pass

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER_EMAIL, APP_PASSWORD)
            server.sendmail(SENDER_EMAIL, [RECEIVER_EMAIL], msg.as_string())

    except Exception as e:
        logger.error(f"Error sending email: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = AutoCADReplacerGUI(root)
    root.mainloop()