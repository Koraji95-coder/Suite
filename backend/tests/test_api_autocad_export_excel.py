from __future__ import annotations

import os
import tempfile
import unittest
from datetime import datetime

from openpyxl import load_workbook

from backend.route_groups.api_autocad_export_excel import export_points_to_excel


class TestApiAutocadExportExcel(unittest.TestCase):
    def test_export_points_to_excel_builds_expected_workbook(self) -> None:
        points = [
            {"name": "B-1", "x": 10.125, "y": 20.5, "z": 0.25, "layer": "B-LAYER"},
            {"name": "A-1", "x": 1.5, "y": 2.75, "z": 3.0, "layer": "A-LAYER"},
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            out_path = export_points_to_excel(
                points,
                precision=2,
                use_corners=False,
                drawing_dir=temp_dir,
                now_fn=lambda: datetime(2026, 1, 2, 3, 4, 5),
            )

            self.assertTrue(os.path.exists(out_path))
            self.assertEqual(os.path.basename(out_path), "coordinates_20260102_030405.xlsx")

            wb = load_workbook(out_path)
            ws = wb["Coordinates"]

            self.assertEqual(ws["A1"].value, "Ground Grid Coordinates")
            self.assertEqual(ws["A2"].value, "Layer: A-LAYER")
            self.assertEqual(ws["A3"].value, "Point ID")
            self.assertEqual(ws["A4"].value, "A-1")
            self.assertEqual(ws["B4"].number_format, "0.00")
            self.assertEqual(ws.freeze_panes, "A3")

    def test_export_points_to_excel_uses_default_exports_folder(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            out_path = export_points_to_excel(
                [{"name": "P1", "x": 0, "y": 0, "z": 0, "layer": "L1"}],
                precision=0,
                use_corners=False,
                drawing_dir=None,
                output_base_dir=temp_dir,
                now_fn=lambda: datetime(2026, 1, 1, 0, 0, 0),
            )

            self.assertTrue(out_path.startswith(os.path.join(temp_dir, "exports")))
            self.assertTrue(os.path.exists(out_path))
            wb = load_workbook(out_path)
            ws = wb["Coordinates"]
            self.assertEqual(ws["B4"].number_format, "0")


if __name__ == "__main__":
    unittest.main()
