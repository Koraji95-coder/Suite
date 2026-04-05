from __future__ import annotations

import os
import re
import shutil
import tempfile
import time
import uuid
from datetime import datetime
from functools import wraps
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

from flask import Blueprint, jsonify, request, send_file
from flask_limiter import Limiter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import safe_join

from backend.runtime_paths import is_absolute_path_value, resolve_runtime_directory, resolve_runtime_path

MAX_DRAWINGS = 50
MAX_SCHEDULE_ROWS = 5000
MAX_APPLY_OPERATIONS = 5000


def create_terminal_authoring_blueprint(
    *,
    limiter: Limiter,
    logger: Any,
    require_supabase_user: Callable,
    is_valid_api_key: Callable[[Optional[str]], bool],
    schedule_cleanup: Callable[[str], None],
    send_autocad_dotnet_command: Optional[Callable[[str, Dict[str, Any]], Dict[str, Any]]] = None,
    send_autocad_acade_command: Optional[Callable[[str, Dict[str, Any]], Dict[str, Any]]] = None,
) -> Blueprint:
    """Create project terminal authoring routes under /api/conduit-route."""

    bp = Blueprint("terminal_authoring_api", __name__, url_prefix="/api/conduit-route")
    generated_reports: Dict[str, Dict[str, str]] = {}

    def require_user_or_api_key(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            provided_key = request.headers.get("X-API-Key")
            if provided_key and is_valid_api_key(provided_key):
                return f(*args, **kwargs)
            wrapped = require_supabase_user(f)
            return wrapped(*args, **kwargs)

        return decorated

    def _normalize_text(value: Any) -> str:
        return str(value or "").strip()

    def _normalize_nullable_text(value: Any) -> str | None:
        normalized = _normalize_text(value)
        return normalized or None

    def _normalize_drawing_key(value: Any) -> str:
        return re.sub(r"[^A-Z0-9]+", "", _normalize_text(value).upper())

    def _request_id() -> str:
        raw = (
            request.headers.get("X-Request-ID")
            or request.headers.get("X-Request-Id")
            or ""
        )
        if not raw:
            payload = request.get_json(silent=True) or {}
            if isinstance(payload, dict):
                raw = str(payload.get("requestId") or payload.get("request_id") or "")
        normalized = re.sub(r"[^A-Za-z0-9._:-]+", "-", str(raw or "").strip())
        if normalized:
            return normalized[:128]
        return f"terminal-authoring-{int(time.time() * 1000)}"

    def _call_dotnet_bridge_action(
        *,
        action: str,
        payload: Dict[str, Any],
        request_id: str,
    ) -> Dict[str, Any]:
        if send_autocad_dotnet_command is None:
            raise RuntimeError("AutoCAD .NET bridge is not configured.")

        response = send_autocad_dotnet_command(
            action,
            {
                **payload,
                "requestId": request_id,
            },
        )
        if not isinstance(response, dict):
            raise RuntimeError("Malformed response from AutoCAD .NET bridge.")
        if not response.get("ok"):
            raise RuntimeError(
                str(response.get("error") or response.get("message") or "Unknown bridge error.")
            )

        result_payload = response.get("result")
        if not isinstance(result_payload, dict):
            raise RuntimeError("Invalid .NET bridge result payload.")
        return result_payload

    def _call_acade_host_action(
        *,
        action: str,
        payload: Dict[str, Any],
        request_id: str,
    ) -> Dict[str, Any]:
        if send_autocad_acade_command is None:
            raise RuntimeError("AutoCAD in-process ACADE host is not configured.")

        response = send_autocad_acade_command(
            action,
            {
                **payload,
                "requestId": request_id,
            },
        )
        if not isinstance(response, dict):
            raise RuntimeError("Malformed response from the AutoCAD in-process ACADE host.")
        if not response.get("ok"):
            raise RuntimeError(
                str(
                    response.get("error")
                    or response.get("message")
                    or "Unknown in-process ACADE host error."
                )
            )

        result_payload = response.get("result")
        if not isinstance(result_payload, dict):
            raise RuntimeError("Invalid in-process ACADE host result payload.")
        return result_payload

    def _is_absolute_windows_path(path: str) -> bool:
        normalized = path.strip()
        return bool(re.match(r"^[A-Za-z]:[\\/]", normalized)) or normalized.startswith("\\\\")

    def _resolve_project_drawings(payload: Dict[str, Any]) -> List[Dict[str, str]]:
        raw_selected = payload.get("selectedDrawingPaths")
        if not isinstance(raw_selected, list) or len(raw_selected) == 0:
            raise ValueError("selectedDrawingPaths must contain at least one drawing path.")
        if len(raw_selected) > MAX_DRAWINGS:
            raise ValueError(f"Too many selected drawings. Maximum is {MAX_DRAWINGS}")

        drawing_root_raw = _normalize_text(payload.get("drawingRootPath"))
        drawing_root = os.path.abspath(drawing_root_raw) if drawing_root_raw else ""
        drawings: List[Dict[str, str]] = []
        seen_paths: set[str] = set()

        for item in raw_selected:
            raw_path = _normalize_text(item)
            if not raw_path:
                continue

            if _is_absolute_windows_path(raw_path):
                absolute_path = os.path.abspath(raw_path)
                relative_path = raw_path
                if drawing_root:
                    try:
                        candidate_relative = os.path.relpath(absolute_path, drawing_root)
                        if not candidate_relative.startswith(".."):
                            relative_path = candidate_relative
                    except ValueError:
                        pass  # Paths on different drives (Windows); keep original path
            else:
                if not drawing_root:
                    raise ValueError(
                        "drawingRootPath is required when selectedDrawingPaths are relative."
                    )
                absolute_path = os.path.abspath(os.path.join(drawing_root, raw_path))
                try:
                    if os.path.commonpath([drawing_root, absolute_path]) != drawing_root:
                        raise ValueError(
                            f"Drawing path '{raw_path}' resolves outside the drawing root."
                        )
                except ValueError:
                    raise ValueError(
                        f"Drawing path '{raw_path}' could not be resolved under the drawing root."
                    )
                relative_path = raw_path

            normalized_key = absolute_path.lower()
            if normalized_key in seen_paths:
                continue
            seen_paths.add(normalized_key)
            drawings.append(
                {
                    "path": absolute_path,
                    "relativePath": relative_path,
                    "drawingName": os.path.basename(absolute_path) or absolute_path,
                    "drawingNumber": Path(relative_path).stem,
                    "drawingKey": _normalize_drawing_key(Path(relative_path).stem),
                }
            )

        if not drawings:
            raise ValueError("No valid project drawings were resolved from the issue set.")
        return drawings

    def _validate_schedule_rows(payload: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        strip_rows = payload.get("stripRows")
        if not isinstance(strip_rows, list) or len(strip_rows) == 0:
            raise ValueError("stripRows must contain at least one TerminalStrips row.")
        connection_rows = payload.get("connectionRows")
        if connection_rows is None:
            normalized_connections: List[Dict[str, Any]] = []
        elif isinstance(connection_rows, list):
            normalized_connections = [entry for entry in connection_rows if isinstance(entry, dict)]
        else:
            raise ValueError("connectionRows must be an array when provided.")

        normalized_strips = [entry for entry in strip_rows if isinstance(entry, dict)]
        if len(normalized_strips) == 0:
            raise ValueError("stripRows must contain at least one valid schedule row.")
        if len(normalized_strips) + len(normalized_connections) > MAX_SCHEDULE_ROWS:
            raise ValueError(f"Too many terminal schedule rows. Maximum is {MAX_SCHEDULE_ROWS}")
        return normalized_strips, normalized_connections

    def _validate_preview_payload(payload: Dict[str, Any]) -> Tuple[List[Dict[str, str]], List[Dict[str, Any]], List[Dict[str, Any]]]:
        if not _normalize_text(payload.get("projectId")):
            raise ValueError("projectId is required.")
        if not _normalize_text(payload.get("issueSetId")):
            raise ValueError("issueSetId is required.")
        if not _normalize_text(payload.get("scheduleSnapshotId")):
            raise ValueError("scheduleSnapshotId is required.")
        drawings = _resolve_project_drawings(payload)
        strip_rows, connection_rows = _validate_schedule_rows(payload)
        return drawings, strip_rows, connection_rows

    def _validate_apply_payload(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
        if not _normalize_text(payload.get("projectId")):
            raise ValueError("projectId is required.")
        if not _normalize_text(payload.get("issueSetId")):
            raise ValueError("issueSetId is required.")
        if not _normalize_text(payload.get("scheduleSnapshotId")):
            raise ValueError("scheduleSnapshotId is required.")

        operations = payload.get("operations")
        if not isinstance(operations, list) or len(operations) == 0:
            raise ValueError("operations must contain at least one approved preview row.")
        normalized = [entry for entry in operations if isinstance(entry, dict)]
        if len(normalized) == 0:
            raise ValueError("operations must contain at least one valid approved preview row.")
        if len(normalized) > MAX_APPLY_OPERATIONS:
            raise ValueError(
                f"Too many terminal authoring operations. Maximum is {MAX_APPLY_OPERATIONS}"
            )
        return normalized

    def _project_root_realpath(project_root: Path) -> str:
        return os.path.realpath(str(project_root))

    def _ensure_under_project_root(project_root: Path, candidate_path: str) -> Path:
        project_root_real = _project_root_realpath(project_root)
        candidate_real = os.path.realpath(candidate_path)
        try:
            if os.path.commonpath([project_root_real, candidate_real]) != project_root_real:
                raise ValueError
        except ValueError as exc:
            raise ValueError("Path resolves outside the project root.") from exc
        return Path(candidate_real)

    def _find_acade_project_file(project_root: str, configured_path: str) -> str | None:
        normalized_path = _normalize_text(configured_path)
        if not normalized_path:
            return None
        normalized_root = _normalize_text(project_root)
        if not normalized_root:
            return None
        if not is_absolute_path_value(normalized_root):
            return None
        root_path = resolve_runtime_directory(normalized_root)
        if root_path is None:
            return None
        try:
            root_path = root_path.resolve()
        except Exception:
            return None

        try:
            if is_absolute_path_value(normalized_path):
                runtime_candidate = resolve_runtime_path(normalized_path)
                if runtime_candidate is None:
                    return None
                resolved_candidate = _ensure_under_project_root(root_path, str(runtime_candidate))
            else:
                candidate_path = safe_join(str(root_path), normalized_path.replace("\\", "/"))
                if candidate_path is None:
                    return None
                resolved_candidate = _ensure_under_project_root(root_path, candidate_path)
        except ValueError:
            return None
        except Exception:
            return None
        if resolved_candidate.suffix.lower() != ".wdp":
            return None
        if not resolved_candidate.exists() or not resolved_candidate.is_file():
            return None
        return str(resolved_candidate)

    def _build_preview_drawings(drawings: List[Any]) -> List[Dict[str, Any]]:
        output: List[Dict[str, Any]] = []
        for drawing in drawings:
            if not isinstance(drawing, dict):
                continue
            drawing_path = _normalize_text(drawing.get("drawingPath"))
            if not drawing_path:
                continue
            output.append(
                {
                    "drawingPath": drawing_path,
                    "drawingName": _normalize_text(drawing.get("drawingName"))
                    or os.path.basename(drawing_path)
                    or drawing_path,
                    "relativePath": _normalize_nullable_text(drawing.get("relativePath")),
                    "operationCount": max(0, int(drawing.get("operationCount") or 0)),
                    "stripUpdateCount": max(0, int(drawing.get("stripUpdateCount") or 0)),
                    "routeUpsertCount": max(0, int(drawing.get("routeUpsertCount") or 0)),
                    "unresolvedCount": max(0, int(drawing.get("unresolvedCount") or 0)),
                    "warnings": [
                        _normalize_text(entry)
                        for entry in (drawing.get("warnings") or [])
                        if _normalize_text(entry)
                    ],
                }
            )
        return output

    def _register_generated_report(report_path: str, report_dir: str) -> Tuple[str, str]:
        report_id = uuid.uuid4().hex
        report_filename = os.path.basename(report_path)
        generated_reports[report_id] = {
            "path": report_path,
            "dir": report_dir,
            "filename": report_filename,
        }
        return report_id, report_filename

    def export_terminal_authoring_report_to_excel(
        changes: List[Dict[str, Any]],
        drawings: List[Dict[str, Any]],
        warnings: List[str],
    ) -> Tuple[str, str]:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = tempfile.mkdtemp(prefix="terminal_authoring_")
        out_path = os.path.join(out_dir, f"terminal_authoring_report_{timestamp}.xlsx")

        workbook = Workbook()
        change_ws = workbook.active
        change_ws.title = "Operations"
        drawing_ws = workbook.create_sheet("Drawings")
        warning_ws = workbook.create_sheet("Warnings")

        title_fill = PatternFill("solid", fgColor="1E4E5F")
        header_fill = PatternFill("solid", fgColor="2F3640")
        alt_even = PatternFill("solid", fgColor="E8ECEF")
        alt_odd = PatternFill("solid", fgColor="DDE3E8")
        title_font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
        header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        body_font = Font(size=10, color="1F2933", name="Arial")
        border_side = Side(style="thin", color="A9B0B8")
        border_all = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        def write_sheet(
            worksheet: Any,
            *,
            title: str,
            headers: List[str],
            rows: List[List[Any]],
            widths: List[int],
        ) -> None:
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = worksheet.cell(row=1, column=1, value=title)
            title_cell.fill = title_fill
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.border = border_all

            for column_index, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=2, column=column_index, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border_all
                worksheet.column_dimensions[get_column_letter(column_index)].width = widths[column_index - 1]

            for row_index, row in enumerate(rows, start=3):
                row_fill = alt_even if (row_index - 3) % 2 == 0 else alt_odd
                for column_index, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=row_index, column=column_index, value=value)
                    cell.fill = row_fill
                    cell.font = body_font
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = border_all

            worksheet.row_dimensions[1].height = 28
            worksheet.row_dimensions[2].height = 22
            worksheet.freeze_panes = "A3"

        change_rows = [
            [
                _normalize_text(change.get("drawingName") or change.get("file")),
                _normalize_text(change.get("relativePath") or change.get("drawingPath")),
                _normalize_text(change.get("operationType")),
                _normalize_text(change.get("source")),
                _normalize_text(change.get("stripId")),
                _normalize_text(change.get("routeRef")),
                _normalize_text(change.get("before")),
                _normalize_text(change.get("after")),
                _normalize_text(change.get("detail")),
                _normalize_text(change.get("status")),
            ]
            for change in changes
        ]
        drawing_rows = [
            [
                _normalize_text(drawing.get("drawingName")),
                _normalize_text(drawing.get("relativePath") or drawing.get("drawingPath")),
                int(drawing.get("updated") or 0),
                int(drawing.get("stripUpdates") or 0),
                int(drawing.get("routeUpserts") or 0),
                " | ".join(
                    [
                        _normalize_text(entry)
                        for entry in (drawing.get("warnings") or [])
                        if _normalize_text(entry)
                    ]
                ),
            ]
            for drawing in drawings
        ]
        warning_rows = [[warning] for warning in warnings if _normalize_text(warning)]

        write_sheet(
            change_ws,
            title="Terminal Authoring Operations",
            headers=[
                "Drawing",
                "Relative Path",
                "Operation",
                "Source",
                "Strip",
                "Route Ref",
                "Before",
                "After",
                "Detail",
                "Status",
            ],
            rows=change_rows or [["", "", "", "", "", "", "", "", "No applied operations were recorded.", ""]],
            widths=[28, 42, 18, 14, 16, 18, 28, 28, 52, 14],
        )
        write_sheet(
            drawing_ws,
            title="Per-Drawing Summary",
            headers=["Drawing", "Relative Path", "Updated", "Strip Writes", "Route Upserts", "Warnings"],
            rows=drawing_rows or [["", "", 0, 0, 0, "No drawings were changed."]],
            widths=[28, 42, 12, 14, 14, 52],
        )
        write_sheet(
            warning_ws,
            title="Warnings",
            headers=["Warning"],
            rows=warning_rows or [["No warnings."]],
            widths=[120],
        )

        workbook.save(out_path)
        return out_path, out_dir

    @bp.route("/terminal-authoring/project-preview", methods=["POST"])
    @require_user_or_api_key
    @limiter.limit("20 per hour")
    def api_terminal_authoring_project_preview():
        request_id = _request_id()
        try:
            payload = request.get_json(silent=True) or {}
            if not isinstance(payload, dict):
                payload = {}
            drawings, strip_rows, connection_rows = _validate_preview_payload(payload)
            warnings: List[str] = []

            acade_project_file = _find_acade_project_file(
                _normalize_text(payload.get("projectRootPath")),
                _normalize_text(payload.get("acadeProjectFilePath")),
            )
            if not acade_project_file:
                warnings.append(
                    "No AutoCAD Electrical .wdp project file was provided under the project root. ACAD writes can continue, but ACADE context could not be verified."
                )

            host_result = _call_acade_host_action(
                action="suite_terminal_authoring_project_preview",
                payload={
                    "projectId": _normalize_text(payload.get("projectId")),
                    "issueSetId": _normalize_text(payload.get("issueSetId")),
                    "scheduleSnapshotId": _normalize_text(payload.get("scheduleSnapshotId")),
                    "drawings": drawings,
                    "stripRows": strip_rows,
                    "connectionRows": connection_rows,
                },
                request_id=request_id,
            )
            if not host_result.get("success", False):
                status_code = 400 if host_result.get("code") == "INVALID_REQUEST" else 503
                return jsonify(host_result), status_code

            data = host_result.get("data") or {}
            host_warnings = host_result.get("warnings") or []
            response_warnings = warnings + [
                _normalize_text(entry)
                for entry in host_warnings
                if _normalize_text(entry)
            ]
            drawings_summary = _build_preview_drawings(data.get("drawings") or [])
            operations = data.get("operations") or []
            return jsonify(
                {
                    "success": True,
                    "requestId": (host_result.get("meta") or {}).get("requestId") or request_id,
                    "scheduleSnapshotId": _normalize_text(payload.get("scheduleSnapshotId")),
                    "operationCount": max(0, int(data.get("operationCount") or len(operations))),
                    "stripUpdateCount": max(0, int(data.get("stripUpdateCount") or 0)),
                    "routeUpsertCount": max(0, int(data.get("routeUpsertCount") or 0)),
                    "unresolvedCount": max(0, int(data.get("unresolvedCount") or 0)),
                    "warnings": response_warnings,
                    "drawings": drawings_summary,
                    "operations": operations,
                    "message": host_result.get("message") or "Project terminal authoring preview completed.",
                }
            )
        except ValueError as exc:
            return jsonify({"success": False, "error": "Invalid request parameters.", "requestId": request_id}), 400
        except Exception as exc:
            logger.exception("Project terminal authoring preview failed")
            return jsonify({"success": False, "error": "An internal error occurred.", "requestId": request_id}), 500

    @bp.route("/terminal-authoring/project-apply", methods=["POST"])
    @require_user_or_api_key
    @limiter.limit("15 per hour")
    def api_terminal_authoring_project_apply():
        request_id = _request_id()
        try:
            payload = request.get_json(silent=True) or {}
            if not isinstance(payload, dict):
                payload = {}
            operations = _validate_apply_payload(payload)
            warnings: List[str] = []

            acade_project_file = _find_acade_project_file(
                _normalize_text(payload.get("projectRootPath")),
                _normalize_text(payload.get("acadeProjectFilePath")),
            )
            if not acade_project_file:
                warnings.append(
                    "No AutoCAD Electrical .wdp project file was provided under the project root. ACAD writes can continue, but ACADE context could not be verified."
                )

            host_result = _call_acade_host_action(
                action="suite_terminal_authoring_project_apply",
                payload={
                    "projectId": _normalize_text(payload.get("projectId")),
                    "issueSetId": _normalize_text(payload.get("issueSetId")),
                    "scheduleSnapshotId": _normalize_text(payload.get("scheduleSnapshotId")),
                    "projectRootPath": _normalize_text(payload.get("projectRootPath")),
                    "operations": operations,
                },
                request_id=request_id,
            )
            if not host_result.get("success", False):
                status_code = 400 if host_result.get("code") == "INVALID_REQUEST" else 503
                return jsonify(host_result), status_code

            data = host_result.get("data") or {}
            host_warnings = host_result.get("warnings") or []
            response_warnings = warnings + [
                _normalize_text(entry)
                for entry in host_warnings
                if _normalize_text(entry)
            ]
            report_path, report_dir = export_terminal_authoring_report_to_excel(
                changes=data.get("changes") or [],
                drawings=data.get("drawings") or [],
                warnings=response_warnings,
            )
            schedule_cleanup(report_dir)
            report_id, report_filename = _register_generated_report(report_path, report_dir)

            return jsonify(
                {
                    "success": True,
                    "requestId": (host_result.get("meta") or {}).get("requestId") or request_id,
                    "changedDrawingCount": max(0, int(data.get("changedDrawingCount") or 0)),
                    "terminalStripUpdateCount": max(0, int(data.get("terminalStripUpdateCount") or 0)),
                    "managedRouteUpsertCount": max(0, int(data.get("managedRouteUpsertCount") or 0)),
                    "reportId": report_id,
                    "reportFilename": report_filename,
                    "downloadUrl": f"/api/conduit-route/reports/{report_id}",
                    "warnings": response_warnings,
                    "drawings": data.get("drawings") or [],
                    "message": host_result.get("message") or "Project terminal authoring apply completed.",
                }
            )
        except ValueError as exc:
            return jsonify({"success": False, "error": "Invalid request parameters.", "requestId": request_id}), 400
        except Exception as exc:
            logger.exception("Project terminal authoring apply failed")
            return jsonify({"success": False, "error": "An internal error occurred.", "requestId": request_id}), 500

    @bp.route("/reports/<report_id>", methods=["GET"])
    @require_user_or_api_key
    @limiter.limit("40 per hour")
    def api_terminal_authoring_report_download(report_id: str):
        report = generated_reports.get(report_id)
        if not report:
            return jsonify({"success": False, "error": "Report not found."}), 404

        report_path = report.get("path") or ""
        if not report_path or not os.path.exists(report_path):
            generated_reports.pop(report_id, None)
            report_dir = report.get("dir")
            if report_dir:
                shutil.rmtree(report_dir, ignore_errors=True)
            return jsonify({"success": False, "error": "Report is no longer available."}), 404

        return send_file(
            report_path,
            as_attachment=True,
            download_name=report.get("filename") or os.path.basename(report_path),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return bp
