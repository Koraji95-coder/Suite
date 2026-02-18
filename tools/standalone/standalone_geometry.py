"""
Standalone Geometry Module for Coordinates Grabber

Pure Python implementation of geometry extraction and processing,
independent of AutoCAD/COM dependencies.

Supports:
- DXF/DWG file reading (via ezdxf library)
- Geometric calculations (centers, corners, transformations)
- Excel export
- Configuration management

No AutoCAD required - works on any system with Python 3.9+

This is the canonical standalone version - kept in tools/ for backup/reference.
"""

from __future__ import annotations

import math
import os
import json
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

try:
    import ezdxf
except ImportError:
    ezdxf = None  # type: ignore

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Type alias
Point3D = Tuple[float, float, float]


# -------------------------
# Configuration
# -------------------------
@dataclass(frozen=True)
class StandaloneConfig:
    """Configuration for standalone geometry processing."""
    input_file: str
    output_excel: str
    target_layer: str
    extraction_mode: str  # "center" or "corners"
    point_prefix: str = "P"
    start_number: int = 1
    decimal_places: int = 3
    use_blocks: bool = True
    include_modelspace: bool = True
    verbose: bool = False


@dataclass
class StandaloneRow:
    """A single extracted point/corner."""
    point_id: str
    east: float
    north: float
    elevation: float
    corner: Optional[str] = None
    source_type: str = "unknown"
    entity_type: str = ""
    layer: str = ""


def bbox_center(points: List[Point3D]) -> Optional[Point3D]:
    """Calculate bounding box center from points."""
    if not points:
        return None
    if len(points) == 1:
        return points[0]
    
    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    zs = [p[2] for p in points]
    
    return (
        (min(xs) + max(xs)) / 2.0,
        (min(ys) + max(ys)) / 2.0,
        (min(zs) + max(zs)) / 2.0,
    )


def bbox_corners(points: List[Point3D]) -> List[Tuple[Point3D, str]]:
    """Extract 4 corners from bounding box of points."""
    if not points:
        return []
    
    if len(points) == 1:
        p = points[0]
        return [(p, "center")]
    
    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    zs = [p[2] for p in points]
    
    minx, maxx = min(xs), max(xs)
    miny, maxy = min(ys), max(ys)
    z_avg = sum(zs) / len(zs) if zs else 0.0
    
    if abs(minx - maxx) < 1e-12 or abs(miny - maxy) < 1e-12:
        center = ((minx + maxx) / 2.0, (miny + maxy) / 2.0, z_avg)
        return [(center, "center")]
    
    return [
        ((minx, maxy, z_avg), "NW"),
        ((maxx, maxy, z_avg), "NE"),
        ((minx, miny, z_avg), "SW"),
        ((maxx, miny, z_avg), "SE"),
    ]


def extract_entity_points(entity: Any) -> List[Point3D]:
    """Extract all vertices/points from an entity."""
    if entity is None:
        return []
    
    points = []
    ent_type = entity.dxftype()
    
    try:
        if ent_type in ("LWPOLYLINE", "POLYLINE"):
            for vertex in entity.get_points():
                x, y = vertex[0], vertex[1]
                z = vertex[2] if len(vertex) > 2 else 0.0
                points.append((float(x), float(y), float(z)))
        
        elif ent_type == "LINE":
            start = entity.dxf.start
            end = entity.dxf.end
            points.append((float(start[0]), float(start[1]), float(start[2])))
            points.append((float(end[0]), float(end[1]), float(end[2])))
        
        elif ent_type in ("CIRCLE", "ARC"):
            center = entity.dxf.center
            points.append((float(center[0]), float(center[1]), float(center[2])))
        
        elif ent_type == "SPLINE":
            for pt in entity.get_control_points():
                points.append((float(pt[0]), float(pt[1]), float(pt[2])))
        
        elif ent_type == "INSERT":
            ins_pt = entity.dxf.insert
            points.append((float(ins_pt[0]), float(ins_pt[1]), float(ins_pt[2])))
        
        elif ent_type in ("REGION", "SOLID"):
            try:
                for pt in entity.get_points():
                    points.append((float(pt[0]), float(pt[1]), float(pt[2])))
            except:
                if hasattr(entity, 'center'):
                    c = entity.center
                    points.append((float(c[0]), float(c[1]), float(c[2])))
        
        else:
            try:
                bbox = entity.bbox
                if bbox:
                    minpt, maxpt = bbox
                    center_x = (minpt[0] + maxpt[0]) / 2.0
                    center_y = (minpt[1] + maxpt[1]) / 2.0
                    center_z = (minpt[2] + maxpt[2]) / 2.0 if len(minpt) > 2 else 0.0
                    points.append((float(center_x), float(center_y), float(center_z)))
            except:
                pass
    
    except Exception as e:
        if False:  # Quiet mode
            print(f"Warning: Could not extract points from {ent_type}: {e}")
    
    return points


def load_cad_file(filepath: str) -> Optional[Any]:
    """Load a DXF or DWG file using ezdxf."""
    if not ezdxf:
        raise ImportError("ezdxf library required. Install with: pip install ezdxf")
    
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")
    
    try:
        doc = ezdxf.readfile(filepath)
        return doc
    except Exception as e:
        print(f"Error loading file: {e}")
        return None


def get_entities_on_layer(doc: Any, layer_name: str) -> List[Any]:
    """Get all entities on a specific layer from a DXF document."""
    if not doc:
        return []
    
    entities = []
    try:
        mspace = doc.modelspace()
        for entity in mspace.query(f'*[layer=="{layer_name}"]'):
            entities.append(entity)
    except Exception as e:
        print(f"Error querying layer: {e}")
    
    return entities


def extract_geometries_from_file(
    config: StandaloneConfig,
    progress_callback=None,
) -> List[StandaloneRow]:
    """Extract geometry from CAD file and generate coordinate rows."""
    if progress_callback:
        progress_callback(0, 100, "Loading file...")
    
    doc = load_cad_file(config.input_file)
    if not doc:
        raise RuntimeError(f"Could not load file: {config.input_file}")
    
    if progress_callback:
        progress_callback(10, 100, "Querying layer geometry...")
    
    entities = get_entities_on_layer(doc, config.target_layer)
    rows: List[StandaloneRow] = []
    counter = config.start_number
    
    if progress_callback:
        progress_callback(20, 100, f"Processing {len(entities)} entities...")
    
    for idx, entity in enumerate(entities):
        if progress_callback and idx % max(1, len(entities) // 10) == 0:
            progress_callback(20 + (idx * 70) // len(entities), 100, 
                            f"Processing {idx}/{len(entities)}...")
        
        points = extract_entity_points(entity)
        if not points:
            continue
        
        ent_type = entity.dxftype()
        
        if config.extraction_mode == "corners":
            corners = bbox_corners(points)
            for corner_pt, corner_name in corners:
                point_id = f"{config.point_prefix}{counter}_{corner_name}"
                counter += 1
                
                row = StandaloneRow(
                    point_id=point_id,
                    east=corner_pt[0],
                    north=corner_pt[1],
                    elevation=corner_pt[2],
                    corner=corner_name,
                    source_type=f"Corner/{ent_type}",
                    entity_type=ent_type,
                    layer=config.target_layer,
                )
                rows.append(row)
        
        else:
            center = bbox_center(points)
            if center:
                point_id = f"{config.point_prefix}{counter}"
                counter += 1
                
                row = StandaloneRow(
                    point_id=point_id,
                    east=center[0],
                    north=center[1],
                    elevation=center[2],
                    corner=None,
                    source_type=f"Center/{ent_type}",
                    entity_type=ent_type,
                    layer=config.target_layer,
                )
                rows.append(row)
    
    if progress_callback:
        progress_callback(90, 100, "Exporting to Excel...")
    
    return rows


def export_to_excel(config: StandaloneConfig, rows: List[StandaloneRow], output_path: Optional[str] = None) -> str:
    """Export extracted geometries to Excel file."""
    out_path = output_path or config.output_excel
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Coordinates"
    
    headers = ["Point ID", "East (X)", "North (Y)", "Elevation (Z)"]
    if config.extraction_mode == "corners":
        headers.insert(1, "Corner")
    headers.extend(["Entity Type", "Layer"])
    
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    for row in rows:
        row_data = [row.point_id]
        
        if config.extraction_mode == "corners":
            row_data.append(row.corner or "")
        
        row_data.extend([row.east, row.north, row.elevation, row.entity_type, row.layer])
        ws.append(row_data)
    
    num_cols = ["East (X)", "North (Y)", "Elevation (Z)"]
    num_format = f"0.{'0' * config.decimal_places}"
    
    for col_idx, header in enumerate(headers, start=1):
        if header in num_cols:
            for row_idx in range(2, len(rows) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_format
                    cell.alignment = Alignment(horizontal="right")
    
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = len(header)
        
        for row_idx in range(2, len(rows) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 50)
    
    ws.freeze_panes = "A2"
    
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    
    wb.save(out_path)
    return out_path


def process_file(
    config: StandaloneConfig,
    progress_callback=None,
) -> Tuple[List[StandaloneRow], str]:
    """Complete processing pipeline: load file → extract → export."""
    rows = extract_geometries_from_file(config, progress_callback)
    
    if progress_callback:
        progress_callback(90, 100, "Exporting to Excel...")
    
    output_path = export_to_excel(config, rows)
    
    if progress_callback:
        progress_callback(100, 100, "Complete!")
    
    return rows, output_path


if __name__ == "__main__":
    import sys
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Coordinates Grabber - Standalone Mode",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  python standalone_geometry.py input.dxf -l Survey -m corners -o output.xlsx
  python standalone_geometry.py drawing.dxf -l "Box Points" -m center -p CTRL
  python standalone_geometry.py file.dwg -c config.json
        """
    )
    
    parser.add_argument("input", help="Input DXF/DWG file")
    parser.add_argument("-l", "--layer", required=True, help="Target layer name to search")
    parser.add_argument("-o", "--output", help="Output Excel file")
    parser.add_argument("-m", "--mode", choices=["center", "corners"], default="center",
                       help="Extraction mode (default: center)")
    parser.add_argument("-p", "--prefix", default="P", help="Point ID prefix (default: P)")
    parser.add_argument("-s", "--start", type=int, default=1, help="Starting point number (default: 1)")
    parser.add_argument("-d", "--decimals", type=int, default=3, help="Decimal places (default: 3)")
    parser.add_argument("-c", "--config", help="Load config from JSON file")
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbose output")
    
    args = parser.parse_args()
    
    if args.config:
        with open(args.config, 'r') as f:
            cfg_dict = json.load(f)
            config = StandaloneConfig(**cfg_dict)
    else:
        output_file = args.output or os.path.splitext(args.input)[0] + "_extracted.xlsx"
        
        config = StandaloneConfig(
            input_file=args.input,
            output_excel=output_file,
            target_layer=args.layer,
            extraction_mode=args.mode,
            point_prefix=args.prefix,
            start_number=args.start,
            decimal_places=args.decimals,
            verbose=args.verbose,
        )
    
    print(f"Loading: {config.input_file}")
    print(f"Layer: {config.target_layer}")
    print(f"Mode: {config.extraction_mode}")
    
    try:
        def progress_cb(current, total, message):
            print(f"[{current:3d}%] {message}")
        
        rows, output_path = process_file(config, progress_callback=progress_cb)
        
        print(f"\n✓ Extraction complete!")
        print(f"  Points extracted: {len(rows)}")
        print(f"  Output: {output_path}")
        
    except Exception as e:
        print(f"\n✗ Error: {e}")
        sys.exit(1)
