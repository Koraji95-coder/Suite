from __future__ import annotations

import os
import shutil
import tempfile
import unittest
from pathlib import Path
from typing import Any, Dict, List, Optional
from unittest.mock import patch


class TestBuildTemporaryIndexWorkbook(unittest.TestCase):
    def test_creates_xlsx_with_document_rows(self) -> None:
        from backend.route_groups.api_transmittal_pdf_analysis import (
            build_temporary_index_workbook,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "sub", "index.xlsx")
            rows = [
                {"drawing_number": "E1-100", "title": "One-Line Diagram", "revision": "2"},
                {"drawing_number": "E2-200", "title": "Panel Schedule", "revision": "A"},
            ]
            result = build_temporary_index_workbook(output_path=output_path, document_rows=rows)

            self.assertEqual(result, output_path)
            self.assertTrue(os.path.exists(output_path))

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            self.assertIn("Document No.", header)
            self.assertIn("Description", header)
            self.assertIn("Revision", header)
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            self.assertEqual(len(data_rows), 2)
            row_values = [list(r) for r in data_rows]
            self.assertIn("E1-100", row_values[0])
            self.assertIn("One-Line Diagram", row_values[0])

    def test_creates_xlsx_with_empty_document_rows(self) -> None:
        from backend.route_groups.api_transmittal_pdf_analysis import (
            build_temporary_index_workbook,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "empty_index.xlsx")
            result = build_temporary_index_workbook(output_path=output_path, document_rows=[])

            self.assertEqual(result, output_path)
            self.assertTrue(os.path.exists(output_path))

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            ws = wb.active
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            self.assertEqual(data_rows, [])

    def test_normalizes_whitespace_in_row_fields(self) -> None:
        from backend.route_groups.api_transmittal_pdf_analysis import (
            build_temporary_index_workbook,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "index.xlsx")
            rows = [{"drawing_number": "  E1-100  ", "title": "  Floor Plan  ", "revision": "  3  "}]
            build_temporary_index_workbook(output_path=output_path, document_rows=rows)

            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            ws = wb.active
            row_values = list(ws.iter_rows(min_row=2, values_only=True))[0]
            self.assertEqual(row_values[0], "E1-100")
            self.assertEqual(row_values[1], "Floor Plan")
            self.assertEqual(row_values[2], "3")


class TestMaterializeDocumentsForRender(unittest.TestCase):
    def _write_fake_pdf(self, path: str) -> None:
        with open(path, "wb") as f:
            f.write(b"%PDF-1.7 fake content")

    def test_copies_and_renames_files_based_on_document_rows(self) -> None:
        from backend.route_groups.api_transmittal_pdf_analysis import (
            materialize_documents_for_render,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            source_dir = os.path.join(temp_dir, "source")
            output_dir = os.path.join(temp_dir, "output")
            os.makedirs(source_dir)

            drawing_file = os.path.join(source_dir, "sheet-01.pdf")
            self._write_fake_pdf(drawing_file)

            rows = [
                {
                    "file_name": "sheet-01.pdf",
                    "drawing_number": "E1-100",
                    "title": "One-Line Diagram",
                    "revision": "2",
                }
            ]
            result = materialize_documents_for_render(
                source_paths=[drawing_file],
                document_rows=rows,
                output_dir=output_dir,
                project_number="PROJ-00001",
            )

            self.assertEqual(len(result), 1)
            output_name = Path(result[0]).name
            self.assertIn("E1-100", output_name)
            self.assertTrue(os.path.exists(result[0]))

    def test_handles_duplicate_file_names_with_counter(self) -> None:
        from backend.route_groups.api_transmittal_pdf_analysis import (
            materialize_documents_for_render,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            source_dir = os.path.join(temp_dir, "source")
            output_dir = os.path.join(temp_dir, "output")
            os.makedirs(source_dir)

            file1 = os.path.join(source_dir, "sheet-01.pdf")
            file2 = os.path.join(source_dir, "sheet-02.pdf")
            self._write_fake_pdf(file1)
            self._write_fake_pdf(file2)

            rows = [
                {"file_name": "sheet-01.pdf", "drawing_number": "E1-100", "title": "Floor Plan", "revision": "1"},
                {"file_name": "sheet-02.pdf", "drawing_number": "E1-100", "title": "Floor Plan", "revision": "1"},
            ]
            result = materialize_documents_for_render(
                source_paths=[file1, file2],
                document_rows=rows,
                output_dir=output_dir,
                project_number="PROJ-00001",
            )

            self.assertEqual(len(result), 2)
            names = {Path(p).name.lower() for p in result}
            self.assertEqual(len(names), 2)

    def test_preserves_files_with_no_matching_row(self) -> None:
        from backend.route_groups.api_transmittal_pdf_analysis import (
            materialize_documents_for_render,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            source_dir = os.path.join(temp_dir, "source")
            output_dir = os.path.join(temp_dir, "output")
            os.makedirs(source_dir)

            file_path = os.path.join(source_dir, "unknown-doc.pdf")
            self._write_fake_pdf(file_path)

            result = materialize_documents_for_render(
                source_paths=[file_path],
                document_rows=[],
                output_dir=output_dir,
                project_number="PROJ-00001",
            )

            self.assertEqual(len(result), 1)
            self.assertEqual(Path(result[0]).name, "unknown-doc.pdf")
            self.assertTrue(os.path.exists(result[0]))

    def test_r3p_prefixed_drawing_number_is_preserved_as_stem(self) -> None:
        from backend.route_groups.api_transmittal_pdf_analysis import (
            materialize_documents_for_render,
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            source_dir = os.path.join(temp_dir, "source")
            output_dir = os.path.join(temp_dir, "output")
            os.makedirs(source_dir)

            file_path = os.path.join(source_dir, "drawing.pdf")
            self._write_fake_pdf(file_path)

            rows = [
                {
                    "file_name": "drawing.pdf",
                    "drawing_number": "R3P-00001-E1-100",
                    "title": "Single Line",
                    "revision": "0",
                }
            ]
            result = materialize_documents_for_render(
                source_paths=[file_path],
                document_rows=rows,
                output_dir=output_dir,
                project_number="PROJ-00001",
            )

            self.assertEqual(len(result), 1)
            self.assertIn("R3P-00001-E1-100", Path(result[0]).name)


class TestAnalyzePdfTitleBlock(unittest.TestCase):
    def _fake_embedded_lines(self) -> Dict[str, Any]:
        return {
            "lines": [
                {"text": "E1-100", "x": 550.0, "y": 20.0},
                {"text": "One-Line Diagram", "x": 450.0, "y": 40.0},
                {"text": "REV: 3", "x": 520.0, "y": 60.0},
                {"text": "DRAWING NO.", "x": 545.0, "y": 10.0},
            ],
            "source": "embedded_text",
            "page_width": 612.0,
            "page_height": 792.0,
        }

    def _passthrough_model_hints(
        self,
        *,
        lines: List[Dict[str, Any]],
        page_width: float,
        page_height: float,
        zone: str,
        current: Dict[str, Any],
    ) -> Dict[str, Any]:
        return current

    def test_returns_expected_fields_with_mocked_extraction(self) -> None:
        from backend.route_groups.api_transmittal_pdf_analysis import analyze_pdf_title_block

        with patch(
            "backend.route_groups.api_transmittal_pdf_analysis._extract_embedded_text_lines",
            return_value=self._fake_embedded_lines(),
        ), patch(
            "backend.route_groups.api_transmittal_pdf_analysis._apply_titleblock_model_hints",
            side_effect=self._passthrough_model_hints,
        ):
            result = analyze_pdf_title_block("/fake/path/drawing.pdf")

        self.assertIn("drawing_number", result)
        self.assertIn("title", result)
        self.assertIn("revision", result)
        self.assertIn("confidence", result)
        self.assertIn("source", result)
        self.assertIn("needs_review", result)
        self.assertIn("accepted", result)
        self.assertIn("recognition", result)
        self.assertIn("fields", result)
        self.assertIsInstance(result["confidence"], float)

    def test_returns_needs_review_when_confidence_below_threshold(self) -> None:
        from backend.route_groups.api_transmittal_pdf_analysis import analyze_pdf_title_block

        with patch(
            "backend.route_groups.api_transmittal_pdf_analysis._extract_embedded_text_lines",
            return_value={"lines": [], "source": "embedded_text", "page_width": 612.0, "page_height": 792.0},
        ), patch(
            "backend.route_groups.api_transmittal_pdf_analysis._apply_titleblock_model_hints",
            side_effect=self._passthrough_model_hints,
        ):
            result = analyze_pdf_title_block("/fake/path/empty.pdf")

        self.assertTrue(result["needs_review"])
        self.assertFalse(result["accepted"])

    def test_result_source_reflects_extraction_source(self) -> None:
        from backend.route_groups.api_transmittal_pdf_analysis import analyze_pdf_title_block

        payload = self._fake_embedded_lines()
        payload["source"] = "embedded_text"
        with patch(
            "backend.route_groups.api_transmittal_pdf_analysis._extract_embedded_text_lines",
            return_value=payload,
        ), patch(
            "backend.route_groups.api_transmittal_pdf_analysis._apply_titleblock_model_hints",
            side_effect=self._passthrough_model_hints,
        ):
            result = analyze_pdf_title_block("/fake/path/drawing.pdf")

        self.assertEqual(result["source"], "embedded_text")


if __name__ == "__main__":
    unittest.main()
