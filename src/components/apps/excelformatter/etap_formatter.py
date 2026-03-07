#!/usr/bin/env python3
"""
ETAP Report Auto-Formatter
Universal tool to clean and professionally format ETAP short-circuit exports.
Handles all sheet types: raw MAX sheets, COMP sheets, SC tables, and summary tables.
"""

import sys
import os
import re
import copy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import pandas as pd


# ── Style Constants ──────────────────────────────────────────────────────────
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(name=FONT_NAME, bold=True, color="1F4E79", size=9)
DATA_FONT = Font(name=FONT_NAME, size=9)
DATA_FONT_BOLD = Font(name=FONT_NAME, size=9, bold=True)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11, color="2E75B6")
META_FONT = Font(name=FONT_NAME, size=9, color="404040")
META_FONT_BOLD = Font(name=FONT_NAME, size=9, bold=True, color="1F4E79")
THIN_BORDER = Side(style="thin", color="B4C6E7")
MEDIUM_BORDER = Side(style="medium", color="1F4E79")
CELL_BORDER = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)
HEADER_BORDER = Border(left=THIN_BORDER, right=THIN_BORDER, top=MEDIUM_BORDER, bottom=MEDIUM_BORDER)
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F7FB")
NUM_FMT_CURRENT = '0.000'
NUM_FMT_XR = '0.0000'
NUM_FMT_KV = '0.00'


def clean_string(val):
    """Strip extra whitespace from string values."""
    if isinstance(val, str):
        return re.sub(r'\s+', ' ', val).strip()
    return val


def extract_metadata(ws):
    """Extract project metadata from ETAP header rows."""
    meta = {}
    for row in ws.iter_rows(min_row=1, max_row=16, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                v = cell.value.strip()
                if 'Project:' in v: meta['Project'] = v.split('Project:')[-1].strip()
                elif 'Location:' in v: meta['Location'] = v.split('Location:')[-1].strip()
                elif 'Engineer:' in v: meta['Engineer'] = v.split('Engineer:')[-1].strip()
                elif 'Date:' in v: meta['Date'] = v.split('Date:')[-1].strip()
                elif 'Study Case:' in v: meta['Study Case'] = v.split('Study Case:')[-1].strip()
                elif 'Filename:' in v: meta['Filename'] = v.split('Filename:')[-1].strip()
                elif 'Config.:' in v: meta['Config'] = v.split('Config.:')[-1].strip()
                elif 'Contract:' in v: meta['Contract'] = v.split('Contract:')[-1].strip()
                elif 'Revision:' in v: meta['Revision'] = v.split('Revision:')[-1].strip()
                elif 'Short-Circuit' in v: meta['Report Title'] = v
                elif 'Cycle' in v and 'Fault' in v: meta['Study Type'] = v
                elif 'Prefault' in v: meta['Prefault'] = v
    return meta


def extract_raw_max_data(ws):
    """Extract data from raw ETAP MAX sheets (MOM_MAX, INT_MAX, 30CYC_MAX)."""
    rows_data = []
    for row in ws.iter_rows(min_row=20, values_only=False):
        cell_c = row[2] if len(row) > 2 else None  # Column C = Bus ID
        cell_f = row[5] if len(row) > 5 else None  # Column F = kV
        if cell_c and cell_c.value and isinstance(cell_c.value, str) and cell_f and isinstance(cell_f.value, (int, float)):
            bus_id = clean_string(cell_c.value)
            kv = cell_f.value
            # 3-Phase: J(Real), O(Imag), S(Mag)
            r3p = row[9].value if len(row) > 9 else None
            i3p = row[14].value if len(row) > 14 else None
            m3p = row[18].value if len(row) > 18 else None
            # LG: Y(Real), AD(Imag), AH(Mag)
            rlg = row[24].value if len(row) > 24 else None
            ilg = row[29].value if len(row) > 29 else None
            mlg = row[33].value if len(row) > 33 else None
            # LL: AM(Real), AR(Imag), AW(Mag)
            rll = row[38].value if len(row) > 38 else None
            ill = row[43].value if len(row) > 43 else None
            mll = row[48].value if len(row) > 48 else None
            # LLG: BB(Real), BF(Imag), BH(Mag)
            rllg = row[53].value if len(row) > 53 else None
            illg = row[57].value if len(row) > 57 else None
            mllg = row[59].value if len(row) > 59 else None
            rows_data.append([bus_id, kv, r3p, i3p, m3p, rlg, ilg, mlg, rll, ill, mll, rllg, illg, mllg])
    cols = ['Bus ID', 'kV', '3Φ Real', '3Φ Imag', '3Φ Mag',
            'LG Real', 'LG Imag', 'LG Mag',
            'LL Real', 'LL Imag', 'LL Mag',
            'LLG Real', 'LLG Imag', 'LLG Mag']
    return pd.DataFrame(rows_data, columns=cols) if rows_data else pd.DataFrame()


def extract_comp_data(ws):
    """Extract data from COMP sheets (A=ID, D=kV, E-G=3ph, I-K=LG, M-O=LL, Q-S=LLG)."""
    rows_data = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        if row[0] and isinstance(row[0], str) and len(row) > 3 and isinstance(row[3], (int, float)):
            bus_id = clean_string(row[0])
            kv = row[3]
            vals = [bus_id, kv]
            for idx in [4, 5, 6, 8, 9, 10, 12, 13, 14, 16, 17, 18]:
                vals.append(row[idx] if len(row) > idx else None)
            rows_data.append(vals)
    if not rows_data:
        return pd.DataFrame()
    cols = ['Bus ID', 'kV', '3Φ Real', '3Φ Imag', '3Φ Mag',
            'LG Real', 'LG Imag', 'LG Mag',
            'LL Real', 'LL Imag', 'LL Mag',
            'LLG Real', 'LLG Imag', 'LLG Mag']
    return pd.DataFrame(rows_data, columns=cols)


def extract_table_data(ws):
    """Extract data from summary table sheets (LV_SC, HV_MV_SC, MAX_MOM)."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return pd.DataFrame(), [], []
    header1 = [clean_string(c) if c else '' for c in rows[0]]
    header2 = [clean_string(c) if c else '' for c in rows[1]]
    data_rows = []
    for row in rows[2:]:
        if row[0] and isinstance(row[0], str):
            data_rows.append([clean_string(c) if isinstance(c, str) else c for c in row])
    max_cols = max(len(header2), max((len(r) for r in data_rows), default=0)) if data_rows else len(header2)
    for r in data_rows:
        while len(r) < max_cols:
            r.append(None)
    header2 = header2 + [''] * (max_cols - len(header2))
    header1 = header1 + [''] * (max_cols - len(header1))
    return pd.DataFrame(data_rows), header1, header2


def apply_table_style(ws, start_row, start_col, df, headers, group_headers=None, num_fmt=NUM_FMT_CURRENT):
    """Write a DataFrame as a professionally formatted table."""
    max_col = start_col + len(headers) - 1
    
    # Group headers (spanning row)
    if group_headers:
        col = start_col
        for label, span in group_headers:
            cell = ws.cell(row=start_row, column=col, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = HEADER_BORDER
            if span > 1:
                ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row, end_column=col + span - 1)
                for c in range(col + 1, col + span):
                    ws.cell(row=start_row, column=c).fill = HEADER_FILL
                    ws.cell(row=start_row, column=c).border = HEADER_BORDER
            col += span
        start_row += 1

    # Column headers
    for i, h in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i, value=h)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = HEADER_BORDER
    start_row += 1

    # Data rows
    for r_idx, (_, row) in enumerate(df.iterrows()):
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)
            cell.font = DATA_FONT_BOLD if c_idx == 0 else DATA_FONT
            cell.border = CELL_BORDER
            if r_idx % 2 == 1:
                cell.fill = ALT_ROW_FILL
            if isinstance(val, float):
                if c_idx == 1 or (headers[c_idx] if c_idx < len(headers) else '').lower().startswith('kv'):
                    cell.number_format = NUM_FMT_KV
                elif 'x/r' in (headers[c_idx] if c_idx < len(headers) else '').lower():
                    cell.number_format = NUM_FMT_XR
                else:
                    cell.number_format = NUM_FMT_CURRENT
                cell.alignment = Alignment(horizontal='center')
            elif c_idx == 0:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')
    
    return start_row + len(df)


def write_metadata_header(ws, meta, title, start_row=1):
    """Write a professional metadata header block."""
    ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
    
    row = start_row + 1
    subtitle = meta.get('Study Type', '')
    if subtitle:
        ws.cell(row=row, column=1, value=subtitle).font = SUBTITLE_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

    prefault = meta.get('Prefault', '')
    if prefault:
        ws.cell(row=row, column=1, value=prefault).font = META_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

    row += 1  # spacer
    meta_items = [('Project', 'Project'), ('Location', 'Location'), ('Engineer', 'Engineer'),
                  ('Date', 'Date'), ('Study Case', 'Study Case'), ('Config', 'Config'),
                  ('Filename', 'Filename'), ('Revision', 'Revision')]
    col_pairs = [(1, 2, 4), (6, 7, 8)]  # (label_col, val_start, val_end)
    idx = 0
    for i in range(0, len(meta_items), 2):
        for j, (key, label) in enumerate(meta_items[i:i+2]):
            if idx >= len(col_pairs):
                break
            lc, vs, ve = col_pairs[j]
            if key in meta:
                ws.cell(row=row, column=lc, value=f"{label}:").font = META_FONT_BOLD
                ws.cell(row=row, column=vs, value=meta[key]).font = META_FONT
                ws.merge_cells(start_row=row, start_column=vs, end_row=row, end_column=ve)
        row += 1
        if i + 2 >= len(meta_items):
            break
    
    # Separator line
    row += 1
    for c in range(1, 16):
        ws.cell(row=row, column=c).border = Border(bottom=Side(style="medium", color="1F4E79"))
    return row + 2


def auto_width(ws, min_width=8, max_width=18):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = min(cell_len + 2, max_width)
        ws.column_dimensions[col_letter].width = max_len


def format_raw_max_sheet(wb_out, sheet_name, df, meta):
    """Format a raw MAX sheet into a clean table."""
    ws = wb_out.create_sheet(sheet_name)
    data_row = write_metadata_header(ws, meta, meta.get('Report Title', 'Short-Circuit Summary Report'))
    
    group_headers = [
        ('', 2),  # Bus ID + kV
        ('3-Phase Fault (kA)', 3),
        ('Line-to-Ground Fault (kA)', 3),
        ('Line-to-Line Fault (kA)', 3),
        ('Line-to-Line-to-Ground (kA)', 3),
    ]
    headers = ['Bus ID', 'kV', 'Real', 'Imag', 'Mag',
               'Real', 'Imag', 'Mag', 'Real', 'Imag', 'Mag',
               'Real', 'Imag', 'Mag']
    
    apply_table_style(ws, data_row, 1, df, headers, group_headers)
    auto_width(ws)
    ws.sheet_properties.tabColor = "1F4E79"


def format_comp_sheet(wb_out, sheet_name, df, meta):
    """Format a COMP sheet."""
    ws = wb_out.create_sheet(sheet_name)
    data_row = write_metadata_header(ws, meta, meta.get('Report Title', 'Short-Circuit Comparison'))
    
    group_headers = [
        ('', 2),
        ('3-Phase Fault (kA)', 3),
        ('Line-to-Ground Fault (kA)', 3),
        ('Line-to-Line Fault (kA)', 3),
        ('Line-to-Line-to-Ground (kA)', 3),
    ]
    headers = list(df.columns)
    apply_table_style(ws, data_row, 1, df, headers, group_headers)
    auto_width(ws)
    ws.sheet_properties.tabColor = "2E75B6"


def format_summary_table(wb_out, sheet_name, df, h1, h2, tab_color="4472C4"):
    """Format a summary table sheet."""
    ws = wb_out.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=sheet_name.replace('_', ' ')).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(len(h2), 14))
    
    # Build group headers from h1
    groups = []
    i = 0
    while i < len(h1):
        label = h1[i]
        span = 1
        while i + span < len(h1) and h1[i + span] == '':
            span += 1
        groups.append((label, span))
        i += span
    
    apply_table_style(ws, 3, 1, df, h2, groups if any(g[0] for g in groups) else None)
    auto_width(ws)
    ws.sheet_properties.tabColor = tab_color


def process_etap_file(input_path, output_path):
    """Main processing: read ETAP file, format all sheets, write output."""
    print(f"Loading {input_path}...")
    wb = load_workbook(input_path, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)  # Remove default sheet
    
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  Processing sheet: {name}")
        
        name_upper = name.upper().replace(' ', '_')
        
        if any(k in name_upper for k in ['MOM_MAX', 'INT_MAX', '30CYC_MAX']):
            meta = extract_metadata(ws)
            df = extract_raw_max_data(ws)
            if not df.empty:
                format_raw_max_sheet(wb_out, name, df, meta)
                print(f"    → {len(df)} rows extracted and formatted")
            else:
                print(f"    → No data found, skipping")
                
        elif 'COMP' in name_upper:
            meta = extract_metadata(ws) if ws.max_column > 10 else {}
            df = extract_comp_data(ws)
            if not df.empty:
                format_comp_sheet(wb_out, name, df, meta)
                print(f"    → {len(df)} rows extracted and formatted")
            else:
                print(f"    → No data found, skipping")
                
        elif 'TABLE' in name_upper or 'SC' in name_upper:
            df, h1, h2 = extract_table_data(ws)
            if not df.empty:
                color = "70AD47" if 'LV' in name_upper else "ED7D31" if 'HV' in name_upper or 'MV' in name_upper else "4472C4"
                format_summary_table(wb_out, name, df, h1, h2, color)
                print(f"    → {len(df)} rows extracted and formatted")
            else:
                print(f"    → No data found, skipping")
        else:
            # Generic: try table extraction
            df, h1, h2 = extract_table_data(ws)
            if not df.empty:
                format_summary_table(wb_out, name, df, h1, h2)
                print(f"    → {len(df)} rows formatted (generic)")
            else:
                print(f"    → Unrecognized format, skipping")
    
    # Add a cover/index sheet
    idx_ws = wb_out.create_sheet("INDEX", 0)
    idx_ws.cell(row=1, column=1, value="ETAP Short-Circuit Report").font = Font(name=FONT_NAME, bold=True, size=18, color="1F4E79")
    idx_ws.merge_cells("A1:F1")
    idx_ws.cell(row=2, column=1, value="Auto-Formatted Report").font = Font(name=FONT_NAME, size=12, color="808080")
    idx_ws.merge_cells("A2:F2")
    
    row = 4
    idx_ws.cell(row=row, column=1, value="Sheet Index").font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E79")
    row += 1
    for i, sn in enumerate(wb_out.sheetnames):
        if sn == "INDEX":
            continue
        cell = idx_ws.cell(row=row, column=1, value=f"{i}.")
        cell.font = DATA_FONT
        cell = idx_ws.cell(row=row, column=2, value=sn)
        cell.font = Font(name=FONT_NAME, size=10, color="2E75B6", underline="single")
        row += 1
    
    auto_width(idx_ws)
    idx_ws.sheet_properties.tabColor = "333333"
    
    wb_out.save(output_path)
    print(f"\nFormatted report saved to: {output_path}")


if __name__ == "__main__":
    inp = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/template.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "/mnt/user-data/outputs/ETAP_Formatted_Report.xlsx"
    os.makedirs(os.path.dirname(out), exist_ok=True)
    process_etap_file(inp, out)
