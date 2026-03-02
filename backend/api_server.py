#!/usr/bin/env python3
# pyright: reportMissingImports=false, reportMissingModuleSource=false
"""
Coordinates Grabber API Server
Flask-based HTTP/WebSocket bridge between React frontend and AutoCAD COM interface

Uses LATE-BOUND COM (dynamic dispatch) to avoid gen_py cache corruption.
Pattern taken from coordtable_excel_always_place_refpoints.py.

This server runs on localhost:5000 and provides:
- AutoCAD process detection (checks for acad.exe)
- COM connection management
- Layer and selection information
- Coordinate extraction from layers
- Real-time status updates

Usage:
    python api_server.py

Requirements:
    pip install flask flask-cors psutil pywin32
"""

from flask import Flask, jsonify, request, send_file, after_this_request, g
from flask_cors import CORS
from flask_sock import Sock
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import psutil
try:
    import pythoncom
    import win32com.client
    import win32com.client.gencache as gencache
    AUTOCAD_COM_AVAILABLE = True
except Exception:
    pythoncom = None
    win32com = None
    gencache = None
    AUTOCAD_COM_AVAILABLE = False
import threading
import time
import json
import math
import os
import sys
import tempfile
import shutil
import subprocess
import re
import traceback
import logging
import hmac
import hashlib
import zipfile
import secrets
from functools import wraps
from datetime import datetime
from typing import Optional, Dict, Any, List, Set, Tuple
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
import jwt
from jwt import PyJWKClient

# ── Logging configuration ────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[
        logging.FileHandler('api_server.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ── gen_py cache fix (from coordtable) ──────────────────────────
# Prevent gen_py from writing wrappers that cause CDispatch issues
if AUTOCAD_COM_AVAILABLE and gencache is not None:
    gencache.is_readonly = True

# ── Environment loading ────────────────────────────────────────
def _load_env_file(path: Path) -> None:
    if not path.exists():
        return
    try:
        for raw in path.read_text().splitlines():
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            key = key.strip()
            if not key:
                continue
            value = value.strip().strip('"').strip("'")
            os.environ.setdefault(key, value)
    except Exception as exc:
        logger.warning("Failed to load env file %s: %s", path, exc)


_env_path = Path(__file__).resolve().parents[1] / ".env"
_load_env_file(_env_path)

app = Flask(__name__)
sock = Sock(app)

# ── Transmittal Builder render helpers ──────────────────────────
TRANSMITTAL_RENDER_AVAILABLE = False
TRANSMITTAL_BUILDER_DIR = Path(__file__).resolve().parent / "Transmittal-Builder"
TRANSMITTAL_TEMPLATE_PATH = (
    TRANSMITTAL_BUILDER_DIR / "R3P-PRJ#-XMTL-001 - DOCUMENT INDEX.docx"
)
TRANSMITTAL_CONFIG_PATH = TRANSMITTAL_BUILDER_DIR / "config.yaml"

TRANSMITTAL_FALLBACK_PROFILES: List[Dict[str, str]] = [
    {
        "id": "sample-engineer",
        "name": "Sample Engineer, PE",
        "title": "Engineering Lead",
        "email": "engineer@example.com",
        "phone": "(000) 000-0000",
    }
]
TRANSMITTAL_FALLBACK_FIRMS = ["TX - Firm #00000"]

TRANSMITTAL_PROFILES_CACHE: Dict[str, Any] = {
    "mtime": None,
    "payload": None,
}
TRANSMITTAL_PROFILES_CACHE_LOCK = threading.Lock()

try:
    transmittal_core_path = TRANSMITTAL_BUILDER_DIR / "core"
    if transmittal_core_path.exists():
        sys.path.append(str(transmittal_core_path))
        from transmittal_render import render_transmittal, render_cid_transmittal  # type: ignore

        TRANSMITTAL_RENDER_AVAILABLE = True
except Exception as exc:
    logger.warning("Transmittal render helpers unavailable: %s", exc)


def _slugify_transmittal_profile_id(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", "-", value.strip().lower())
    normalized = normalized.strip("-")
    return normalized[:64]


def _normalize_transmittal_profile(
    row: Dict[str, Any], fallback_index: int
) -> Optional[Dict[str, str]]:
    name = str(row.get("name") or "").strip()
    if not name:
        return None

    profile_id = str(row.get("id") or "").strip()
    if not profile_id:
        profile_id = _slugify_transmittal_profile_id(name) or f"profile-{fallback_index}"
    else:
        profile_id = _slugify_transmittal_profile_id(profile_id) or f"profile-{fallback_index}"

    title = str(row.get("title") or "").strip()[:120]
    email = str(row.get("email") or "").strip().lower()[:254]
    if email and not _is_valid_email(email):
        email = ""
    phone = str(row.get("phone") or "").strip()[:64]

    return {
        "id": profile_id,
        "name": name[:120],
        "title": title,
        "email": email,
        "phone": phone,
    }


def _load_transmittal_profiles_payload() -> Dict[str, Any]:
    cfg_mtime = None
    try:
        if TRANSMITTAL_CONFIG_PATH.exists():
            cfg_mtime = TRANSMITTAL_CONFIG_PATH.stat().st_mtime
    except Exception:
        cfg_mtime = None

    with TRANSMITTAL_PROFILES_CACHE_LOCK:
        cached = TRANSMITTAL_PROFILES_CACHE.get("payload")
        if cached and TRANSMITTAL_PROFILES_CACHE.get("mtime") == cfg_mtime:
            return cached

        raw_cfg: Dict[str, Any] = {}
        if TRANSMITTAL_CONFIG_PATH.exists():
            try:
                raw_text = TRANSMITTAL_CONFIG_PATH.read_text(encoding="utf-8")
                if raw_text.strip():
                    try:
                        import yaml  # type: ignore

                        loaded = yaml.safe_load(raw_text)
                        if isinstance(loaded, dict):
                            raw_cfg = loaded
                    except Exception:
                        parsed = json.loads(raw_text)
                        if isinstance(parsed, dict):
                            raw_cfg = parsed
            except Exception as exc:
                logger.warning("Failed to load transmittal config yaml: %s", exc)

        business = raw_cfg.get("business", {})
        ui = raw_cfg.get("ui", {})

        raw_profiles = business.get("pe_profiles", [])
        normalized_profiles: List[Dict[str, str]] = []
        seen_ids: Set[str] = set()
        if isinstance(raw_profiles, list):
            for index, row in enumerate(raw_profiles, start=1):
                if not isinstance(row, dict):
                    continue
                normalized = _normalize_transmittal_profile(row, index)
                if not normalized:
                    continue
                base_id = normalized["id"]
                dedupe_id = base_id
                suffix = 2
                while dedupe_id in seen_ids:
                    dedupe_id = f"{base_id}-{suffix}"
                    suffix += 1
                normalized["id"] = dedupe_id
                seen_ids.add(dedupe_id)
                normalized_profiles.append(normalized)

        if not normalized_profiles:
            normalized_profiles = [dict(item) for item in TRANSMITTAL_FALLBACK_PROFILES]

        raw_firms = business.get("firm_numbers", [])
        firm_numbers: List[str] = []
        seen_firms: Set[str] = set()
        if isinstance(raw_firms, list):
            for value in raw_firms:
                firm = str(value or "").strip()[:80]
                if not firm or firm in seen_firms:
                    continue
                seen_firms.add(firm)
                firm_numbers.append(firm)

        if not firm_numbers:
            firm_numbers = list(TRANSMITTAL_FALLBACK_FIRMS)

        default_profile = str(ui.get("default_pe") or "").strip()
        default_profile_id = ""
        if default_profile:
            for profile in normalized_profiles:
                if default_profile in {profile["id"], profile["name"]}:
                    default_profile_id = profile["id"]
                    break
        if not default_profile_id:
            default_profile_id = normalized_profiles[0]["id"]

        default_firm = str(ui.get("default_firm") or "").strip()
        if default_firm not in firm_numbers:
            default_firm = firm_numbers[0]

        payload = {
            "profiles": normalized_profiles,
            "firm_numbers": firm_numbers,
            "defaults": {
                "profile_id": default_profile_id,
                "firm": default_firm,
            },
            "source": str(TRANSMITTAL_CONFIG_PATH),
        }

        TRANSMITTAL_PROFILES_CACHE["mtime"] = cfg_mtime
        TRANSMITTAL_PROFILES_CACHE["payload"] = payload
        return payload


def _parse_csv_env(var_name: str, fallback: List[str]) -> List[str]:
    raw = os.environ.get(var_name, "")
    if not raw.strip():
        return fallback
    return [item.strip() for item in raw.split(",") if item.strip()]


def _parse_int_env(var_name: str, fallback: int, minimum: int = 1) -> int:
    raw = os.environ.get(var_name)
    if raw is None:
        return fallback
    try:
        value = int(raw)
        return max(value, minimum)
    except ValueError:
        logger.warning("Invalid %s=%r; using fallback %s", var_name, raw, fallback)
        return fallback


def _parse_bool_env(var_name: str, fallback: bool = False) -> bool:
    raw = os.environ.get(var_name)
    if raw is None:
        return fallback
    normalized = str(raw).strip().lower()
    if normalized in {"1", "true", "yes", "on"}:
        return True
    if normalized in {"0", "false", "no", "off"}:
        return False
    logger.warning("Invalid %s=%r; using fallback %s", var_name, raw, fallback)
    return fallback


EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PAIRING_CODE_PATTERN = re.compile(r"^\d{6}$")
PAIRING_CHALLENGE_ID_PATTERN = re.compile(r"^[A-Za-z0-9_-]{20,200}$")
PASSKEY_CALLBACK_STATE_PATTERN = re.compile(r"^[A-Za-z0-9_-]{20,200}$")
PASSKEY_CALLBACK_SIGNATURE_PATTERN = re.compile(r"^[A-Fa-f0-9]{64}$")
PASSKEY_CALLBACK_TIMESTAMP_PATTERN = re.compile(r"^\d{10,13}$")


app.config['MAX_CONTENT_LENGTH'] = _parse_int_env(
    'API_MAX_CONTENT_LENGTH',
    104857600  # 100 MB
)

# CORS configuration - restrict to specific origins for security
# In production, replace with actual frontend domain
ALLOWED_ORIGINS = [
    "http://localhost:5173",  # Vite dev server
    "http://localhost:3000",  # Alternative dev port
    "http://127.0.0.1:5173",
    "http://127.0.0.1:3000",
]
ALLOWED_ORIGINS = _parse_csv_env('API_ALLOWED_ORIGINS', ALLOWED_ORIGINS)
AUTH_ALLOWED_REDIRECT_ORIGINS = _parse_csv_env(
    "AUTH_ALLOWED_REDIRECT_ORIGINS",
    ALLOWED_ORIGINS,
)
AUTH_EMAIL_REDIRECT_URL = (
    (os.environ.get("AUTH_EMAIL_REDIRECT_URL") or "").strip()
    or (os.environ.get("VITE_AUTH_REDIRECT_URL") or "").strip()
)

CORS(app, 
     origins=ALLOWED_ORIGINS,
     supports_credentials=True,
     methods=["GET", "POST", "OPTIONS"],
     allow_headers=["Content-Type", "Authorization", "X-API-Key"])

# ── Rate Limiting ────────────────────────────────────────────────
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=[
        os.environ.get('API_RATE_LIMIT_DAY', '200 per day'),
        os.environ.get('API_RATE_LIMIT_HOUR', '50 per hour'),
    ],
    storage_uri="memory://",
    strategy="fixed-window"
)

# ── Security Headers ─────────────────────────────────────────────
@app.after_request
def add_security_headers(response):
    """Add security headers to all responses."""
    # Prevent clickjacking attacks
    response.headers['X-Frame-Options'] = 'DENY'
    # Prevent MIME-type sniffing
    response.headers['X-Content-Type-Options'] = 'nosniff'
    # Enable XSS protection
    response.headers['X-XSS-Protection'] = '1; mode=block'
    # Content Security Policy
    response.headers['Content-Security-Policy'] = "default-src 'self'"
    # Referrer policy
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # Strict Transport Security (HTTPS only - comment out for localhost)
    # response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# ── API Authentication ───────────────────────────────────────────
# SECURITY: API_KEY must be explicitly set in environment. No defaults allowed.
API_KEY = (os.environ.get('API_KEY') or '').strip()
if not API_KEY:
    fallback_key = (os.environ.get('VITE_API_KEY') or '').strip()
    if fallback_key:
        API_KEY = fallback_key
        logger.warning(
            "API_KEY not set; using VITE_API_KEY for local development. "
            "Set API_KEY explicitly for production."
        )
    else:
        raise RuntimeError(
            "FATAL: API_KEY environment variable is not set.\n"
            "Please set your API key before starting the server:\n"
            "  export API_KEY='your-secure-api-key-here'\n"
            "Then start the server again."
        )
if len(API_KEY) < 16:
    logger.warning("API_KEY length is under 16 characters; use a longer key for production.")

BATCH_SESSION_COOKIE = "bfr_session"
BATCH_SESSION_TTL_SECONDS = _parse_int_env("BATCH_SESSION_TTL_SECONDS", 6 * 60 * 60, minimum=300)

# ── Agent Broker + Supabase Auth ────────────────────────────────
SUPABASE_URL = (os.environ.get("SUPABASE_URL") or "").strip()
if not SUPABASE_URL:
    dev_supabase_url = (os.environ.get("VITE_SUPABASE_URL") or "").strip()
    if dev_supabase_url:
        SUPABASE_URL = dev_supabase_url
        logger.warning(
            "SUPABASE_URL not set; using VITE_SUPABASE_URL for development. "
            "Set SUPABASE_URL in backend env for production."
        )

SUPABASE_SERVICE_ROLE_KEY = (os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or "").strip()
SUPABASE_ANON_KEY = (os.environ.get("SUPABASE_ANON_KEY") or "").strip()
SUPABASE_JWT_SECRET = (os.environ.get("SUPABASE_JWT_SECRET") or "").strip()
SUPABASE_API_KEY = SUPABASE_SERVICE_ROLE_KEY or SUPABASE_ANON_KEY
if SUPABASE_API_KEY == SUPABASE_ANON_KEY and SUPABASE_API_KEY:
    logger.warning(
        "Using SUPABASE_ANON_KEY for backend auth lookups. "
        "Prefer SUPABASE_SERVICE_ROLE_KEY in production."
    )

SUPABASE_JWKS_URL = (
    SUPABASE_URL.rstrip("/") + "/auth/v1/keys"
    if SUPABASE_URL
    else ""
)
_SUPABASE_JWKS_CLIENT: Optional[PyJWKClient] = None

AGENT_GATEWAY_URL = (
    (os.environ.get("AGENT_GATEWAY_URL") or "").strip()
    or (os.environ.get("VITE_AGENT_GATEWAY_URL") or "").strip()
    or "http://127.0.0.1:3000"
)
AGENT_WEBHOOK_SECRET = (os.environ.get("AGENT_WEBHOOK_SECRET") or "").strip()
AGENT_REQUIRE_WEBHOOK_SECRET = (
    (os.environ.get("AGENT_REQUIRE_WEBHOOK_SECRET") or "true").strip().lower()
    != "false"
)
if not AGENT_WEBHOOK_SECRET:
    dev_webhook_secret = (os.environ.get("VITE_AGENT_WEBHOOK_SECRET") or "").strip()
    if dev_webhook_secret:
        AGENT_WEBHOOK_SECRET = dev_webhook_secret
        logger.warning(
            "AGENT_WEBHOOK_SECRET not set; using VITE_AGENT_WEBHOOK_SECRET for development. "
            "Set AGENT_WEBHOOK_SECRET in backend env for production."
        )

AGENT_SESSION_COOKIE = (os.environ.get("AGENT_SESSION_COOKIE") or "suite_agent_session").strip()
AGENT_SESSION_SAMESITE = (os.environ.get("AGENT_SESSION_SAMESITE") or "Strict").strip()
AGENT_SESSION_SECURE = (os.environ.get("AGENT_SESSION_SECURE") or "false").strip().lower() == "true"
AGENT_SESSION_TTL_SECONDS = _parse_int_env(
    "AGENT_SESSION_TTL_SECONDS",
    6 * 60 * 60,
    minimum=300,
)
AGENT_DEFAULT_TIMEOUT_SECONDS = _parse_int_env("AGENT_TIMEOUT_SECONDS", 30, minimum=3)
AGENT_MAX_TIMEOUT_SECONDS = _parse_int_env("AGENT_MAX_TIMEOUT_SECONDS", 300, minimum=30)

AGENT_SESSIONS: Dict[str, Dict[str, Any]] = {}
AGENT_PAIRING_CHALLENGES: Dict[str, Dict[str, Any]] = {}
AGENT_PAIRING_CHALLENGE_LOCK = threading.Lock()
AGENT_PAIRING_CHALLENGE_TTL_SECONDS = _parse_int_env(
    "AGENT_PAIRING_CHALLENGE_TTL_SECONDS",
    15 * 60,
    minimum=60,
)
AGENT_PAIRING_CHALLENGE_MAX_ENTRIES = _parse_int_env(
    "AGENT_PAIRING_CHALLENGE_MAX_ENTRIES",
    10_000,
    minimum=100,
)
AGENT_PAIRING_REDIRECT_PATH = (
    (os.environ.get("AGENT_PAIRING_REDIRECT_PATH") or "/app/agent").strip()
    or "/app/agent"
)
AUTH_PASSKEY_ENABLED = _parse_bool_env("AUTH_PASSKEY_ENABLED", False)
AUTH_PASSKEY_PROVIDER = (
    (os.environ.get("AUTH_PASSKEY_PROVIDER") or "supabase").strip().lower()
    or "supabase"
)
if AUTH_PASSKEY_PROVIDER not in {"supabase", "external"}:
    logger.warning(
        "Unsupported AUTH_PASSKEY_PROVIDER=%r; falling back to 'supabase'.",
        AUTH_PASSKEY_PROVIDER,
    )
    AUTH_PASSKEY_PROVIDER = "supabase"
AUTH_PASSKEY_EXTERNAL_NAME = (
    (os.environ.get("AUTH_PASSKEY_EXTERNAL_NAME") or "External IdP").strip()
    or "External IdP"
)
AUTH_PASSKEY_EXTERNAL_DISCOVERY_URL = (
    (os.environ.get("AUTH_PASSKEY_EXTERNAL_DISCOVERY_URL") or "").strip()
)
AUTH_PASSKEY_EXTERNAL_SIGNIN_URL = (
    (os.environ.get("AUTH_PASSKEY_EXTERNAL_SIGNIN_URL") or "").strip()
)
AUTH_PASSKEY_EXTERNAL_ENROLL_URL = (
    (os.environ.get("AUTH_PASSKEY_EXTERNAL_ENROLL_URL") or "").strip()
)
AUTH_PASSKEY_CALLBACK_STATE_TTL_SECONDS = _parse_int_env(
    "AUTH_PASSKEY_CALLBACK_STATE_TTL_SECONDS",
    15 * 60,
    minimum=60,
)
AUTH_PASSKEY_CALLBACK_STATE_MAX_ENTRIES = _parse_int_env(
    "AUTH_PASSKEY_CALLBACK_STATE_MAX_ENTRIES",
    10_000,
    minimum=100,
)
AUTH_PASSKEY_REQUIRE_SIGNED_CALLBACK = _parse_bool_env(
    "AUTH_PASSKEY_REQUIRE_SIGNED_CALLBACK",
    AUTH_PASSKEY_PROVIDER == "external",
)
AUTH_PASSKEY_CALLBACK_SIGNING_SECRET = (
    (os.environ.get("AUTH_PASSKEY_CALLBACK_SIGNING_SECRET") or "").strip()
)
AUTH_PASSKEY_CALLBACK_SIGNATURE_MAX_AGE_SECONDS = _parse_int_env(
    "AUTH_PASSKEY_CALLBACK_SIGNATURE_MAX_AGE_SECONDS",
    300,
    minimum=30,
)
AUTH_PASSKEY_CALLBACK_SIGNATURE_MAX_CLOCK_SKEW_SECONDS = _parse_int_env(
    "AUTH_PASSKEY_CALLBACK_SIGNATURE_MAX_CLOCK_SKEW_SECONDS",
    90,
    minimum=0,
)
AGENT_PAIRING_ACTION_WINDOW_SECONDS = _parse_int_env(
    "AGENT_PAIRING_ACTION_WINDOW_SECONDS",
    900,
    minimum=60,
)
AGENT_PAIRING_ACTION_MAX_ATTEMPTS = _parse_int_env(
    "AGENT_PAIRING_ACTION_MAX_ATTEMPTS",
    8,
    minimum=1,
)
AGENT_PAIRING_ACTION_MIN_INTERVAL_SECONDS = _parse_int_env(
    "AGENT_PAIRING_ACTION_MIN_INTERVAL_SECONDS",
    20,
    minimum=0,
)
AGENT_PAIRING_ACTION_BLOCK_SECONDS = _parse_int_env(
    "AGENT_PAIRING_ACTION_BLOCK_SECONDS",
    1800,
    minimum=60,
)
AGENT_PAIRING_CONFIRM_FAILURE_WINDOW_SECONDS = _parse_int_env(
    "AGENT_PAIRING_CONFIRM_FAILURE_WINDOW_SECONDS",
    900,
    minimum=60,
)
AGENT_PAIRING_CONFIRM_FAILURE_MAX_ATTEMPTS = _parse_int_env(
    "AGENT_PAIRING_CONFIRM_FAILURE_MAX_ATTEMPTS",
    6,
    minimum=1,
)
AGENT_PAIRING_CONFIRM_FAILURE_BLOCK_SECONDS = _parse_int_env(
    "AGENT_PAIRING_CONFIRM_FAILURE_BLOCK_SECONDS",
    1800,
    minimum=60,
)

# ── Auth Email Abuse Controls ────────────────────────────────────
AUTH_EMAIL_WINDOW_SECONDS = _parse_int_env("AUTH_EMAIL_WINDOW_SECONDS", 900, minimum=60)
AUTH_EMAIL_MAX_ATTEMPTS = _parse_int_env("AUTH_EMAIL_MAX_ATTEMPTS", 6, minimum=1)
AUTH_EMAIL_MIN_INTERVAL_SECONDS = _parse_int_env("AUTH_EMAIL_MIN_INTERVAL_SECONDS", 15, minimum=0)
AUTH_EMAIL_BLOCK_SECONDS = _parse_int_env("AUTH_EMAIL_BLOCK_SECONDS", 1800, minimum=60)
AUTH_EMAIL_IP_WINDOW_SECONDS = _parse_int_env(
    "AUTH_EMAIL_IP_WINDOW_SECONDS",
    900,
    minimum=60,
)
AUTH_EMAIL_IP_MAX_ATTEMPTS = _parse_int_env(
    "AUTH_EMAIL_IP_MAX_ATTEMPTS",
    30,
    minimum=1,
)
AUTH_EMAIL_IP_BLOCK_SECONDS = _parse_int_env(
    "AUTH_EMAIL_IP_BLOCK_SECONDS",
    1800,
    minimum=60,
)
AUTH_EMAIL_MIN_RESPONSE_MS = _parse_int_env("AUTH_EMAIL_MIN_RESPONSE_MS", 450, minimum=0)
AUTH_EMAIL_RESPONSE_JITTER_MS = _parse_int_env(
    "AUTH_EMAIL_RESPONSE_JITTER_MS",
    120,
    minimum=0,
)
AUTH_EMAIL_HONEYPOT_FIELD = (os.environ.get("AUTH_EMAIL_HONEYPOT_FIELD") or "company").strip() or "company"
AUTH_EMAIL_TURNSTILE_SECRET = (os.environ.get("AUTH_EMAIL_TURNSTILE_SECRET") or "").strip()
AUTH_EMAIL_TURNSTILE_VERIFY_URL = (
    (os.environ.get("AUTH_EMAIL_TURNSTILE_VERIFY_URL") or "").strip()
    or "https://challenges.cloudflare.com/turnstile/v0/siteverify"
)
AUTH_EMAIL_TURNSTILE_TIMEOUT_SECONDS = _parse_int_env(
    "AUTH_EMAIL_TURNSTILE_TIMEOUT_SECONDS",
    5,
    minimum=1,
)
_default_require_turnstile = "true" if AUTH_EMAIL_TURNSTILE_SECRET else "false"
AUTH_EMAIL_REQUIRE_TURNSTILE = (
    (os.environ.get("AUTH_EMAIL_REQUIRE_TURNSTILE") or _default_require_turnstile)
    .strip()
    .lower()
    != "false"
)

AUTH_EMAIL_WINDOW: Dict[str, List[float]] = {}
AUTH_EMAIL_LAST_ATTEMPT: Dict[str, float] = {}
AUTH_EMAIL_BLOCKED_UNTIL: Dict[str, float] = {}
AUTH_EMAIL_IP_WINDOW: Dict[str, List[float]] = {}
AUTH_EMAIL_IP_BLOCKED_UNTIL: Dict[str, float] = {}
AUTH_EMAIL_ABUSE_LOCK = threading.Lock()
PASSKEY_CALLBACK_STATES: Dict[str, Dict[str, Any]] = {}
PASSKEY_CALLBACK_STATES_LOCK = threading.Lock()
AGENT_PAIRING_ACTION_WINDOW: Dict[str, List[float]] = {}
AGENT_PAIRING_ACTION_LAST_ATTEMPT: Dict[str, float] = {}
AGENT_PAIRING_ACTION_BLOCKED_UNTIL: Dict[str, float] = {}
AGENT_PAIRING_ACTION_ABUSE_LOCK = threading.Lock()
AGENT_PAIRING_CONFIRM_FAILURE_WINDOW: Dict[str, List[float]] = {}
AGENT_PAIRING_CONFIRM_BLOCKED_UNTIL: Dict[str, float] = {}
AGENT_PAIRING_CONFIRM_ABUSE_LOCK = threading.Lock()


def _purge_expired_agent_sessions() -> None:
    now = time.time()
    expired = [sid for sid, session in AGENT_SESSIONS.items() if session["expires_at"] <= now]
    for sid in expired:
        AGENT_SESSIONS.pop(sid, None)


def _purge_expired_agent_pairing_challenges(now: Optional[float] = None) -> None:
    ts = time.time() if now is None else now
    expired = [
        cid
        for cid, challenge in AGENT_PAIRING_CHALLENGES.items()
        if challenge.get("expires_at", 0) <= ts
    ]
    for cid in expired:
        AGENT_PAIRING_CHALLENGES.pop(cid, None)

    overflow = len(AGENT_PAIRING_CHALLENGES) - AGENT_PAIRING_CHALLENGE_MAX_ENTRIES
    if overflow <= 0:
        return

    oldest = sorted(
        AGENT_PAIRING_CHALLENGES.items(),
        key=lambda item: item[1].get("created_at", 0),
    )[:overflow]
    for cid, _ in oldest:
        AGENT_PAIRING_CHALLENGES.pop(cid, None)


def _create_agent_pairing_challenge(
    action: str,
    user_id: str,
    email: str,
    pairing_code: str,
    client_ip: str,
) -> Tuple[str, int]:
    now = time.time()
    expires_at = int(now) + AGENT_PAIRING_CHALLENGE_TTL_SECONDS
    challenge_id = secrets.token_urlsafe(32)
    payload = {
        "action": action,
        "user_id": user_id,
        "email": email.strip().lower(),
        "pairing_code": pairing_code,
        "created_at": now,
        "expires_at": expires_at,
        "client_ip": client_ip,
    }

    with AGENT_PAIRING_CHALLENGE_LOCK:
        _purge_expired_agent_pairing_challenges(now)
        AGENT_PAIRING_CHALLENGES[challenge_id] = payload

    return challenge_id, expires_at


def _consume_agent_pairing_challenge(
    challenge_id: str,
    user_id: str,
    email: str,
) -> Tuple[Optional[Dict[str, Any]], str]:
    now = time.time()
    normalized_email = email.strip().lower()

    with AGENT_PAIRING_CHALLENGE_LOCK:
        _purge_expired_agent_pairing_challenges(now)
        challenge = AGENT_PAIRING_CHALLENGES.get(challenge_id)
        if not challenge:
            return None, "missing"

        if challenge.get("expires_at", 0) <= now:
            AGENT_PAIRING_CHALLENGES.pop(challenge_id, None)
            return None, "expired"

        if challenge.get("user_id") != user_id:
            return None, "user-mismatch"

        if normalized_email and challenge.get("email") != normalized_email:
            return None, "email-mismatch"

        AGENT_PAIRING_CHALLENGES.pop(challenge_id, None)
        return challenge, "ok"


def _purge_expired_passkey_callback_states(now: Optional[float] = None) -> None:
    ts = time.time() if now is None else now
    expired = [
        state
        for state, entry in PASSKEY_CALLBACK_STATES.items()
        if entry.get("expires_at", 0) <= ts
    ]
    for state in expired:
        PASSKEY_CALLBACK_STATES.pop(state, None)

    overflow = len(PASSKEY_CALLBACK_STATES) - AUTH_PASSKEY_CALLBACK_STATE_MAX_ENTRIES
    if overflow <= 0:
        return

    oldest = sorted(
        PASSKEY_CALLBACK_STATES.items(),
        key=lambda item: item[1].get("created_at", 0),
    )[:overflow]
    for state, _ in oldest:
        PASSKEY_CALLBACK_STATES.pop(state, None)


def _create_passkey_callback_state(
    intent: str,
    client_ip: str,
    client_redirect_to: str = "",
    user_id: str = "",
    email: str = "",
) -> Tuple[str, int]:
    now = time.time()
    expires_at = int(now) + AUTH_PASSKEY_CALLBACK_STATE_TTL_SECONDS
    state = secrets.token_urlsafe(32)
    payload = {
        "intent": intent,
        "user_id": user_id.strip(),
        "email": email.strip().lower(),
        "client_ip": client_ip.strip(),
        "client_redirect_to": client_redirect_to.strip(),
        "created_at": now,
        "expires_at": expires_at,
    }

    with PASSKEY_CALLBACK_STATES_LOCK:
        _purge_expired_passkey_callback_states(now)
        PASSKEY_CALLBACK_STATES[state] = payload

    return state, expires_at


def _consume_passkey_callback_state(state: str) -> Tuple[Optional[Dict[str, Any]], str]:
    now = time.time()
    with PASSKEY_CALLBACK_STATES_LOCK:
        _purge_expired_passkey_callback_states(now)
        payload = PASSKEY_CALLBACK_STATES.get(state)
        if not payload:
            return None, "missing"

        if payload.get("expires_at", 0) <= now:
            PASSKEY_CALLBACK_STATES.pop(state, None)
            return None, "expired"

        PASSKEY_CALLBACK_STATES.pop(state, None)
        return payload, "ok"


def _get_passkey_callback_state(state: str) -> Tuple[Optional[Dict[str, Any]], str]:
    now = time.time()
    with PASSKEY_CALLBACK_STATES_LOCK:
        _purge_expired_passkey_callback_states(now)
        payload = PASSKEY_CALLBACK_STATES.get(state)
        if not payload:
            return None, "missing"

        if payload.get("expires_at", 0) <= now:
            PASSKEY_CALLBACK_STATES.pop(state, None)
            return None, "expired"

        return dict(payload), "ok"


def _normalize_passkey_callback_timestamp(raw_value: str) -> Optional[int]:
    value = str(raw_value or "").strip()
    if not PASSKEY_CALLBACK_TIMESTAMP_PATTERN.match(value):
        return None

    parsed = int(value)
    if parsed > 10_000_000_000:
        parsed = parsed // 1000

    if parsed <= 0:
        return None
    return parsed


def _build_passkey_callback_signature_payload(
    state: str,
    intent: str,
    status: str,
    email: str,
    error_message: str,
    timestamp: int,
) -> str:
    return "\n".join(
        [
            state.strip(),
            intent.strip().lower(),
            status.strip().lower(),
            email.strip().lower(),
            error_message.replace("\r", " ").replace("\n", " ").strip(),
            str(int(timestamp)),
        ]
    )


def _verify_passkey_callback_signature(
    state: str,
    intent: str,
    status: str,
    email: str,
    error_message: str,
    signature: str,
    timestamp_raw: str,
) -> Tuple[bool, str]:
    if not AUTH_PASSKEY_REQUIRE_SIGNED_CALLBACK:
        return True, "disabled"

    if not AUTH_PASSKEY_CALLBACK_SIGNING_SECRET:
        return False, "missing-secret"

    normalized_signature = str(signature or "").strip().lower()
    if not PASSKEY_CALLBACK_SIGNATURE_PATTERN.match(normalized_signature):
        return False, "invalid-signature-format"

    timestamp = _normalize_passkey_callback_timestamp(timestamp_raw)
    if timestamp is None:
        return False, "invalid-timestamp-format"

    now = int(time.time())
    if timestamp > (now + AUTH_PASSKEY_CALLBACK_SIGNATURE_MAX_CLOCK_SKEW_SECONDS):
        return False, "timestamp-in-future"

    max_age = AUTH_PASSKEY_CALLBACK_SIGNATURE_MAX_AGE_SECONDS
    if max_age > 0 and (now - timestamp) > max_age:
        return False, "timestamp-expired"

    payload = _build_passkey_callback_signature_payload(
        state=state,
        intent=intent,
        status=status,
        email=email,
        error_message=error_message,
        timestamp=timestamp,
    )
    expected_signature = hmac.new(
        AUTH_PASSKEY_CALLBACK_SIGNING_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()

    if not hmac.compare_digest(expected_signature, normalized_signature):
        return False, "signature-mismatch"

    return True, "ok"


def _agent_pairing_action_key(user_id: str, action: str) -> str:
    return f"{user_id}:{action}"


def _compact_agent_pairing_action_state(now: float) -> None:
    if len(AGENT_PAIRING_ACTION_LAST_ATTEMPT) < 5000:
        return

    stale_before = now - max(300.0, AGENT_PAIRING_ACTION_WINDOW_SECONDS * 4)
    for key, last_seen in list(AGENT_PAIRING_ACTION_LAST_ATTEMPT.items()):
        blocked_until = AGENT_PAIRING_ACTION_BLOCKED_UNTIL.get(key, 0.0)
        if last_seen >= stale_before or blocked_until > now:
            continue
        AGENT_PAIRING_ACTION_LAST_ATTEMPT.pop(key, None)
        AGENT_PAIRING_ACTION_WINDOW.pop(key, None)
        AGENT_PAIRING_ACTION_BLOCKED_UNTIL.pop(key, None)


def _is_agent_pairing_action_allowed(user_id: str, action: str) -> Tuple[bool, str, int]:
    key = _agent_pairing_action_key(user_id, action)
    now = time.time()

    with AGENT_PAIRING_ACTION_ABUSE_LOCK:
        _compact_agent_pairing_action_state(now)

        blocked_until = AGENT_PAIRING_ACTION_BLOCKED_UNTIL.get(key, 0.0)
        if blocked_until > now:
            retry_after = max(1, int(math.ceil(blocked_until - now)))
            return False, "blocked", retry_after

        window = [
            ts
            for ts in AGENT_PAIRING_ACTION_WINDOW.get(key, [])
            if (now - ts) <= AGENT_PAIRING_ACTION_WINDOW_SECONDS
        ]
        last_attempt = AGENT_PAIRING_ACTION_LAST_ATTEMPT.get(key, 0.0)

        if (
            AGENT_PAIRING_ACTION_MIN_INTERVAL_SECONDS > 0
            and last_attempt > 0
            and (now - last_attempt) < AGENT_PAIRING_ACTION_MIN_INTERVAL_SECONDS
        ):
            window.append(now)
            AGENT_PAIRING_ACTION_WINDOW[key] = window
            AGENT_PAIRING_ACTION_LAST_ATTEMPT[key] = now
            if len(window) > AGENT_PAIRING_ACTION_MAX_ATTEMPTS:
                AGENT_PAIRING_ACTION_BLOCKED_UNTIL[key] = now + AGENT_PAIRING_ACTION_BLOCK_SECONDS
                return False, "window-limit", AGENT_PAIRING_ACTION_BLOCK_SECONDS
            retry_after = max(
                1,
                int(
                    math.ceil(
                        AGENT_PAIRING_ACTION_MIN_INTERVAL_SECONDS - (now - last_attempt)
                    )
                ),
            )
            return False, "min-interval", retry_after

        window.append(now)
        AGENT_PAIRING_ACTION_WINDOW[key] = window
        AGENT_PAIRING_ACTION_LAST_ATTEMPT[key] = now

        if len(window) > AGENT_PAIRING_ACTION_MAX_ATTEMPTS:
            AGENT_PAIRING_ACTION_BLOCKED_UNTIL[key] = now + AGENT_PAIRING_ACTION_BLOCK_SECONDS
            return False, "window-limit", AGENT_PAIRING_ACTION_BLOCK_SECONDS

    return True, "ok", 0


def _agent_pairing_confirm_key(user_id: str, client_ip: str) -> str:
    return f"{user_id}:{client_ip or 'unknown'}"


def _compact_agent_pairing_confirm_state(now: float) -> None:
    if len(AGENT_PAIRING_CONFIRM_FAILURE_WINDOW) < 5000:
        return

    stale_before = now - max(300.0, AGENT_PAIRING_CONFIRM_FAILURE_WINDOW_SECONDS * 4)
    for key, attempts in list(AGENT_PAIRING_CONFIRM_FAILURE_WINDOW.items()):
        blocked_until = AGENT_PAIRING_CONFIRM_BLOCKED_UNTIL.get(key, 0.0)
        recent_attempts = [
            ts
            for ts in attempts
            if (now - ts) <= AGENT_PAIRING_CONFIRM_FAILURE_WINDOW_SECONDS
        ]
        if recent_attempts or blocked_until > now:
            AGENT_PAIRING_CONFIRM_FAILURE_WINDOW[key] = recent_attempts
            continue
        AGENT_PAIRING_CONFIRM_FAILURE_WINDOW.pop(key, None)
        AGENT_PAIRING_CONFIRM_BLOCKED_UNTIL.pop(key, None)


def _is_agent_pairing_confirm_blocked(user_id: str, client_ip: str) -> Tuple[bool, int]:
    key = _agent_pairing_confirm_key(user_id, client_ip)
    now = time.time()
    with AGENT_PAIRING_CONFIRM_ABUSE_LOCK:
        _compact_agent_pairing_confirm_state(now)
        blocked_until = AGENT_PAIRING_CONFIRM_BLOCKED_UNTIL.get(key, 0.0)
        if blocked_until <= now:
            return False, 0
        retry_after = max(1, int(math.ceil(blocked_until - now)))
        return True, retry_after


def _register_agent_pairing_confirm_failure(user_id: str, client_ip: str) -> Tuple[bool, int]:
    key = _agent_pairing_confirm_key(user_id, client_ip)
    now = time.time()

    with AGENT_PAIRING_CONFIRM_ABUSE_LOCK:
        _compact_agent_pairing_confirm_state(now)

        blocked_until = AGENT_PAIRING_CONFIRM_BLOCKED_UNTIL.get(key, 0.0)
        if blocked_until > now:
            retry_after = max(1, int(math.ceil(blocked_until - now)))
            return True, retry_after

        attempts = [
            ts
            for ts in AGENT_PAIRING_CONFIRM_FAILURE_WINDOW.get(key, [])
            if (now - ts) <= AGENT_PAIRING_CONFIRM_FAILURE_WINDOW_SECONDS
        ]
        attempts.append(now)
        AGENT_PAIRING_CONFIRM_FAILURE_WINDOW[key] = attempts

        if len(attempts) > AGENT_PAIRING_CONFIRM_FAILURE_MAX_ATTEMPTS:
            AGENT_PAIRING_CONFIRM_BLOCKED_UNTIL[key] = now + AGENT_PAIRING_CONFIRM_FAILURE_BLOCK_SECONDS
            return True, AGENT_PAIRING_CONFIRM_FAILURE_BLOCK_SECONDS

    return False, 0


def _clear_agent_pairing_confirm_failures(user_id: str, client_ip: str) -> None:
    key = _agent_pairing_confirm_key(user_id, client_ip)
    with AGENT_PAIRING_CONFIRM_ABUSE_LOCK:
        AGENT_PAIRING_CONFIRM_FAILURE_WINDOW.pop(key, None)
        AGENT_PAIRING_CONFIRM_BLOCKED_UNTIL.pop(key, None)


def _looks_like_uuid(value: str) -> bool:
    if not value:
        return False
    if len(value) != 36:
        return False
    if value.count("-") != 4:
        return False
    return True


def _get_supabase_jwks_client() -> Optional[PyJWKClient]:
    global _SUPABASE_JWKS_CLIENT
    if not SUPABASE_JWKS_URL:
        return None
    if _SUPABASE_JWKS_CLIENT is None:
        _SUPABASE_JWKS_CLIENT = PyJWKClient(SUPABASE_JWKS_URL)
    return _SUPABASE_JWKS_CLIENT


def _agent_broker_config_status() -> Dict[str, Any]:
    missing: List[str] = []
    warnings: List[str] = []

    if not SUPABASE_URL:
        missing.append("SUPABASE_URL")
    if not AGENT_GATEWAY_URL:
        missing.append("AGENT_GATEWAY_URL")
    if AGENT_REQUIRE_WEBHOOK_SECRET and not AGENT_WEBHOOK_SECRET:
        missing.append("AGENT_WEBHOOK_SECRET")

    if SUPABASE_JWT_SECRET and _looks_like_uuid(SUPABASE_JWT_SECRET):
        warnings.append(
            "SUPABASE_JWT_SECRET looks like a key ID. For ECC keys, leave it empty and use JWKS."
        )

    if SUPABASE_API_KEY and SUPABASE_API_KEY == SUPABASE_ANON_KEY:
        warnings.append("Using SUPABASE_ANON_KEY for backend auth. Prefer service role key.")

    return {
        "ok": len(missing) == 0,
        "missing": missing,
        "warnings": warnings,
        "require_webhook_secret": AGENT_REQUIRE_WEBHOOK_SECRET,
    }


def _auth_passkey_capability() -> Dict[str, Any]:
    config_missing: List[str] = []
    warnings: List[str] = []
    provider_label = "Supabase"
    handlers_ready = False

    if AUTH_PASSKEY_PROVIDER == "supabase":
        provider_label = "Supabase"
        if not SUPABASE_URL:
            config_missing.append("SUPABASE_URL")
    elif AUTH_PASSKEY_PROVIDER == "external":
        provider_label = AUTH_PASSKEY_EXTERNAL_NAME
        if not AUTH_PASSKEY_EXTERNAL_SIGNIN_URL:
            config_missing.append("AUTH_PASSKEY_EXTERNAL_SIGNIN_URL")
        elif not _normalize_absolute_http_url(AUTH_PASSKEY_EXTERNAL_SIGNIN_URL):
            config_missing.append("AUTH_PASSKEY_EXTERNAL_SIGNIN_URL (must be absolute http(s) URL)")
        if AUTH_PASSKEY_REQUIRE_SIGNED_CALLBACK and not AUTH_PASSKEY_CALLBACK_SIGNING_SECRET:
            config_missing.append("AUTH_PASSKEY_CALLBACK_SIGNING_SECRET")
        if not AUTH_PASSKEY_REQUIRE_SIGNED_CALLBACK:
            warnings.append(
                "AUTH_PASSKEY_REQUIRE_SIGNED_CALLBACK=false; external callback trust is reduced."
            )
        if (
            AUTH_PASSKEY_EXTERNAL_DISCOVERY_URL
            and not AUTH_PASSKEY_EXTERNAL_DISCOVERY_URL.startswith(("http://", "https://"))
        ):
            warnings.append("AUTH_PASSKEY_EXTERNAL_DISCOVERY_URL must be an absolute http(s) URL.")

    if AUTH_PASSKEY_PROVIDER == "supabase":
        warnings.append(
            "Passkey enrollment/login handlers are not wired in this build yet."
        )
    else:
        warnings.append("External provider redirect flow is enabled when configured.")

    config_ready = len(config_missing) == 0
    if not AUTH_PASSKEY_ENABLED:
        rollout_state = "disabled"
        next_step = "Set AUTH_PASSKEY_ENABLED=true and restart backend."
    elif not config_ready:
        rollout_state = "needs-config"
        missing_list = ", ".join(config_missing)
        next_step = f"Set missing passkey config: {missing_list}."
    else:
        if AUTH_PASSKEY_PROVIDER == "external":
            handlers_ready = True
            rollout_state = "ready"
            next_step = "External provider passkey start handlers are ready."
        else:
            rollout_state = "planned"
            next_step = "Provider selected, but passkey handlers are not available in this build."

    return {
        "enabled": AUTH_PASSKEY_ENABLED,
        "provider": AUTH_PASSKEY_PROVIDER,
        "provider_label": provider_label,
        "rollout_state": rollout_state,
        "handlers_ready": handlers_ready,
        "signed_callback_required": AUTH_PASSKEY_REQUIRE_SIGNED_CALLBACK,
        "config_ready": config_ready,
        "config_missing": config_missing,
        "warnings": warnings,
        "next_step": next_step,
    }


def _get_supabase_user_id(user: Dict[str, Any]) -> Optional[str]:
    return (user.get("id") or user.get("sub") or "").strip() or None


def _get_supabase_user_email(user: Dict[str, Any]) -> Optional[str]:
    email = str(user.get("email") or "").strip().lower()
    if not _is_valid_email(email):
        return None
    return email


def _get_bearer_token() -> Optional[str]:
    auth = request.headers.get("Authorization", "")
    if not auth:
        return None
    if auth.lower().startswith("bearer "):
        return auth.split(" ", 1)[1].strip()
    return None


def _get_request_ip() -> str:
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if forwarded_for:
        first_hop = forwarded_for.split(",", 1)[0].strip()
        if first_hop:
            return first_hop
    return (request.remote_addr or "").strip() or "unknown"


def _email_fingerprint(email: str, length: int = 12) -> str:
    digest = hashlib.sha256(email.encode("utf-8")).hexdigest()
    return digest[: max(6, length)]


def _auth_email_key(email: str, client_ip: str) -> str:
    return f"{client_ip}:{_email_fingerprint(email, length=24)}"


def _auth_email_ip_key(client_ip: str) -> str:
    return client_ip or "unknown"


def _compact_auth_email_state(now: float) -> None:
    if len(AUTH_EMAIL_LAST_ATTEMPT) < 5000:
        if len(AUTH_EMAIL_IP_WINDOW) < 5000:
            return

    stale_before = now - max(300.0, AUTH_EMAIL_WINDOW_SECONDS * 4)
    for key, last_seen in list(AUTH_EMAIL_LAST_ATTEMPT.items()):
        blocked_until = AUTH_EMAIL_BLOCKED_UNTIL.get(key, 0.0)
        if last_seen >= stale_before or blocked_until > now:
            continue
        AUTH_EMAIL_LAST_ATTEMPT.pop(key, None)
        AUTH_EMAIL_WINDOW.pop(key, None)
        AUTH_EMAIL_BLOCKED_UNTIL.pop(key, None)

    ip_stale_before = now - max(300.0, AUTH_EMAIL_IP_WINDOW_SECONDS * 4)
    for key, attempts in list(AUTH_EMAIL_IP_WINDOW.items()):
        blocked_until = AUTH_EMAIL_IP_BLOCKED_UNTIL.get(key, 0.0)
        recent_attempts = [
            ts
            for ts in attempts
            if (now - ts) <= AUTH_EMAIL_IP_WINDOW_SECONDS
        ]
        if recent_attempts or blocked_until > now:
            AUTH_EMAIL_IP_WINDOW[key] = recent_attempts
            continue
        if attempts and attempts[-1] >= ip_stale_before:
            AUTH_EMAIL_IP_WINDOW[key] = recent_attempts
            continue
        AUTH_EMAIL_IP_WINDOW.pop(key, None)
        AUTH_EMAIL_IP_BLOCKED_UNTIL.pop(key, None)


def _is_auth_email_request_allowed(email: str, client_ip: str) -> Tuple[bool, str]:
    key = _auth_email_key(email, client_ip)
    ip_key = _auth_email_ip_key(client_ip)
    now = time.time()

    with AUTH_EMAIL_ABUSE_LOCK:
        _compact_auth_email_state(now)

        blocked_until = AUTH_EMAIL_BLOCKED_UNTIL.get(key, 0.0)
        if blocked_until > now:
            return False, "blocked"

        ip_blocked_until = AUTH_EMAIL_IP_BLOCKED_UNTIL.get(ip_key, 0.0)
        if ip_blocked_until > now:
            return False, "ip-blocked"

        window = [
            ts
            for ts in AUTH_EMAIL_WINDOW.get(key, [])
            if (now - ts) <= AUTH_EMAIL_WINDOW_SECONDS
        ]
        last_attempt = AUTH_EMAIL_LAST_ATTEMPT.get(key, 0.0)

        if (
            AUTH_EMAIL_MIN_INTERVAL_SECONDS > 0
            and last_attempt > 0
            and (now - last_attempt) < AUTH_EMAIL_MIN_INTERVAL_SECONDS
        ):
            AUTH_EMAIL_WINDOW[key] = window
            AUTH_EMAIL_LAST_ATTEMPT[key] = now
            if len(window) >= AUTH_EMAIL_MAX_ATTEMPTS:
                AUTH_EMAIL_BLOCKED_UNTIL[key] = now + AUTH_EMAIL_BLOCK_SECONDS
            return False, "min-interval"

        window.append(now)
        AUTH_EMAIL_WINDOW[key] = window
        AUTH_EMAIL_LAST_ATTEMPT[key] = now

        if len(window) > AUTH_EMAIL_MAX_ATTEMPTS:
            AUTH_EMAIL_BLOCKED_UNTIL[key] = now + AUTH_EMAIL_BLOCK_SECONDS
            return False, "window-limit"

        ip_window = [
            ts
            for ts in AUTH_EMAIL_IP_WINDOW.get(ip_key, [])
            if (now - ts) <= AUTH_EMAIL_IP_WINDOW_SECONDS
        ]
        ip_window.append(now)
        AUTH_EMAIL_IP_WINDOW[ip_key] = ip_window
        if len(ip_window) > AUTH_EMAIL_IP_MAX_ATTEMPTS:
            AUTH_EMAIL_IP_BLOCKED_UNTIL[ip_key] = now + AUTH_EMAIL_IP_BLOCK_SECONDS
            return False, "ip-window-limit"

    return True, "ok"


def _auth_email_generic_response() -> Dict[str, Any]:
    return {
        "ok": True,
        "message": "If the email is eligible, a link has been sent.",
    }


def _apply_auth_email_response_floor(start_time: float) -> None:
    target_ms = AUTH_EMAIL_MIN_RESPONSE_MS
    if AUTH_EMAIL_RESPONSE_JITTER_MS > 0:
        target_ms += secrets.randbelow(AUTH_EMAIL_RESPONSE_JITTER_MS + 1)

    if target_ms <= 0:
        return

    elapsed_ms = (time.perf_counter() - start_time) * 1000.0
    remaining_ms = target_ms - elapsed_ms
    if remaining_ms > 0:
        time.sleep(remaining_ms / 1000.0)


def _verify_turnstile_token(token: str, client_ip: str) -> bool:
    if not AUTH_EMAIL_TURNSTILE_SECRET:
        return True

    if not token:
        return False

    try:
        response = requests.post(
            AUTH_EMAIL_TURNSTILE_VERIFY_URL,
            data={
                "secret": AUTH_EMAIL_TURNSTILE_SECRET,
                "response": token,
                "remoteip": client_ip,
            },
            timeout=AUTH_EMAIL_TURNSTILE_TIMEOUT_SECONDS,
        )
        if response.status_code != 200:
            logger.warning(
                "Turnstile verification failed with status=%s",
                response.status_code,
            )
            return False

        payload = response.json() if response.content else {}
        if payload.get("success") is True:
            return True

        logger.warning(
            "Turnstile verification rejected request: codes=%s",
            payload.get("error-codes"),
        )
        return False
    except Exception as exc:
        logger.warning("Turnstile verification error: %s", exc)
        return False


def _is_valid_email(value: str) -> bool:
    if not value:
        return False
    return bool(EMAIL_PATTERN.match(value))


def _normalize_origin(candidate: str) -> Optional[str]:
    try:
        parsed = urlparse(candidate)
    except Exception:
        return None

    if parsed.scheme not in ("http", "https"):
        return None
    if not parsed.netloc:
        return None
    return f"{parsed.scheme}://{parsed.netloc}"


def _normalize_absolute_http_url(candidate: str) -> Optional[str]:
    try:
        parsed = urlparse(candidate)
    except Exception:
        return None

    if parsed.scheme not in ("http", "https"):
        return None
    if not parsed.netloc:
        return None

    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, parsed.query, parsed.fragment))


def _build_auth_redirect_url(
    path: str,
    client_redirect_to: str = "",
    query_params: Optional[Dict[str, str]] = None,
) -> Optional[str]:
    safe_path = path if path.startswith("/") else f"/{path}"
    allowed = {
        origin
        for origin in (
            _normalize_origin(entry.strip())
            for entry in AUTH_ALLOWED_REDIRECT_ORIGINS
        )
        if origin
    }

    candidates = [
        client_redirect_to,
        AUTH_EMAIL_REDIRECT_URL,
        request.headers.get("Origin", "").strip(),
        request.headers.get("Referer", "").strip(),
    ]

    for candidate in candidates:
        origin = _normalize_origin(candidate.strip())
        if not origin:
            continue
        if allowed and origin not in allowed:
            logger.warning("Rejected auth redirect origin outside allowlist: %s", origin)
            continue

        parsed = urlparse(origin)
        query = ""
        if query_params:
            normalized_query = {
                str(key): str(value)
                for key, value in query_params.items()
                if str(key).strip() and str(value).strip()
            }
            if normalized_query:
                query = urlencode(normalized_query)
        return urlunparse((parsed.scheme, parsed.netloc, safe_path, "", query, ""))

    return None


def _build_external_passkey_redirect(
    intent: str,
    state_token: str,
    client_redirect_to: str = "",
) -> Optional[str]:
    normalized_intent = intent.strip().lower()
    if normalized_intent not in {"sign-in", "enroll"}:
        return None
    if not PASSKEY_CALLBACK_STATE_PATTERN.match(state_token):
        return None

    base_url = AUTH_PASSKEY_EXTERNAL_SIGNIN_URL
    if normalized_intent == "enroll":
        base_url = AUTH_PASSKEY_EXTERNAL_ENROLL_URL or AUTH_PASSKEY_EXTERNAL_SIGNIN_URL

    normalized_base = _normalize_absolute_http_url(base_url)
    if not normalized_base:
        return None

    return_path = "/login" if normalized_intent == "sign-in" else "/app/settings"
    suite_return_to = _build_auth_redirect_url(
        return_path,
        client_redirect_to,
        query_params={
            "passkey_state": state_token,
            "passkey_intent": normalized_intent,
        },
    )
    callback_api = _build_auth_redirect_url(
        "/api/auth/passkey/callback/complete",
        client_redirect_to,
    )

    parsed = urlparse(normalized_base)
    query_pairs = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query_pairs["suite_intent"] = normalized_intent
    query_pairs["suite_state"] = state_token
    query_pairs["suite_callback_sig_required"] = (
        "1" if AUTH_PASSKEY_REQUIRE_SIGNED_CALLBACK else "0"
    )
    query_pairs["suite_callback_sig_alg"] = "hmac-sha256"
    query_pairs["suite_callback_sig_payload"] = (
        "state,intent,status,email,error,timestamp"
    )
    query_pairs["suite_callback_sig_max_age_seconds"] = str(
        AUTH_PASSKEY_CALLBACK_SIGNATURE_MAX_AGE_SECONDS
    )
    query_pairs["suite_claims_required"] = "1"
    query_pairs["suite_claims_format"] = "jwt"
    query_pairs["suite_claims_alg"] = "HS256"
    if suite_return_to:
        query_pairs["suite_return_to"] = suite_return_to
    if callback_api:
        query_pairs["suite_callback_api"] = callback_api

    query = urlencode(query_pairs)
    return urlunparse(
        (parsed.scheme, parsed.netloc, parsed.path, parsed.params, query, parsed.fragment)
    )


def _send_supabase_email_link(
    email: str,
    flow: str,
    client_redirect_to: str = "",
    redirect_path: str = "/login",
    redirect_query: Optional[Dict[str, str]] = None,
) -> None:
    if not SUPABASE_URL or not SUPABASE_API_KEY:
        raise RuntimeError("Supabase auth is not configured for backend email auth.")

    if flow not in {"signin", "signup"}:
        raise ValueError("Unsupported email auth flow.")

    redirect_to = _build_auth_redirect_url(
        redirect_path,
        client_redirect_to,
        query_params=redirect_query,
    )

    headers = {
        "Authorization": f"Bearer {SUPABASE_API_KEY}",
        "apikey": SUPABASE_API_KEY,
        "Content-Type": "application/json",
    }

    endpoint = f"{SUPABASE_URL.rstrip('/')}/auth/v1/otp"
    payload: Dict[str, Any] = {
        "email": email,
        "create_user": flow == "signup",
    }
    if redirect_to:
        payload["email_redirect_to"] = redirect_to

    response = requests.post(endpoint, headers=headers, json=payload, timeout=8)
    if response.status_code >= 400:
        raise RuntimeError(
            f"Supabase email auth request failed ({response.status_code})"
        )


def _generate_supabase_magic_link_url(
    email: str,
    client_redirect_to: str = "",
    redirect_path: str = "/login",
) -> str:
    if not SUPABASE_URL or not SUPABASE_SERVICE_ROLE_KEY:
        raise RuntimeError("Supabase service role key is required for magic-link generation.")

    redirect_to = _build_auth_redirect_url(
        redirect_path,
        client_redirect_to,
    )

    headers = {
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "apikey": SUPABASE_SERVICE_ROLE_KEY,
        "Content-Type": "application/json",
    }

    endpoint = f"{SUPABASE_URL.rstrip('/')}/auth/v1/admin/generate_link"
    payload: Dict[str, Any] = {
        "type": "magiclink",
        "email": email,
    }

    params: Dict[str, str] = {}
    if redirect_to:
        params["redirect_to"] = redirect_to

    response = requests.post(
        endpoint,
        headers=headers,
        json=payload,
        params=params if params else None,
        timeout=8,
    )
    if response.status_code >= 400:
        raise RuntimeError(
            f"Supabase admin generate_link failed ({response.status_code})"
        )

    body = response.json() if response.content else {}
    action_link = str(body.get("action_link") or "").strip()
    if not action_link:
        raise RuntimeError("Supabase generate_link response missing action_link.")
    return action_link


def _verify_supabase_user_token(token: str) -> Optional[Dict[str, Any]]:
    if not token:
        return None

    if SUPABASE_JWT_SECRET and not _looks_like_uuid(SUPABASE_JWT_SECRET):
        try:
            payload = jwt.decode(
                token,
                SUPABASE_JWT_SECRET,
                algorithms=["HS256"],
                options={"verify_aud": False},
            )
            return payload
        except Exception as exc:
            logger.warning("Supabase JWT validation failed (HS256): %s", exc)

    if SUPABASE_URL and SUPABASE_API_KEY:
        try:
            url = SUPABASE_URL.rstrip("/") + "/auth/v1/user"
            response = requests.get(
                url,
                headers={
                    "Authorization": f"Bearer {token}",
                    "apikey": SUPABASE_API_KEY,
                },
                timeout=5,
            )
            if response.status_code != 200:
                logger.warning(
                    "Supabase auth lookup failed: %s %s",
                    response.status_code,
                    response.text,
                )
                return None
            return response.json()
        except Exception as exc:
            logger.warning("Supabase auth lookup error: %s", exc)
            return None

    if SUPABASE_URL:
        try:
            jwks_client = _get_supabase_jwks_client()
            if jwks_client is not None:
                signing_key = jwks_client.get_signing_key_from_jwt(token)
                algorithm = getattr(signing_key, "algorithm", None)
                algorithms = [algorithm] if algorithm else ["ES256", "RS256", "ES384", "RS384"]
                payload = jwt.decode(
                    token,
                    signing_key.key,
                    algorithms=algorithms,
                    options={"verify_aud": False},
                )
                return payload
        except Exception as exc:
            logger.warning("Supabase JWT validation failed (JWKS): %s", exc)

    logger.warning(
        "Supabase auth is not configured. Set SUPABASE_URL for JWKS verification or provide SUPABASE_JWT_SECRET/SUPABASE_SERVICE_ROLE_KEY."
    )
    return None


def require_supabase_user(f):
    """Decorator to require a valid Supabase access token."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = _get_bearer_token()
        if not token:
            return jsonify({"error": "Authorization bearer token required"}), 401

        user = _verify_supabase_user_token(token)
        if not user:
            return jsonify({"error": "Invalid or expired Supabase token"}), 401

        g.supabase_user = user
        return f(*args, **kwargs)
    return decorated_function


def _is_admin_user(user: Dict[str, Any]) -> bool:
    app_meta = user.get("app_metadata") or {}
    role = app_meta.get("role")
    if isinstance(role, str) and role.strip().lower() == "admin":
        return True

    roles = app_meta.get("roles")
    if isinstance(roles, list):
        return any(
            isinstance(entry, str) and entry.strip().lower() == "admin"
            for entry in roles
        )

    return False


def _is_agent_task_allowed(task_name: str, user: Dict[str, Any]) -> bool:
    if _is_admin_user(user):
        return True
    return task_name == "chat"


def _create_agent_session(token: str, user_id: str) -> Tuple[str, int]:
    _purge_expired_agent_sessions()
    session_id = secrets.token_urlsafe(32)
    expires_at = int(time.time()) + AGENT_SESSION_TTL_SECONDS
    AGENT_SESSIONS[session_id] = {
        "token": token,
        "user_id": user_id,
        "expires_at": expires_at,
    }
    return session_id, expires_at


def _get_agent_session() -> Optional[Dict[str, Any]]:
    _purge_expired_agent_sessions()
    session_id = request.cookies.get(AGENT_SESSION_COOKIE)
    if not session_id:
        return None
    session = AGENT_SESSIONS.get(session_id)
    if not session:
        return None
    if session["expires_at"] <= time.time():
        AGENT_SESSIONS.pop(session_id, None)
        return None
    return session


def _clear_agent_session_for_request() -> None:
    session_id = request.cookies.get(AGENT_SESSION_COOKIE)
    if session_id:
        AGENT_SESSIONS.pop(session_id, None)


def require_agent_session(f):
    """Decorator to require a valid agent session cookie."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        session = _get_agent_session()
        user = getattr(g, "supabase_user", None)
        user_id = _get_supabase_user_id(user or {})
        if not session or not user_id or session.get("user_id") != user_id:
            return jsonify({"error": "Agent session required"}), 401

        g.agent_session = session
        return f(*args, **kwargs)
    return decorated_function


def _create_batch_session_token() -> str:
    timestamp = int(time.time())
    ts_bytes = str(timestamp).encode("utf-8")
    signature = hmac.new(API_KEY.encode("utf-8"), ts_bytes, hashlib.sha256).hexdigest()
    return f"{timestamp}.{signature}"


def _is_valid_batch_session(token: Optional[str]) -> bool:
    if not token:
        return False
    try:
        ts_str, signature = token.split(".", 1)
        timestamp = int(ts_str)
    except Exception:
        return False

    if timestamp <= 0:
        return False
    if (time.time() - timestamp) > BATCH_SESSION_TTL_SECONDS:
        return False

    expected = hmac.new(API_KEY.encode("utf-8"), ts_str.encode("utf-8"), hashlib.sha256).hexdigest()
    return hmac.compare_digest(signature, expected)


def is_valid_api_key(provided_key: Optional[str]) -> bool:
    if not provided_key:
        return False
    return hmac.compare_digest(provided_key, API_KEY)

def require_api_key(f):
    """Decorator to require API key authentication for protected routes."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        provided_key = request.headers.get('X-API-Key')
        
        # Log all API requests for audit trail
        logger.info(
            f"API Request: {request.method} {request.path} from {request.remote_addr} "
            f"- Auth: {'Valid' if is_valid_api_key(provided_key) else 'Invalid/Missing'}"
        )
        
        if not provided_key:
            logger.warning(f"Unauthorized request (no API key): {request.path} from {request.remote_addr}")
            return jsonify({"error": "API key required", "code": "AUTH_REQUIRED"}), 401
        
        if not is_valid_api_key(provided_key):
            logger.warning(f"Unauthorized request (invalid API key): {request.path} from {request.remote_addr}")
            return jsonify({"error": "Invalid API key", "code": "AUTH_INVALID"}), 401
        
        return f(*args, **kwargs)
    return decorated_function


def require_batch_session_or_api_key(f):
    """Decorator allowing either API key header or a signed batch session cookie."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        provided_key = request.headers.get('X-API-Key')
        if provided_key and is_valid_api_key(provided_key):
            return f(*args, **kwargs)

        token = request.cookies.get(BATCH_SESSION_COOKIE)
        if not _is_valid_batch_session(token):
            logger.warning(
                "Unauthorized batch request: %s from %s",
                request.path,
                request.remote_addr,
            )
            return jsonify({"error": "Batch session required", "code": "AUTH_REQUIRED"}), 401

        return f(*args, **kwargs)
    return decorated_function

# ── Input Validation ─────────────────────────────────────────────
def validate_layer_config(config: Any) -> Dict[str, Any]:
    """
    Validate and sanitize layer extraction configuration.
    Prevents injection attacks and ensures data integrity.
    """
    if not isinstance(config, dict):
        raise ValueError("Config must be a JSON object")
    
    # Validate and sanitize layers
    layers = config.get('layers', [])
    if not isinstance(layers, list):
        raise ValueError("'layers' must be an array")
    if len(layers) > 100:  # Prevent DoS via excessive layers
        raise ValueError("Maximum 100 layers allowed")
    
    sanitized_layers = []
    for layer in layers:
        if not isinstance(layer, str):
            continue
        # Remove potentially dangerous characters, allow alphanumeric, dash, underscore, space
        sanitized = re.sub(r'[^a-zA-Z0-9\-_ ]', '', layer.strip())
        if sanitized and len(sanitized) <= 255:
            sanitized_layers.append(sanitized)
    
    # Validate block reference path if provided
    ref_dwg = config.get('ref_dwg', '')
    if ref_dwg:
        if not isinstance(ref_dwg, str):
            raise ValueError("'ref_dwg' must be a string")
        # Prevent path traversal attacks
        if '..' in ref_dwg or ref_dwg.startswith(('/', '\\\\')):
            raise ValueError("Invalid reference path")
        # Ensure .dwg extension
        if not ref_dwg.lower().endswith('.dwg'):
            raise ValueError ("'ref_dwg' must have .dwg extension")
    
    # Validate block name if provided
    block_name = config.get('block_name', '')
    if block_name:
        if not isinstance(block_name, str):
            raise ValueError("'block_name' must be a string")
        # Sanitize block name
        block_name = re.sub(r'[^a-zA-Z0-9\-_]', '', block_name.strip())
        if len(block_name) > 255:
            raise ValueError("Block name too long")
    
    return {
        'layers': sanitized_layers,
        'ref_dwg': ref_dwg.strip() if ref_dwg else '',
        'block_name': block_name,
        'export_excel': bool(config.get('export_excel', False))
    }


# ── Transmittal Builder helpers ─────────────────────────────────
def _parse_json_field(name: str, default):
    raw = request.form.get(name)
    if not raw:
        return default
    try:
        return json.loads(raw)
    except Exception:
        return default


def _save_upload(file_storage, dest_dir: str, filename: Optional[str] = None) -> str:
    if file_storage is None:
        raise ValueError("Missing file upload")
    safe_name = secure_filename(filename or file_storage.filename or "upload")
    if not safe_name:
        safe_name = "upload"
    path = os.path.join(dest_dir, safe_name)
    file_storage.save(path)
    return path


def _schedule_cleanup(path: str) -> None:
    """Ensure temporary directories are removed after the response is sent."""
    if not path:
        return

    @after_this_request
    def _cleanup(response):
        try:
            shutil.rmtree(path, ignore_errors=True)
        except Exception as exc:
            logger.warning("Failed to cleanup temp dir %s: %s", path, exc)
        return response


def _convert_docx_to_pdf(docx_path: str, output_dir: str) -> Tuple[Optional[str], str]:
    """Convert a DOCX file to PDF. Returns (pdf_path, error_message)."""
    errors: List[str] = []

    # Attempt conversion with docx2pdf (requires Word on Windows or macOS)
    try:
        from docx2pdf import convert  # type: ignore

        convert(docx_path, output_dir)
        pdf_path = os.path.join(
            output_dir, f"{Path(docx_path).stem}.pdf"
        )
        if os.path.exists(pdf_path):
            return pdf_path, ""
        errors.append("docx2pdf did not produce a PDF file.")
    except Exception as exc:
        errors.append(f"docx2pdf failed: {exc}")

    # Attempt conversion with LibreOffice if available
    for cmd in ("soffice", "libreoffice"):
        exe = shutil.which(cmd)
        if not exe:
            continue
        try:
            result = subprocess.run(
                [
                    exe,
                    "--headless",
                    "--convert-to",
                    "pdf",
                    "--outdir",
                    output_dir,
                    docx_path,
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            if result.returncode == 0:
                pdf_path = os.path.join(
                    output_dir, f"{Path(docx_path).stem}.pdf"
                )
                if os.path.exists(pdf_path):
                    return pdf_path, ""
            errors.append(
                f"{cmd} conversion failed: {(result.stderr or result.stdout).strip()}"
            )
        except Exception as exc:
            errors.append(f"{cmd} conversion error: {exc}")

    return None, "; ".join([e for e in errors if e]) or "No PDF converter available."


# ── Batch Find & Replace helpers ───────────────────────────────
MAX_BATCH_FILES = 50
MAX_BATCH_RULES = 100
MAX_PREVIEW_MATCHES = 500
MAX_APPLY_CHANGE_ROWS = 5000


def _parse_batch_rules() -> List[Dict[str, Any]]:
    raw_rules = request.form.get("rules", "")
    if not raw_rules:
        raise ValueError("Missing rules payload")

    try:
        parsed = json.loads(raw_rules)
    except Exception:
        raise ValueError("Rules payload is not valid JSON")

    if not isinstance(parsed, list):
        raise ValueError("Rules payload must be an array")
    if len(parsed) == 0:
        raise ValueError("At least one rule is required")
    if len(parsed) > MAX_BATCH_RULES:
        raise ValueError(f"Too many rules. Maximum is {MAX_BATCH_RULES}")

    rules: List[Dict[str, Any]] = []
    for idx, item in enumerate(parsed):
        if not isinstance(item, dict):
            continue
        find_text = str(item.get("find", "")).strip()
        if not find_text:
            continue

        rules.append(
            {
                "id": str(item.get("id", f"rule-{idx + 1}")),
                "find": find_text,
                "replace": str(item.get("replace", "")),
                "use_regex": bool(item.get("useRegex", False)),
                "match_case": bool(item.get("matchCase", False)),
            }
        )

    if not rules:
        raise ValueError("No valid rules provided")

    return rules


def _decode_uploaded_text(raw_bytes: bytes) -> str:
    if b"\x00" in raw_bytes:
        raise ValueError("Binary files are not supported")

    for encoding in ("utf-8", "utf-16", "latin-1"):
        try:
            return raw_bytes.decode(encoding)
        except Exception:
            continue

    raise ValueError("Unable to decode file as text")


def _build_batch_pattern(rule: Dict[str, Any]) -> re.Pattern[str]:
    flags = 0 if rule["match_case"] else re.IGNORECASE
    if rule["use_regex"]:
        try:
            return re.compile(rule["find"], flags)
        except re.error as exc:
            raise ValueError(f"Invalid regex for rule '{rule['id']}': {exc}")

    return re.compile(re.escape(rule["find"]), flags)


def _apply_rule_to_lines(
    lines: List[str],
    pattern: re.Pattern[str],
    replacement: str,
    rule_id: str,
    file_name: str,
    preview_matches: List[Dict[str, Any]],
    max_matches: int,
) -> Tuple[List[str], int]:
    changed_count = 0
    next_lines: List[str] = []

    for line_number, line in enumerate(lines, start=1):
        updated_line, replaced = pattern.subn(replacement, line)
        if replaced > 0:
            changed_count += replaced
            if len(preview_matches) < max_matches:
                preview_matches.append(
                    {
                        "file": file_name,
                        "line": line_number,
                        "before": line[:500],
                        "after": updated_line[:500],
                        "ruleId": rule_id,
                    }
                )
        next_lines.append(updated_line)

    return next_lines, changed_count


def _process_batch_files(preview_only: bool) -> Dict[str, Any]:
    uploaded_files = request.files.getlist("files")
    if not uploaded_files:
        raise ValueError("No files uploaded")
    if len(uploaded_files) > MAX_BATCH_FILES:
        raise ValueError(f"Too many files. Maximum is {MAX_BATCH_FILES}")

    rules = _parse_batch_rules()
    max_matches = MAX_PREVIEW_MATCHES if preview_only else MAX_APPLY_CHANGE_ROWS
    preview_matches: List[Dict[str, Any]] = []
    updated_files: List[Dict[str, str]] = []
    files_changed = 0
    replacements_total = 0

    for file_storage in uploaded_files:
        file_name = secure_filename(file_storage.filename or "upload.txt")
        if not file_name:
            file_name = "upload.txt"

        try:
            raw_bytes = file_storage.read()
        finally:
            try:
                file_storage.close()
            except Exception:
                pass
        content = _decode_uploaded_text(raw_bytes)

        line_break = "\r\n" if "\r\n" in content else "\n"
        lines = content.splitlines()
        file_replacements = 0

        for rule in rules:
            pattern = _build_batch_pattern(rule)
            lines, changed = _apply_rule_to_lines(
                lines,
                pattern,
                rule["replace"],
                rule["id"],
                file_name,
                preview_matches,
                max_matches,
            )
            file_replacements += changed

        if file_replacements > 0:
            files_changed += 1
            replacements_total += file_replacements

        if not preview_only:
            updated_files.append(
                {
                    "file": file_name,
                    "content": line_break.join(lines),
                }
            )

    return {
        "matches": preview_matches,
        "files_changed": files_changed,
        "replacements": replacements_total,
        "files_processed": len(uploaded_files),
        "updated_files": updated_files,
    }


def export_batch_changes_to_excel(changes: List[Dict[str, Any]]) -> Tuple[str, str]:
    """
    Export batch find/replace changes to an Excel report.
    Styled similarly to Ground Grid coordinate export workbook.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = tempfile.mkdtemp(prefix="batch_find_replace_")
    out_path = os.path.join(out_dir, f"batch_find_replace_changes_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Changes"
    summary_ws = wb.create_sheet("Summary")

    headers = ["File", "Line", "Rule ID", "Before", "After"]

    title_fill = PatternFill("solid", fgColor="2B6CB5")
    title_font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    header_fill = PatternFill("solid", fgColor="3A3F47")
    header_font = Font(bold=True, color="F0F0F0", size=11, name="Arial")
    alt_fill_even = PatternFill("solid", fgColor="E8E6E2")
    alt_fill_odd = PatternFill("solid", fgColor="D4D1CC")
    data_font = Font(size=10, color="2A2A2A", name="Arial")
    border_side = Side(style="thin", color="B0ADA8")
    all_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    header_border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=Side(style="medium", color="3A3F47"),
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="Batch Find & Replace Change Report")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = all_border
    for col_idx in range(2, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = title_fill
        c.border = all_border

    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = header_border

    for row_idx, change in enumerate(changes, start=3):
        row_fill = alt_fill_even if (row_idx - 3) % 2 == 0 else alt_fill_odd
        row_values = [
            str(change.get("file", "")),
            int(change.get("line", 0) or 0),
            str(change.get("ruleId", "")),
            str(change.get("before", "")),
            str(change.get("after", "")),
        ]

        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.border = all_border
            cell.font = data_font
            if col_idx == 2:
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.column_dimensions[get_column_letter(1)].width = 42
    ws.column_dimensions[get_column_letter(2)].width = 10
    ws.column_dimensions[get_column_letter(3)].width = 16
    ws.column_dimensions[get_column_letter(4)].width = 60
    ws.column_dimensions[get_column_letter(5)].width = 60

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    summary_headers = ["Metric", "Value"]
    summary_title_fill = PatternFill("solid", fgColor="2B6CB5")
    summary_title_font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    summary_header_fill = PatternFill("solid", fgColor="3A3F47")
    summary_header_font = Font(bold=True, color="F0F0F0", size=11, name="Arial")
    summary_data_font = Font(size=10, color="2A2A2A", name="Arial")

    summary_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    summary_title_cell = summary_ws.cell(row=1, column=1, value="Batch Find & Replace Summary")
    summary_title_cell.font = summary_title_font
    summary_title_cell.fill = summary_title_fill
    summary_title_cell.alignment = Alignment(horizontal="center", vertical="center")
    summary_title_cell.border = all_border
    summary_ws.cell(row=1, column=2).fill = summary_title_fill
    summary_ws.cell(row=1, column=2).border = all_border

    for col_idx, header in enumerate(summary_headers, start=1):
        c = summary_ws.cell(row=2, column=col_idx, value=header)
        c.font = summary_header_font
        c.fill = summary_header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = header_border

    file_counts: Dict[str, int] = {}
    rule_counts: Dict[str, int] = {}
    for change in changes:
        file_name = str(change.get("file", "")) or "(unknown)"
        rule_id = str(change.get("ruleId", "")) or "(unknown)"
        file_counts[file_name] = file_counts.get(file_name, 0) + 1
        rule_counts[rule_id] = rule_counts.get(rule_id, 0) + 1

    summary_rows: List[Tuple[str, Any]] = [
        ("Total changes", len(changes)),
        ("Files with changes", len(file_counts)),
        ("Rules with changes", len(rule_counts)),
    ]

    current_row = 3
    for idx, (metric, value) in enumerate(summary_rows):
        row_fill = alt_fill_even if idx % 2 == 0 else alt_fill_odd
        metric_cell = summary_ws.cell(row=current_row, column=1, value=metric)
        value_cell = summary_ws.cell(row=current_row, column=2, value=value)
        for c in (metric_cell, value_cell):
            c.fill = row_fill
            c.border = all_border
            c.font = summary_data_font
            c.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

    current_row += 1
    section_header = summary_ws.cell(row=current_row, column=1, value="By File")
    section_header.font = summary_header_font
    section_header.fill = summary_header_fill
    section_header.alignment = Alignment(horizontal="left", vertical="center")
    section_header.border = header_border
    summary_ws.cell(row=current_row, column=2, value="Changes")
    summary_ws.cell(row=current_row, column=2).font = summary_header_font
    summary_ws.cell(row=current_row, column=2).fill = summary_header_fill
    summary_ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    summary_ws.cell(row=current_row, column=2).border = header_border
    current_row += 1

    for idx, (file_name, count) in enumerate(sorted(file_counts.items(), key=lambda item: item[0].lower())):
        row_fill = alt_fill_even if idx % 2 == 0 else alt_fill_odd
        file_cell = summary_ws.cell(row=current_row, column=1, value=file_name)
        count_cell = summary_ws.cell(row=current_row, column=2, value=count)
        for c in (file_cell, count_cell):
            c.fill = row_fill
            c.border = all_border
            c.font = summary_data_font
        file_cell.alignment = Alignment(horizontal="left", vertical="center")
        count_cell.alignment = Alignment(horizontal="right", vertical="center")
        current_row += 1

    current_row += 1
    section_header = summary_ws.cell(row=current_row, column=1, value="By Rule")
    section_header.font = summary_header_font
    section_header.fill = summary_header_fill
    section_header.alignment = Alignment(horizontal="left", vertical="center")
    section_header.border = header_border
    summary_ws.cell(row=current_row, column=2, value="Changes")
    summary_ws.cell(row=current_row, column=2).font = summary_header_font
    summary_ws.cell(row=current_row, column=2).fill = summary_header_fill
    summary_ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    summary_ws.cell(row=current_row, column=2).border = header_border
    current_row += 1

    for idx, (rule_id, count) in enumerate(sorted(rule_counts.items(), key=lambda item: item[0].lower())):
        row_fill = alt_fill_even if idx % 2 == 0 else alt_fill_odd
        rule_cell = summary_ws.cell(row=current_row, column=1, value=rule_id)
        count_cell = summary_ws.cell(row=current_row, column=2, value=count)
        for c in (rule_cell, count_cell):
            c.fill = row_fill
            c.border = all_border
            c.font = summary_data_font
        rule_cell.alignment = Alignment(horizontal="left", vertical="center")
        count_cell.alignment = Alignment(horizontal="right", vertical="center")
        current_row += 1

    summary_ws.column_dimensions[get_column_letter(1)].width = 52
    summary_ws.column_dimensions[get_column_letter(2)].width = 16
    summary_ws.row_dimensions[1].height = 28
    summary_ws.row_dimensions[2].height = 22
    summary_ws.freeze_panes = "A3"

    wb.save(out_path)
    return out_path, out_dir

# Global AutoCAD manager instance
_manager = None
FOUNDATION_SOURCE_TYPE = "Foundation Coordinates"


# ── Late-bound COM helpers (from coordtable) ────────────────────
def dyn(obj: Any) -> Any:
    """
    Force late-bound dynamic dispatch on a COM object.
    Avoids stale gen_py wrappers and CDispatch type errors.
    """
    try:
        if type(obj).__name__ == "CDispatch":
            return obj
    except Exception:
        pass

    try:
        ole = obj._oleobj_
    except Exception:
        ole = obj

    try:
        if not AUTOCAD_COM_AVAILABLE:
            return obj
        disp = ole.QueryInterface(pythoncom.IID_IDispatch)
        return win32com.client.dynamic.Dispatch(disp)
    except Exception:
        try:
            return win32com.client.dynamic.Dispatch(obj)
        except Exception:
            return obj


def connect_autocad() -> Any:
    """Connect to AutoCAD using late-bound dynamic dispatch (no gen_py)."""
    if not AUTOCAD_COM_AVAILABLE:
        raise RuntimeError("AutoCAD COM bridge unavailable on this platform. Run backend on Windows with pywin32 installed.")
    acad = win32com.client.dynamic.Dispatch("AutoCAD.Application")
    if acad is None:
        raise RuntimeError("Could not connect to AutoCAD.Application")
    return dyn(acad)


def com_call_with_retry(callable_func, max_retries: int = 25, initial_delay: float = 0.03):
    """Retry COM calls that get RPC_E_CALL_REJECTED (AutoCAD busy)."""
    delay = initial_delay
    for _ in range(max_retries):
        try:
            return callable_func()
        except pythoncom.com_error as e:
            if e.args and e.args[0] == -2147418111:  # RPC_E_CALL_REJECTED
                time.sleep(delay)
                delay = min(delay * 1.5, 0.5)
                continue
            raise
    raise RuntimeError("AutoCAD COM call failed: RPC busy too long")


def pt(x: float, y: float, z: float = 0.0):
    if not AUTOCAD_COM_AVAILABLE:
        raise RuntimeError("AutoCAD COM bridge unavailable on this platform")
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (float(x), float(y), float(z)))


def ensure_layer(doc: Any, layer_name: str) -> None:
    doc = dyn(doc)
    try:
        layers = dyn(doc.Layers)
        try:
            layers.Item(layer_name)
        except Exception:
            layers.Add(layer_name)
    except Exception:
        pass


def wait_for_command_finish(doc: Any, timeout_s: float = 10.0) -> bool:
    doc = dyn(doc)
    t0 = time.time()
    while (time.time() - t0) < timeout_s:
        try:
            names = ""
            if hasattr(doc, "GetVariable"):
                names = str(doc.GetVariable("CMDNAMES") or "")
            if not names.strip():
                return True
        except Exception:
            pass
        time.sleep(0.15)
    return False


_REF_IMPORT_CACHE: Dict[str, str] = {}


def ensure_block_exists(doc: Any, block_name: str, dwg_path: str) -> str:
    doc = dyn(doc)
    dwg_path = os.path.abspath(dwg_path)
    try:
        doc.Blocks.Item(block_name)
        return block_name
    except Exception:
        logger.info(f"Block '{block_name}' not found. Importing via Xref-Bind...")

    if not os.path.exists(dwg_path):
        raise RuntimeError(f"External file not found: {dwg_path}")

    ms = dyn(doc.ModelSpace)
    origin = pt(0, 0, 0)
    xref_name = block_name

    def _attach(name: str):
        if hasattr(ms, "AttachExternalReference"):
            return ms.AttachExternalReference(dwg_path, name, origin, 1.0, 1.0, 1.0, 0.0, False)
        if hasattr(ms, "AttachXref"):
            return ms.AttachXref(dwg_path, name, origin, 1.0, 1.0, 1.0, 0.0, False)
        raise RuntimeError("Neither AttachExternalReference nor AttachXref available.")

    try:
        try:
            xref_obj = com_call_with_retry(lambda: _attach(xref_name))
        except Exception:
            xref_name = f"TEMP_IMPORT_{block_name}_{int(time.time())}"
            xref_obj = com_call_with_retry(lambda: _attach(xref_name))

        cmd = f'_.-XREF _B "{xref_name}" \\n'
        com_call_with_retry(lambda: doc.SendCommand(cmd))
        wait_for_command_finish(doc, timeout_s=20.0)

        try:
            if xref_obj is not None:
                dyn(xref_obj).Delete()
        except Exception:
            pass

        try:
            doc.Blocks.Item(block_name)
            return block_name
        except Exception:
            try:
                doc.Blocks.Item(xref_name)
                return xref_name
            except Exception as exc:
                raise RuntimeError(
                    f"Xref bind completed but block not found. Tried: '{block_name}', '{xref_name}'."
                ) from exc

    except Exception as exc:
        raise RuntimeError(
            f"Failed to import reference DWG.\nDWG: {dwg_path}\nBlock: {block_name}\nDetails: {exc}"
        ) from exc


def insert_reference_block(doc, ms, ref_dwg_path, layer_name, x, y, z, scale, rotation_deg):
    doc = dyn(doc)
    ms = dyn(ms)
    ref_dwg_path = os.path.abspath(ref_dwg_path)

    if not os.path.exists(ref_dwg_path):
        raise RuntimeError(
            f"Reference DWG not found: {ref_dwg_path}\n"
            "Put 'Coordinate Reference Point.dwg' in an 'assets' folder next to api_server.py."
        )

    block_name = os.path.splitext(os.path.basename(ref_dwg_path))[0]
    cache_key = os.path.normcase(ref_dwg_path)

    if cache_key in _REF_IMPORT_CACHE:
        insert_name = _REF_IMPORT_CACHE[cache_key]
    else:
        insert_name = ensure_block_exists(doc, block_name, ref_dwg_path)
        _REF_IMPORT_CACHE[cache_key] = insert_name

    ensure_layer(doc, layer_name)

    def _insert():
        return ms.InsertBlock(
            pt(x, y, z), insert_name,
            float(scale), float(scale), float(scale),
            math.radians(float(rotation_deg)),
        )

    br = com_call_with_retry(_insert)
    br = dyn(br)
    try:
        br.Layer = layer_name
    except Exception:
        pass
    return br


def add_point_label(ms, layer_name, label_text, x, y, z, scale):
    text_height = max(scale * 1.5, 0.5)
    x_offset = scale * 3.0

    def _add():
        return ms.AddText(label_text, pt(x + x_offset, y, z), text_height)

    txt = com_call_with_retry(_add)
    txt = dyn(txt)
    try:
        txt.Layer = layer_name
    except Exception:
        pass
    try:
        txt.Alignment = 0  # acAlignmentLeft
    except Exception:
        pass
    return txt


def default_ref_dwg_path() -> str:
    base = os.path.dirname(os.path.abspath(__file__))
    cand = os.path.join(base, "assets", "Coordinate Reference Point.dwg")
    if os.path.exists(cand):
        return cand
    return os.path.join(base, "Coordinate Reference Point.dwg")


def _entity_bbox(ent):
    ent = dyn(ent)
    try:
        mn, mx = ent.GetBoundingBox()
        minx, miny = float(mn[0]), float(mn[1])
        maxx, maxy = float(mx[0]), float(mx[1])
        minz = float(mn[2]) if len(mn) > 2 else 0.0
        maxz = float(mx[2]) if len(mx) > 2 else 0.0
        if maxx < minx:
            minx, maxx = maxx, minx
        if maxy < miny:
            miny, maxy = maxy, miny
        if maxz < minz:
            minz, maxz = maxz, minz
        return (minx, miny, minz, maxx, maxy, maxz)
    except Exception:
        return None


def _poly_centroid(ent):
    ent = dyn(ent)
    try:
        obj_name = str(ent.ObjectName)
    except Exception:
        obj_name = ''

    coords = []
    try:
        raw = list(ent.Coordinates)
        if obj_name == 'AcDb3dPolyline':
            for i in range(0, len(raw), 3):
                if i + 2 < len(raw):
                    coords.append((float(raw[i]), float(raw[i+1]), float(raw[i+2])))
        else:
            elev = 0.0
            try:
                elev = float(ent.Elevation)
            except Exception:
                pass
            for i in range(0, len(raw), 2):
                if i + 1 < len(raw):
                    coords.append((float(raw[i]), float(raw[i+1]), elev))
    except Exception:
        try:
            n = int(ent.NumberOfVertices)
            elev = 0.0
            try:
                elev = float(ent.Elevation)
            except Exception:
                pass
            for i in range(n):
                p = ent.Coordinate(i)
                z = float(p[2]) if len(p) > 2 else elev
                coords.append((float(p[0]), float(p[1]), z))
        except Exception:
            return None

    if not coords:
        return None

    n = len(coords)
    return (
        sum(p[0] for p in coords) / n,
        sum(p[1] for p in coords) / n,
        sum(p[2] for p in coords) / n,
    )


def _entity_center(ent):
    ent = dyn(ent)
    try:
        obj_name = str(ent.ObjectName)
    except Exception:
        obj_name = ''

    if obj_name in ('AcDbPolyline', 'AcDb2dPolyline', 'AcDb3dPolyline'):
        result = _poly_centroid(ent)
        if result:
            return result

    bbox = _entity_bbox(ent)
    if bbox:
        minx, miny, minz, maxx, maxy, maxz = bbox
        return ((minx + maxx) / 2.0, (miny + maxy) / 2.0, (minz + maxz) / 2.0)

    return None


def export_points_to_excel(points, precision, use_corners, drawing_dir=None):
    """
    Export coordinates to Excel with points organized by layer.
    Each layer gets its own table section with a 2-row gap between layers.
    """
    if drawing_dir:
        out_dir = drawing_dir
    else:
        out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")
    os.makedirs(out_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"coordinates_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Coordinates"

    headers = ["Point ID", "East (X)", "North (Y)", "Elevation (Z)", "Layer"]

    # ── Style definitions ──
    # Row 1: Title banner -- R3P logo blue (#2B6CB5)
    title_fill = PatternFill("solid", fgColor="2B6CB5")
    title_font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    # Layer section headers -- medium blue
    layer_header_fill = PatternFill("solid", fgColor="5B9BD5")
    layer_header_font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    # Column headers -- dark charcoal gray
    header_fill = PatternFill("solid", fgColor="3A3F47")
    header_font = Font(bold=True, color="F0F0F0", size=11, name="Arial")
    # Data rows: alternating warm neutrals
    alt_fill_even = PatternFill("solid", fgColor="E8E6E2")
    alt_fill_odd = PatternFill("solid", fgColor="D4D1CC")
    data_font = Font(size=10, color="2A2A2A", name="Arial")
    # Borders: visible but not heavy
    border_side = Side(style="thin", color="B0ADA8")
    all_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    # Thicker border under header
    header_border = Border(
        left=border_side, right=border_side,
        top=border_side,
        bottom=Side(style="medium", color="3A3F47"),
    )

    # Row 1: merged title "Ground Grid Coordinates"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="Ground Grid Coordinates")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = all_border
    for col_idx in range(2, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = title_fill
        c.border = all_border

    num_fmt = "0" if precision <= 0 else "0." + ("0" * precision)
    numeric_cols = {"East (X)", "North (Y)", "Elevation (Z)"}

    # ── Group points by layer ──
    from collections import defaultdict
    points_by_layer = defaultdict(list)
    for p in points:
        layer_name = p.get('layer', 'Default')
        points_by_layer[layer_name].append(p)

    # Sort layers alphabetically for consistent output
    sorted_layers = sorted(points_by_layer.keys())

    current_row = 2  # Start after title row

    # ── Write each layer as a separate table ──
    for layer_idx, layer_name in enumerate(sorted_layers):
        layer_points = points_by_layer[layer_name]
        
        # Layer section header (merged across all columns)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        layer_header_cell = ws.cell(row=current_row, column=1, value=f"Layer: {layer_name}")
        layer_header_cell.font = layer_header_font
        layer_header_cell.fill = layer_header_fill
        layer_header_cell.alignment = Alignment(horizontal="left", vertical="center")
        layer_header_cell.border = all_border
        for col_idx in range(2, len(headers) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.fill = layer_header_fill
            c.border = all_border
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Column headers for this layer section
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = header_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Data rows for this layer
        for idx, p in enumerate(layer_points):
            row = [
                p['name'],
                p['x'],
                p['y'],
                p['z'],
                p.get('layer', ''),
            ]
            row_fill = alt_fill_even if idx % 2 == 0 else alt_fill_odd
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = all_border
                cell.font = data_font
                if headers[col_idx - 1] in numeric_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = num_fmt
                elif col_idx == 1:
                    cell.font = Font(bold=True, size=10, color="2A2A2A", name="Arial")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

        # Add 2-row gap between layer sections (except after last layer)
        if layer_idx < len(sorted_layers) - 1:
            current_row += 2

    # Auto-fit column widths based on all data
    for col_idx, h in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = len(h)
        for p in points:
            for field_idx, field in enumerate([p['name'], p['x'], p['y'], p['z'], p.get('layer', '')]):
                if field_idx + 1 == col_idx:
                    width = max(width, len(str(field)))
        ws.column_dimensions[col_letter].width = min(max(width + 3, 14), 70)
    
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"
    
    wb.save(out_path)
    return out_path


class AutoCADManager:
    """
    Thread-safe AutoCAD connection manager
    Uses late-bound COM (dynamic dispatch) to avoid gen_py cache issues.
    """
    
    def __init__(self):
        self.start_time = time.time()
        self._lock = threading.Lock()
        self._cached_status = None
        self._cache_ttl = 2.0  # Cache status for 2 seconds
        self.last_check_time = 0
        
        print("[AutoCADManager] Initialized")
    
    def is_autocad_process_running(self) -> Tuple[bool, Optional[str]]:
        """
        Check if acad.exe process is running on Windows
        Returns: (is_running, process_exe_path)
        """
        try:
            for proc in psutil.process_iter(['name', 'exe']):
                try:
                    proc_name = proc.info.get('name', '').lower()
                    if proc_name == 'acad.exe':
                        return (True, proc.info.get('exe'))
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
        except Exception as e:
            print(f"[AutoCADManager] Error checking process: {e}")
        
        return (False, None)
    
    def _fresh_com_connection(self) -> Tuple[Any, Any, bool, Optional[str], Optional[str]]:
        """
        Get a FRESH late-bound COM connection every time.
        Never caches COM objects across calls (avoids stale ref issues).
        Returns: (acad, doc, drawing_open, drawing_name, error_message)
        """
        if not AUTOCAD_COM_AVAILABLE:
            return (None, None, False, None, "AutoCAD COM is unavailable in this environment (Windows + pywin32 required)")
        try:
            acad = connect_autocad()
            
            try:
                doc = dyn(acad.ActiveDocument)
                if doc is None:
                    return (acad, None, False, None, "No drawing is open")
                
                try:
                    drawing_name = str(doc.Name)
                except Exception:
                    drawing_name = "Unknown"
                
                return (acad, doc, True, drawing_name, None)
                
            except Exception as e:
                return (acad, None, False, None, f"Cannot access ActiveDocument: {str(e)}")
                
        except Exception as e:
            return (None, None, False, None, f"Cannot connect to AutoCAD: {str(e)}")
    
    def get_status(self, force_refresh: bool = False) -> Dict[str, Any]:
        """
        Get comprehensive AutoCAD status.
        Uses process-level caching only; COM refs are always fresh.
        """
        with self._lock:
            current_time = time.time()
            
            # Return cached status if still valid
            if not force_refresh and self._cached_status is not None:
                if current_time - self._cached_status['timestamp'] < self._cache_ttl:
                    return self._cached_status
            
            if not AUTOCAD_COM_AVAILABLE:
                status = {
                    'connected': False,
                    'autocad_running': False,
                    'drawing_open': False,
                    'drawing_name': None,
                    'autocad_path': None,
                    'error': 'AutoCAD COM unavailable (run on Windows with pywin32 and AutoCAD)',
                    'checks': {'process': False, 'com': False, 'document': False},
                    'backend_uptime': current_time - self.start_time,
                    'timestamp': current_time,
                    'degraded_mode': True,
                }
                self._cached_status = status
                self.last_check_time = current_time
                return status

            process_running, acad_path = self.is_autocad_process_running()
            
            if not process_running:
                status = {
                    'connected': False,
                    'autocad_running': False,
                    'drawing_open': False,
                    'drawing_name': None,
                    'autocad_path': None,
                    'error': 'AutoCAD process (acad.exe) not detected',
                    'checks': {'process': False, 'com': False, 'document': False},
                    'backend_uptime': current_time - self.start_time,
                    'timestamp': current_time
                }
            else:
                # Fresh COM connection every time (no stale refs)
                try:
                    pythoncom.CoInitialize()
                    acad, doc, drawing_ok, drawing_name, error = self._fresh_com_connection()
                    com_ok = acad is not None
                except Exception as e:
                    com_ok, drawing_ok, drawing_name, error = False, False, None, str(e)
                finally:
                    try:
                        pythoncom.CoUninitialize()
                    except:
                        pass
                
                status = {
                    'connected': com_ok,
                    'autocad_running': process_running,
                    'drawing_open': drawing_ok,
                    'drawing_name': drawing_name,
                    'autocad_path': acad_path,
                    'error': error,
                    'checks': {
                        'process': process_running,
                        'com': com_ok,
                        'document': drawing_ok
                    },
                    'backend_uptime': current_time - self.start_time,
                    'timestamp': current_time
                }
            
            self._cached_status = status
            self.last_check_time = current_time
            return status
    
    def get_layers(self) -> Tuple[bool, List[str], Optional[str]]:
        """
        Get list of layer names from active drawing.
        Uses fresh late-bound COM connection every call.
        """
        status = self.get_status()
        
        if not status['drawing_open']:
            return (False, [], status.get('error', 'No drawing open'))
        
        try:
            pythoncom.CoInitialize()
            
            acad = connect_autocad()
            doc = dyn(acad.ActiveDocument)
            
            if doc is None:
                return (False, [], 'Document reference lost')
            
            layers = []
            layer_collection = dyn(doc.Layers)
            for i in range(int(layer_collection.Count)):
                layer = dyn(layer_collection.Item(i))
                layers.append(str(layer.Name))
            
            return (True, sorted(layers), None)
            
        except Exception as e:
            return (False, [], f'COM error: {str(e)}')
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    
    def execute_layer_search(self, config: Dict) -> Dict[str, Any]:
        """
        Execute layer search matching the desktop coordinatesgrabber.py logic:
        - Find entities on target layer in ModelSpace
        - Compute ONE center point per entity (not per vertex)
        - Insert reference blocks at each point
        - Export Excel and auto-open it
        """
        try:
            pythoncom.CoInitialize()

            acad = connect_autocad()
            doc = dyn(acad.ActiveDocument)
            ms = dyn(doc.ModelSpace)

            if doc is None or ms is None:
                raise RuntimeError('Cannot access AutoCAD document or modelspace')

            raw_layers = config.get('layer_search_names')
            requested_layers = []
            if isinstance(raw_layers, list):
                requested_layers.extend(
                    [str(layer).strip() for layer in raw_layers if str(layer).strip()]
                )

            fallback_layers_raw = str(config.get('layer_search_name', '')).strip()
            if fallback_layers_raw:
                for part in re.split(r'[;,\n]+', fallback_layers_raw):
                    layer_name_part = part.strip()
                    if layer_name_part:
                        requested_layers.append(layer_name_part)

            requested_layers = list(dict.fromkeys(requested_layers))

            if not requested_layers:
                return {
                    'success': False,
                    'points': [],
                    'count': 0,
                    'layers': [],
                    'excel_path': '',
                    'blocks_inserted': 0,
                    'error': 'No layer names provided'
                }

            requested_layer_lookup = {layer.strip().lower() for layer in requested_layers}
            prefix = config.get('prefix', 'P')
            start_num = int(config.get('initial_number', 1))
            precision = int(config.get('precision', 3))
            use_corners = config.get('layer_search_use_corners', False)

            points = []
            point_num = start_num
            entities_scanned = 0

            entity_count = int(ms.Count)
            for idx in range(entity_count):
                try:
                    ent = dyn(ms.Item(idx))

                    try:
                        ent_layer = str(ent.Layer)
                    except Exception:
                        continue

                    ent_layer_normalized = ent_layer.strip().lower()
                    if ent_layer_normalized not in requested_layer_lookup:
                        continue

                    entities_scanned += 1

                    if use_corners:
                        bbox = _entity_bbox(ent)
                        if not bbox:
                            continue
                        minx, miny, minz, maxx, maxy, maxz = bbox
                        z_val = (minz + maxz) / 2.0
                        corner_defs = [
                            (minx, maxy, 'NW'),
                            (maxx, maxy, 'NE'),
                            (minx, miny, 'SW'),
                            (maxx, miny, 'SE'),
                        ]
                        for cx, cy, corner_name in corner_defs:
                            points.append({
                                'name': f'{prefix}{point_num}_{corner_name}',
                                'x': round(cx, precision),
                                'y': round(cy, precision),
                                'z': round(z_val, precision),
                                'corner': corner_name,
                                'source_type': FOUNDATION_SOURCE_TYPE,
                                'layer': ent_layer.strip(),
                            })
                            point_num += 1
                    else:
                        center = _entity_center(ent)
                        if not center:
                            continue
                        cx, cy, cz = center
                        points.append({
                            'name': f'{prefix}{point_num}',
                            'x': round(cx, precision),
                            'y': round(cy, precision),
                            'z': round(cz, precision),
                            'source_type': FOUNDATION_SOURCE_TYPE,
                            'layer': ent_layer.strip(),
                        })
                        point_num += 1

                except Exception as e:
                    print(f"[execute] Entity {idx} error: {e}")
                    continue

            print(f"[execute] Scanned {entities_scanned} entities across layers {requested_layers}, extracted {len(points)} points")

            if not points:
                return {
                    'success': False,
                    'points': [],
                    'count': 0,
                    'layers': requested_layers,
                    'excel_path': '',
                    'blocks_inserted': 0,
                    'error': f'No entities found on requested layers: {", ".join(requested_layers)}'
                }

            ref_dwg = config.get('ref_dwg_path', '').strip()
            if not ref_dwg:
                ref_dwg = default_ref_dwg_path()
            ref_layer = config.get('ref_layer_name', 'Coordinate Reference Point')
            ref_scale = float(config.get('ref_scale', 1.0))
            ref_rotation = float(config.get('ref_rotation_deg', 0))

            blocks_inserted = 0
            block_errors = []
            if os.path.exists(ref_dwg):
                print(f"[execute] Inserting reference blocks from: {ref_dwg}")
                for p in points:
                    try:
                        insert_reference_block(
                            doc, ms, ref_dwg, ref_layer,
                            p['x'], p['y'], p['z'],
                            ref_scale, ref_rotation
                        )
                        try:
                            add_point_label(
                                ms, ref_layer, p['name'],
                                p['x'], p['y'], p['z'],
                                ref_scale,
                            )
                        except Exception as label_err:
                            print(f"[execute] Label at {p['name']}: {label_err}")
                        blocks_inserted += 1
                    except Exception as e:
                        block_errors.append(f"Block at {p['name']}: {e}")
                        print(f"[execute] Block insert error at {p['name']}: {e}")

                try:
                    doc.Regen(1)
                except Exception:
                    pass

                if blocks_inserted > 0:
                    print(f"[execute] Inserted {blocks_inserted} reference blocks")
            else:
                block_errors.append(f"Reference DWG not found: {ref_dwg}")
                print(f"[execute] WARNING: Reference DWG not found at {ref_dwg}, skipping block insertion")

            drawing_dir = None
            try:
                drawing_path = str(doc.FullName)
                if drawing_path:
                    drawing_dir = os.path.dirname(drawing_path)
            except Exception:
                pass

            excel_path = ''
            try:
                excel_path = export_points_to_excel(points, precision, use_corners, drawing_dir)
                print(f"[execute] Excel exported to: {excel_path}")
                try:
                    os.startfile(excel_path)
                except Exception:
                    pass
            except Exception as e:
                block_errors.append(f"Excel export: {e}")
                print(f"[execute] Excel export error: {e}")

            return {
                'success': True,
                'points': points,
                'count': len(points),
                'layers': requested_layers,
                'excel_path': excel_path,
                'blocks_inserted': blocks_inserted,
                'block_errors': block_errors if block_errors else None,
                'error': None
            }

        except Exception as e:
            traceback.print_exc()
            return {
                'success': False,
                'points': [],
                'count': 0,
                'layers': [],
                'excel_path': '',
                'blocks_inserted': 0,
                'error': str(e)
            }
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass


# Initialize manager
def get_manager() -> AutoCADManager:
    global _manager
    if _manager is None:
        _manager = AutoCADManager()
    return _manager


# ========== API ENDPOINTS ==========

# ========== AGENT BROKER ENDPOINTS ==========

@app.route('/api/auth/passkey-capability', methods=['GET'])
@limiter.limit("60 per hour")
def api_auth_passkey_capability():
    """Expose passkey rollout capability status for frontend gating."""
    payload = _auth_passkey_capability()
    return jsonify({
        "ok": True,
        "passkey": payload,
        "server_time": datetime.utcnow().isoformat() + "Z",
    }), 200


@app.route('/api/auth/passkey/sign-in', methods=['POST'])
@limiter.limit("20 per hour")
def api_auth_passkey_sign_in():
    """Start a passkey sign-in flow."""
    payload = request.get_json(silent=True) if request.is_json else {}
    payload = payload or {}
    client_redirect_to = str(
        payload.get("redirectTo") or payload.get("redirect_to") or ""
    ).strip()
    client_ip = _get_request_ip()

    capability = _auth_passkey_capability()
    if not capability.get("enabled"):
        return jsonify({
            "error": "Passkey sign-in is disabled.",
            "code": "passkey-disabled",
        }), 503

    if not capability.get("handlers_ready"):
        return jsonify({
            "error": "Passkey sign-in is not available for the configured provider.",
            "code": "passkey-provider-unavailable",
            "provider": capability.get("provider"),
            "next_step": capability.get("next_step"),
        }), 501

    state_token, expires_at = _create_passkey_callback_state(
        intent="sign-in",
        client_ip=client_ip,
        client_redirect_to=client_redirect_to,
    )
    redirect_url = _build_external_passkey_redirect(
        "sign-in",
        state_token,
        client_redirect_to,
    )
    if not redirect_url:
        with PASSKEY_CALLBACK_STATES_LOCK:
            PASSKEY_CALLBACK_STATES.pop(state_token, None)
        return jsonify({
            "error": "Passkey sign-in provider URL is invalid or missing.",
            "code": "passkey-config-invalid",
        }), 503

    return jsonify({
        "ok": True,
        "method": "passkey",
        "mode": "redirect",
        "provider": capability.get("provider"),
        "provider_label": capability.get("provider_label"),
        "state": state_token,
        "state_expires_at": datetime.utcfromtimestamp(expires_at).isoformat() + "Z",
        "redirect_url": redirect_url,
        "message": "Continue passkey sign-in with your identity provider.",
    }), 200


@app.route('/api/auth/passkey/enroll', methods=['POST'])
@require_supabase_user
@limiter.limit("20 per hour")
def api_auth_passkey_enroll():
    """Start a passkey enrollment flow for an authenticated user."""
    payload = request.get_json(silent=True) if request.is_json else {}
    payload = payload or {}
    client_redirect_to = str(
        payload.get("redirectTo") or payload.get("redirect_to") or ""
    ).strip()
    client_ip = _get_request_ip()

    user = getattr(g, "supabase_user", {}) or {}
    user_id = _get_supabase_user_id(user) or ""
    user_email = _get_supabase_user_email(user) or ""

    capability = _auth_passkey_capability()
    if not capability.get("enabled"):
        return jsonify({
            "error": "Passkey enrollment is disabled.",
            "code": "passkey-disabled",
        }), 503

    if not capability.get("handlers_ready"):
        return jsonify({
            "error": "Passkey enrollment is not available for the configured provider.",
            "code": "passkey-provider-unavailable",
            "provider": capability.get("provider"),
            "next_step": capability.get("next_step"),
        }), 501

    state_token, expires_at = _create_passkey_callback_state(
        intent="enroll",
        client_ip=client_ip,
        client_redirect_to=client_redirect_to,
        user_id=user_id,
        email=user_email,
    )
    redirect_url = _build_external_passkey_redirect(
        "enroll",
        state_token,
        client_redirect_to,
    )
    if not redirect_url:
        with PASSKEY_CALLBACK_STATES_LOCK:
            PASSKEY_CALLBACK_STATES.pop(state_token, None)
        return jsonify({
            "error": "Passkey enrollment provider URL is invalid or missing.",
            "code": "passkey-config-invalid",
        }), 503

    return jsonify({
        "ok": True,
        "method": "passkey",
        "mode": "redirect",
        "provider": capability.get("provider"),
        "provider_label": capability.get("provider_label"),
        "state": state_token,
        "state_expires_at": datetime.utcfromtimestamp(expires_at).isoformat() + "Z",
        "redirect_url": redirect_url,
        "message": "Continue passkey enrollment with your identity provider.",
    }), 200


@app.route('/api/auth/passkey/callback/complete', methods=['POST'])
@limiter.limit("60 per hour")
def api_auth_passkey_callback_complete():
    """Complete a passkey callback by consuming one-time state and issuing the next auth step."""
    if not request.is_json:
        return jsonify({"error": "Expected JSON payload."}), 400

    payload = request.get_json(silent=True) or {}
    state_token = str(
        payload.get("state")
        or payload.get("passkey_state")
        or ""
    ).strip()
    if not state_token:
        return jsonify({"error": "state is required."}), 400
    if not PASSKEY_CALLBACK_STATE_PATTERN.match(state_token):
        return jsonify({"error": "Invalid state format."}), 400

    status = str(payload.get("status") or payload.get("passkey_status") or "").strip().lower()
    if status not in {"success", "failed"}:
        return jsonify({"error": "status must be success or failed."}), 400

    callback_state, reason = _get_passkey_callback_state(state_token)
    if not callback_state:
        if reason == "expired":
            return jsonify({"error": "Passkey callback state expired. Start again."}), 410
        return jsonify({"error": "Invalid passkey callback state."}), 400

    intent = str(callback_state.get("intent") or "").strip().lower()
    payload_intent = str(
        payload.get("intent") or payload.get("passkey_intent") or ""
    ).strip().lower()
    if payload_intent and payload_intent != intent:
        return jsonify({
            "error": "Passkey callback intent mismatch.",
            "code": "passkey-intent-mismatch",
        }), 400

    provider_error = str(
        payload.get("error") or payload.get("passkey_error") or ""
    ).strip()
    email = str(payload.get("email") or payload.get("passkey_email") or "").strip().lower()
    callback_signature = str(
        payload.get("signature")
        or payload.get("passkey_signature")
        or payload.get("provider_signature")
        or ""
    ).strip().lower()
    callback_timestamp_raw = str(
        payload.get("timestamp")
        or payload.get("passkey_timestamp")
        or payload.get("provider_timestamp")
        or ""
    ).strip()

    if status == "success" and intent == "sign-in" and not _is_valid_email(email):
        return jsonify({
            "error": "A valid email is required to complete passkey sign-in.",
            "code": "passkey-email-required",
        }), 400

    signature_ok, signature_reason = _verify_passkey_callback_signature(
        state=state_token,
        intent=intent,
        status=status,
        email=email,
        error_message=provider_error,
        signature=callback_signature,
        timestamp_raw=callback_timestamp_raw,
    )
    if not signature_ok:
        state_fingerprint = hashlib.sha256(state_token.encode("utf-8")).hexdigest()[:12]
        logger.warning(
            "Rejected passkey callback signature: reason=%s intent=%s state_hash=%s ip=%s",
            signature_reason,
            intent,
            state_fingerprint,
            _get_request_ip(),
        )
        return jsonify({
            "error": "Passkey callback signature validation failed.",
            "code": "passkey-callback-signature-invalid",
            "reason": signature_reason,
        }), 401

    callback_state, reason = _consume_passkey_callback_state(state_token)
    if not callback_state:
        if reason == "expired":
            return jsonify({"error": "Passkey callback state expired. Start again."}), 410
        return jsonify({
            "error": "Passkey callback state has already been used.",
            "code": "passkey-callback-state-used",
        }), 409

    client_redirect_to = str(callback_state.get("client_redirect_to") or "").strip()
    expected_email = str(callback_state.get("email") or "").strip().lower()
    if intent == "enroll" and expected_email and email and expected_email != email:
        return jsonify({
            "error": "Passkey enrollment email mismatch.",
            "code": "passkey-email-mismatch",
        }), 400

    if status == "failed":
        message = provider_error or "Passkey verification was not completed."
        return jsonify({
            "ok": True,
            "completed": False,
            "intent": intent,
            "status": "failed",
            "message": message,
        }), 200

    if intent == "sign-in":
        if SUPABASE_SERVICE_ROLE_KEY:
            try:
                magic_link = _generate_supabase_magic_link_url(
                    email,
                    client_redirect_to=client_redirect_to,
                    redirect_path="/login",
                )
                return jsonify({
                    "ok": True,
                    "completed": True,
                    "intent": intent,
                    "session_mode": "magic-link-direct",
                    "resume_url": magic_link,
                    "message": "Passkey verified. Continuing sign-in.",
                }), 200
            except Exception as exc:
                logger.warning("Passkey callback direct magic-link generation failed: %s", exc)

        try:
            _send_supabase_email_link(
                email,
                "signin",
                client_redirect_to=client_redirect_to,
                redirect_path="/login",
            )
        except Exception as exc:
            logger.warning("Passkey callback email-link fallback failed: %s", exc)
            return jsonify({
                "error": "Passkey callback completed, but sign-in continuation failed.",
                "code": "passkey-continuation-failed",
            }), 502

        return jsonify({
            "ok": True,
            "completed": True,
            "intent": intent,
            "session_mode": "email-link-fallback",
            "message": "Passkey verified. Check your email to finish sign-in.",
        }), 200

    if intent == "enroll":
        settings_url = _build_auth_redirect_url("/app/settings", client_redirect_to)
        return jsonify({
            "ok": True,
            "completed": True,
            "intent": intent,
            "status": "success",
            "redirect_to": settings_url,
            "message": "Passkey enrollment verified. Return to settings.",
        }), 200

    return jsonify({
        "error": "Unsupported passkey callback intent.",
        "code": "passkey-intent-invalid",
    }), 400


@app.route('/api/auth/email-link', methods=['POST'])
@limiter.limit("12 per hour")
def api_auth_email_link():
    """Request a Supabase email link for passwordless sign-in/sign-up."""
    started_at = time.perf_counter()

    def _finalize(payload: Dict[str, Any], status: int):
        _apply_auth_email_response_floor(started_at)
        return jsonify(payload), status

    if not SUPABASE_URL or not SUPABASE_API_KEY:
        return _finalize({
            "error": "Email authentication backend is not configured.",
            "missing": ["SUPABASE_URL", "SUPABASE_SERVICE_ROLE_KEY or SUPABASE_ANON_KEY"],
        }, 503)

    if not request.is_json:
        return _finalize({"error": "Expected JSON payload."}, 400)

    payload = request.get_json(silent=True) or {}
    email = str(payload.get("email") or "").strip().lower()
    flow = str(payload.get("flow") or "signin").strip().lower()
    client_redirect_to = str(
        payload.get("redirectTo") or payload.get("redirect_to") or ""
    ).strip()
    captcha_token = str(
        payload.get("captchaToken") or payload.get("turnstileToken") or ""
    ).strip()
    honeypot_value = str(payload.get(AUTH_EMAIL_HONEYPOT_FIELD) or "").strip()

    if flow not in {"signin", "signup"}:
        return _finalize({"error": "Invalid flow. Use signin or signup."}, 400)

    if not _is_valid_email(email):
        return _finalize({"error": "Enter a valid email address."}, 400)

    if honeypot_value:
        logger.warning(
            "Auth email honeypot triggered for flow=%s ip=%s",
            flow,
            _get_request_ip(),
        )
        return _finalize(_auth_email_generic_response(), 202)

    client_ip = _get_request_ip()
    allowed, reason = _is_auth_email_request_allowed(email, client_ip)
    if not allowed:
        logger.warning(
            "Auth email throttled flow=%s reason=%s ip=%s email_hash=%s",
            flow,
            reason,
            client_ip,
            _email_fingerprint(email),
        )
        return _finalize(_auth_email_generic_response(), 202)

    if AUTH_EMAIL_TURNSTILE_SECRET:
        captcha_ok = _verify_turnstile_token(captcha_token, client_ip)
        if not captcha_ok:
            logger.warning(
                "Auth email captcha verification failed flow=%s ip=%s email_hash=%s",
                flow,
                client_ip,
                _email_fingerprint(email),
            )
            if AUTH_EMAIL_REQUIRE_TURNSTILE:
                return _finalize(_auth_email_generic_response(), 202)

    try:
        _send_supabase_email_link(email, flow, client_redirect_to=client_redirect_to)
    except Exception as exc:
        # Do not leak delivery failures to clients to reduce account enumeration signal.
        logger.warning("Email auth request failed for flow=%s: %s", flow, exc)

    return _finalize(_auth_email_generic_response(), 202)


@app.route('/api/agent/pairing-challenge', methods=['POST'])
@require_supabase_user
@limiter.limit("12 per hour")
def api_agent_pairing_challenge():
    """Request an email link to authorize a pair/unpair action."""
    config_status = _agent_broker_config_status()
    if not config_status["ok"]:
        return jsonify({
            "error": "Agent broker misconfigured",
            "missing": config_status["missing"],
            "warnings": config_status["warnings"],
        }), 503

    if not request.is_json:
        return jsonify({"error": "Expected JSON payload"}), 400

    payload = request.get_json(silent=True) or {}
    action = str(payload.get("action") or "").strip().lower()
    if action not in {"pair", "unpair"}:
        return jsonify({"error": "Invalid action. Use pair or unpair."}), 400

    user = getattr(g, "supabase_user", {}) or {}
    user_id = _get_supabase_user_id(user)
    user_email = _get_supabase_user_email(user)
    if not user_id or not user_email:
        return jsonify({"error": "Authenticated user must have a valid email address."}), 400

    pairing_code = ""
    if action == "pair":
        pairing_code = str(
            payload.get("pairing_code")
            or payload.get("pairingCode")
            or ""
        ).strip()
        if not pairing_code:
            return jsonify({"error": "Pairing code required for pair action."}), 400
        if not PAIRING_CODE_PATTERN.match(pairing_code):
            return jsonify({"error": "Pairing code must be a 6-digit value."}), 400

    client_redirect_to = str(
        payload.get("redirectTo") or payload.get("redirect_to") or ""
    ).strip()
    client_ip = _get_request_ip()
    allowed, reason, retry_after_seconds = _is_agent_pairing_action_allowed(
        user_id,
        action,
    )
    if not allowed:
        logger.warning(
            "Pairing challenge throttled action=%s reason=%s user=%s ip=%s",
            action,
            reason,
            _email_fingerprint(user_email),
            client_ip,
        )
        response = jsonify({
            "error": "Too many verification requests. Please wait and try again.",
            "reason": reason,
            "retry_after_seconds": retry_after_seconds,
        })
        if retry_after_seconds > 0:
            response.headers["Retry-After"] = str(retry_after_seconds)
        return response, 429

    challenge_id, expires_at = _create_agent_pairing_challenge(
        action=action,
        user_id=user_id,
        email=user_email,
        pairing_code=pairing_code,
        client_ip=client_ip,
    )

    try:
        _send_supabase_email_link(
            user_email,
            "signin",
            client_redirect_to=client_redirect_to,
            redirect_path=AGENT_PAIRING_REDIRECT_PATH,
            redirect_query={
                "agent_action": action,
                "agent_challenge": challenge_id,
            },
        )
    except Exception as exc:
        logger.warning(
            "Agent pairing challenge email failed action=%s user=%s ip=%s: %s",
            action,
            _email_fingerprint(user_email),
            client_ip,
            exc,
        )
        with AGENT_PAIRING_CHALLENGE_LOCK:
            AGENT_PAIRING_CHALLENGES.pop(challenge_id, None)
        return jsonify({
            "error": "Unable to send verification email right now. Please retry.",
        }), 502

    return jsonify({
        "ok": True,
        "action": action,
        "message": "Verification link sent to your email.",
        "expires_at": datetime.utcfromtimestamp(expires_at).isoformat() + "Z",
    }), 202


@app.route('/api/agent/pairing-confirm', methods=['POST'])
@require_supabase_user
@limiter.limit("20 per hour")
def api_agent_pairing_confirm():
    """Confirm a pair/unpair action using the emailed challenge link."""
    config_status = _agent_broker_config_status()
    if not config_status["ok"]:
        return jsonify({
            "error": "Agent broker misconfigured",
            "missing": config_status["missing"],
            "warnings": config_status["warnings"],
        }), 503

    if not request.is_json:
        return jsonify({"error": "Expected JSON payload"}), 400

    payload = request.get_json(silent=True) or {}
    challenge_id = str(
        payload.get("challenge_id")
        or payload.get("challengeId")
        or ""
    ).strip()
    if not challenge_id:
        return jsonify({"error": "challenge_id is required"}), 400
    if not PAIRING_CHALLENGE_ID_PATTERN.match(challenge_id):
        return jsonify({"error": "Invalid challenge_id format."}), 400

    user = getattr(g, "supabase_user", {}) or {}
    user_id = _get_supabase_user_id(user)
    user_email = _get_supabase_user_email(user)
    if not user_id or not user_email:
        return jsonify({"error": "Authenticated user must have a valid email address."}), 400
    client_ip = _get_request_ip()

    blocked, retry_after_seconds = _is_agent_pairing_confirm_blocked(user_id, client_ip)
    if blocked:
        response = jsonify({
            "error": "Too many invalid verification attempts. Please request a new link and try later.",
            "retry_after_seconds": retry_after_seconds,
        })
        response.headers["Retry-After"] = str(retry_after_seconds)
        return response, 429

    challenge, reason = _consume_agent_pairing_challenge(
        challenge_id=challenge_id,
        user_id=user_id,
        email=user_email,
    )
    if not challenge:
        should_count_failure = reason in {"missing", "user-mismatch", "email-mismatch"}
        if should_count_failure:
            blocked_after_failure, blocked_retry_after = _register_agent_pairing_confirm_failure(
                user_id,
                client_ip,
            )
            if blocked_after_failure:
                response = jsonify({
                    "error": "Too many invalid verification attempts. Please request a new link and try later.",
                    "retry_after_seconds": blocked_retry_after,
                })
                response.headers["Retry-After"] = str(blocked_retry_after)
                return response, 429
        if reason == "expired":
            return jsonify({"error": "Verification link expired. Request a new one."}), 410
        if reason in {"user-mismatch", "email-mismatch"}:
            return jsonify({"error": "Verification link does not match this user."}), 403
        return jsonify({"error": "Invalid verification link."}), 400

    action = str(challenge.get("action") or "").strip().lower()
    if action == "pair":
        pairing_code = str(challenge.get("pairing_code") or "").strip()
        if not pairing_code:
            _register_agent_pairing_confirm_failure(user_id, client_ip)
            return jsonify({"error": "Pairing challenge is missing code."}), 400
        response = _pair_agent_session_for_user(
            pairing_code,
            user_id,
            extra_payload={
                "verified": True,
                "action": "pair",
            },
        )
        if response[1] < 400:
            _clear_agent_pairing_confirm_failures(user_id, client_ip)
        return response

    if action == "unpair":
        session = _get_agent_session()
        session_token = ""
        if session and session.get("user_id") == user_id:
            session_token = str(session.get("token") or "")

        revoke_response, revoke_status = _revoke_gateway_agent_token(session_token)
        if revoke_status >= 500:
            return revoke_response, revoke_status

        _clear_agent_session_for_request()
        payload = revoke_response.get_json(silent=True) or {}
        payload.update({
            "paired": False,
            "verified": True,
            "action": "unpair",
        })
        _clear_agent_pairing_confirm_failures(user_id, client_ip)
        resp = jsonify(payload)
        resp.delete_cookie(AGENT_SESSION_COOKIE, path="/")
        return resp, 200

    _register_agent_pairing_confirm_failure(user_id, client_ip)
    return jsonify({"error": "Unsupported challenge action."}), 400


@app.route('/api/agent/health', methods=['GET'])
@require_supabase_user
def api_agent_health():
    """Proxy ZeroClaw health check through the backend."""
    config_status = _agent_broker_config_status()
    if not config_status["ok"]:
        return jsonify({
            "error": "Agent broker misconfigured",
            "missing": config_status["missing"],
            "warnings": config_status["warnings"],
        }), 503
    try:
        response = requests.get(
            f"{AGENT_GATEWAY_URL.rstrip('/')}/health",
            timeout=3,
        )
        return jsonify(response.json()), response.status_code
    except Exception as exc:
        logger.warning("Agent health proxy failed: %s", exc)
        return jsonify({"error": "Agent gateway unavailable"}), 503


@app.route('/api/agent/config', methods=['GET'])
@require_supabase_user
def api_agent_config():
    """Expose broker configuration readiness (no secrets)."""
    return jsonify(_agent_broker_config_status()), 200


def _pair_agent_session_for_user(
    pairing_code: str,
    user_id: str,
    extra_payload: Optional[Dict[str, Any]] = None,
):
    try:
        response = requests.post(
            f"{AGENT_GATEWAY_URL.rstrip('/')}/pair",
            headers={"X-Pairing-Code": pairing_code},
            timeout=10,
        )
    except Exception as exc:
        logger.warning("Pairing proxy failed: %s", exc)
        return jsonify({"error": "Agent gateway unavailable"}), 503

    if response.status_code != 200:
        return jsonify({
            "error": "Pairing failed",
            "details": response.text,
        }), response.status_code

    try:
        data = response.json()
    except Exception:
        data = {}

    token = data.get("token")
    if not token:
        return jsonify({"error": "Gateway did not return a token"}), 502

    session_id, expires_at = _create_agent_session(token, user_id)
    response_payload = {
        "paired": True,
        "expires_at": datetime.utcfromtimestamp(expires_at).isoformat() + "Z",
    }
    if extra_payload:
        response_payload.update(extra_payload)
    resp = jsonify(response_payload)
    resp.set_cookie(
        AGENT_SESSION_COOKIE,
        session_id,
        httponly=True,
        samesite=AGENT_SESSION_SAMESITE,
        secure=AGENT_SESSION_SECURE,
        max_age=AGENT_SESSION_TTL_SECONDS,
        path="/",
    )
    return resp, 200


def _revoke_gateway_agent_token(token: str):
    if not token:
        return jsonify({"revoked": False, "paired": False}), 200

    try:
        response = requests.post(
            f"{AGENT_GATEWAY_URL.rstrip('/')}/unpair",
            headers={
                "Authorization": f"Bearer {token}",
            },
            timeout=10,
        )
    except Exception as exc:
        logger.warning("Unpair proxy failed: %s", exc)
        return jsonify({"error": "Agent gateway unavailable"}), 503

    if response.status_code in (200, 401, 403, 404):
        payload: Dict[str, Any] = {}
        try:
            payload = response.json()
        except Exception:
            payload = {}
        return jsonify({
            "revoked": response.status_code == 200 or payload.get("revoked") is True,
            "gateway_status": response.status_code,
            "paired": bool(payload.get("paired")),
            "pairing_code": payload.get("pairing_code"),
        }), 200

    return jsonify({
        "error": "Gateway unpair failed",
        "details": response.text,
    }), response.status_code


@app.route('/api/agent/session', methods=['GET'])
@require_supabase_user
def api_agent_session():
    """Return whether a valid agent session cookie exists."""
    config_status = _agent_broker_config_status()
    if not config_status["ok"]:
        return jsonify({
            "paired": False,
            "error": "Agent broker misconfigured",
            "missing": config_status["missing"],
            "warnings": config_status["warnings"],
        }), 503
    session = _get_agent_session()
    user_id = _get_supabase_user_id(getattr(g, "supabase_user", {}) or {})
    if session and user_id and session.get("user_id") == user_id:
        return jsonify({
            "paired": True,
            "expires_at": datetime.utcfromtimestamp(session["expires_at"]).isoformat() + "Z",
        })
    return jsonify({"paired": False})


@app.route('/api/agent/pair', methods=['POST'])
@require_supabase_user
@limiter.limit("10 per hour")
def api_agent_pair():
    return jsonify({
        "error": "Direct pair is disabled. Request email verification first.",
        "next": [
            "POST /api/agent/pairing-challenge",
            "POST /api/agent/pairing-confirm",
        ],
    }), 428


@app.route('/api/agent/unpair', methods=['POST'])
@require_supabase_user
def api_agent_unpair():
    return jsonify({
        "error": "Direct unpair is disabled. Request email verification first.",
        "next": [
            "POST /api/agent/pairing-challenge",
            "POST /api/agent/pairing-confirm",
        ],
    }), 428


@app.route('/api/agent/session/clear', methods=['POST'])
@require_supabase_user
def api_agent_session_clear():
    _clear_agent_session_for_request()

    resp = jsonify({"paired": False})
    resp.delete_cookie(AGENT_SESSION_COOKIE, path="/")
    return resp


@app.route('/api/agent/webhook', methods=['POST'])
@require_supabase_user
@require_agent_session
def api_agent_webhook():
    """Proxy webhook requests to ZeroClaw using the server-side session token."""
    config_status = _agent_broker_config_status()
    if not config_status["ok"]:
        return jsonify({
            "error": "Agent broker misconfigured",
            "missing": config_status["missing"],
            "warnings": config_status["warnings"],
        }), 503
    if not request.is_json:
        return jsonify({"error": "Expected JSON payload"}), 400

    if AGENT_REQUIRE_WEBHOOK_SECRET and not AGENT_WEBHOOK_SECRET:
        return jsonify({
            "error": "Agent webhook secret is required but not configured.",
            "missing": ["AGENT_WEBHOOK_SECRET"],
        }), 503

    payload = request.get_json(silent=False)
    raw_message = payload.get("message")
    task_name = ""
    try:
        if isinstance(raw_message, str):
            parsed = json.loads(raw_message)
            if isinstance(parsed, dict):
                task_name = str(parsed.get("task") or "").strip()
                timeout_ms = parsed.get("timeout")
            else:
                timeout_ms = None
        else:
            timeout_ms = None
    except Exception:
        timeout_ms = None

    user = getattr(g, "supabase_user", {}) or {}
    if not task_name and not _is_admin_user(user):
        return jsonify({"error": "Agent task not allowed for this user"}), 403
    if task_name and not _is_agent_task_allowed(task_name, user):
        return jsonify({"error": "Agent task not allowed for this user"}), 403

    timeout_seconds = AGENT_DEFAULT_TIMEOUT_SECONDS
    if isinstance(timeout_ms, (int, float)) and timeout_ms > 0:
        timeout_seconds = min(
            max(int(timeout_ms / 1000), AGENT_DEFAULT_TIMEOUT_SECONDS),
            AGENT_MAX_TIMEOUT_SECONDS,
        )

    session = getattr(g, "agent_session", {})
    token = session.get("token")
    if not token:
        return jsonify({"error": "Agent session missing"}), 401

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    if AGENT_WEBHOOK_SECRET:
        headers["X-Webhook-Secret"] = AGENT_WEBHOOK_SECRET

    try:
        response = requests.post(
            f"{AGENT_GATEWAY_URL.rstrip('/')}/webhook",
            headers=headers,
            json=payload,
            timeout=timeout_seconds,
        )
    except Exception as exc:
        logger.warning("Agent webhook proxy failed: %s", exc)
        return jsonify({"error": "Agent gateway unavailable"}), 503

    if response.status_code in (401, 403):
        # Clear invalid session
        session_id = request.cookies.get(AGENT_SESSION_COOKIE)
        if session_id:
            AGENT_SESSIONS.pop(session_id, None)
        resp = jsonify({"error": "Agent session expired. Please pair again."})
        resp.delete_cookie(AGENT_SESSION_COOKIE, path="/")
        return resp, 401

    try:
        data = response.json()
        return jsonify(data), response.status_code
    except Exception:
        return (response.text or ""), response.status_code

@app.route('/api/status', methods=['GET'])
@require_api_key
def api_status():
    """
    Health check endpoint - returns detailed AutoCAD connection status
    """
    manager = get_manager()
    status = manager.get_status()
    status['backend_id'] = 'coordinates-grabber-api'
    status['backend_version'] = '1.0.0'

    http_code = 200 if status['autocad_running'] else 503

    return jsonify(status), http_code


@app.route('/api/layers', methods=['GET'])
@require_api_key
def api_layers():
    """
    List available layers in the active AutoCAD drawing
    
    Response:
    {
        "success": bool,
        "layers": [str],  # Array of layer names
        "count": int,
        "error": str|null
    }
    """
    manager = get_manager()
    success, layers, error = manager.get_layers()
    
    response = {
        'success': success,
        'layers': layers,
        'count': len(layers),
        'error': error
    }
    
    return jsonify(response), 200 if success else 503


@app.route('/api/selection-count', methods=['GET'])
@require_api_key
@limiter.limit("120 per hour")
def api_selection_count():
    """Get count of currently selected objects in AutoCAD (fresh COM)."""
    manager = get_manager()
    status = manager.get_status()
    
    if not status['drawing_open']:
        return jsonify({'success': False, 'count': 0, 'error': 'No drawing open'}), 503
    
    try:
        pythoncom.CoInitialize()
        acad = connect_autocad()
        if acad is None:
            return jsonify({'success': False, 'count': 0, 'error': 'Cannot connect to AutoCAD'}), 503
        
        doc = dyn(acad.ActiveDocument)
        
        # Try to delete an old temp selection set first
        try:
            old_ss = doc.SelectionSets.Item("TEMP_COUNT")
            old_ss.Delete()
        except Exception:
            pass
        
        ss = doc.SelectionSets.Add("TEMP_COUNT")
        ss.SelectOnScreen()
        count = ss.Count
        ss.Delete()
        
        return jsonify({'success': True, 'count': count, 'error': None})
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'count': 0, 'error': f'COM error: {str(e)}'}), 500
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


@app.route('/api/execute', methods=['POST'])
@require_api_key
@limiter.limit("30 per hour")
def api_execute():
    """
    Execute coordinate extraction based on provided configuration.
    Uses late-bound COM via the manager's execute_layer_search method.
    """
    manager = get_manager()
    status = manager.get_status()
    
    if not status['drawing_open']:
        return jsonify({
            'success': False,
            'message': 'No drawing open in AutoCAD',
            'points_created': 0,
            'error_details': 'Please open a drawing before executing'
        }), 400
    
    try:
        if not request.is_json:
            raise ValueError('Expected application/json payload')

        raw_config = request.get_json(silent=False)
        if not raw_config:
            raise ValueError('No configuration provided')
        
        # Validate and sanitize input
        config = validate_layer_config(raw_config)
        
        start_time = time.time()
        
        result = manager.execute_layer_search(config)
        
        duration = time.time() - start_time
        
        if result['success']:
            blocks_inserted = result.get('blocks_inserted', 0)
            block_errors = result.get('block_errors')
            layers = result.get('layers', [])
            if layers:
                msg = f'Extracted {result["count"]} points from {len(layers)} layer(s): {", ".join(layers)}'
            else:
                msg = f'Extracted {result["count"]} points'
            if blocks_inserted > 0:
                msg += f', inserted {blocks_inserted} reference blocks'
            if block_errors:
                msg += f' (warnings: {len(block_errors)})'
            return jsonify({
                'success': True,
                'message': msg,
                'points_created': result['count'],
                'blocks_inserted': blocks_inserted,
                'excel_path': result.get('excel_path', ''),
                'duration_seconds': round(duration, 2),
                'points': result['points'],
                'block_errors': block_errors,
                'error_details': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'message': result.get('error', 'No entities found'),
                'points_created': 0,
                'blocks_inserted': 0,
                'excel_path': '',
                'duration_seconds': round(duration, 2),
                'points': [],
                'error_details': result.get('error')
            }), 400
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Execution failed: {str(e)}',
            'points_created': 0,
            'error_details': str(e)
        }), 500


@app.route('/api/trigger-selection', methods=['POST'])
@require_api_key
@limiter.limit("120 per hour")
def api_trigger_selection():
    """Bring AutoCAD to foreground (fresh COM)."""
    manager = get_manager()
    status = manager.get_status()
    
    if not status['drawing_open']:
        return jsonify({'success': False, 'message': 'No drawing open'}), 503
    
    try:
        pythoncom.CoInitialize()
        acad = connect_autocad()
        if acad is None:
            return jsonify({'success': False, 'message': 'Cannot connect to AutoCAD'}), 503
        
        acad.Visible = True
        acad.WindowState = 1  # Restore if minimized
        
        return jsonify({'success': True, 'message': 'AutoCAD activated'})
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


@app.route('/api/transmittal/render', methods=['POST'])
@require_api_key
@limiter.limit("30 per hour")
def api_transmittal_render():
    """
    Render a transmittal document (standard or CID) and return a DOCX file.
    Expects multipart/form-data with:
      - type: "standard" | "cid"
      - mode: "preview" | "generate" (optional)
      - format: "docx" | "pdf" | "both" (optional)
      - template: DOCX file
      - index: Excel file (standard only)
      - documents: PDF/CID files (standard only)
      - cid_files: CID files (cid only)
      - fields: JSON string
      - checks: JSON string
      - contacts: JSON string
      - cid_index_data: JSON string (cid only)
    """
    if not TRANSMITTAL_RENDER_AVAILABLE:
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Transmittal render helpers not available on server.",
                }
            ),
            503,
        )

    try:
        transmittal_type = request.form.get("type", "standard").lower()
        mode = request.form.get("mode", "generate").lower()
        output_format = request.form.get("format", "docx").lower()
        if output_format not in {"docx", "pdf", "both"}:
            output_format = "docx"
        fields = _parse_json_field("fields", {}) or {}
        if not isinstance(fields, dict):
            fields = {}
        checks = _parse_json_field("checks", {}) or {}
        if not isinstance(checks, dict):
            checks = {}
        contacts = _parse_json_field("contacts", []) or []
        if not isinstance(contacts, list):
            contacts = []
        cid_index_data = _parse_json_field("cid_index_data", []) or []
        if not isinstance(cid_index_data, list):
            cid_index_data = []

        profile_options = _load_transmittal_profiles_payload()
        available_profiles = profile_options.get("profiles", [])
        available_firms = set(profile_options.get("firm_numbers", []))
        defaults = (
            profile_options.get("defaults", {})
            if isinstance(profile_options.get("defaults"), dict)
            else {}
        )
        requested_profile_id = str(
            fields.get("from_profile_id")
            or fields.get("fromProfileId")
            or ""
        ).strip()
        if requested_profile_id:
            selected_profile = next(
                (p for p in available_profiles if p.get("id") == requested_profile_id),
                None,
            )
            if not selected_profile:
                return jsonify(
                    {
                        "success": False,
                        "message": "Invalid transmittal profile selection.",
                    }
                ), 400

            fields["from_profile_id"] = requested_profile_id
            fields["from_name"] = selected_profile.get("name", "")
            fields["from_title"] = selected_profile.get("title", "")
            fields["from_email"] = selected_profile.get("email", "")
            fields["from_phone"] = selected_profile.get("phone", "")

        firm_value = str(fields.get("firm") or "").strip()
        if firm_value and available_firms and firm_value not in available_firms:
            return jsonify(
                {
                    "success": False,
                    "message": "Invalid firm selection.",
                }
            ), 400
        if not firm_value:
            default_firm = str(defaults.get("firm") or "").strip()
            if default_firm:
                fields["firm"] = default_firm

        # Normalize checks with defaults
        default_checks = {
            "trans_pdf": False,
            "trans_cad": False,
            "trans_originals": False,
            "via_email": False,
            "via_ftp": False,
            "ci_approval": False,
            "ci_bid": False,
            "ci_construction": False,
            "ci_asbuilt": False,
            "ci_reference": False,
            "ci_preliminary": False,
            "ci_info": False,
            "ci_fab": False,
            "ci_const": False,
            "ci_record": False,
            "ci_ref": False,
            "vr_approved": False,
            "vr_approved_noted": False,
            "vr_rejected": False,
        }
        merged_checks = {**default_checks, **checks}
        if not merged_checks.get("ci_const"):
            merged_checks["ci_const"] = merged_checks.get("ci_construction", False)
        if not merged_checks.get("ci_ref"):
            merged_checks["ci_ref"] = merged_checks.get("ci_reference", False)

        # Normalize contacts to expected keys
        normalized_contacts = []
        for c in contacts:
            if not isinstance(c, dict):
                continue
            normalized_contacts.append(
                {
                    "name": str(c.get("name", "")).strip(),
                    "company": str(c.get("company", "")).strip(),
                    "email": str(c.get("email", "")).strip(),
                    "phone": str(c.get("phone", "")).strip(),
                }
            )

        work_dir = tempfile.mkdtemp(prefix="transmittal_")
        _schedule_cleanup(work_dir)
        template_file = request.files.get("template")
        if not template_file:
            return jsonify({"success": False, "message": "Template file is required"}), 400

        template_path = _save_upload(template_file, work_dir, "template.docx")

        project_num = str(fields.get("job_num", "")).strip() or "UNKNOWN"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_stem = "Transmittal"

        if transmittal_type == "cid":
            cid_files = request.files.getlist("cid_files")
            if not cid_files:
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "CID files are required for CID transmittal",
                        }
                    ),
                    400,
                )
            if not cid_index_data:
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "CID document index data is required",
                        }
                    ),
                    400,
                )

            cid_dir = os.path.join(work_dir, "cid_files")
            os.makedirs(cid_dir, exist_ok=True)
            for f in cid_files:
                _save_upload(f, cid_dir)

            output_stem = f"CID_Transmittal_{project_num}_{timestamp}"
            output_name = f"{output_stem}.docx"
            out_path = os.path.join(work_dir, output_name)

            render_cid_transmittal(
                template_path,
                cid_dir,
                cid_index_data,
                fields,
                merged_checks,
                normalized_contacts,
                out_path,
            )
        else:
            index_file = request.files.get("index")
            if not index_file:
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "Drawing index (Excel) file is required",
                        }
                    ),
                    400,
                )
            index_path = _save_upload(index_file, work_dir, "index.xlsx")

            document_files = request.files.getlist("documents")
            if not document_files:
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "Document files are required for standard transmittal",
                        }
                    ),
                    400,
                )

            docs_dir = os.path.join(work_dir, "documents")
            os.makedirs(docs_dir, exist_ok=True)
            for f in document_files:
                _save_upload(f, docs_dir)

            output_stem = f"Transmittal_{project_num}_{timestamp}"
            output_name = f"{output_stem}.docx"
            out_path = os.path.join(work_dir, output_name)

            render_transmittal(
                template_path,
                docs_dir,
                index_path,
                fields,
                merged_checks,
                normalized_contacts,
                out_path,
                None,
            )

        pdf_path = None
        if output_format in {"pdf", "both"}:
            pdf_path, pdf_error = _convert_docx_to_pdf(out_path, work_dir)
            if not pdf_path:
                return (
                    jsonify(
                        {
                            "success": False,
                            "message": "PDF conversion failed.",
                            "detail": pdf_error,
                        }
                    ),
                    500,
                )

        if output_format == "pdf" and pdf_path:
            return send_file(
                pdf_path,
                as_attachment=True,
                download_name=f"{output_stem}.pdf",
                mimetype="application/pdf",
            )

        if output_format == "both" and pdf_path:
            zip_name = f"{output_stem}.zip"
            zip_path = os.path.join(work_dir, zip_name)
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.write(out_path, arcname=os.path.basename(out_path))
                zf.write(pdf_path, arcname=os.path.basename(pdf_path))
            return send_file(
                zip_path,
                as_attachment=True,
                download_name=zip_name,
                mimetype="application/zip",
            )

        return send_file(
            out_path,
            as_attachment=True,
            download_name=output_name,
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )

    except Exception as e:
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/transmittal/profiles', methods=['GET'])
@require_api_key
@limiter.limit("120 per hour")
def api_transmittal_profiles():
    """Return allowed sender profiles and firm names for transmittal generation."""
    payload = _load_transmittal_profiles_payload()
    return jsonify(
        {
            "success": True,
            "profiles": payload.get("profiles", []),
            "firm_numbers": payload.get("firm_numbers", []),
            "defaults": payload.get("defaults", {}),
        }
    )


@app.route('/api/transmittal/template', methods=['GET'])
@require_api_key
@limiter.limit("60 per hour")
def api_transmittal_template():
    """Download the example transmittal DOCX template bundled with the repo."""
    template_path = TRANSMITTAL_TEMPLATE_PATH
    if not template_path.exists():
        return (
            jsonify(
                {
                    "success": False,
                    "message": "Example template not found on server.",
                }
            ),
            404,
        )
    return send_file(
        str(template_path),
        as_attachment=True,
        download_name=template_path.name,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


@app.route('/api/batch-find-replace/session', methods=['POST'])
@limiter.limit("60 per hour")
def api_batch_find_replace_session():
    token = _create_batch_session_token()
    response = jsonify({"success": True})
    response.set_cookie(
        BATCH_SESSION_COOKIE,
        token,
        httponly=True,
        samesite="Strict",
        secure=request.is_secure,
        max_age=BATCH_SESSION_TTL_SECONDS,
        path="/api/batch-find-replace/",
    )
    return response


@app.route('/api/batch-find-replace/preview', methods=['POST'])
@require_batch_session_or_api_key
@limiter.limit("30 per hour")
def api_batch_find_replace_preview():
    try:
        result = _process_batch_files(preview_only=True)
        return jsonify(
            {
                "success": True,
                "matches": result["matches"],
                "files_processed": result["files_processed"],
                "files_changed": result["files_changed"],
                "replacements": result["replacements"],
                "message": (
                    f"Preview completed: {result['replacements']} replacement(s) "
                    f"across {result['files_changed']} file(s)."
                ),
            }
        )
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:
        logger.exception("Batch preview failed")
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route('/api/batch-find-replace/apply', methods=['POST'])
@require_batch_session_or_api_key
@limiter.limit("20 per hour")
def api_batch_find_replace_apply():
    try:
        result = _process_batch_files(preview_only=False)
        report_path, report_dir = export_batch_changes_to_excel(result["matches"])
        _schedule_cleanup(report_dir)

        if hasattr(os, "startfile"):
            try:
                os.startfile(report_path)
            except Exception as exc:
                logger.warning("Could not auto-open Excel report: %s", exc)

        return send_file(
            report_path,
            as_attachment=True,
            download_name=os.path.basename(report_path),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:
        logger.exception("Batch apply failed")
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route('/health', methods=['GET'])
def health():
    """Simple health check for backend server"""
    return jsonify({
        'status': 'running',
        'server': 'Coordinates Grabber API',
        'backend_id': 'coordinates-grabber-api',
        'version': '1.0.0',
        'timestamp': time.time()
    })


@sock.route('/ws')
def websocket_status_bridge(ws):
    """WebSocket status stream for frontend real-time backend/AutoCAD connectivity updates."""
    provided_key = request.args.get('api_key')
    if not is_valid_api_key(provided_key):
        try:
            ws.send(json.dumps({
                'type': 'error',
                'message': 'Invalid API key',
                'code': 'AUTH_INVALID'
            }))
        finally:
            try:
                ws.close()
            except Exception:
                pass
        logger.warning("Unauthorized websocket connection attempt from %s", request.remote_addr)
        return

    logger.info("WebSocket connected from %s", request.remote_addr)

    try:
        ws.send(json.dumps({
            'type': 'connected',
            'backend_id': 'coordinates-grabber-api',
            'backend_version': '1.0.0',
            'timestamp': time.time(),
        }))

        while True:
            manager = get_manager()
            status = manager.get_status(force_refresh=True)

            ws.send(json.dumps({
                'type': 'status',
                'backend_id': 'coordinates-grabber-api',
                'backend_version': '1.0.0',
                'connected': bool(status.get('connected')),
                'autocad_running': bool(status.get('autocad_running')),
                'drawing_open': bool(status.get('drawing_open')),
                'drawing_name': status.get('drawing_name'),
                'error': status.get('error'),
                'checks': status.get('checks', {}),
                'timestamp': time.time(),
            }))

            try:
                incoming = ws.receive(timeout=0.1)
                if incoming is None:
                    pass
            except TypeError:
                pass
            except Exception:
                pass

            time.sleep(2.0)

    except Exception as exc:
        logger.info("WebSocket disconnected from %s (%s)", request.remote_addr, exc)


# ========== MAIN ==========

if __name__ == '__main__':
    api_host = os.environ.get('API_HOST', '127.0.0.1').strip() or '127.0.0.1'
    api_port = _parse_int_env('API_PORT', 5000, minimum=1)

    print("=" * 60)
    print("🚀 Coordinates Grabber API Server")
    print("=" * 60)
    print(f"Server starting on: http://{api_host}:{api_port}")
    print(f"Health check: http://{api_host}:{api_port}/health")
    print(f"Status endpoint: http://{api_host}:{api_port}/api/status")
    print("")
    print("📋 Prerequisites:")
    print("  - AutoCAD must be running")
    print("  - A drawing must be open in AutoCAD")
    print("  - React frontend should connect to localhost:5000")
    print("")
    print("Press Ctrl+C to stop the server")
    print("=" * 60)
    
    # Initialize manager to show initial status
    manager = get_manager()
    initial_status = manager.get_status()
    
    if initial_status['autocad_running']:
        print(f"✅ AutoCAD detected: {initial_status['autocad_path']}")
        if initial_status['drawing_open']:
            print(f"✅ Drawing open: {initial_status['drawing_name']}")
        else:
            print("⚠️  No drawing is currently open")
    else:
        print("❌ AutoCAD not detected - waiting for it to start...")
    
    print("=" * 60)
    print("")
    
    # Run Flask server
    app.run(
        host=api_host,
        port=api_port,
        debug=False,  # Set to True for development
        threaded=True
    )
