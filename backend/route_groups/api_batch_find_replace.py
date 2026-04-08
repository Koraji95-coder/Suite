from __future__ import annotations

import hashlib
import hmac
import json
import os
import re
import tempfile
import time
import uuid
from datetime import datetime
from functools import wraps
from typing import Any, Callable, Dict, List, Optional, Tuple

from flask import Blueprint, jsonify, request, send_file
from ..response_helpers import make_error_response, make_response
from flask_limiter import Limiter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

MAX_BATCH_FILES = 50
MAX_BATCH_RULES = 100
MAX_PREVIEW_MATCHES = 500
MAX_APPLY_CHANGE_ROWS = 5000


def create_batch_find_replace_blueprint(
    *,
    limiter: Limiter,
    logger: Any,
    is_valid_api_key: Callable[[Optional[str]], bool],
    api_key: str,
    schedule_cleanup: Callable[[str], None],
    batch_session_cookie: str,
    batch_session_ttl_seconds: int,
    send_autocad_dotnet_command: Optional[Callable[[str, Dict[str, Any]], Dict[str, Any]]] = None,
    send_autocad_acade_command: Optional[Callable[[str, Dict[str, Any]], Dict[str, Any]]] = None,
) -> Blueprint:
    """Create /api/batch-find-replace route group blueprint."""
    bp = Blueprint("batch_find_replace_api", __name__, url_prefix="/api/batch-find-replace")
    generated_reports: Dict[str, Dict[str, str]] = {}

    def _create_batch_session_token() -> str:
        timestamp = int(time.time())
        ts_bytes = str(timestamp).encode("utf-8")
        signature = hmac.new(api_key.encode("utf-8"), ts_bytes, hashlib.sha256).hexdigest()
        return f"{timestamp}.{signature}"

    def _is_valid_batch_session(token: Optional[str]) -> bool:
        if not token:
            return False
        try:
            ts_str, signature = token.split(".", 1)
            timestamp = int(ts_str)
        except Exception:
            return False

        if timestamp <= 0:
            return False
        if (time.time() - timestamp) > batch_session_ttl_seconds:
            return False

        expected = hmac.new(api_key.encode("utf-8"), ts_str.encode("utf-8"), hashlib.sha256).hexdigest()
        return hmac.compare_digest(signature, expected)

    def require_batch_session_or_api_key(f):
        """Allow either API key header or a signed batch session cookie."""

        @wraps(f)
        def decorated_function(*args, **kwargs):
            provided_key = request.headers.get("X-API-Key")
            if provided_key and is_valid_api_key(provided_key):
                return f(*args, **kwargs)

            token = request.cookies.get(batch_session_cookie)
            if not _is_valid_batch_session(token):
                logger.warning(
                    "Unauthorized batch request: %s from %s",
                    request.path,
                    request.remote_addr,
                )
                return make_error_response("Batch session required", code="AUTH_REQUIRED", status=401)

            return f(*args, **kwargs)

        return decorated_function

    def _parse_batch_rules() -> List[Dict[str, Any]]:
        raw_rules = request.form.get("rules", "")
        if not raw_rules:
            raise ValueError("Missing rules payload")

        try:
            parsed = json.loads(raw_rules)
        except Exception:
            raise ValueError("Rules payload is not valid JSON")

        if not isinstance(parsed, list):
            raise ValueError("Rules payload must be an array")
        if len(parsed) == 0:
            raise ValueError("At least one rule is required")
        if len(parsed) > MAX_BATCH_RULES:
            raise ValueError(f"Too many rules. Maximum is {MAX_BATCH_RULES}")

        rules: List[Dict[str, Any]] = []
        for idx, item in enumerate(parsed):
            if not isinstance(item, dict):
                continue
            find_text = str(item.get("find", "")).strip()
            if not find_text:
                continue

            rules.append(
                {
                    "id": str(item.get("id", f"rule-{idx + 1}")),
                    "find": find_text,
                    "replace": str(item.get("replace", "")),
                    "use_regex": bool(item.get("useRegex", False)),
                    "match_case": bool(item.get("matchCase", False)),
                }
            )

        if not rules:
            raise ValueError("No valid rules provided")

        return rules

    def _decode_uploaded_text(raw_bytes: bytes) -> str:
        if b"\x00" in raw_bytes:
            raise ValueError("Binary files are not supported")

        for encoding in ("utf-8", "utf-16", "latin-1"):
            try:
                return raw_bytes.decode(encoding)
            except Exception:
                continue

        raise ValueError("Unable to decode file as text")

    def _build_batch_pattern(rule: Dict[str, Any]) -> re.Pattern[str]:
        flags = 0 if rule["match_case"] else re.IGNORECASE
        if rule["use_regex"]:
            try:
                return re.compile(rule["find"], flags)
            except re.error as exc:
                raise ValueError(f"Invalid regex for rule '{rule['id']}': {exc}")

        return re.compile(re.escape(rule["find"]), flags)

    def _apply_rule_to_lines(
        lines: List[str],
        pattern: re.Pattern[str],
        replacement: str,
        rule_id: str,
        file_name: str,
        preview_matches: List[Dict[str, Any]],
        max_matches: int,
    ) -> Tuple[List[str], int]:
        changed_count = 0
        next_lines: List[str] = []

        for line_number, line in enumerate(lines, start=1):
            updated_line, replaced = pattern.subn(replacement, line)
            if replaced > 0:
                changed_count += replaced
                if len(preview_matches) < max_matches:
                    preview_matches.append(
                        {
                            "file": file_name,
                            "line": line_number,
                            "before": line[:500],
                            "after": updated_line[:500],
                            "ruleId": rule_id,
                        }
                    )
            next_lines.append(updated_line)

        return next_lines, changed_count

    def _process_batch_files(preview_only: bool) -> Dict[str, Any]:
        uploaded_files = request.files.getlist("files")
        if not uploaded_files:
            raise ValueError("No files uploaded")
        if len(uploaded_files) > MAX_BATCH_FILES:
            raise ValueError(f"Too many files. Maximum is {MAX_BATCH_FILES}")

        rules = _parse_batch_rules()
        max_matches = MAX_PREVIEW_MATCHES if preview_only else MAX_APPLY_CHANGE_ROWS
        preview_matches: List[Dict[str, Any]] = []
        updated_files: List[Dict[str, str]] = []
        files_changed = 0
        replacements_total = 0

        for file_storage in uploaded_files:
            file_name = secure_filename(file_storage.filename or "upload.txt")
            if not file_name:
                file_name = "upload.txt"

            try:
                raw_bytes = file_storage.read()
            finally:
                try:
                    file_storage.close()
                except Exception:
                    pass  # Best-effort cleanup of uploaded file handle
            content = _decode_uploaded_text(raw_bytes)

            line_break = "\r\n" if "\r\n" in content else "\n"
            lines = content.splitlines()
            file_replacements = 0

            for rule in rules:
                pattern = _build_batch_pattern(rule)
                lines, changed = _apply_rule_to_lines(
                    lines,
                    pattern,
                    rule["replace"],
                    rule["id"],
                    file_name,
                    preview_matches,
                    max_matches,
                )
                file_replacements += changed

            if file_replacements > 0:
                files_changed += 1
                replacements_total += file_replacements

            if not preview_only:
                updated_files.append(
                    {
                        "file": file_name,
                        "content": line_break.join(lines),
                    }
                )

        return {
            "matches": preview_matches,
            "files_changed": files_changed,
            "replacements": replacements_total,
            "files_processed": len(uploaded_files),
            "updated_files": updated_files,
        }

    def _next_request_id() -> str:
        return f"batch-{int(time.time() * 1000)}"

    def _current_request_id(prefix: str = "batch") -> str:
        header_value = str(request.headers.get("X-Request-ID") or "").strip()
        if header_value:
            return header_value
        return f"{prefix}-{int(time.time() * 1000)}"

    def _call_dotnet_bridge_action(action: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        if send_autocad_dotnet_command is None:
            raise RuntimeError("AutoCAD .NET bridge is not configured.")

        request_id = str(payload.get("requestId") or _next_request_id()).strip() or _next_request_id()
        response = send_autocad_dotnet_command(
            action,
            {
                **payload,
                "requestId": request_id,
            },
        )
        if not isinstance(response, dict):
            raise RuntimeError("Malformed response from AutoCAD .NET bridge.")
        if not response.get("ok"):
            raise RuntimeError(
                str(response.get("error") or response.get("message") or "Unknown bridge error.")
            )
        result_payload = response.get("result")
        if not isinstance(result_payload, dict):
            raise RuntimeError("Invalid .NET bridge result payload.")
        return result_payload

    def _call_acade_host_action(action: str, payload: Dict[str, Any]) -> Dict[str, Any]:
        if send_autocad_acade_command is None:
            raise RuntimeError("AutoCAD in-process ACADE host is not configured.")

        request_id = str(payload.get("requestId") or _next_request_id()).strip() or _next_request_id()
        response = send_autocad_acade_command(
            action,
            {
                **payload,
                "requestId": request_id,
            },
        )
        if not isinstance(response, dict):
            raise RuntimeError("Malformed response from the AutoCAD in-process ACADE host.")
        if not response.get("ok"):
            raise RuntimeError(
                str(
                    response.get("error")
                    or response.get("message")
                    or "Unknown in-process ACADE host error."
                )
            )

        result_payload = response.get("result")
        if not isinstance(result_payload, dict):
            raise RuntimeError("Invalid in-process ACADE host result payload.")
        return result_payload

    def _cleanup_status_code_for_host_result(result: Dict[str, Any]) -> int:
        code = str(result.get("code") or "").strip().upper()
        if code == "INVALID_REQUEST":
            return 400
        if code == "AUTOCAD_NOT_READY":
            return 503
        if code.endswith("_FAILED"):
            return 503
        return 422

    def _cleanup_error_response(
        *,
        code: str,
        message: str,
        request_id: str,
        status_code: int,
        warnings: Optional[List[str]] = None,
        meta: Optional[Dict[str, Any]] = None,
    ):
        return (
            jsonify(
                {
                    "success": False,
                    "code": code,
                    "message": message,
                    "requestId": request_id,
                    "data": None,
                    "warnings": warnings or [],
                    "meta": {"requestId": request_id, **(meta or {})},
                }
            ),
            status_code,
        )

    def _batch_error_response(*, message: str, status_code: int):
        return make_error_response(message, status=status_code)

    def _safe_batch_validation_message(exc: ValueError, fallback: str) -> str:
        message = str(exc or "").strip()
        canonical_messages = {
            "Missing rules payload": "Rules payload is required.",
            "Rules payload is not valid JSON": "Rules payload is not valid JSON.",
            "Rules payload must be an array": "Rules payload must be an array.",
            "At least one rule is required": "At least one rule is required.",
            "No valid rules provided": "No valid rules were provided.",
            "Binary files are not supported": "Binary files are not supported.",
            "Unable to decode file as text": "Unable to decode file as text.",
            "No files uploaded": "At least one file is required.",
            "rules must be an array": "rules must be an array.",
            "selectedDrawingPaths must contain at least one drawing path.": (
                "selectedDrawingPaths must contain at least one drawing path."
            ),
            "No valid project drawings were resolved from the issue set.": (
                "No valid project drawings were resolved from the issue set."
            ),
            "matches must contain at least one preview row.": (
                "matches must contain at least one preview row."
            ),
            "matches must contain at least one project preview row.": (
                "matches must contain at least one project preview row."
            ),
        }
        if message in canonical_messages:
            return canonical_messages[message]
        if message.startswith("Invalid regex for rule "):
            return "Invalid regex rule."
        if message.startswith("Too many rules."):
            return "Too many rules were provided."
        if message.startswith("Too many files."):
            return "Too many files were provided."
        if message.startswith("Too many selected drawings."):
            return "Too many selected drawings were provided."
        if message.startswith("matches must contain at least one preview row."):
            return "matches must contain at least one preview row."
        if message.startswith("matches must contain at least one project preview row."):
            return "matches must contain at least one project preview row."
        return fallback

    def _normalize_cleanup_host_result(
        result: Dict[str, Any], request_id: str
    ) -> Dict[str, Any]:
        normalized = dict(result)
        meta = dict(normalized.get("meta") or {})
        if not meta.get("requestId"):
            meta["requestId"] = request_id
        normalized["meta"] = meta
        normalized["requestId"] = str(normalized.get("requestId") or meta.get("requestId") or request_id)
        normalized.setdefault("warnings", [])
        normalized.setdefault("data", None)
        return normalized

    def _parse_batch_rules_from_json(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
        parsed = payload.get("rules")
        if not isinstance(parsed, list):
            raise ValueError("rules must be an array")

        if len(parsed) == 0:
            raise ValueError("At least one rule is required")
        if len(parsed) > MAX_BATCH_RULES:
            raise ValueError(f"Too many rules. Maximum is {MAX_BATCH_RULES}")

        rules: List[Dict[str, Any]] = []
        for idx, item in enumerate(parsed):
            if not isinstance(item, dict):
                continue
            find_text = str(item.get("find", "")).strip()
            if not find_text:
                continue
            rules.append(
                {
                    "id": str(item.get("id", f"rule-{idx + 1}")),
                    "find": find_text,
                    "replace": str(item.get("replace", "")),
                    "useRegex": bool(item.get("useRegex", False)),
                    "matchCase": bool(item.get("matchCase", False)),
                }
            )

        if not rules:
            raise ValueError("No valid rules provided")
        return rules

    def _is_absolute_windows_path(path: str) -> bool:
        normalized = path.strip()
        return bool(re.match(r"^[A-Za-z]:[\\/]", normalized)) or normalized.startswith("\\\\")

    def _resolve_project_drawings(payload: Dict[str, Any]) -> List[Dict[str, str]]:
        raw_selected = payload.get("selectedDrawingPaths")
        if not isinstance(raw_selected, list) or len(raw_selected) == 0:
            raise ValueError("selectedDrawingPaths must contain at least one drawing path.")
        if len(raw_selected) > MAX_BATCH_FILES:
            raise ValueError(f"Too many selected drawings. Maximum is {MAX_BATCH_FILES}")

        drawing_root_raw = str(payload.get("drawingRootPath") or "").strip()
        drawing_root = os.path.abspath(drawing_root_raw) if drawing_root_raw else ""
        drawings: List[Dict[str, str]] = []
        seen_paths: set[str] = set()

        for item in raw_selected:
            raw_path = str(item or "").strip()
            if not raw_path:
                continue

            if _is_absolute_windows_path(raw_path):
                absolute_path = os.path.abspath(raw_path)
                relative_path = raw_path
                if drawing_root:
                    try:
                        candidate_relative = os.path.relpath(absolute_path, drawing_root)
                        if not candidate_relative.startswith(".."):
                            relative_path = candidate_relative
                    except ValueError:
                        pass  # Paths on different drives (Windows); keep original path
            else:
                if not drawing_root:
                    raise ValueError(
                        "drawingRootPath is required when selectedDrawingPaths are relative."
                    )
                absolute_path = os.path.abspath(os.path.join(drawing_root, raw_path))
                try:
                    if os.path.commonpath([drawing_root, absolute_path]) != drawing_root:
                        raise ValueError(
                            f"Drawing path '{raw_path}' resolves outside the drawing root."
                        )
                except ValueError:
                    raise ValueError(
                        f"Drawing path '{raw_path}' could not be resolved under the drawing root."
                    )
                relative_path = raw_path

            normalized_key = absolute_path.lower()
            if normalized_key in seen_paths:
                continue
            seen_paths.add(normalized_key)
            drawings.append(
                {
                    "path": absolute_path,
                    "relativePath": relative_path,
                    "drawingName": os.path.basename(absolute_path) or absolute_path,
                }
            )

        if not drawings:
            raise ValueError("No valid project drawings were resolved from the issue set.")

        return drawings

    def _build_project_preview_drawings(
        drawings: List[Dict[str, str]],
        matches: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        counts: Dict[str, int] = {}
        for match in matches:
            drawing_path = str(match.get("drawingPath") or "").strip()
            if not drawing_path:
                continue
            counts[drawing_path] = counts.get(drawing_path, 0) + 1

        result: List[Dict[str, Any]] = []
        for drawing in drawings:
            drawing_path = drawing["path"]
            result.append(
                {
                    "drawingPath": drawing_path,
                    "drawingName": drawing["drawingName"],
                    "relativePath": drawing["relativePath"],
                    "matchCount": counts.get(drawing_path, 0),
                }
            )
        return result

    def _normalize_project_preview_matches(
        matches: List[Any],
        drawings: List[Dict[str, str]],
    ) -> List[Dict[str, Any]]:
        drawing_meta = {
            drawing["path"].lower(): drawing for drawing in drawings if drawing.get("path")
        }
        normalized: List[Dict[str, Any]] = []

        for index, item in enumerate(matches):
            if not isinstance(item, dict):
                continue
            drawing_path = str(item.get("drawingPath") or "").strip()
            drawing_name = str(item.get("drawingName") or "").strip()
            relative_path = str(item.get("relativePath") or "").strip()
            meta = drawing_meta.get(drawing_path.lower()) if drawing_path else None
            if meta:
                if not drawing_path:
                    drawing_path = meta["path"]
                if not drawing_name:
                    drawing_name = meta["drawingName"]
                if not relative_path:
                    relative_path = meta["relativePath"]

            file_name = str(item.get("file") or "").strip() or drawing_name or "Drawing"
            rule_id = str(item.get("ruleId") or "").strip()
            handle = str(item.get("handle") or "").strip()
            attribute_tag = str(item.get("attributeTag") or "").strip()
            before = str(item.get("before") or "")
            after = str(item.get("after") or "")
            group_key = str(item.get("groupKey") or drawing_path or file_name).strip()
            match_key = str(item.get("matchKey") or "").strip() or "::".join(
                [
                    group_key,
                    handle,
                    attribute_tag,
                    rule_id,
                    before,
                    after,
                    str(index),
                ]
            )

            normalized.append(
                {
                    **item,
                    "file": file_name,
                    "drawingPath": drawing_path or None,
                    "drawingName": drawing_name or file_name,
                    "relativePath": relative_path or None,
                    "groupKey": group_key,
                    "matchKey": match_key,
                }
            )

        return normalized

    def _register_generated_report(report_path: str, report_dir: str) -> Tuple[str, str]:
        report_id = uuid.uuid4().hex
        report_filename = os.path.basename(report_path)
        generated_reports[report_id] = {
            "path": report_path,
            "dir": report_dir,
            "filename": report_filename,
        }
        return report_id, report_filename

    def export_batch_changes_to_excel(changes: List[Dict[str, Any]]) -> Tuple[str, str]:
        """Export batch find/replace changes to a styled Excel report."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_dir = tempfile.mkdtemp(prefix="batch_find_replace_")
        out_path = os.path.join(out_dir, f"batch_find_replace_changes_{timestamp}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Changes"
        summary_ws = wb.create_sheet("Summary")

        headers = ["File", "Line", "Rule ID", "Before", "After"]

        title_fill = PatternFill("solid", fgColor="2B6CB5")
        title_font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
        header_fill = PatternFill("solid", fgColor="3A3F47")
        header_font = Font(bold=True, color="F0F0F0", size=11, name="Arial")
        alt_fill_even = PatternFill("solid", fgColor="E8E6E2")
        alt_fill_odd = PatternFill("solid", fgColor="D4D1CC")
        data_font = Font(size=10, color="2A2A2A", name="Arial")
        border_side = Side(style="thin", color="B0ADA8")
        all_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        header_border = Border(
            left=border_side,
            right=border_side,
            top=border_side,
            bottom=Side(style="medium", color="3A3F47"),
        )

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value="Batch Find & Replace Change Report")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = all_border
        for col_idx in range(2, len(headers) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = title_fill
            c.border = all_border

        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=2, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = header_border

        for row_idx, change in enumerate(changes, start=3):
            row_fill = alt_fill_even if (row_idx - 3) % 2 == 0 else alt_fill_odd
            row_values = [
                str(change.get("file", "")),
                int(change.get("line", 0) or 0),
                str(change.get("ruleId", "")),
                str(change.get("before", "")),
                str(change.get("after", "")),
            ]

            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = all_border
                cell.font = data_font
                if col_idx == 2:
                    cell.alignment = Alignment(horizontal="right", vertical="top")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        ws.column_dimensions[get_column_letter(1)].width = 42
        ws.column_dimensions[get_column_letter(2)].width = 10
        ws.column_dimensions[get_column_letter(3)].width = 16
        ws.column_dimensions[get_column_letter(4)].width = 60
        ws.column_dimensions[get_column_letter(5)].width = 60

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 22
        ws.freeze_panes = "A3"

        summary_headers = ["Metric", "Value"]
        summary_title_fill = PatternFill("solid", fgColor="2B6CB5")
        summary_title_font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
        summary_header_fill = PatternFill("solid", fgColor="3A3F47")
        summary_header_font = Font(bold=True, color="F0F0F0", size=11, name="Arial")
        summary_data_font = Font(size=10, color="2A2A2A", name="Arial")

        summary_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        summary_title_cell = summary_ws.cell(row=1, column=1, value="Batch Find & Replace Summary")
        summary_title_cell.font = summary_title_font
        summary_title_cell.fill = summary_title_fill
        summary_title_cell.alignment = Alignment(horizontal="center", vertical="center")
        summary_title_cell.border = all_border
        summary_ws.cell(row=1, column=2).fill = summary_title_fill
        summary_ws.cell(row=1, column=2).border = all_border

        for col_idx, header in enumerate(summary_headers, start=1):
            c = summary_ws.cell(row=2, column=col_idx, value=header)
            c.font = summary_header_font
            c.fill = summary_header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = header_border

        file_counts: Dict[str, int] = {}
        rule_counts: Dict[str, int] = {}
        for change in changes:
            file_name = str(change.get("file", "")) or "(unknown)"
            rule_id = str(change.get("ruleId", "")) or "(unknown)"
            file_counts[file_name] = file_counts.get(file_name, 0) + 1
            rule_counts[rule_id] = rule_counts.get(rule_id, 0) + 1

        summary_rows: List[Tuple[str, Any]] = [
            ("Total changes", len(changes)),
            ("Files with changes", len(file_counts)),
            ("Rules with changes", len(rule_counts)),
        ]

        current_row = 3
        for idx, (metric, value) in enumerate(summary_rows):
            row_fill = alt_fill_even if idx % 2 == 0 else alt_fill_odd
            metric_cell = summary_ws.cell(row=current_row, column=1, value=metric)
            value_cell = summary_ws.cell(row=current_row, column=2, value=value)
            for c in (metric_cell, value_cell):
                c.fill = row_fill
                c.border = all_border
                c.font = summary_data_font
                c.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

        current_row += 1
        section_header = summary_ws.cell(row=current_row, column=1, value="By File")
        section_header.font = summary_header_font
        section_header.fill = summary_header_fill
        section_header.alignment = Alignment(horizontal="left", vertical="center")
        section_header.border = header_border
        summary_ws.cell(row=current_row, column=2, value="Changes")
        summary_ws.cell(row=current_row, column=2).font = summary_header_font
        summary_ws.cell(row=current_row, column=2).fill = summary_header_fill
        summary_ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        summary_ws.cell(row=current_row, column=2).border = header_border
        current_row += 1

        for idx, (file_name, count) in enumerate(sorted(file_counts.items(), key=lambda item: item[0].lower())):
            row_fill = alt_fill_even if idx % 2 == 0 else alt_fill_odd
            file_cell = summary_ws.cell(row=current_row, column=1, value=file_name)
            count_cell = summary_ws.cell(row=current_row, column=2, value=count)
            for c in (file_cell, count_cell):
                c.fill = row_fill
                c.border = all_border
                c.font = summary_data_font
            file_cell.alignment = Alignment(horizontal="left", vertical="center")
            count_cell.alignment = Alignment(horizontal="right", vertical="center")
            current_row += 1

        current_row += 1
        section_header = summary_ws.cell(row=current_row, column=1, value="By Rule")
        section_header.font = summary_header_font
        section_header.fill = summary_header_fill
        section_header.alignment = Alignment(horizontal="left", vertical="center")
        section_header.border = header_border
        summary_ws.cell(row=current_row, column=2, value="Changes")
        summary_ws.cell(row=current_row, column=2).font = summary_header_font
        summary_ws.cell(row=current_row, column=2).fill = summary_header_fill
        summary_ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        summary_ws.cell(row=current_row, column=2).border = header_border
        current_row += 1

        for idx, (rule_id, count) in enumerate(sorted(rule_counts.items(), key=lambda item: item[0].lower())):
            row_fill = alt_fill_even if idx % 2 == 0 else alt_fill_odd
            rule_cell = summary_ws.cell(row=current_row, column=1, value=rule_id)
            count_cell = summary_ws.cell(row=current_row, column=2, value=count)
            for c in (rule_cell, count_cell):
                c.fill = row_fill
                c.border = all_border
                c.font = summary_data_font
            rule_cell.alignment = Alignment(horizontal="left", vertical="center")
            count_cell.alignment = Alignment(horizontal="right", vertical="center")
            current_row += 1

        summary_ws.column_dimensions[get_column_letter(1)].width = 52
        summary_ws.column_dimensions[get_column_letter(2)].width = 16
        summary_ws.row_dimensions[1].height = 28
        summary_ws.row_dimensions[2].height = 22
        summary_ws.freeze_panes = "A3"

        wb.save(out_path)
        return out_path, out_dir

    @bp.route("/session", methods=["POST"])
    @limiter.limit("60 per hour")
    def api_batch_find_replace_session():
        token = _create_batch_session_token()
        response, _status = make_response()
        response.set_cookie(
            batch_session_cookie,
            token,
            httponly=True,
            samesite="Strict",
            secure=request.is_secure,
            max_age=batch_session_ttl_seconds,
            path="/api/batch-find-replace/",
        )
        return response

    @bp.route("/preview", methods=["POST"])
    @require_batch_session_or_api_key
    @limiter.limit("30 per hour")
    def api_batch_find_replace_preview():
        try:
            result = _process_batch_files(preview_only=True)
            return jsonify(
                {
                    "success": True,
                    "matches": result["matches"],
                    "files_processed": result["files_processed"],
                    "files_changed": result["files_changed"],
                    "replacements": result["replacements"],
                    "message": (
                        f"Preview completed: {result['replacements']} replacement(s) "
                        f"across {result['files_changed']} file(s)."
                    ),
                }
            )
        except ValueError as exc:
            return _batch_error_response(
                message=_safe_batch_validation_message(
                    exc,
                    "Invalid batch preview request.",
                ),
                status_code=400,
            )
        except Exception:
            logger.exception("Batch preview failed")
            return _batch_error_response(
                message="Batch preview failed.",
                status_code=500,
            )

    @bp.route("/apply", methods=["POST"])
    @require_batch_session_or_api_key
    @limiter.limit("20 per hour")
    def api_batch_find_replace_apply():
        try:
            result = _process_batch_files(preview_only=False)
            report_path, report_dir = export_batch_changes_to_excel(result["matches"])
            schedule_cleanup(report_dir)

            if hasattr(os, "startfile"):
                try:
                    os.startfile(report_path)
                except Exception as exc:
                    logger.warning("Could not auto-open Excel report: %s", exc)

            return send_file(
                report_path,
                as_attachment=True,
                download_name=os.path.basename(report_path),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except ValueError as exc:
            return _batch_error_response(
                message=_safe_batch_validation_message(
                    exc,
                    "Invalid batch apply request.",
                ),
                status_code=400,
            )
        except Exception:
            logger.exception("Batch apply failed")
            return _batch_error_response(
                message="Batch apply failed.",
                status_code=500,
            )

    @bp.route("/cad/preview", methods=["POST"])
    @require_batch_session_or_api_key
    @limiter.limit("30 per hour")
    def api_batch_find_replace_cad_preview():
        try:
            payload = request.get_json(silent=True) or {}
            rules = _parse_batch_rules_from_json(payload)
            host_result = _call_acade_host_action(
                "suite_batch_find_replace_preview",
                {
                    "rules": rules,
                    "blockNameHint": str(payload.get("blockNameHint") or "").strip(),
                },
            )
            if not host_result.get("success", False):
                status_code = 400 if host_result.get("code") == "INVALID_REQUEST" else 503
                return jsonify(host_result), status_code

            data = host_result.get("data") or {}
            matches = data.get("matches") or []
            return jsonify(
                {
                    "success": True,
                    "requestId": host_result.get("meta", {}).get("requestId"),
                    "matches": matches,
                    "matchCount": len(matches),
                    "warnings": host_result.get("warnings") or [],
                    "drawingName": data.get("drawingName"),
                    "message": host_result.get("message") or "CAD preview completed.",
                }
            )
        except ValueError as exc:
            return _batch_error_response(
                message=_safe_batch_validation_message(
                    exc,
                    "Invalid CAD batch preview request.",
                ),
                status_code=400,
            )
        except Exception:
            logger.exception("CAD batch preview failed")
            return _batch_error_response(
                message="CAD batch preview failed.",
                status_code=500,
            )

    @bp.route("/cad/apply", methods=["POST"])
    @require_batch_session_or_api_key
    @limiter.limit("20 per hour")
    def api_batch_find_replace_cad_apply():
        try:
            payload = request.get_json(silent=True) or {}
            matches = payload.get("matches")
            if not isinstance(matches, list) or len(matches) == 0:
                raise ValueError("matches must contain at least one preview row.")

            host_result = _call_acade_host_action(
                "suite_batch_find_replace_apply",
                {
                    "matches": matches,
                    "blockNameHint": str(payload.get("blockNameHint") or "").strip(),
                },
            )
            if not host_result.get("success", False):
                status_code = 400 if host_result.get("code") == "INVALID_REQUEST" else 503
                return jsonify(host_result), status_code

            data = host_result.get("data") or {}
            change_rows = data.get("changes") or []
            report_path, report_dir = export_batch_changes_to_excel(change_rows)
            schedule_cleanup(report_dir)

            if hasattr(os, "startfile"):
                try:
                    os.startfile(report_path)
                except Exception as exc:
                    logger.warning("Could not auto-open Excel report: %s", exc)

            return send_file(
                report_path,
                as_attachment=True,
                download_name=os.path.basename(report_path),
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except ValueError as exc:
            return _batch_error_response(
                message=_safe_batch_validation_message(
                    exc,
                    "Invalid CAD batch apply request.",
                ),
                status_code=400,
            )
        except Exception:
            logger.exception("CAD batch apply failed")
            return _batch_error_response(
                message="CAD batch apply failed.",
                status_code=500,
            )

    @bp.route("/cad/cleanup-preview", methods=["POST"])
    @require_batch_session_or_api_key
    @limiter.limit("20 per hour")
    def api_batch_find_replace_cad_cleanup_preview():
        request_id = _current_request_id("cleanup")
        if not request.is_json:
            return _cleanup_error_response(
                code="INVALID_REQUEST",
                message="Expected application/json payload.",
                request_id=request_id,
                status_code=400,
                meta={"stage": "drawing_cleanup.preview.validation"},
            )

        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            return _cleanup_error_response(
                code="INVALID_REQUEST",
                message="Request payload must be a JSON object.",
                request_id=request_id,
                status_code=400,
                meta={"stage": "drawing_cleanup.preview.validation"},
            )

        try:
            host_result = _normalize_cleanup_host_result(
                _call_acade_host_action(
                    "suite_drawing_cleanup_preview",
                    {
                        "requestId": request_id,
                        "entryMode": str(payload.get("entryMode") or "").strip(),
                        "preset": str(payload.get("preset") or "").strip(),
                        "sourcePath": str(payload.get("sourcePath") or "").strip(),
                        "saveDrawing": bool(payload.get("saveDrawing", False)),
                        "timeoutMs": payload.get("timeoutMs", 90000),
                    },
                ),
                request_id,
            )
            status_code = 200 if host_result.get("success") else _cleanup_status_code_for_host_result(host_result)
            return jsonify(host_result), status_code
        except Exception:
            logger.exception("CAD drawing cleanup preview failed")
            return _cleanup_error_response(
                code="ACADE_HOST_FAILED",
                message="Drawing cleanup preview failed.",
                request_id=request_id,
                status_code=503,
                meta={"stage": "drawing_cleanup.preview.host"},
            )

    @bp.route("/cad/cleanup-apply", methods=["POST"])
    @require_batch_session_or_api_key
    @limiter.limit("15 per hour")
    def api_batch_find_replace_cad_cleanup_apply():
        request_id = _current_request_id("cleanup")
        if not request.is_json:
            return _cleanup_error_response(
                code="INVALID_REQUEST",
                message="Expected application/json payload.",
                request_id=request_id,
                status_code=400,
                meta={"stage": "drawing_cleanup.apply.validation"},
            )

        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            return _cleanup_error_response(
                code="INVALID_REQUEST",
                message="Request payload must be a JSON object.",
                request_id=request_id,
                status_code=400,
                meta={"stage": "drawing_cleanup.apply.validation"},
            )

        try:
            host_result = _normalize_cleanup_host_result(
                _call_acade_host_action(
                    "suite_drawing_cleanup_apply",
                    {
                        "requestId": request_id,
                        "entryMode": str(payload.get("entryMode") or "").strip(),
                        "preset": str(payload.get("preset") or "").strip(),
                        "sourcePath": str(payload.get("sourcePath") or "").strip(),
                        "saveDrawing": bool(payload.get("saveDrawing", False)),
                        "timeoutMs": payload.get("timeoutMs", 90000),
                        "selectedFixIds": payload.get("selectedFixIds") or [],
                        "approvedReviewIds": payload.get("approvedReviewIds") or [],
                    },
                ),
                request_id,
            )
            status_code = 200 if host_result.get("success") else _cleanup_status_code_for_host_result(host_result)
            return jsonify(host_result), status_code
        except Exception:
            logger.exception("CAD drawing cleanup apply failed")
            return _cleanup_error_response(
                code="ACADE_HOST_FAILED",
                message="Drawing cleanup apply failed.",
                request_id=request_id,
                status_code=503,
                meta={"stage": "drawing_cleanup.apply.host"},
            )

    @bp.route("/cad/project-preview", methods=["POST"])
    @require_batch_session_or_api_key
    @limiter.limit("20 per hour")
    def api_batch_find_replace_cad_project_preview():
        try:
            payload = request.get_json(silent=True) or {}
            rules = _parse_batch_rules_from_json(payload)
            drawings = _resolve_project_drawings(payload)
            host_result = _call_acade_host_action(
                "suite_batch_find_replace_project_preview",
                {
                    "rules": rules,
                    "drawings": drawings,
                    "blockNameHint": str(payload.get("blockNameHint") or "").strip(),
                },
            )
            if not host_result.get("success", False):
                status_code = 400 if host_result.get("code") == "INVALID_REQUEST" else 503
                return jsonify(host_result), status_code

            data = host_result.get("data") or {}
            matches = _normalize_project_preview_matches(data.get("matches") or [], drawings)
            drawing_summaries = _build_project_preview_drawings(drawings, matches)
            affected_drawings = sum(
                1 for summary in drawing_summaries if int(summary.get("matchCount") or 0) > 0
            )
            return jsonify(
                {
                    "success": True,
                    "requestId": host_result.get("meta", {}).get("requestId"),
                    "matches": matches,
                    "matchCount": len(matches),
                    "drawings": drawing_summaries,
                    "warnings": host_result.get("warnings") or [],
                    "message": host_result.get("message")
                    or (
                        f"Project CAD preview completed: {len(matches)} replacement(s) "
                        f"across {affected_drawings} drawing(s)."
                    ),
                }
            )
        except ValueError as exc:
            return _batch_error_response(
                message=_safe_batch_validation_message(
                    exc,
                    "Invalid project CAD preview request.",
                ),
                status_code=400,
            )
        except Exception:
            logger.exception("Project CAD batch preview failed")
            return _batch_error_response(
                message="Project CAD batch preview failed.",
                status_code=500,
            )

    @bp.route("/cad/project-apply", methods=["POST"])
    @require_batch_session_or_api_key
    @limiter.limit("15 per hour")
    def api_batch_find_replace_cad_project_apply():
        try:
            payload = request.get_json(silent=True) or {}
            matches = payload.get("matches")
            if not isinstance(matches, list) or len(matches) == 0:
                raise ValueError("matches must contain at least one project preview row.")
            if len(matches) > MAX_APPLY_CHANGE_ROWS:
                raise ValueError(
                    f"Too many project CAD apply rows. Maximum is {MAX_APPLY_CHANGE_ROWS}"
                )

            host_result = _call_acade_host_action(
                "suite_batch_find_replace_project_apply",
                {
                    "matches": matches,
                    "blockNameHint": str(payload.get("blockNameHint") or "").strip(),
                },
            )
            if not host_result.get("success", False):
                status_code = 400 if host_result.get("code") == "INVALID_REQUEST" else 503
                return jsonify(host_result), status_code

            data = host_result.get("data") or {}
            change_rows = data.get("changes") or []
            report_path, report_dir = export_batch_changes_to_excel(change_rows)
            schedule_cleanup(report_dir)
            report_id, report_filename = _register_generated_report(report_path, report_dir)

            return jsonify(
                {
                    "success": True,
                    "requestId": host_result.get("meta", {}).get("requestId"),
                    "updated": int(data.get("updated") or 0),
                    "changedDrawingCount": int(data.get("changedDrawingCount") or 0),
                    "changedItemCount": int(
                        data.get("changedItemCount") or data.get("updated") or 0
                    ),
                    "drawings": data.get("drawings") or [],
                    "warnings": host_result.get("warnings") or [],
                    "reportId": report_id,
                    "reportFilename": report_filename,
                    "downloadUrl": f"/api/batch-find-replace/reports/{report_id}",
                    "message": host_result.get("message")
                    or "Project CAD apply completed.",
                }
            )
        except ValueError as exc:
            return _batch_error_response(
                message=_safe_batch_validation_message(
                    exc,
                    "Invalid project CAD apply request.",
                ),
                status_code=400,
            )
        except Exception:
            logger.exception("Project CAD batch apply failed")
            return _batch_error_response(
                message="Project CAD batch apply failed.",
                status_code=500,
            )

    @bp.route("/reports/<report_id>", methods=["GET"])
    @require_batch_session_or_api_key
    @limiter.limit("40 per hour")
    def api_batch_find_replace_report_download(report_id: str):
        report = generated_reports.get(report_id)
        if not report:
            return make_error_response("Report not found.", status=404)

        report_path = report.get("path") or ""
        if not report_path or not os.path.exists(report_path):
            generated_reports.pop(report_id, None)
            return make_error_response("Report is no longer available.", status=404)

        return send_file(
            report_path,
            as_attachment=True,
            download_name=report.get("filename") or os.path.basename(report_path),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    return bp
