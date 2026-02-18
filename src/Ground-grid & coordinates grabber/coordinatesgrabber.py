# coordinatesgrabber.py
"""
Coordinates Grabber (AutoCAD -> Excel exporter) + ALWAYS places the
"Coordinate Reference Point" block at every exported point.

This is a PySide6 rewrite of your original Tkinter-based coordtable.py,
wired into the shared UI system:

- ui/theme.py        (apply_theme, Tokens)
- ui/icons.py        (IconManager)
- ui/components.py   (CardSection, make_button)

Modes:
- polylines: export every polyline vertex (places refpoint at every vertex)
- blocks: export center of selected block references (places refpoint at each block center)
- layer_search: export ONE point PER BLOCK REFERENCE by finding geometry on a chosen layer
               INSIDE each block definition (and nested blocks), then computing center from
               those found items.

Requirements:
- Windows + AutoCAD running
- Python 3.9+
- pip install pywin32 openpyxl PySide6

Run:
    python coordinatesgrabber.py
"""

from __future__ import annotations

import math
import os
import re
import sys
import time
import traceback
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pythoncom
import pywintypes
import win32com.client
import win32com.client.gencache as gencache

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from PySide6.QtCore import Qt
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (
    QApplication,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QSizePolicy,
    QSpinBox,
    QStatusBar,
    QTextEdit,
    QVBoxLayout,
    QWidget,
    QProgressBar,
)

from ui.theme import apply_theme
from ui.icons import IconManager
from ui.components import CardSection, make_button


DEBUG = False

# Configure gencache to avoid corruption issues
# See: https://support.pyxll.com/hc/en-gb/articles/360058200414
try:
    gencache.EnsureModule("{00020813-0000-0000-C000-000000000046}", 0, 1, 9)  # Excel
except Exception:
    pass  # If cache is corrupted or doesn't exist, fall back to late binding

Point3D = Tuple[float, float, float]

# Cache: normalized DWG path -> block name inserted in this drawing
_REF_IMPORT_CACHE: Dict[str, str] = {}

# Cache for layer_search: (block_def_name_lower, layer_lower) -> list of local points in block-def coords
_BLOCKDEF_LAYERPTS_CACHE: Dict[Tuple[str, str], List[Point3D]] = {}


# -------------------------
# Data models
# -------------------------
@dataclass(frozen=True)
class TableOptions:
    segment: bool
    elevation: bool
    distance: bool
    distance_3d: bool
    bearing_quadrant: bool
    azimuth_from_north: bool


@dataclass(frozen=True)
class ExportOptions:
    excel_path: str
    replace_previous: bool
    auto_increment: bool


@dataclass(frozen=True)
class RefBlockOptions:
    ref_dwg_path: str
    layer_name: str
    scale: float
    rotation_deg: float


@dataclass(frozen=True)
class Config:
    mode: str  # "polylines" | "blocks" | "layer_search"
    precision: int
    prefix: str
    initial_number: int
    table_options: TableOptions
    export: ExportOptions
    refblock: RefBlockOptions
    block_name_filter: str  # optional, for blocks/layer_search mode
    layer_search_name: str  # layer name to search for in layer_search mode
    layer_search_use_selection: bool  # if True, only scan selected entities (blocks + modelspace)
    layer_search_include_modelspace: bool  # if True, also find geometry on the layer in ModelSpace (outside blocks)
    layer_search_use_corners: bool  # if True, place blocks at 4 corners instead of center
    add_to_selection: bool


@dataclass(frozen=True)
class Row:
    point_name: str
    east: float
    north: float
    elev: float
    segment_name: str
    dist_2d: Optional[float]
    dist_3d: Optional[float]
    bearing: str
    azimuth: str
    source_type: str
    source_handle: str = ""
    source_name: str = ""
    source_index: int = -1
    corner_name: Optional[str] = None
    source_handle: str
    source_name: str
    source_index: int
    corner_name: Optional[str] = None  # NW, NE, SW, SE for 4-corner mode


# -------------------------
# Logging / errors
# -------------------------
def log(msg: str) -> None:
    print(msg, flush=True)


def format_com_error(exc: BaseException) -> str:
    if isinstance(exc, pywintypes.com_error):
        try:
            hr, ctx, details, argerr = exc.args  # type: ignore[misc]
        except Exception:
            return f"{exc!r}"
        return f"pywintypes.com_error(hr={hr}, ctx={ctx}, details={details}, argerr={argerr})"
    return repr(exc)


def format_exception_text(exc: BaseException, context: str) -> str:
    return (
        "\n" + "=" * 80 + "\n"
        f"ERROR CONTEXT: {context}\n"
        f"ERROR: {format_com_error(exc)}\n"
        "TRACEBACK:\n"
        f"{traceback.format_exc().rstrip()}\n"
        + "=" * 80 + "\n"
    )


# -------------------------
# COM helpers (late-bound only)
# -------------------------
def dyn(obj: Any) -> Any:
    try:
        if type(obj).__name__ == "CDispatch":
            return obj
    except Exception:
        pass

    try:
        ole = obj._oleobj_  # type: ignore[attr-defined]
    except Exception:
        ole = obj

    try:
        disp = ole.QueryInterface(pythoncom.IID_IDispatch)
        return win32com.client.dynamic.Dispatch(disp)
    except Exception:
        try:
            return win32com.client.dynamic.Dispatch(obj)
        except Exception:
            return obj


def pt(x: float, y: float, z: float = 0.0):
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (float(x), float(y), float(z)))


def com_call_with_retry(callable_func, max_retries: int = 25, initial_delay: float = 0.03):
    delay = initial_delay
    for _ in range(max_retries):
        try:
            return callable_func()
        except pythoncom.com_error as e:
            if e.args and e.args[0] == -2147418111:  # RPC_E_CALL_REJECTED
                time.sleep(delay)
                delay = min(delay * 1.5, 0.5)
                continue
            raise
    raise RuntimeError("AutoCAD COM call failed: RPC busy too long")


def connect_autocad() -> Any:
    acad = win32com.client.dynamic.Dispatch("AutoCAD.Application")
    if acad is None:
        raise RuntimeError("Could not connect to AutoCAD.Application (is AutoCAD running?).")
    return dyn(acad)


def prompt(doc: Any, msg: str) -> None:
    doc = dyn(doc)
    try:
        doc.Utility.Prompt(msg)
    except Exception:
        log(msg)


def get_or_create_selection_set(doc: Any, name: str) -> Any:
    doc = dyn(doc)
    try:
        ss = doc.SelectionSets.Item(name)
        ss.Clear()
        return dyn(ss)
    except Exception:
        pass
    try:
        return dyn(doc.SelectionSets.Add(name))
    except Exception:
        try:
            ss = doc.SelectionSets.Item(name)
            ss.Delete()
        except Exception:
            pass
        return dyn(doc.SelectionSets.Add(name))


def resolve_by_handle(doc: Any, handle: str) -> Any:
    doc = dyn(doc)
    return dyn(doc.HandleToObject(str(handle)))


def ensure_layer(doc: Any, layer_name: str) -> None:
    """Create layer if missing (so br.Layer assignment doesn't fail)."""
    doc = dyn(doc)
    try:
        layers = dyn(doc.Layers)
        try:
            layers.Item(layer_name)
        except Exception:
            layers.Add(layer_name)
    except Exception:
        pass


def wait_for_command_finish(doc: Any, timeout_s: float = 10.0) -> bool:
    """
    SendCommand is async. Poll CMDNAMES until empty.
    Returns True if finished within timeout.
    """
    doc = dyn(doc)
    start = time.time()
    while time.time() - start < timeout_s:
        try:
            cmd = com_call_with_retry(lambda: str(doc.GetVariable("CMDNAMES") or ""))
        except Exception:
            cmd = ""
        if not cmd.strip():
            return True
        time.sleep(0.05)
    return False


# -------------------------
# Geometry / formatting
# -------------------------
def dms_int(deg: float) -> str:
    deg = float(deg) % 360.0
    d = int(deg)
    m_float = (deg - d) * 60.0
    m = int(m_float)
    s = int(round((m_float - m) * 60.0))
    if s >= 60:
        s -= 60
        m += 1
    if m >= 60:
        m -= 60
        d += 1
    return f"{d}°{m:02d}\'{s:02d}\""


def azimuth_deg_clockwise_from_north(dx: float, dy: float) -> float:
    ang = math.degrees(math.atan2(dx, dy))  # 0 at +Y, clockwise
    if ang < 0:
        ang += 360.0
    return ang


def quadrant_bearing(dx: float, dy: float) -> Tuple[float, str]:
    eps = 1e-12
    if abs(dx) < eps and abs(dy) < eps:
        return 0.0, "NE"

    ns = "N" if dy >= 0 else "S"
    ew = "E" if dx >= 0 else "W"

    adx, ady = abs(dx), abs(dy)
    if ady < eps:
        angle = 90.0
    elif adx < eps:
        angle = 0.0
    else:
        angle = math.degrees(math.atan2(adx, ady))  # 0..90
    return angle, f"{ns}{ew}"


def dist2d(a: Point3D, b: Point3D) -> float:
    return math.hypot(b[0] - a[0], b[1] - a[1])


def dist3d(a: Point3D, b: Point3D) -> float:
    dx, dy, dz = (b[0] - a[0]), (b[1] - a[1]), (b[2] - a[2])
    return math.sqrt(dx * dx + dy * dy + dz * dz)


# -------------------------
# Entity inspection
# -------------------------
def object_name(ent: Any) -> str:
    try:
        return str(dyn(ent).ObjectName)
    except Exception:
        return "Unknown"


def is_polyline(ent: Any) -> bool:
    return object_name(ent) in {"AcDbPolyline", "AcDb2dPolyline", "AcDb3dPolyline"}


def is_3d_poly(ent: Any) -> bool:
    return object_name(ent) == "AcDb3dPolyline"


def is_block_reference(ent: Any) -> bool:
    return object_name(ent) in {"AcDbBlockReference", "AcDbMInsertBlock"}


def block_effective_name(ent: Any) -> str:
    ent = dyn(ent)
    for prop in ("EffectiveName", "Name"):
        try:
            v = str(getattr(ent, prop))
            if v:
                return v
        except Exception:
            continue
    return ""


def matches_name_filter(name: str, filt: str) -> bool:
    f = (filt or "").strip()
    if not f:
        return True

    name_l = (name or "").lower()
    f_l = f.lower()

    if any(ch in f_l for ch in "*?"):
        pat = "^" + re.escape(f_l).replace(r"\\*", ".*").replace(r"\\?", ".") + "$"
        return re.match(pat, name_l) is not None

    return f_l in name_l


def _entity_layer(ent: Any) -> str:
    ent = dyn(ent)
    try:
        return str(ent.Layer).strip()
    except Exception:
        return ""


# -------------------------
# Polyline vertices + centroid
# -------------------------
def _try_coords(ent: Any) -> Optional[List[float]]:
    ent = dyn(ent)
    try:
        coords = list(ent.Coordinates)
        return [float(x) for x in coords]
    except Exception:
        return None


def _try_vertices_indexed(ent: Any, elev_fallback: float) -> Optional[List[Point3D]]:
    ent = dyn(ent)
    try:
        n = int(ent.NumberOfVertices)
        out: List[Point3D] = []
        for i in range(n):
            p = ent.Coordinate(i)
            if len(p) >= 3:
                out.append((float(p[0]), float(p[1]), float(p[2])))
            else:
                out.append((float(p[0]), float(p[1]), float(elev_fallback)))
        return out
    except Exception:
        return None


def poly_vertices(ent: Any) -> List[Point3D]:
    ent = dyn(ent)
    name = object_name(ent)

    elev = 0.0
    try:
        elev = float(getattr(ent, "Elevation", 0.0))
    except Exception:
        elev = 0.0

    coords = _try_coords(ent)
    if coords:
        if name == "AcDb3dPolyline":
            if len(coords) % 3 != 0:
                raise RuntimeError("3D polyline Coordinates not divisible by 3.")
            return [(coords[i], coords[i + 1], coords[i + 2]) for i in range(0, len(coords), 3)]
        if len(coords) % 2 == 0:
            return [(coords[i], coords[i + 1], elev) for i in range(0, len(coords), 2)]

    pts = _try_vertices_indexed(ent, elev)
    if pts is not None:
        return pts

    raise RuntimeError(f"Polyline '{name}' does not expose Coordinates/Coordinate(i).")


def is_closed(ent: Any) -> bool:
    ent = dyn(ent)
    try:
        return bool(ent.Closed)
    except Exception:
        return False


def poly_centroid(ent: Any) -> Point3D:
    pts = poly_vertices(ent)
    if not pts:
        return bbox_center(ent)
    n = len(pts)
    cx = sum(p[0] for p in pts) / n
    cy = sum(p[1] for p in pts) / n
    cz = sum(p[2] for p in pts) / n
    return (cx, cy, cz)


# -------------------------
# Bounding box center
# -------------------------
def bbox_center(ent: Any) -> Point3D:
    ent = dyn(ent)

    try:
        mn, mx = ent.GetBoundingBox()
        x = (float(mn[0]) + float(mx[0])) / 2.0
        y = (float(mn[1]) + float(mx[1])) / 2.0
        z = (float(mn[2]) + float(mx[2])) / 2.0 if len(mn) > 2 and len(mx) > 2 else 0.0
        return (x, y, z)
    except Exception:
        pass

    try:
        ip = ent.InsertionPoint
        return (float(ip[0]), float(ip[1]), float(ip[2]) if len(ip) > 2 else 0.0)
    except Exception:
        return (0.0, 0.0, 0.0)


def _center_of_entity(ent: Any) -> Optional[Point3D]:
    ent = dyn(ent)
    if is_polyline(ent):
        return poly_centroid(ent)
    return bbox_center(ent)


# -------------------------
# 2D transform helpers (for nested blocks) + z passthrough
# -------------------------
def _bref_transform_components(bref: Any) -> Tuple[float, float, float, float, float, float]:
    """Return (sx, sy, sz, rot_radians, tx, ty, tz)."""
    bref = dyn(bref)
    try:
        ip = bref.InsertionPoint
        tx = float(ip[0])
        ty = float(ip[1])
        tz = float(ip[2]) if len(ip) > 2 else 0.0
    except Exception:
        tx, ty, tz = 0.0, 0.0, 0.0

    try:
        sx = float(bref.XScaleFactor)
    except Exception:
        sx = 1.0
    try:
        sy = float(bref.YScaleFactor)
    except Exception:
        sy = 1.0
    try:
        sz = float(bref.ZScaleFactor)
    except Exception:
        sz = 1.0
    try:
        rot = float(bref.Rotation)
    except Exception:
        rot = 0.0

    return sx, sy, sz, rot, tx, ty, tz


def _apply_bref_local_transform(bref: Any, p: Point3D) -> Point3D:
    """
    Apply a block reference transform to a point in the referenced block's local coords,
    producing a point in the parent coordinate system (or world, depending on context).
    """
    sx, sy, sz, rot, tx, ty, tz = _bref_transform_components(bref)
    lx, ly, lz = p

    x = lx * sx
    y = ly * sy
    z = lz * sz

    c = math.cos(rot)
    s = math.sin(rot)

    rx = x * c - y * s
    ry = x * s + y * c

    return (tx + rx, ty + ry, tz + z)


def _center_from_points(points: List[Point3D]) -> Optional[Point3D]:
    """Compute a stable center from a set of points (bbox center)."""
    if not points:
        return None
    if len(points) == 1:
        return points[0]
    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    zs = [p[2] for p in points]
    return ((min(xs) + max(xs)) / 2.0, (min(ys) + max(ys)) / 2.0, (min(zs) + max(zs)) / 2.0)


def _bbox_corners_from_points(points: List[Point3D]) -> List[Tuple[Point3D, str]]:
    """
    Extract 4 corners from bounding box of points (NW, NE, SW, SE).
    Returns list of (point, corner_name) tuples.
    NW = top-left, NE = top-right, SW = bottom-left, SE = bottom-right
    (In CAD/surveying convention: max-Y is North, max-X is East)
    """
    if not points:
        return []
    if len(points) == 1:
        p = points[0]
        return [(p, "center")]
    
    xs = [p[0] for p in points]
    ys = [p[1] for p in points]
    zs = [p[2] for p in points]
    
    minx, maxx = min(xs), max(xs)
    miny, maxy = min(ys), max(ys)
    z_avg = sum(zs) / len(zs) if zs else 0.0
    
    corners = [
        ((minx, maxy, z_avg), "NW"),  # NW = min-X, max-Y
        ((maxx, maxy, z_avg), "NE"),  # NE = max-X, max-Y
        ((minx, miny, z_avg), "SW"),  # SW = min-X, min-Y
        ((maxx, miny, z_avg), "SE"),  # SE = max-X, min-Y
    ]
    return corners


def _bbox_corners_from_entity(ent: Any) -> List[Tuple[Point3D, str]]:
    """
    Extract 4 corners from an entity's bounding box (NW, NE, SW, SE).
    Returns list of (point, corner_name) tuples.
    """
    ent = dyn(ent)
    try:
        mn, mx = ent.GetBoundingBox()
        minx = float(mn[0])
        miny = float(mn[1])
        minz = float(mn[2]) if len(mn) > 2 else 0.0
        maxx = float(mx[0])
        maxy = float(mx[1])
        maxz = float(mx[2]) if len(mx) > 2 else 0.0
        z_avg = (minz + maxz) / 2.0
    except Exception:
        # Fallback: use center and return single point
        c = bbox_center(ent)
        return [(c, "center")]
    
    if abs(minx - maxx) < 1e-12 or abs(miny - maxy) < 1e-12:
        # Degenerate bbox (point or line), return center
        c = ((minx + maxx) / 2.0, (miny + maxy) / 2.0, z_avg)
        return [(c, "center")]
    
    corners = [
        ((minx, maxy, z_avg), "NW"),  # NW = min-X, max-Y
        ((maxx, maxy, z_avg), "NE"),  # NE = max-X, max-Y
        ((minx, miny, z_avg), "SW"),  # SW = min-X, min-Y
        ((maxx, miny, z_avg), "SE"),  # SE = max-X, min-Y
    ]
    return corners


# -------------------------
# Layer Search: cache points on layer in a BLOCK DEFINITION (local coords)
# -------------------------
def _blockdef_is_layout(blk_def: Any) -> bool:
    try:
        return bool(getattr(dyn(blk_def), "IsLayout", False))
    except Exception:
        return False


def _points_on_layer_in_blockdef(doc: Any, block_def_name: str, target_layer_lower: str, visiting: Optional[set] = None) -> List[Point3D]:
    """
    Returns points (centers) of entities on target layer INSIDE a block definition, expressed
    in that block definition's coordinate system.

    Includes nested blocks: recursively pulls points from nested block defs and transforms them
    through the nested blockref transform inside this block definition.

    Uses cache so repeated instances of the same block definition are fast.
    """
    doc = dyn(doc)
    key = (block_def_name.lower(), target_layer_lower)
    if key in _BLOCKDEF_LAYERPTS_CACHE:
        return _BLOCKDEF_LAYERPTS_CACHE[key]

    if visiting is None:
        visiting = set()
    if block_def_name.lower() in visiting:
        # protect against circular references
        return []

    visiting.add(block_def_name.lower())

    out: List[Point3D] = []
    try:
        blk_def = dyn(doc.Blocks.Item(block_def_name))
    except Exception:
        visiting.remove(block_def_name.lower())
        _BLOCKDEF_LAYERPTS_CACHE[key] = []
        return []

    # Skip layout blocks / paper space blocks
    if _blockdef_is_layout(blk_def):
        visiting.remove(block_def_name.lower())
        _BLOCKDEF_LAYERPTS_CACHE[key] = []
        return []

    try:
        count = int(blk_def.Count)
    except Exception:
        count = 0

    for i in range(count):
        try:
            ent = dyn(blk_def.Item(i))
        except Exception:
            continue

        if is_block_reference(ent):
            nbname = block_effective_name(ent)
            if not nbname:
                continue
            child_pts = _points_on_layer_in_blockdef(doc, nbname, target_layer_lower, visiting)
            if child_pts:
                for p in child_pts:
                    out.append(_apply_bref_local_transform(ent, p))
            continue

        if _entity_layer(ent).strip().lower() == target_layer_lower:
            c = _center_of_entity(ent)
            if c:
                out.append(c)

    visiting.remove(block_def_name.lower())
    _BLOCKDEF_LAYERPTS_CACHE[key] = out
    return out


def _blockrefs_in_modelspace(doc: Any) -> List[Any]:
    doc = dyn(doc)
    ms = dyn(doc.ModelSpace)
    res: List[Any] = []
    try:
        n = int(ms.Count)
    except Exception:
        n = 0
    for i in range(n):
        try:
            ent = dyn(ms.Item(i))
        except Exception:
            continue
        if is_block_reference(ent):
            res.append(ent)
    return res



# -------------------------
# Layer Search: ModelSpace layer geometry (outside blocks)
# -------------------------
def _layer_entities_in_modelspace(doc: Any, target_layer_lower: str, handles: Optional[Sequence[str]] = None, log_cb=None) -> List[Any]:
    """
    Return entities that live directly in ModelSpace on a given layer.
    If handles is provided, only those entities are considered (and filtered by layer).
    Block references are excluded (outside-block scan).
    """
    doc = dyn(doc)
    target_layer_lower = target_layer_lower.strip().lower()
    res: List[Any] = []

    def _consider(ent: Any) -> None:
        ent = dyn(ent)
        if is_block_reference(ent):
            return
        try:
            lyr = str(ent.Layer).strip().lower()
        except Exception:
            lyr = ""
        if lyr == target_layer_lower:
            res.append(ent)

    if handles:
        for h in handles:
            try:
                ent = resolve_by_handle(doc, h)
            except Exception:
                continue
            _consider(ent)
        if log_cb:
            log_cb(f"[LayerSearch/ModelSpace] Using selection-only: {len(res)} entities on layer.")
        return res

    ms = dyn(doc.ModelSpace)
    try:
        n = int(ms.Count)
    except Exception:
        n = 0
    if log_cb:
        log_cb(f"[LayerSearch/ModelSpace] Scanning ModelSpace for layer entities (total ents={n}) ...")

    for i in range(n):
        try:
            ent = dyn(ms.Item(i))
        except Exception:
            continue
        _consider(ent)

    if log_cb:
        log_cb(f"[LayerSearch/ModelSpace] Found {len(res)} ModelSpace entities on layer.")
    return res


def _entity_bbox_3d(ent: Any) -> Optional[Tuple[float, float, float, float, float, float]]:
    """Return (minx, miny, minz, maxx, maxy, maxz) or None."""
    ent = dyn(ent)
    try:
        mn, mx = ent.GetBoundingBox()
        minx, miny = float(mn[0]), float(mn[1])
        maxx, maxy = float(mx[0]), float(mx[1])
        minz = float(mn[2]) if len(mn) > 2 else 0.0
        maxz = float(mx[2]) if len(mx) > 2 else 0.0
        # normalize
        if maxx < minx:
            minx, maxx = maxx, minx
        if maxy < miny:
            miny, maxy = maxy, miny
        if maxz < minz:
            minz, maxz = maxz, minz
        return (minx, miny, minz, maxx, maxy, maxz)
    except Exception:
        return None


def _bbox_center_3d(bb: Tuple[float, float, float, float, float, float]) -> Point3D:
    minx, miny, minz, maxx, maxy, maxz = bb
    return ((minx + maxx) / 2.0, (miny + maxy) / 2.0, (minz + maxz) / 2.0)


def _bboxes_overlap_2d(a: Tuple[float, float, float, float], b: Tuple[float, float, float, float], tol: float) -> bool:
    aminx, aminy, amaxx, amaxy = a
    bminx, bminy, bmaxx, bmaxy = b
    return (aminx - tol) <= (bmaxx + tol) and (bminx - tol) <= (amaxx + tol) and (aminy - tol) <= (bmaxy + tol) and (bminy - tol) <= (amaxy + tol)


def _cluster_bboxes(bboxes: List[Tuple[float, float, float, float, float, float]], log_cb=None) -> List[List[int]]:
    """
    Cluster bboxes that overlap in XY (with tolerance), returning a list of clusters
    as lists of indices into bboxes.
    """
    n = len(bboxes)
    if n == 0:
        return []

    # tolerance based on typical size of objects
    sizes = []
    for bb in bboxes:
        w = max(0.0, bb[3] - bb[0])
        h = max(0.0, bb[4] - bb[1])
        sizes.append(max(w, h))
    typical = sorted(sizes)[n // 2] if sizes else 0.0
    tol = max(1e-6, typical * 0.02)

    # grid hashing to avoid O(n^2) on large layers
    cell = max(typical * 1.5, 1.0)
    grid: Dict[Tuple[int, int], List[int]] = {}
    def _cells_for(bb):
        minx, miny, _, maxx, maxy, _ = bb
        ix0 = int(math.floor(minx / cell))
        ix1 = int(math.floor(maxx / cell))
        iy0 = int(math.floor(miny / cell))
        iy1 = int(math.floor(maxy / cell))
        for ix in range(ix0, ix1 + 1):
            for iy in range(iy0, iy1 + 1):
                yield (ix, iy)

    for i, bb in enumerate(bboxes):
        for key in _cells_for(bb):
            grid.setdefault(key, []).append(i)

    parent = list(range(n))
    rank = [0] * n

    def find(i):
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra == rb:
            return
        if rank[ra] < rank[rb]:
            parent[ra] = rb
        elif rank[ra] > rank[rb]:
            parent[rb] = ra
        else:
            parent[rb] = ra
            rank[ra] += 1

    checked_pairs = 0
    for cell_key, idxs in grid.items():
        if len(idxs) <= 1:
            continue
        # compare within cell
        for i_pos in range(len(idxs)):
            i = idxs[i_pos]
            ai = (bboxes[i][0], bboxes[i][1], bboxes[i][3], bboxes[i][4])
            for j_pos in range(i_pos + 1, len(idxs)):
                j = idxs[j_pos]
                if find(i) == find(j):
                    continue
                aj = (bboxes[j][0], bboxes[j][1], bboxes[j][3], bboxes[j][4])
                checked_pairs += 1
                if _bboxes_overlap_2d(ai, aj, tol):
                    union(i, j)

    clusters: Dict[int, List[int]] = {}
    for i in range(n):
        clusters.setdefault(find(i), []).append(i)

    if log_cb:
        log_cb(f"[LayerSearch/ModelSpace] Clustered {n} items into {len(clusters)} cluster(s) (tol≈{tol:g}, checked_pairs={checked_pairs}).")

    return list(clusters.values())


def build_rows_layer_search_modelspace(
    cfg: Config,
    ents: Sequence[Any],
    target_layer: str,
    start_number: int = 1,
    use_corners: bool = False,
    progress_cb=None,
    log_cb=None,
) -> Tuple[List[Row], bool, int]:
    """
    OUTSIDE-BLOCK layer search (ModelSpace).

    Given a list of entities already filtered to target_layer, compute a center for each
    and return one Row per entity.

    This is intended for users who draw "search boxes" directly in ModelSpace (e.g. rectangle polylines)
    and want a reference point placed at the center of each.
    """
    rows: List[Row] = []
    counter = start_number
    any_3d = False

    total = len(ents)
    if log_cb:
        log_cb(f"[LayerSearch/ModelSpace] Entities on '{target_layer}': {total}")

    max_entity_logs = 30
    suppressed = False

    for i, ent in enumerate(ents, start=1):
        ent = dyn(ent)
        
        src_handle = str(getattr(ent, "Handle", ""))
        src_name = object_name(ent)

        if use_corners:
            # Extract 4 corner points from entity bbox
            corners_list = _bbox_corners_from_entity(ent)
            for corner_pt, corner_name in corners_list:
                if corner_pt:
                    any_3d = any_3d or abs(corner_pt[2]) > 1e-12
                    point_name = f"{cfg.prefix}{counter}_{corner_name}"
                    counter += 1
                    
                    rows.append(
                        Row(
                            point_name=point_name,
                            east=corner_pt[0],
                            north=corner_pt[1],
                            elev=corner_pt[2],
                            segment_name="",
                            dist_2d=None,
                            dist_3d=None,
                            bearing="",
                            azimuth="",
                            source_type=f"LayerSearchCorner/ModelSpace(layer={target_layer})",
                            source_handle=src_handle,
                            source_name=src_name,
                            source_index=i - 1,
                            corner_name=corner_name,
                        )
                    )
        else:
            # Original center-only mode
            c = _center_of_entity(ent)
            if not c:
                if progress_cb:
                    progress_cb(i, total, src_name, 0)
                continue

            any_3d = any_3d or abs(c[2]) > 1e-12
            point_name = f"{cfg.prefix}{counter}"
            counter += 1

            rows.append(
                Row(
                    point_name=point_name,
                    east=c[0],
                    north=c[1],
                    elev=c[2],
                    segment_name="",
                    dist_2d=None,
                    dist_3d=None,
                    bearing="",
                    azimuth="",
                    source_type=f"LayerSearchCenter/ModelSpace(layer={target_layer})",
                    source_handle=src_handle,
                    source_name=src_name,
                    source_index=i - 1,
                )
            )

        if log_cb:
            if len(rows) <= max_entity_logs:
                if use_corners:
                    log_cb(
                        f"[LayerSearch/ModelSpace] + 4 corners from {src_name} (handle={src_handle})"
                    )
                else:
                    log_cb(
                        f"[LayerSearch/ModelSpace] + {point_name} from {src_name} (handle={src_handle}) @ "
                        f"({c[0]:.3f}, {c[1]:.3f}, {c[2]:.3f})"
                    )
            elif not suppressed:
                suppressed = True
                log_cb(f"[LayerSearch/ModelSpace] (More than {max_entity_logs} entities; additional per-entity logs suppressed)")

        if progress_cb and (i % 25 == 0 or i == total):
            progress_cb(i, total, src_name, 0)

    if log_cb:
        log_cb(f"[LayerSearch/ModelSpace] Done. Points created={len(rows)}")
    return rows, any_3d, counter
def build_rows_polylines(cfg: Config, polylines: Sequence[Any]) -> Tuple[List[Row], bool]:
    rows: List[Row] = []
    counter = cfg.initial_number
    any_3d = False

    for ent in polylines:
        ent = dyn(ent)
        pts = poly_vertices(ent)
        closed = is_closed(ent)

        src_handle = str(getattr(ent, "Handle", ""))
        src_name = object_name(ent)

        if is_3d_poly(ent) or any(abs(p[2]) > 1e-12 for p in pts):
            any_3d = True

        names = [f"{cfg.prefix}{counter + i}" for i in range(len(pts))]
        counter += len(pts)

        for i, p in enumerate(pts):
            if i + 1 < len(pts):
                j = i + 1
            else:
                j = 0 if closed and len(pts) > 1 else -1

            if j == -1:
                seg = ""
                d2 = None
                d3 = None
                bear = ""
                az = ""
            else:
                q = pts[j]
                seg = f"{names[i]}-{names[j]}"
                d2 = dist2d(p, q)
                d3 = dist3d(p, q)

                dx, dy = (q[0] - p[0]), (q[1] - p[1])
                b_ang, quad = quadrant_bearing(dx, dy)
                bear = f"{dms_int(b_ang)} {quad}"
                az = dms_int(azimuth_deg_clockwise_from_north(dx, dy))

            rows.append(
                Row(
                    point_name=names[i],
                    east=p[0],
                    north=p[1],
                    elev=p[2],
                    segment_name=seg,
                    dist_2d=d2,
                    dist_3d=d3,
                    bearing=bear,
                    azimuth=az,
                    source_type="PolylineVertex",
                    source_handle=src_handle,
                    source_name=src_name,
                    source_index=i,
                )
            )

    return rows, any_3d


def build_rows_blocks(cfg: Config, blocks: Sequence[Any]) -> Tuple[List[Row], bool]:
    rows: List[Row] = []
    counter = cfg.initial_number
    any_3d = False

    for ent in blocks:
        ent = dyn(ent)
        name = block_effective_name(ent)
        if not matches_name_filter(name, cfg.block_name_filter):
            continue

        c = bbox_center(ent)
        any_3d = any_3d or abs(c[2]) > 1e-12

        src_handle = str(getattr(ent, "Handle", ""))
        src_name = name or object_name(ent)

        point_name = f"{cfg.prefix}{counter}"
        counter += 1

        rows.append(
            Row(
                point_name=point_name,
                east=c[0],
                north=c[1],
                elev=c[2],
                segment_name="",
                dist_2d=None,
                dist_3d=None,
                bearing="",
                azimuth="",
                source_type="BlockCenter",
                source_handle=src_handle,
                source_name=src_name,
                source_index=0,
            )
        )

    return rows, any_3d


def build_rows_layer_search_per_block(
    cfg: Config,
    doc: Any,
    blockrefs: Sequence[Any],
    target_layer: str,
    start_number: int = 1,
    use_corners: bool = False,
    progress_cb=None,
    log_cb=None,
) -> Tuple[List[Row], bool, int]:
    """
    INSIDE-BLOCK layer search.

    For each block reference, find entities on target_layer inside its block definition (and nested defs),
    compute a single center from those hits, and produce ONE Row per block reference.

    - Uses an in-memory cache so repeated block instances are cheap.
    - Progress/log callbacks are optional and are used by the UI for live feedback.
    """
    doc = dyn(doc)
    rows: List[Row] = []
    counter = start_number
    any_3d = False

    target_l = (target_layer or "").strip().lower()
    total = len(blockrefs)

    # Track which block definitions we've had to compute (cache misses) so logs are useful but not spammy.
    computed_defs: set[str] = set()

    for idx, bref in enumerate(blockrefs, start=1):
        bref = dyn(bref)

        bname = block_effective_name(bref)
        if not bname:
            if progress_cb:
                progress_cb(idx, total, 0, "<unnamed>")
            continue
        if not matches_name_filter(bname, cfg.block_name_filter):
            if progress_cb:
                progress_cb(idx, total, 0, bname)
            continue

        cache_key = (bname.lower(), target_l)
        cache_hit = cache_key in _BLOCKDEF_LAYERPTS_CACHE

        if (not cache_hit) and (bname.lower() not in computed_defs):
            computed_defs.add(bname.lower())
            if log_cb:
                log_cb(f"[LayerSearch/Blocks] Inspecting block def '{bname}' for layer '{target_layer}' (cache miss)")

        local_hits = _points_on_layer_in_blockdef(doc, bname, target_l)

        if (not cache_hit) and (bname.lower() in computed_defs):
            # Only log after compute; local_hits is now cached.
            if log_cb:
                log_cb(f"[LayerSearch/Blocks] Block def '{bname}' layer hits: {len(local_hits)}")

        if not local_hits:
            if progress_cb:
                progress_cb(idx, total, 0, bname)
            # Log a few "no hits" early so users can diagnose layer problems without flooding.
            if log_cb and idx <= 10:
                log_cb(f"[LayerSearch/Blocks] '{bname}' instance {idx}/{total}: no geometry found on '{target_layer}'")
            continue

        # Transform all local hits into world using this blockref's transform
        world_hits = [_apply_bref_local_transform(bref, p) for p in local_hits]
        src_handle = str(getattr(bref, "Handle", ""))

        if use_corners:
            # Extract 4 corner points
            corners = _bbox_corners_from_points(world_hits)
            for corner_pt, corner_name in corners:
                any_3d = any_3d or abs(corner_pt[2]) > 1e-12
                point_name = f"{cfg.prefix}{counter}_{corner_name}"
                counter += 1
                rows.append(
                    Row(
                        point_name=point_name,
                        east=corner_pt[0],
                        north=corner_pt[1],
                        elev=corner_pt[2],
                        segment_name="",
                        dist_2d=None,
                        dist_3d=None,
                        bearing="",
                        azimuth="",
                        source_type=f"LayerSearchCorner/Blocks(layer={target_layer})",
                        source_handle=src_handle,
                        source_name=bname,
                        source_index=idx - 1,
                        corner_name=corner_name,
                    )
                )
        else:
            # Original center-only mode
            center = _center_from_points(world_hits)
            if not center:
                if progress_cb:
                    progress_cb(idx, total, len(local_hits), bname)
                continue

            any_3d = any_3d or abs(center[2]) > 1e-12
            point_name = f"{cfg.prefix}{counter}"
            counter += 1

            rows.append(
                Row(
                    point_name=point_name,
                    east=center[0],
                    north=center[1],
                    elev=center[2],
                    segment_name="",
                    dist_2d=None,
                    dist_3d=None,
                    bearing="",
                    azimuth="",
                    source_type=f"LayerSearchCenter/Blocks(layer={target_layer})",
                    source_handle=src_handle,
                    source_name=bname,
                    source_index=idx - 1,
                    corner_name=None,
                )
            )

        if log_cb and len(rows) <= 10:
            if use_corners:
                # For corners mode, log the first corner
                if rows:
                    first_row = rows[-4] if len(rows) >= 4 else rows[-1]  # Log the NW corner or last row
                    log_cb(
                        f"[LayerSearch/Blocks] + 4 corners from '{bname}' (handle={src_handle}) using {len(local_hits)} layer hits"
                    )
            else:
                log_cb(
                    f"[LayerSearch/Blocks] + {point_name} from '{bname}' (handle={src_handle}) @ "
                    f"({center[0]:.3f}, {center[1]:.3f}, {center[2]:.3f}) using {len(local_hits)} layer hits"
                )

        if progress_cb:
            progress_cb(idx, total, len(local_hits), bname)

    if log_cb:
        log_cb(f"[LayerSearch/Blocks] Done. Block instances scanned={total}, points created={len(rows)}")
    return rows, any_3d, counter
def normalize_xlsx_path(path: str) -> str:
    path = path.strip()
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    return path


def next_available_path(base_path: str) -> str:
    base_path = normalize_xlsx_path(base_path)
    if not os.path.exists(base_path):
        return base_path

    folder = os.path.dirname(base_path) or "."
    stem = os.path.splitext(os.path.basename(base_path))[0]
    ext = ".xlsx"

    m = re.match(r"^(.*)_(\\d{3})$", stem)
    if m:
        stem_root = m.group(1)
        start_n = int(m.group(2)) + 1
    else:
        stem_root = stem
        start_n = 1

    for n in range(start_n, 1000):
        candidate = os.path.join(folder, f"{stem_root}_{n:03d}{ext}")
        if not os.path.exists(candidate):
            return candidate

    raise RuntimeError("Could not find an available filename (001-999).")


def get_active_excel_app() -> Optional[Any]:
    try:
        obj = pythoncom.GetActiveObject("Excel.Application")
        return dyn(obj)
    except Exception:
        return None


def try_close_workbook_in_excel(target_path: str) -> bool:
    target = os.path.abspath(target_path).lower()
    excel = get_active_excel_app()
    if not excel:
        return False

    try:
        prev_alerts = bool(excel.DisplayAlerts)
    except Exception:
        prev_alerts = True

    try:
        excel.DisplayAlerts = False
    except Exception:
        pass

    closed = False
    try:
        wbs = excel.Workbooks
        for i in range(1, int(wbs.Count) + 1):
            wb = dyn(wbs.Item(i))
            try:
                full = os.path.abspath(str(wb.FullName)).lower()
            except Exception:
                continue
            if full == target:
                try:
                    wb.Close(SaveChanges=False)
                except Exception:
                    try:
                        wb.Close(False)
                    except Exception:
                        pass
                closed = True
                break
    finally:
        try:
            excel.DisplayAlerts = prev_alerts
        except Exception:
            pass

    return closed


def ensure_deleted(path: str, attempts: int = 10, delay_s: float = 0.15) -> None:
    if not os.path.exists(path):
        return
    last_exc: Optional[Exception] = None
    for _ in range(attempts):
        try:
            os.remove(path)
            return
        except Exception as exc:
            last_exc = exc
            time.sleep(delay_s)
    if last_exc:
        raise last_exc


def excel_number_format(precision: int) -> str:
    return "0" if precision <= 0 else "0." + ("0" * precision)


def headers_for_export(cfg: Config, has_3d: bool, has_corners: bool = False) -> List[str]:
    cols = ["Point ID", "East (X)", "North (Y)"]
    o = cfg.table_options

    if has_corners:
        cols.insert(1, "Corner")  # Insert after Point ID
    
    if o.segment and cfg.mode == "polylines":
        cols.append("Segment (PointA–PointB)")
    if o.elevation:
        cols.append("Elevation (Z)")
    if o.distance and cfg.mode == "polylines":
        cols.append("Distance 2D (XY)")
    if o.distance_3d and has_3d and cfg.mode == "polylines":
        cols.append("Distance 3D (XYZ)")
    if o.bearing_quadrant and cfg.mode == "polylines":
        cols.append("Quadrant Bearing (0–90° + NE/SE/SW/NW)")
    if o.azimuth_from_north and cfg.mode == "polylines":
        cols.append("Azimuth (0–360°, clockwise from North)")

    cols.extend(["Source Type", "Source Handle", "Source Name", "Source Index"])
    return cols


def row_values_for_export(cfg: Config, r: Row, has_3d: bool, has_corners: bool = False) -> List[Any]:
    o = cfg.table_options
    out: List[Any] = [r.point_name]
    
    if has_corners and r.corner_name:
        out.append(r.corner_name)
    elif has_corners:
        out.append("")
    
    out.extend([r.east, r.north])

    if o.segment and cfg.mode == "polylines":
        out.append(r.segment_name)
    if o.elevation:
        out.append(r.elev)
    if o.distance and cfg.mode == "polylines":
        out.append(None if r.dist_2d is None else r.dist_2d)
    if o.distance_3d and has_3d and cfg.mode == "polylines":
        out.append(None if r.dist_3d is None else r.dist_3d)
    if o.bearing_quadrant and cfg.mode == "polylines":
        out.append(r.bearing)
    if o.azimuth_from_north and cfg.mode == "polylines":
        out.append(r.azimuth)

    out.extend([r.source_type, r.source_handle, r.source_name, r.source_index])
    return out


def export_excel(cfg: Config, rows: List[Row], has_3d: bool, last_export_path: Optional[str], has_corners: bool = False) -> str:
    base_path = normalize_xlsx_path(cfg.export.excel_path)
    os.makedirs(os.path.dirname(base_path) or ".", exist_ok=True)

    if cfg.export.replace_previous:
        target = normalize_xlsx_path(last_export_path or base_path)
        if os.path.exists(target):
            try_close_workbook_in_excel(target)
            ensure_deleted(target)
        out_path = target
    else:
        out_path = base_path
        if cfg.export.auto_increment and os.path.exists(out_path):
            out_path = next_available_path(out_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Coordinates"

    headers = headers_for_export(cfg, has_3d, has_corners)
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    for r in rows:
        ws.append(row_values_for_export(cfg, r, has_3d, has_corners))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    num_fmt = excel_number_format(cfg.precision)
    numeric_headers = {"East (X)", "North (Y)", "Elevation (Z)", "Distance 2D (XY)", "Distance 3D (XYZ)"}

    for col_idx, h in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = len(h)
        for row_idx in range(2, len(rows) + 2):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            width = max(width, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(width + 2, 12), 70)

        if h in numeric_headers:
            for row_idx in range(2, len(rows) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal="right")

    wb.save(out_path)
    return out_path


# -------------------------
# Reference block import: Xref attach -> bind -> delete instance
# -------------------------
def ensure_block_exists(doc: Any, block_name: str, dwg_path: str) -> str:
    """
    Ensure a block definition exists in doc.

    If not found, attach dwg_path as an XREF, bind it (creating a local block definition),
    then delete the temporary xref instance left in ModelSpace.

    Returns the block name you should insert.
    """
    doc = dyn(doc)
    dwg_path = os.path.abspath(dwg_path)

    # Already exists?
    try:
        doc.Blocks.Item(block_name)
        return block_name
    except Exception:
        log(f"Block '{block_name}' not found. Importing via Xref-Bind workaround...")

    if not os.path.exists(dwg_path):
        raise RuntimeError(f"External file not found: {dwg_path}")

    ms = dyn(doc.ModelSpace)
    origin = pt(0, 0, 0)

    preferred_xref_name = block_name
    xref_name = preferred_xref_name
    xref_obj = None

    def _attach(name: str):
        if hasattr(ms, "AttachExternalReference"):
            return ms.AttachExternalReference(dwg_path, name, origin, 1.0, 1.0, 1.0, 0.0, False)
        if hasattr(ms, "AttachXref"):
            return ms.AttachXref(dwg_path, name, origin, 1.0, 1.0, 1.0, 0.0, False)
        raise RuntimeError("Neither ModelSpace.AttachExternalReference nor ModelSpace.AttachXref is available.")

    try:
        try:
            xref_obj = com_call_with_retry(lambda: _attach(xref_name))
        except Exception:
            xref_name = f"TEMP_IMPORT_{block_name}_{int(time.time())}"
            xref_obj = com_call_with_retry(lambda: _attach(xref_name))

        cmd = f'_.-XREF _B "{xref_name}" \\n'
        com_call_with_retry(lambda: doc.SendCommand(cmd))
        if not wait_for_command_finish(doc, timeout_s=20.0):
            log("WARNING: Timed out waiting for -XREF BIND to finish. Continuing anyway...")

        try:
            if xref_obj is not None:
                dyn(xref_obj).Delete()
        except Exception:
            pass

        try:
            doc.Blocks.Item(block_name)
            log(f"Block '{block_name}' successfully imported.")
            return block_name
        except Exception:
            try:
                doc.Blocks.Item(xref_name)
                log(f"Block '{xref_name}' successfully imported (bound name).")
                return xref_name
            except Exception as exc:
                raise RuntimeError(
                    "Xref bind completed but expected block definition was not found.\\n"
                    f"Tried: '{block_name}' and '{xref_name}'."
                ) from exc

    except Exception as exc:
        raise RuntimeError(
            "Failed to import reference DWG via Xref-Bind workaround.\\n"
            f"DWG: {dwg_path}\\n"
            f"Block desired: {block_name}\\n"
            f"Details: {exc}"
        ) from exc


def insert_reference_block(
    doc: Any,
    ms: Any,
    ref_dwg_path: str,
    layer_name: str,
    x: float,
    y: float,
    z: float,
    scale: float,
    rotation_deg: float,
) -> Any:
    doc = dyn(doc)
    ms = dyn(ms)
    ref_dwg_path = os.path.abspath(ref_dwg_path)

    try:
        current = os.path.abspath(str(getattr(doc, "FullName", "")) or "")
    except Exception:
        current = ""

    if current and os.path.normcase(current) == os.path.normcase(ref_dwg_path):
        raise RuntimeError(
            "You are running the script while the active drawing IS the reference DWG.\\n"
            "Open your target drawing, then run the script again."
        )

    if not os.path.exists(ref_dwg_path):
        raise RuntimeError(
            f"Reference DWG not found: {ref_dwg_path}\n"
            "Tip: put 'Coordinate Reference Point.dwg' in an 'assets' folder next to this script, "
            "or browse to it in the UI."
        )

    block_name = os.path.splitext(os.path.basename(ref_dwg_path))[0]
    cache_key = os.path.normcase(ref_dwg_path)

    if cache_key in _REF_IMPORT_CACHE:
        insert_name = _REF_IMPORT_CACHE[cache_key]
    else:
        insert_name = ensure_block_exists(doc, block_name, ref_dwg_path)
        _REF_IMPORT_CACHE[cache_key] = insert_name

    ensure_layer(doc, layer_name)

    def _insert():
        return ms.InsertBlock(
            pt(x, y, z),
            insert_name,
            float(scale),
            float(scale),
            float(scale),
            math.radians(float(rotation_deg)),
        )

    br = com_call_with_retry(_insert)
    br = dyn(br)
    try:
        br.Layer = layer_name
    except Exception:
        pass
    return br


# -------------------------
# Layer listing (for dropdown)
# -------------------------
def list_layers(doc: Any) -> List[str]:
    doc = dyn(doc)
    layers = dyn(doc.Layers)
    names: List[str] = []
    try:
        n = int(layers.Count)
    except Exception:
        n = 0
    for i in range(n):
        try:
            layer = dyn(layers.Item(i))
            nm = str(getattr(layer, "Name", "")).strip()
            if nm:
                names.append(nm)
        except Exception:
            continue
    # Unique + sort
    return sorted({n for n in names}, key=lambda s: s.lower())


# -------------------------
# Qt UI
# -------------------------
class CoordinatesGrabberWindow(QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("Coordinates Grabber — AutoCAD → Excel + Reference Points")
        self.setMinimumSize(1120, 740)

        # COM / AutoCAD
        self.acad = connect_autocad()
        self.doc = dyn(self.acad.ActiveDocument)
        self.ms = dyn(self.doc.ModelSpace)
        self.ss = get_or_create_selection_set(self.doc, "CC_COORDGRABBER_SS")

        self.handles: List[str] = []
        self.last_export_path: Optional[str] = None

        # UI resources
        self.icons = IconManager()

        # Build UI
        self._build_ui()
        self._sync_mode_widgets()
        self.refresh_layers()

        # Always on top like your Tk app (toggle off during selection)
        self.setWindowFlag(Qt.WindowStaysOnTopHint, True)

    # ---------- UI building ----------
    def _build_ui(self) -> None:
        root = QWidget()
        root.setObjectName("AppRoot")
        self.setCentralWidget(root)

        root_layout = QHBoxLayout(root)
        root_layout.setContentsMargins(16, 16, 16, 16)
        root_layout.setSpacing(16)

        # Sidebar
        sidebar = QWidget()
        sidebar.setObjectName("Sidebar")
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(12, 12, 12, 12)
        sidebar_layout.setSpacing(12)

        brand = QLabel("Root3Power Tools")
        brand.setObjectName("BrandTitle")
        sidebar_layout.addWidget(brand)

        subtitle = QLabel("Coordinates Grabber")
        subtitle.setObjectName("SidebarSubtitle")
        subtitle.setWordWrap(True)
        sidebar_layout.addWidget(subtitle)

        # Actions
        self.btn_select = make_button("Select / Reselect", variant="primary", icon=self.icons.icon("select"))
        self.btn_select.clicked.connect(self.on_select)
        sidebar_layout.addWidget(self.btn_select)

        self.btn_run_layer_search = make_button("Run Layer Search", variant="secondary", icon=self.icons.icon("search"))
        self.btn_run_layer_search.clicked.connect(self.on_layer_search)
        sidebar_layout.addWidget(self.btn_run_layer_search)

        self.btn_quit = make_button("Quit", variant="ghost", icon=self.icons.icon("close"))
        self.btn_quit.clicked.connect(self.close)
        sidebar_layout.addWidget(self.btn_quit)

        # Activity log
        log_title = QLabel("Activity")
        log_title.setObjectName("SectionLabel")
        sidebar_layout.addWidget(log_title)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setObjectName("ActivityLog")
        self.log_box.setMinimumHeight(260)
        sidebar_layout.addWidget(self.log_box, 1)

        root_layout.addWidget(sidebar, 0)

        # Content (scrollable)
        content = QWidget()
        content.setObjectName("ContentArea")
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(12)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)
        scroll.setWidget(content)
        root_layout.addWidget(scroll, 1)

        # --- Mode ---
        mode_card = CardSection("Mode")
        content_layout.addWidget(mode_card)
        mode_layout = QVBoxLayout()
        mode_layout.setSpacing(8)

        self.rb_polylines = QRadioButton("Polyline vertices (ref point at EVERY vertex)")
        self.rb_blocks = QRadioButton("Selected block centers (ref point at each block center)")
        self.rb_layer_search = QRadioButton("Layer Search inside blocks (ref point at center of layer-geometry per block)")

        self.rb_polylines.setChecked(True)

        self.mode_group = QButtonGroup(self)
        for rb in (self.rb_polylines, self.rb_blocks, self.rb_layer_search):
            self.mode_group.addButton(rb)
            rb.toggled.connect(self._sync_mode_widgets)

        mode_layout.addWidget(self.rb_polylines)
        mode_layout.addWidget(self.rb_blocks)
        mode_layout.addWidget(self.rb_layer_search)

        mode_card.setContentLayout(mode_layout)

        # --- Selection ---
        sel_card = CardSection("Selection")
        content_layout.addWidget(sel_card)
        sel_layout = QVBoxLayout()
        sel_layout.setSpacing(10)

        row = QHBoxLayout()
        row.setSpacing(10)
        row.addWidget(QLabel("Selection stored in AutoCAD SelectionSet:"))
        self.lbl_selected = QLabel("Selected: 0 entities")
        self.lbl_selected.setObjectName("MutedLabel")
        row.addWidget(self.lbl_selected, 1)
        sel_layout.addLayout(row)

        self.chk_add_to_selection = QCheckBox("Add to current selection (instead of replacing)")
        sel_layout.addWidget(self.chk_add_to_selection)

        # Block filter in a container for proper visibility management
        filter_container = QWidget()
        self.block_filter_row = QHBoxLayout(filter_container)
        self.block_filter_row.setSpacing(10)
        self.block_filter_row.setContentsMargins(0, 0, 0, 0)
        self.lbl_block_filter = QLabel("Block name filter (optional, supports * and ?):") 
        self.block_filter_row.addWidget(self.lbl_block_filter)
        self.txt_block_filter = QLineEdit()
        self.txt_block_filter.setPlaceholderText("e.g. GRID*  or  ?POINT")
        self.txt_block_filter.setMaximumWidth(360)
        self.block_filter_row.addWidget(self.txt_block_filter, 1)
        self.block_filter_container = filter_container
        sel_layout.addWidget(filter_container)

        # --- Layer Search ---
        self.layer_card = CardSection("Layer Search Configuration")
        content_layout.addWidget(self.layer_card)
        layer_layout = QVBoxLayout()
        layer_layout.setSpacing(12)

        # Layer selection
        row = QHBoxLayout()
        row.setSpacing(10)
        row.addWidget(QLabel("Select layer:"))
        self.cmb_layers = QComboBox()
        self.cmb_layers.setMinimumWidth(280)
        row.addWidget(self.cmb_layers)
        self.btn_refresh_layers = make_button("", variant="secondary", icon=self.icons.icon("refresh"))
        self.btn_refresh_layers.clicked.connect(self.refresh_layers)
        self.btn_refresh_layers.setMaximumWidth(40)
        row.addWidget(self.btn_refresh_layers)
        row.addStretch(1)
        layer_layout.addLayout(row)

        # Result style selector (unified: center vs corners)
        style_label = QLabel("Reference block placement:")
        style_label.setObjectName("FieldLabel")
        layer_layout.addWidget(style_label)

        self.rb_result_center = QRadioButton("Single block at geometry center")
        self.rb_result_corners = QRadioButton("Four blocks at geometry corners (NW, NE, SW, SE)")
        self.rb_result_center.setChecked(True)

        self.result_style_group = QButtonGroup(self)
        self.result_style_group.addButton(self.rb_result_center)
        self.result_style_group.addButton(self.rb_result_corners)

        layer_layout.addWidget(self.rb_result_center)
        layer_layout.addWidget(self.rb_result_corners)

        # Scan options (compact, no overbearing checkboxes)
        layer_layout.addSpacing(8)
        sep = QLabel("")
        sep.setStyleSheet("border-top: 1px solid rgba(255,255,255,0.1); height: 0px;")
        layer_layout.addWidget(sep)
        layer_layout.addSpacing(8)

        self.chk_layer_use_selection = QCheckBox("Scan selected entities only")
        layer_layout.addWidget(self.chk_layer_use_selection)

        self.chk_layer_include_modelspace = QCheckBox("Also include ModelSpace geometry (outside blocks)")
        self.chk_layer_include_modelspace.setChecked(True)
        layer_layout.addWidget(self.chk_layer_include_modelspace)

        # Progress indicator
        self.pbar = QProgressBar()
        self.pbar.setRange(0, 100)
        self.pbar.setValue(0)
        self.pbar.setMaximumHeight(6)
        layer_layout.addWidget(self.pbar)

        self.lbl_progress = QLabel("")
        self.lbl_progress.setObjectName("MutedLabel")
        self.lbl_progress.setWordWrap(True)
        self.lbl_progress.setStyleSheet("font-size: 11px;")
        layer_layout.addWidget(self.lbl_progress)

        self.layer_card.setContentLayout(layer_layout)

        # --- Point naming + columns ---
        cols_card = CardSection("Point naming + columns")
        content_layout.addWidget(cols_card)

        cols_layout = QVBoxLayout()
        cols_layout.setSpacing(10)

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(10)

        grid.addWidget(QLabel("Point ID prefix:"), 0, 0)
        self.txt_prefix = QLineEdit("P")
        self.txt_prefix.setMaximumWidth(120)
        grid.addWidget(self.txt_prefix, 0, 1)

        grid.addWidget(QLabel("Start #:"), 0, 2)
        self.spin_start = QSpinBox()
        self.spin_start.setRange(1, 1_000_000)
        self.spin_start.setValue(1)
        self.spin_start.setMaximumWidth(140)
        grid.addWidget(self.spin_start, 0, 3)

        grid.addWidget(QLabel("Decimal places:"), 0, 4)
        self.spin_precision = QSpinBox()
        self.spin_precision.setRange(0, 12)
        self.spin_precision.setValue(3)
        self.spin_precision.setMaximumWidth(140)
        grid.addWidget(self.spin_precision, 0, 5)

        cols_layout.addLayout(grid)

        self.chk_segment = QCheckBox("Segment (PointA–PointB)")
        self.chk_segment.setChecked(True)
        self.chk_elev = QCheckBox("Elevation (Z)")
        self.chk_elev.setChecked(True)
        self.chk_dist = QCheckBox("Distance 2D (XY)")
        self.chk_dist.setChecked(True)
        self.chk_dist3 = QCheckBox("Distance 3D (XYZ)")
        self.chk_dist3.setChecked(False)
        self.chk_bearing = QCheckBox("Quadrant Bearing (0–90° + NE/SE/SW/NW)")
        self.chk_bearing.setChecked(False)
        self.chk_az = QCheckBox("Azimuth (0–360°, clockwise from North)")
        self.chk_az.setChecked(True)

        cols_layout.addWidget(self.chk_segment)
        cols_layout.addWidget(self.chk_elev)
        cols_layout.addWidget(self.chk_dist)
        cols_layout.addWidget(self.chk_dist3)
        cols_layout.addWidget(self.chk_bearing)
        cols_layout.addWidget(self.chk_az)

        cols_card.setContentLayout(cols_layout)

        # --- Reference block ---
        ref_card = CardSection("Reference point block (ALWAYS placed)")
        content_layout.addWidget(ref_card)

        ref_layout = QFormLayout()
        ref_layout.setLabelAlignment(Qt.AlignLeft)
        ref_layout.setFormAlignment(Qt.AlignLeft)
        ref_layout.setHorizontalSpacing(12)
        ref_layout.setVerticalSpacing(10)

        self.txt_ref_dwg = QLineEdit(self._default_ref_dwg_path())
        browse_ref = make_button("Browse…", variant="secondary", icon=self.icons.icon("folder"))
        browse_ref.clicked.connect(self.on_browse_ref_dwg)

        ref_row = QHBoxLayout()
        ref_row.addWidget(self.txt_ref_dwg, 1)
        ref_row.addWidget(browse_ref)
        ref_layout.addRow("Reference DWG:", ref_row)

        self.txt_ref_layer = QLineEdit("Coordinate Reference Point")
        self.txt_ref_layer.setMaximumWidth(320)
        ref_layout.addRow("Layer name:", self.txt_ref_layer)

        self.spin_ref_scale = QDoubleSpinBox()
        self.spin_ref_scale.setRange(0.0001, 1_000_000.0)
        self.spin_ref_scale.setDecimals(6)
        self.spin_ref_scale.setValue(1.0)
        self.spin_ref_scale.setMaximumWidth(200)
        ref_layout.addRow("Scale:", self.spin_ref_scale)

        self.spin_ref_rot = QDoubleSpinBox()
        self.spin_ref_rot.setRange(-360.0, 360.0)
        self.spin_ref_rot.setDecimals(3)
        self.spin_ref_rot.setValue(0.0)
        self.spin_ref_rot.setMaximumWidth(200)
        ref_layout.addRow("Rotation (deg):", self.spin_ref_rot)

        ref_card.setContentLayout(ref_layout)

        # --- Excel output ---
        exp_card = CardSection("Excel output")
        content_layout.addWidget(exp_card)

        exp_layout = QVBoxLayout()
        exp_layout.setSpacing(10)

        path_row = QHBoxLayout()
        path_row.setSpacing(10)
        path_row.addWidget(QLabel("Save as:"))
        self.txt_xlsx = QLineEdit(self._default_xlsx_path())
        path_row.addWidget(self.txt_xlsx, 1)
        self.btn_browse_xlsx = make_button("Browse…", variant="secondary", icon=self.icons.icon("folder"))
        self.btn_browse_xlsx.clicked.connect(self.on_browse_xlsx)
        path_row.addWidget(self.btn_browse_xlsx)
        exp_layout.addLayout(path_row)

        self.chk_replace_previous = QCheckBox(
            "Replace previous file (try close workbook in Excel, delete, regenerate)"
        )
        self.chk_replace_previous.setChecked(True)
        self.chk_replace_previous.stateChanged.connect(self._sync_export_mode)

        self.chk_autoinc = QCheckBox("Keep versions (auto-increment filename if it exists)")
        self.chk_autoinc.setChecked(False)
        self.chk_autoinc.stateChanged.connect(self._sync_export_mode)

        exp_layout.addWidget(self.chk_replace_previous)
        exp_layout.addWidget(self.chk_autoinc)

        exp_card.setContentLayout(exp_layout)

        content_layout.addStretch(1)

        # Status bar
        sb = QStatusBar()
        sb.setObjectName("StatusBar")
        self.setStatusBar(sb)
        self.lbl_status = QLabel("Ready.")
        sb.addWidget(self.lbl_status, 1)

    def _default_xlsx_path(self) -> str:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        return os.path.join(desktop, "CoordinateTable.xlsx")

    def _default_ref_dwg_path(self) -> str:
        """Default path for the reference DWG (block to insert).

        Looks for:
        1) ./assets/Coordinate Reference Point.dwg (recommended)
        2) ./Coordinate Reference Point.dwg
        """
        base = os.path.dirname(os.path.abspath(__file__))
        cand = os.path.join(base, "assets", "Coordinate Reference Point.dwg")
        if os.path.exists(cand):
            return cand
        return os.path.join(base, "Coordinate Reference Point.dwg")

    # ---------- UI state sync ----------
    def _sync_export_mode(self) -> None:
        if self.chk_replace_previous.isChecked() and self.chk_autoinc.isChecked():
            # keep the same mutually exclusive behavior
            self.chk_autoinc.setChecked(False)

    def _current_mode(self) -> str:
        if self.rb_blocks.isChecked():
            return "blocks"
        if self.rb_layer_search.isChecked():
            return "layer_search"
        return "polylines"

    def _sync_mode_widgets(self) -> None:
        mode = self._current_mode()
        is_poly = mode == "polylines"

        for cb in (self.chk_segment, self.chk_dist, self.chk_dist3, self.chk_bearing, self.chk_az):
            cb.setEnabled(is_poly)

        # block filter visible for blocks + layer_search
        show_filter = mode in ("blocks", "layer_search")
        if hasattr(self, 'block_filter_container'):
            self.block_filter_container.setVisible(show_filter)

        # layer search card only for layer_search mode
        self.layer_card.setVisible(mode == "layer_search")
        self.btn_run_layer_search.setEnabled(mode == "layer_search")

    # ---------- helpers ----------
    def _append_log(self, msg: str) -> None:
        self.log_box.append(msg)
        self.log_box.ensureCursorVisible()

    def _set_status(self, msg: str) -> None:
        self.lbl_status.setText(msg)

    def _message_box(self, title: str, text: str, icon=QMessageBox.Warning) -> None:
        box = QMessageBox(self)
        box.setWindowTitle(title)
        box.setText(text)
        box.setIcon(icon)
        box.exec()

    def _cfg_from_ui(self) -> Config:
        mode = self._current_mode()
        precision = int(self.spin_precision.value())
        prefix = (self.txt_prefix.text() or "P").strip()
        start = int(self.spin_start.value())

        opts = TableOptions(
            segment=bool(self.chk_segment.isChecked()),
            elevation=bool(self.chk_elev.isChecked()),
            distance=bool(self.chk_dist.isChecked()),
            distance_3d=bool(self.chk_dist3.isChecked()),
            bearing_quadrant=bool(self.chk_bearing.isChecked()),
            azimuth_from_north=bool(self.chk_az.isChecked()),
        )

        export = ExportOptions(
            excel_path=str(self.txt_xlsx.text()),
            replace_previous=bool(self.chk_replace_previous.isChecked()),
            auto_increment=bool(self.chk_autoinc.isChecked()),
        )

        refblock = RefBlockOptions(
            ref_dwg_path=str(self.txt_ref_dwg.text()),
            layer_name=str(self.txt_ref_layer.text() or "Coordinate Reference Point"),
            scale=float(self.spin_ref_scale.value()),
            rotation_deg=float(self.spin_ref_rot.value()),
        )

        return Config(
            mode=mode,
            precision=precision,
            prefix=prefix,
            initial_number=start,
            table_options=opts,
            export=export,
            refblock=refblock,
            block_name_filter=str(self.txt_block_filter.text() or "").strip(),
            layer_search_name=str(self.cmb_layers.currentText() or "").strip(),
            layer_search_use_selection=bool(self.chk_layer_use_selection.isChecked()),
            layer_search_include_modelspace=bool(getattr(self, 'chk_layer_include_modelspace', None).isChecked() if hasattr(self, 'chk_layer_include_modelspace') else True),
            layer_search_use_corners=bool(self.rb_result_corners.isChecked()) if hasattr(self, 'rb_result_corners') else False,
            add_to_selection=bool(self.chk_add_to_selection.isChecked()),
        )

    def _collect_selection_handles(self) -> List[str]:
        new_handles: List[str] = []
        for i in range(int(self.ss.Count)):
            try:
                ent = dyn(self.ss.Item(i))
                h = str(ent.Handle)
                new_handles.append(h)
            except Exception:
                continue
        return new_handles

    def _place_refpoints_for_rows(self, cfg: Config, rows: List[Row]) -> None:
        ref_path = os.path.abspath(cfg.refblock.ref_dwg_path or "")
        
        # If configured path doesn't exist, try to auto-detect
        if not ref_path or not os.path.exists(ref_path):
            self._append_log(f"[PlaceRefPoints] Path from UI not found: '{ref_path}'")
            auto_ref_path = self._default_ref_dwg_path()
            if auto_ref_path and os.path.exists(auto_ref_path):
                self._append_log(f"[PlaceRefPoints] Auto-detected: '{auto_ref_path}'")
                ref_path = auto_ref_path
            else:
                self._append_log(
                    f"[PlaceRefPoints] WARNING: Reference DWG not found. "
                    f"Tried '{cfg.refblock.ref_dwg_path}' and auto-detect. "
                    "Skipping block placement (export will still proceed)."
                )
                return

        self._append_log(f"[PlaceRefPoints] Using: '{ref_path}'")
        placed = 0
        for r in rows:
            try:
                insert_reference_block(
                    doc=self.doc,
                    ms=self.ms,
                    ref_dwg_path=ref_path,
                    layer_name=cfg.refblock.layer_name,
                    x=r.east,
                    y=r.north,
                    z=r.elev,
                    scale=cfg.refblock.scale,
                    rotation_deg=cfg.refblock.rotation_deg,
                )
                placed += 1
            except BaseException as exc:
                # Keep going; log the failure and continue.
                self._append_log(f"[PlaceRefPoints] Failed at {r.point_name}: {format_com_error(exc)}")

        try:
            self.doc.Regen(1)
        except Exception:
            pass

        self._append_log(f"[PlaceRefPoints] Placed reference points: {placed}/{len(rows)}")
    def on_browse_xlsx(self) -> None:
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel file",
            self.txt_xlsx.text() or self._default_xlsx_path(),
            "Excel Workbook (*.xlsx)",
        )
        if path:
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self.txt_xlsx.setText(path)

    def on_browse_ref_dwg(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Pick reference DWG to insert as block",
            self.txt_ref_dwg.text() or self._default_ref_dwg_path(),
            "DWG (*.dwg);;All Files (*)",
        )
        if path:
            self.txt_ref_dwg.setText(path)

    # ---------- layer listing ----------
    def refresh_layers(self) -> None:
        try:
            layers = list_layers(self.doc)
            self.cmb_layers.clear()
            self.cmb_layers.addItems(layers)
            if layers:
                self.cmb_layers.setCurrentIndex(0)
            self._set_status(f"Loaded {len(layers)} layers from current drawing.")
            self._append_log(f"[Layers] Loaded {len(layers)} layers.")
        except BaseException as exc:
            self._append_log(format_exception_text(exc, "refresh_layers"))
            self._set_status("Failed to load layers (see activity log).")

    # ---------- selection + exports ----------
    def _set_topmost(self, enabled: bool) -> None:
        # Changing this flag requires a re-show on Qt
        self.setWindowFlag(Qt.WindowStaysOnTopHint, enabled)
        self.show()

    def on_select(self) -> None:
        cfg = self._cfg_from_ui()
        mode = cfg.mode

        try:
            self._set_topmost(False)
        except Exception:
            pass

        try:
            self.ss.Clear()
            if mode == "polylines":
                prompt(self.doc, "\\nSelect polylines, then press Enter...\\n")
            else:
                prompt(self.doc, "\\nSelect blocks, then press Enter...\\n")
            self.ss.SelectOnScreen()
        except pywintypes.com_error as exc:
            self._append_log(format_exception_text(exc, "SelectionSet.SelectOnScreen"))
            self._set_status("Selection cancelled (or COM error). No export.")
            return
        finally:
            try:
                self._set_topmost(True)
            except Exception:
                pass

        new_handles = self._collect_selection_handles()
        if not new_handles:
            self.lbl_selected.setText("Selected: 0 entities")
            self._set_status("Nothing selected.")
            return

        if cfg.add_to_selection:
            combined = self.handles + new_handles
        else:
            combined = new_handles

        seen = set()
        self.handles = []
        for h in combined:
            if h not in seen:
                seen.add(h)
                self.handles.append(h)

        entities: List[Any] = []
        for h in self.handles:
            try:
                ent = resolve_by_handle(self.doc, h)
            except Exception:
                continue
            if mode == "polylines" and is_polyline(ent):
                entities.append(ent)
            elif mode in ("blocks", "layer_search") and is_block_reference(ent):
                entities.append(ent)

        self.lbl_selected.setText(f"Selected: {len(entities)} entities")
        self._set_status("Selection stored. (Layer Search can optionally use selection only.)")
        self._append_log(f"[Select] Stored {len(entities)} entities for mode={mode}.")

        # Auto-run export for polylines/blocks mode (keeps your existing flow)
        if mode in ("polylines", "blocks"):
            if not entities:
                self._set_status("Selected entities did not match mode. No export.")
                return
            self._run_export_for_entities(entities)

    def _run_export_for_entities(self, entities: Sequence[Any]) -> None:
        cfg = self._cfg_from_ui()
        try:
            if cfg.mode == "polylines":
                rows, has_3d = build_rows_polylines(cfg, entities)
            else:
                rows, has_3d = build_rows_blocks(cfg, entities)

            if not rows:
                self._set_status("No points found (or no blocks matched filter). No export.")
                return

            self._set_status("Placing reference points...")
            QApplication.processEvents()

            self._place_refpoints_for_rows(cfg, rows)

            self._set_status("Exporting Excel...")
            QApplication.processEvents()

            out_path = export_excel(cfg, rows, has_3d, self.last_export_path)
            self.last_export_path = out_path
            self.txt_xlsx.setText(out_path)

            self._append_log(f"[Excel] Exported: {out_path}")
            try:
                os.startfile(out_path)  # type: ignore[attr-defined]
            except Exception:
                pass

            self._set_status(f"Exported + opened: {out_path}")

        except BaseException as exc:
            self._append_log(format_exception_text(exc, "Export pipeline (build rows -> place blocks -> export excel)"))
            self._set_status("Export failed. See activity log for full error.")
            self._message_box("Export failed", "Export failed. Check the Activity log for details.")

    def on_layer_search(self) -> None:
        cfg = self._cfg_from_ui()
        layer_name = (cfg.layer_search_name or "").strip()
        if not layer_name:
            self.lbl_status.setText("Pick a layer first.")
            return

        # Clear per-run cache so changing layers gives correct results
        _BLOCKDEF_LAYERPTS_CACHE.clear()

        self._append_log("=" * 70)
        self._append_log(f"[LayerSearch] START | layer='{layer_name}' | selection_only={cfg.layer_search_use_selection} | include_modelspace={cfg.layer_search_include_modelspace}")
        if cfg.block_name_filter:
            self._append_log(f"[LayerSearch] Block name filter: '{cfg.block_name_filter}'")
        self.lbl_status.setText(f"Layer Search running for layer: '{layer_name}' ...")
        self.lbl_progress.setText("Preparing...")
        QApplication.processEvents()

        # Determine selection filter (if enabled)
        sel_handles: Optional[List[str]] = None
        if cfg.layer_search_use_selection and self.handles:
            sel_handles = list(self.handles)
            self._append_log(f"[LayerSearch] Using selection-only: {len(sel_handles)} selected handle(s).")

        # Determine which blockrefs to scan (inside-block search)
        blockrefs: List[Any] = []
        try:
            if sel_handles:
                for h in sel_handles:
                    try:
                        ent = resolve_by_handle(self.doc, h)
                    except Exception:
                        continue
                    if is_block_reference(ent):
                        blockrefs.append(ent)
            else:
                blockrefs = _blockrefs_in_modelspace(self.doc)
        except BaseException as exc:
            self._append_log(f"[LayerSearch] ERROR collecting block references: {format_com_error(exc)}")
            blockrefs = []

        self._append_log(f"[LayerSearch] Block references to scan: {len(blockrefs)}")

        # Pre-scan ModelSpace layer entities (outside-block search) so we know progress totals.
        ms_ents: List[Any] = []
        if cfg.layer_search_include_modelspace:
            try:
                ms_ents = _layer_entities_in_modelspace(
                    self.doc,
                    layer_name.strip().lower(),
                    handles=sel_handles,
                    log_cb=self._append_log,
                )
            except BaseException as exc:
                self._append_log(f"[LayerSearch] ERROR scanning ModelSpace entities: {format_com_error(exc)}")
                ms_ents = []

        total_work = max(1, len(blockrefs) + (len(ms_ents) if cfg.layer_search_include_modelspace else 0))
        self.pbar.setRange(0, total_work)
        self.pbar.setValue(0)

        def log_cb(msg: str) -> None:
            self._append_log(msg)

        def prog_blocks(i: int, tot: int, hits: int, bname: str) -> None:
            # i runs 1..tot
            self.pbar.setValue(min(i, total_work))
            self.lbl_progress.setText(f"Blocks: {i}/{tot} | current='{bname}' | layer hits in def={hits}")
            if i % 25 == 0 or i == tot:
                QApplication.processEvents()

        def prog_ms(i: int, tot: int, hits: int, label: str) -> None:
            done = len(blockrefs) + i
            self.pbar.setValue(min(done, total_work))
            self.lbl_progress.setText(f"ModelSpace: {i}/{tot} | {label} | items={hits}")
            if i % 25 == 0 or i == tot:
                QApplication.processEvents()

        try:
            # 1) Inside blocks
            rows_in, has_3d_in, next_num = build_rows_layer_search_per_block(
                cfg=cfg,
                doc=self.doc,
                blockrefs=blockrefs,
                target_layer=layer_name,
                start_number=cfg.initial_number,
                use_corners=cfg.layer_search_use_corners,
                progress_cb=prog_blocks,
                log_cb=log_cb,
            )

            # 2) Outside blocks (ModelSpace geometry on layer)
            rows_out: List[Row] = []
            has_3d_out = False
            if cfg.layer_search_include_modelspace:
                rows_out, has_3d_out, next_num = build_rows_layer_search_modelspace(
                    cfg=cfg,
                    ents=ms_ents,
                    target_layer=layer_name,
                    start_number=next_num,
                    use_corners=cfg.layer_search_use_corners,
                    progress_cb=prog_ms,
                    log_cb=log_cb,
                )

            rows = rows_in + rows_out
            has_3d = has_3d_in or has_3d_out

            self._append_log(
                f"[LayerSearch] Results: inside_blocks={len(rows_in)} point(s), modelspace={len(rows_out)} point(s), total={len(rows)}"
            )

            if not rows:
                self.lbl_status.setText(
                    f"No geometry found on layer '{layer_name}' (inside blocks or ModelSpace). "
                    "Verify your 'box' geometry truly lives on that layer."
                )
                self.lbl_progress.setText("")
                return

            # Place reference blocks (if ref DWG exists) and export
            self._append_log("[LayerSearch] Placing reference point blocks ...")
            self._place_refpoints_for_rows(cfg, rows)

            self._append_log("[LayerSearch] Exporting Excel ...")
            out_path = export_excel(cfg, rows, has_3d, self.last_export_path, has_corners=cfg.layer_search_use_corners)
            self.last_export_path = out_path
            self.txt_xlsx.setText(out_path)

            self._append_log(f"[LayerSearch] DONE → Exported: {out_path}")
            try:
                os.startfile(out_path)  # type: ignore[attr-defined]
            except Exception:
                pass

            self.lbl_status.setText(f"Layer Search done → Exported: {out_path}")
        except BaseException as exc:
            self._append_log(exc, "Layer search pipeline")
            self._append_log("[LayerSearch] FAILED (see traceback above).")
            self.lbl_status.setText("Layer search failed. See Activity log / terminal for details.")
        finally:
            self.lbl_progress.setText("")
            QApplication.processEvents()


def main() -> int:
    pythoncom.CoInitialize()
    try:
        app = QApplication(sys.argv)
        apply_theme(app)

        win = CoordinatesGrabberWindow()
        win.show()
        return app.exec()
    except KeyboardInterrupt:
        return 130
    except pywintypes.com_error as exc:
        # Check if this might be a gencache corruption issue
        msg = format_exception_text(exc, "main() top-level COM error")
        if "gen_py" in str(exc).lower() or "cache" in str(exc).lower():
            gen_py_hint = os.path.join(os.path.expanduser("~"), "AppData", "Local", "Temp", "gen_py")
            msg += (
                "\n\nNOTE: This might be a gen_py cache corruption issue.\n"
                "Try clearing the cache by deleting the folder at:\n"
                f"  {gen_py_hint}\n"
                "Then restart the script.\n"
                "See: https://support.pyxll.com/hc/en-gb/articles/360058200414"
            )
        print(msg)
        return 2
    except Exception as exc:
        print(format_exception_text(exc, "main() top-level error"))
        return 3
    finally:
        pythoncom.CoUninitialize()



if __name__ == "__main__":
    raise SystemExit(main())