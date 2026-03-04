from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime
from typing import Any, Iterable, Mapping, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def export_points_to_excel(
    points: Iterable[Mapping[str, Any]],
    precision: int,
    use_corners: bool,
    drawing_dir: Optional[str] = None,
    *,
    output_base_dir: Optional[str] = None,
    now_fn: Any = datetime.now,
) -> str:
    del use_corners  # maintained for API compatibility

    if drawing_dir:
        out_dir = drawing_dir
    else:
        base_dir = output_base_dir or os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(base_dir, "exports")
    os.makedirs(out_dir, exist_ok=True)

    timestamp = now_fn().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"coordinates_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Coordinates"

    headers = ["Point ID", "East (X)", "North (Y)", "Elevation (Z)", "Layer"]

    title_fill = PatternFill("solid", fgColor="2B6CB5")
    title_font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    layer_header_fill = PatternFill("solid", fgColor="5B9BD5")
    layer_header_font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    header_fill = PatternFill("solid", fgColor="3A3F47")
    header_font = Font(bold=True, color="F0F0F0", size=11, name="Arial")
    alt_fill_even = PatternFill("solid", fgColor="E8E6E2")
    alt_fill_odd = PatternFill("solid", fgColor="D4D1CC")
    data_font = Font(size=10, color="2A2A2A", name="Arial")
    border_side = Side(style="thin", color="B0ADA8")
    all_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    header_border = Border(
        left=border_side,
        right=border_side,
        top=border_side,
        bottom=Side(style="medium", color="3A3F47"),
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="Ground Grid Coordinates")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = all_border
    for col_idx in range(2, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = title_fill
        c.border = all_border

    num_fmt = "0" if precision <= 0 else "0." + ("0" * precision)
    numeric_cols = {"East (X)", "North (Y)", "Elevation (Z)"}

    normalized_points = [dict(p) for p in points]
    points_by_layer = defaultdict(list)
    for p in normalized_points:
        layer_name = p.get("layer", "Default")
        points_by_layer[layer_name].append(p)

    sorted_layers = sorted(points_by_layer.keys())
    current_row = 2

    for layer_idx, layer_name in enumerate(sorted_layers):
        layer_points = points_by_layer[layer_name]

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        layer_header_cell = ws.cell(row=current_row, column=1, value=f"Layer: {layer_name}")
        layer_header_cell.font = layer_header_font
        layer_header_cell.fill = layer_header_fill
        layer_header_cell.alignment = Alignment(horizontal="left", vertical="center")
        layer_header_cell.border = all_border
        for col_idx in range(2, len(headers) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.fill = layer_header_fill
            c.border = all_border
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = header_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for idx, p in enumerate(layer_points):
            row = [
                p["name"],
                p["x"],
                p["y"],
                p["z"],
                p.get("layer", ""),
            ]
            row_fill = alt_fill_even if idx % 2 == 0 else alt_fill_odd
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = all_border
                cell.font = data_font
                if headers[col_idx - 1] in numeric_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = num_fmt
                elif col_idx == 1:
                    cell.font = Font(bold=True, size=10, color="2A2A2A", name="Arial")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

        if layer_idx < len(sorted_layers) - 1:
            current_row += 2

    for col_idx, h in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = len(h)
        for p in normalized_points:
            for field_idx, field in enumerate([p["name"], p["x"], p["y"], p["z"], p.get("layer", "")]):
                if field_idx + 1 == col_idx:
                    width = max(width, len(str(field)))
        ws.column_dimensions[col_letter].width = min(max(width + 3, 14), 70)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"

    wb.save(out_path)
    return out_path
