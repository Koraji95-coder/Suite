#!/usr/bin/env python3
# pyright: reportMissingImports=false, reportMissingModuleSource=false
"""
Coordinates Grabber API Server
Flask-based HTTP/WebSocket bridge between React frontend and AutoCAD COM interface

Uses LATE-BOUND COM (dynamic dispatch) to avoid gen_py cache corruption.
Pattern taken from coordtable_excel_always_place_refpoints.py.

This server runs on localhost:5000 and provides:
- AutoCAD process detection (checks for acad.exe)
- COM connection management
- Layer and selection information
- Coordinate extraction from layers
- Real-time status updates

Usage:
    python api_server.py

Requirements:
    pip install flask flask-cors psutil pywin32
"""

from flask import Flask, jsonify, request
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import psutil
import pythoncom
import win32com.client
import win32com.client.gencache as gencache
import threading
import time
import json
import math
import os
import re
import traceback
import logging
from functools import wraps
from datetime import datetime
from typing import Optional, Dict, Any, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Logging configuration ────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
    handlers=[
        logging.FileHandler('api_server.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ── gen_py cache fix (from coordtable) ──────────────────────────
# Prevent gen_py from writing wrappers that cause CDispatch issues
gencache.is_readonly = True

app = Flask(__name__)

# CORS configuration - restrict to specific origins for security
# In production, replace with actual frontend domain
ALLOWED_ORIGINS = [
    "http://localhost:5173",  # Vite dev server
    "http://localhost:3000",  # Alternative dev port
    "http://127.0.0.1:5173",
    "http://127.0.0.1:3000",
]

CORS(app, 
     origins=ALLOWED_ORIGINS,
     supports_credentials=True,
     methods=["GET", "POST", "OPTIONS"],
     allow_headers=["Content-Type", "Authorization", "X-API-Key"])

# ── Rate Limiting ────────────────────────────────────────────────
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://",
    strategy="fixed-window"
)

# ── Security Headers ─────────────────────────────────────────────
@app.after_request
def add_security_headers(response):
    """Add security headers to all responses."""
    # Prevent clickjacking attacks
    response.headers['X-Frame-Options'] = 'DENY'
    # Prevent MIME-type sniffing
    response.headers['X-Content-Type-Options'] = 'nosniff'
    # Enable XSS protection
    response.headers['X-XSS-Protection'] = '1; mode=block'
    # Content Security Policy
    response.headers['Content-Security-Policy'] = "default-src 'self'"
    # Referrer policy
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # Strict Transport Security (HTTPS only - comment out for localhost)
    # response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# ── API Authentication ───────────────────────────────────────────
# SECURITY: API_KEY must be explicitly set in environment. No defaults allowed.
API_KEY = os.environ.get('API_KEY')
if not API_KEY:
    raise RuntimeError(
        "FATAL: API_KEY environment variable is not set.\n"
        "Please set your API key before starting the server:\n"
        "  export API_KEY='your-secure-api-key-here'\n"
        "Then start the server again."
    )

def require_api_key(f):
    """Decorator to require API key authentication for protected routes."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        provided_key = request.headers.get('X-API-Key')
        
        # Log all API requests for audit trail
        logger.info(
            f"API Request: {request.method} {request.path} from {request.remote_addr} "
            f"- Auth: {'Valid' if provided_key == API_KEY else 'Invalid/Missing'}"
        )
        
        if not provided_key:
            logger.warning(f"Unauthorized request (no API key): {request.path} from {request.remote_addr}")
            return jsonify({"error": "API key required", "code": "AUTH_REQUIRED"}), 401
        
        if provided_key != API_KEY:
            logger.warning(f"Unauthorized request (invalid API key): {request.path} from {request.remote_addr}")
            return jsonify({"error": "Invalid API key", "code": "AUTH_INVALID"}), 401
        
        return f(*args, **kwargs)
    return decorated_function


# ── Input Validation ─────────────────────────────────────────────
def validate_layer_config(config: Any) -> Dict[str, Any]:
    """
    Validate and sanitize layer extraction configuration.
    Prevents injection attacks and ensures data integrity.
    """
    if not isinstance(config, dict):
        raise ValueError("Config must be a JSON object")
    
    # Validate and sanitize layers
    layers = config.get('layers', [])
    if not isinstance(layers, list):
        raise ValueError("'layers' must be an array")
    if len(layers) > 100:  # Prevent DoS via excessive layers
        raise ValueError("Maximum 100 layers allowed")
    
    sanitized_layers = []
    for layer in layers:
        if not isinstance(layer, str):
            continue
        # Remove potentially dangerous characters, allow alphanumeric, dash, underscore, space
        sanitized = re.sub(r'[^a-zA-Z0-9\-_ ]', '', layer.strip())
        if sanitized and len(sanitized) <= 255:
            sanitized_layers.append(sanitized)
    
    # Validate block reference path if provided
    ref_dwg = config.get('ref_dwg', '')
    if ref_dwg:
        if not isinstance(ref_dwg, str):
            raise ValueError("'ref_dwg' must be a string")
        # Prevent path traversal attacks
        if '..' in ref_dwg or ref_dwg.startswith(('/', '\\\\')):
            raise ValueError("Invalid reference path")
        # Ensure .dwg extension
        if not ref_dwg.lower().endswith('.dwg'):
            raise ValueError ("'ref_dwg' must have .dwg extension")
    
    # Validate block name if provided
    block_name = config.get('block_name', '')
    if block_name:
        if not isinstance(block_name, str):
            raise ValueError("'block_name' must be a string")
        # Sanitize block name
        block_name = re.sub(r'[^a-zA-Z0-9\-_]', '', block_name.strip())
        if len(block_name) > 255:
            raise ValueError("Block name too long")
    
    return {
        'layers': sanitized_layers,
        'ref_dwg': ref_dwg.strip() if ref_dwg else '',
        'block_name': block_name,
        'export_excel': bool(config.get('export_excel', False))
    }

# Global AutoCAD manager instance
_manager = None
FOUNDATION_SOURCE_TYPE = "Foundation Coordinates"


# ── Late-bound COM helpers (from coordtable) ────────────────────
def dyn(obj: Any) -> Any:
    """
    Force late-bound dynamic dispatch on a COM object.
    Avoids stale gen_py wrappers and CDispatch type errors.
    """
    try:
        if type(obj).__name__ == "CDispatch":
            return obj
    except Exception:
        pass

    try:
        ole = obj._oleobj_
    except Exception:
        ole = obj

    try:
        disp = ole.QueryInterface(pythoncom.IID_IDispatch)
        return win32com.client.dynamic.Dispatch(disp)
    except Exception:
        try:
            return win32com.client.dynamic.Dispatch(obj)
        except Exception:
            return obj


def connect_autocad() -> Any:
    """Connect to AutoCAD using late-bound dynamic dispatch (no gen_py)."""
    acad = win32com.client.dynamic.Dispatch("AutoCAD.Application")
    if acad is None:
        raise RuntimeError("Could not connect to AutoCAD.Application")
    return dyn(acad)


def com_call_with_retry(callable_func, max_retries: int = 25, initial_delay: float = 0.03):
    """Retry COM calls that get RPC_E_CALL_REJECTED (AutoCAD busy)."""
    delay = initial_delay
    for _ in range(max_retries):
        try:
            return callable_func()
        except pythoncom.com_error as e:
            if e.args and e.args[0] == -2147418111:  # RPC_E_CALL_REJECTED
                time.sleep(delay)
                delay = min(delay * 1.5, 0.5)
                continue
            raise
    raise RuntimeError("AutoCAD COM call failed: RPC busy too long")


def pt(x: float, y: float, z: float = 0.0):
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (float(x), float(y), float(z)))


def ensure_layer(doc: Any, layer_name: str) -> None:
    doc = dyn(doc)
    try:
        layers = dyn(doc.Layers)
        try:
            layers.Item(layer_name)
        except Exception:
            layers.Add(layer_name)
    except Exception:
        pass


def wait_for_command_finish(doc: Any, timeout_s: float = 10.0) -> bool:
    doc = dyn(doc)
    t0 = time.time()
    while (time.time() - t0) < timeout_s:
        try:
            names = ""
            if hasattr(doc, "GetVariable"):
                names = str(doc.GetVariable("CMDNAMES") or "")
            if not names.strip():
                return True
        except Exception:
            pass
        time.sleep(0.15)
    return False


_REF_IMPORT_CACHE: Dict[str, str] = {}


def ensure_block_exists(doc: Any, block_name: str, dwg_path: str) -> str:
    doc = dyn(doc)
    dwg_path = os.path.abspath(dwg_path)
    try:
        doc.Blocks.Item(block_name)
        return block_name
    except Exception:
        logger.info(f"Block '{block_name}' not found. Importing via Xref-Bind...")

    if not os.path.exists(dwg_path):
        raise RuntimeError(f"External file not found: {dwg_path}")

    ms = dyn(doc.ModelSpace)
    origin = pt(0, 0, 0)
    xref_name = block_name

    def _attach(name: str):
        if hasattr(ms, "AttachExternalReference"):
            return ms.AttachExternalReference(dwg_path, name, origin, 1.0, 1.0, 1.0, 0.0, False)
        if hasattr(ms, "AttachXref"):
            return ms.AttachXref(dwg_path, name, origin, 1.0, 1.0, 1.0, 0.0, False)
        raise RuntimeError("Neither AttachExternalReference nor AttachXref available.")

    try:
        try:
            xref_obj = com_call_with_retry(lambda: _attach(xref_name))
        except Exception:
            xref_name = f"TEMP_IMPORT_{block_name}_{int(time.time())}"
            xref_obj = com_call_with_retry(lambda: _attach(xref_name))

        cmd = f'_.-XREF _B "{xref_name}" \\n'
        com_call_with_retry(lambda: doc.SendCommand(cmd))
        wait_for_command_finish(doc, timeout_s=20.0)

        try:
            if xref_obj is not None:
                dyn(xref_obj).Delete()
        except Exception:
            pass

        try:
            doc.Blocks.Item(block_name)
            return block_name
        except Exception:
            try:
                doc.Blocks.Item(xref_name)
                return xref_name
            except Exception as exc:
                raise RuntimeError(
                    f"Xref bind completed but block not found. Tried: '{block_name}', '{xref_name}'."
                ) from exc

    except Exception as exc:
        raise RuntimeError(
            f"Failed to import reference DWG.\nDWG: {dwg_path}\nBlock: {block_name}\nDetails: {exc}"
        ) from exc


def insert_reference_block(doc, ms, ref_dwg_path, layer_name, x, y, z, scale, rotation_deg):
    doc = dyn(doc)
    ms = dyn(ms)
    ref_dwg_path = os.path.abspath(ref_dwg_path)

    if not os.path.exists(ref_dwg_path):
        raise RuntimeError(
            f"Reference DWG not found: {ref_dwg_path}\n"
            "Put 'Coordinate Reference Point.dwg' in an 'assets' folder next to api_server.py."
        )

    block_name = os.path.splitext(os.path.basename(ref_dwg_path))[0]
    cache_key = os.path.normcase(ref_dwg_path)

    if cache_key in _REF_IMPORT_CACHE:
        insert_name = _REF_IMPORT_CACHE[cache_key]
    else:
        insert_name = ensure_block_exists(doc, block_name, ref_dwg_path)
        _REF_IMPORT_CACHE[cache_key] = insert_name

    ensure_layer(doc, layer_name)

    def _insert():
        return ms.InsertBlock(
            pt(x, y, z), insert_name,
            float(scale), float(scale), float(scale),
            math.radians(float(rotation_deg)),
        )

    br = com_call_with_retry(_insert)
    br = dyn(br)
    try:
        br.Layer = layer_name
    except Exception:
        pass
    return br


def add_point_label(ms, layer_name, label_text, x, y, z, scale):
    text_height = max(scale * 1.5, 0.5)
    x_offset = scale * 3.0

    def _add():
        return ms.AddText(label_text, pt(x + x_offset, y, z), text_height)

    txt = com_call_with_retry(_add)
    txt = dyn(txt)
    try:
        txt.Layer = layer_name
    except Exception:
        pass
    try:
        txt.Alignment = 0  # acAlignmentLeft
    except Exception:
        pass
    return txt


def default_ref_dwg_path() -> str:
    base = os.path.dirname(os.path.abspath(__file__))
    cand = os.path.join(base, "assets", "Coordinate Reference Point.dwg")
    if os.path.exists(cand):
        return cand
    return os.path.join(base, "Coordinate Reference Point.dwg")


def _entity_bbox(ent):
    ent = dyn(ent)
    try:
        mn, mx = ent.GetBoundingBox()
        minx, miny = float(mn[0]), float(mn[1])
        maxx, maxy = float(mx[0]), float(mx[1])
        minz = float(mn[2]) if len(mn) > 2 else 0.0
        maxz = float(mx[2]) if len(mx) > 2 else 0.0
        if maxx < minx:
            minx, maxx = maxx, minx
        if maxy < miny:
            miny, maxy = maxy, miny
        if maxz < minz:
            minz, maxz = maxz, minz
        return (minx, miny, minz, maxx, maxy, maxz)
    except Exception:
        return None


def _poly_centroid(ent):
    ent = dyn(ent)
    try:
        obj_name = str(ent.ObjectName)
    except Exception:
        obj_name = ''

    coords = []
    try:
        raw = list(ent.Coordinates)
        if obj_name == 'AcDb3dPolyline':
            for i in range(0, len(raw), 3):
                if i + 2 < len(raw):
                    coords.append((float(raw[i]), float(raw[i+1]), float(raw[i+2])))
        else:
            elev = 0.0
            try:
                elev = float(ent.Elevation)
            except Exception:
                pass
            for i in range(0, len(raw), 2):
                if i + 1 < len(raw):
                    coords.append((float(raw[i]), float(raw[i+1]), elev))
    except Exception:
        try:
            n = int(ent.NumberOfVertices)
            elev = 0.0
            try:
                elev = float(ent.Elevation)
            except Exception:
                pass
            for i in range(n):
                p = ent.Coordinate(i)
                z = float(p[2]) if len(p) > 2 else elev
                coords.append((float(p[0]), float(p[1]), z))
        except Exception:
            return None

    if not coords:
        return None

    n = len(coords)
    return (
        sum(p[0] for p in coords) / n,
        sum(p[1] for p in coords) / n,
        sum(p[2] for p in coords) / n,
    )


def _entity_center(ent):
    ent = dyn(ent)
    try:
        obj_name = str(ent.ObjectName)
    except Exception:
        obj_name = ''

    if obj_name in ('AcDbPolyline', 'AcDb2dPolyline', 'AcDb3dPolyline'):
        result = _poly_centroid(ent)
        if result:
            return result

    bbox = _entity_bbox(ent)
    if bbox:
        minx, miny, minz, maxx, maxy, maxz = bbox
        return ((minx + maxx) / 2.0, (miny + maxy) / 2.0, (minz + maxz) / 2.0)

    return None


def export_points_to_excel(points, precision, use_corners, drawing_dir=None):
    """
    Export coordinates to Excel with points organized by layer.
    Each layer gets its own table section with a 2-row gap between layers.
    """
    if drawing_dir:
        out_dir = drawing_dir
    else:
        out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "exports")
    os.makedirs(out_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(out_dir, f"coordinates_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Coordinates"

    headers = ["Point ID", "East (X)", "North (Y)", "Elevation (Z)", "Layer"]

    # ── Style definitions ──
    # Row 1: Title banner -- R3P logo blue (#2B6CB5)
    title_fill = PatternFill("solid", fgColor="2B6CB5")
    title_font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    # Layer section headers -- medium blue
    layer_header_fill = PatternFill("solid", fgColor="5B9BD5")
    layer_header_font = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    # Column headers -- dark charcoal gray
    header_fill = PatternFill("solid", fgColor="3A3F47")
    header_font = Font(bold=True, color="F0F0F0", size=11, name="Arial")
    # Data rows: alternating warm neutrals
    alt_fill_even = PatternFill("solid", fgColor="E8E6E2")
    alt_fill_odd = PatternFill("solid", fgColor="D4D1CC")
    data_font = Font(size=10, color="2A2A2A", name="Arial")
    # Borders: visible but not heavy
    border_side = Side(style="thin", color="B0ADA8")
    all_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    # Thicker border under header
    header_border = Border(
        left=border_side, right=border_side,
        top=border_side,
        bottom=Side(style="medium", color="3A3F47"),
    )

    # Row 1: merged title "Ground Grid Coordinates"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="Ground Grid Coordinates")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = all_border
    for col_idx in range(2, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = title_fill
        c.border = all_border

    num_fmt = "0" if precision <= 0 else "0." + ("0" * precision)
    numeric_cols = {"East (X)", "North (Y)", "Elevation (Z)"}

    # ── Group points by layer ──
    from collections import defaultdict
    points_by_layer = defaultdict(list)
    for p in points:
        layer_name = p.get('layer', 'Default')
        points_by_layer[layer_name].append(p)

    # Sort layers alphabetically for consistent output
    sorted_layers = sorted(points_by_layer.keys())

    current_row = 2  # Start after title row

    # ── Write each layer as a separate table ──
    for layer_idx, layer_name in enumerate(sorted_layers):
        layer_points = points_by_layer[layer_name]
        
        # Layer section header (merged across all columns)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        layer_header_cell = ws.cell(row=current_row, column=1, value=f"Layer: {layer_name}")
        layer_header_cell.font = layer_header_font
        layer_header_cell.fill = layer_header_fill
        layer_header_cell.alignment = Alignment(horizontal="left", vertical="center")
        layer_header_cell.border = all_border
        for col_idx in range(2, len(headers) + 1):
            c = ws.cell(row=current_row, column=col_idx)
            c.fill = layer_header_fill
            c.border = all_border
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        # Column headers for this layer section
        for col_idx, h in enumerate(headers, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = header_border
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Data rows for this layer
        for idx, p in enumerate(layer_points):
            row = [
                p['name'],
                p['x'],
                p['y'],
                p['z'],
                p.get('layer', ''),
            ]
            row_fill = alt_fill_even if idx % 2 == 0 else alt_fill_odd
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = all_border
                cell.font = data_font
                if headers[col_idx - 1] in numeric_cols:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = num_fmt
                elif col_idx == 1:
                    cell.font = Font(bold=True, size=10, color="2A2A2A", name="Arial")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

        # Add 2-row gap between layer sections (except after last layer)
        if layer_idx < len(sorted_layers) - 1:
            current_row += 2

    # Auto-fit column widths based on all data
    for col_idx, h in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = len(h)
        for p in points:
            for field_idx, field in enumerate([p['name'], p['x'], p['y'], p['z'], p.get('layer', '')]):
                if field_idx + 1 == col_idx:
                    width = max(width, len(str(field)))
        ws.column_dimensions[col_letter].width = min(max(width + 3, 14), 70)
    
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"
    
    wb.save(out_path)
    return out_path


class AutoCADManager:
    """
    Thread-safe AutoCAD connection manager
    Uses late-bound COM (dynamic dispatch) to avoid gen_py cache issues.
    """
    
    def __init__(self):
        self.start_time = time.time()
        self._lock = threading.Lock()
        self._cached_status = None
        self._cache_ttl = 2.0  # Cache status for 2 seconds
        self.last_check_time = 0
        
        print("[AutoCADManager] Initialized")
    
    def is_autocad_process_running(self) -> Tuple[bool, Optional[str]]:
        """
        Check if acad.exe process is running on Windows
        Returns: (is_running, process_exe_path)
        """
        try:
            for proc in psutil.process_iter(['name', 'exe']):
                try:
                    proc_name = proc.info.get('name', '').lower()
                    if proc_name == 'acad.exe':
                        return (True, proc.info.get('exe'))
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    continue
        except Exception as e:
            print(f"[AutoCADManager] Error checking process: {e}")
        
        return (False, None)
    
    def _fresh_com_connection(self) -> Tuple[Any, Any, bool, Optional[str], Optional[str]]:
        """
        Get a FRESH late-bound COM connection every time.
        Never caches COM objects across calls (avoids stale ref issues).
        Returns: (acad, doc, drawing_open, drawing_name, error_message)
        """
        try:
            acad = connect_autocad()
            
            try:
                doc = dyn(acad.ActiveDocument)
                if doc is None:
                    return (acad, None, False, None, "No drawing is open")
                
                try:
                    drawing_name = str(doc.Name)
                except Exception:
                    drawing_name = "Unknown"
                
                return (acad, doc, True, drawing_name, None)
                
            except Exception as e:
                return (acad, None, False, None, f"Cannot access ActiveDocument: {str(e)}")
                
        except Exception as e:
            return (None, None, False, None, f"Cannot connect to AutoCAD: {str(e)}")
    
    def get_status(self, force_refresh: bool = False) -> Dict[str, Any]:
        """
        Get comprehensive AutoCAD status.
        Uses process-level caching only; COM refs are always fresh.
        """
        with self._lock:
            current_time = time.time()
            
            # Return cached status if still valid
            if not force_refresh and self._cached_status is not None:
                if current_time - self._cached_status['timestamp'] < self._cache_ttl:
                    return self._cached_status
            
            process_running, acad_path = self.is_autocad_process_running()
            
            if not process_running:
                status = {
                    'connected': False,
                    'autocad_running': False,
                    'drawing_open': False,
                    'drawing_name': None,
                    'autocad_path': None,
                    'error': 'AutoCAD process (acad.exe) not detected',
                    'checks': {'process': False, 'com': False, 'document': False},
                    'backend_uptime': current_time - self.start_time,
                    'timestamp': current_time
                }
            else:
                # Fresh COM connection every time (no stale refs)
                try:
                    pythoncom.CoInitialize()
                    acad, doc, drawing_ok, drawing_name, error = self._fresh_com_connection()
                    com_ok = acad is not None
                except Exception as e:
                    com_ok, drawing_ok, drawing_name, error = False, False, None, str(e)
                finally:
                    try:
                        pythoncom.CoUninitialize()
                    except:
                        pass
                
                status = {
                    'connected': com_ok,
                    'autocad_running': process_running,
                    'drawing_open': drawing_ok,
                    'drawing_name': drawing_name,
                    'autocad_path': acad_path,
                    'error': error,
                    'checks': {
                        'process': process_running,
                        'com': com_ok,
                        'document': drawing_ok
                    },
                    'backend_uptime': current_time - self.start_time,
                    'timestamp': current_time
                }
            
            self._cached_status = status
            self.last_check_time = current_time
            return status
    
    def get_layers(self) -> Tuple[bool, List[str], Optional[str]]:
        """
        Get list of layer names from active drawing.
        Uses fresh late-bound COM connection every call.
        """
        status = self.get_status()
        
        if not status['drawing_open']:
            return (False, [], status.get('error', 'No drawing open'))
        
        try:
            pythoncom.CoInitialize()
            
            acad = connect_autocad()
            doc = dyn(acad.ActiveDocument)
            
            if doc is None:
                return (False, [], 'Document reference lost')
            
            layers = []
            layer_collection = dyn(doc.Layers)
            for i in range(int(layer_collection.Count)):
                layer = dyn(layer_collection.Item(i))
                layers.append(str(layer.Name))
            
            return (True, sorted(layers), None)
            
        except Exception as e:
            return (False, [], f'COM error: {str(e)}')
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass
    
    def execute_layer_search(self, config: Dict) -> Dict[str, Any]:
        """
        Execute layer search matching the desktop coordinatesgrabber.py logic:
        - Find entities on target layer in ModelSpace
        - Compute ONE center point per entity (not per vertex)
        - Insert reference blocks at each point
        - Export Excel and auto-open it
        """
        try:
            pythoncom.CoInitialize()

            acad = connect_autocad()
            doc = dyn(acad.ActiveDocument)
            ms = dyn(doc.ModelSpace)

            if doc is None or ms is None:
                raise RuntimeError('Cannot access AutoCAD document or modelspace')

            raw_layers = config.get('layer_search_names')
            requested_layers = []
            if isinstance(raw_layers, list):
                requested_layers.extend(
                    [str(layer).strip() for layer in raw_layers if str(layer).strip()]
                )

            fallback_layers_raw = str(config.get('layer_search_name', '')).strip()
            if fallback_layers_raw:
                for part in re.split(r'[;,\n]+', fallback_layers_raw):
                    layer_name_part = part.strip()
                    if layer_name_part:
                        requested_layers.append(layer_name_part)

            requested_layers = list(dict.fromkeys(requested_layers))

            if not requested_layers:
                return {
                    'success': False,
                    'points': [],
                    'count': 0,
                    'layers': [],
                    'excel_path': '',
                    'blocks_inserted': 0,
                    'error': 'No layer names provided'
                }

            requested_layer_lookup = {layer.strip().lower() for layer in requested_layers}
            prefix = config.get('prefix', 'P')
            start_num = int(config.get('initial_number', 1))
            precision = int(config.get('precision', 3))
            use_corners = config.get('layer_search_use_corners', False)

            points = []
            point_num = start_num
            entities_scanned = 0

            entity_count = int(ms.Count)
            for idx in range(entity_count):
                try:
                    ent = dyn(ms.Item(idx))

                    try:
                        ent_layer = str(ent.Layer)
                    except Exception:
                        continue

                    ent_layer_normalized = ent_layer.strip().lower()
                    if ent_layer_normalized not in requested_layer_lookup:
                        continue

                    entities_scanned += 1

                    if use_corners:
                        bbox = _entity_bbox(ent)
                        if not bbox:
                            continue
                        minx, miny, minz, maxx, maxy, maxz = bbox
                        z_val = (minz + maxz) / 2.0
                        corner_defs = [
                            (minx, maxy, 'NW'),
                            (maxx, maxy, 'NE'),
                            (minx, miny, 'SW'),
                            (maxx, miny, 'SE'),
                        ]
                        for cx, cy, corner_name in corner_defs:
                            points.append({
                                'name': f'{prefix}{point_num}_{corner_name}',
                                'x': round(cx, precision),
                                'y': round(cy, precision),
                                'z': round(z_val, precision),
                                'corner': corner_name,
                                'source_type': FOUNDATION_SOURCE_TYPE,
                                'layer': ent_layer.strip(),
                            })
                            point_num += 1
                    else:
                        center = _entity_center(ent)
                        if not center:
                            continue
                        cx, cy, cz = center
                        points.append({
                            'name': f'{prefix}{point_num}',
                            'x': round(cx, precision),
                            'y': round(cy, precision),
                            'z': round(cz, precision),
                            'source_type': FOUNDATION_SOURCE_TYPE,
                            'layer': ent_layer.strip(),
                        })
                        point_num += 1

                except Exception as e:
                    print(f"[execute] Entity {idx} error: {e}")
                    continue

            print(f"[execute] Scanned {entities_scanned} entities across layers {requested_layers}, extracted {len(points)} points")

            if not points:
                return {
                    'success': False,
                    'points': [],
                    'count': 0,
                    'layers': requested_layers,
                    'excel_path': '',
                    'blocks_inserted': 0,
                    'error': f'No entities found on requested layers: {", ".join(requested_layers)}'
                }

            ref_dwg = config.get('ref_dwg_path', '').strip()
            if not ref_dwg:
                ref_dwg = default_ref_dwg_path()
            ref_layer = config.get('ref_layer_name', 'Coordinate Reference Point')
            ref_scale = float(config.get('ref_scale', 1.0))
            ref_rotation = float(config.get('ref_rotation_deg', 0))

            blocks_inserted = 0
            block_errors = []
            if os.path.exists(ref_dwg):
                print(f"[execute] Inserting reference blocks from: {ref_dwg}")
                for p in points:
                    try:
                        insert_reference_block(
                            doc, ms, ref_dwg, ref_layer,
                            p['x'], p['y'], p['z'],
                            ref_scale, ref_rotation
                        )
                        try:
                            add_point_label(
                                ms, ref_layer, p['name'],
                                p['x'], p['y'], p['z'],
                                ref_scale,
                            )
                        except Exception as label_err:
                            print(f"[execute] Label at {p['name']}: {label_err}")
                        blocks_inserted += 1
                    except Exception as e:
                        block_errors.append(f"Block at {p['name']}: {e}")
                        print(f"[execute] Block insert error at {p['name']}: {e}")

                try:
                    doc.Regen(1)
                except Exception:
                    pass

                if blocks_inserted > 0:
                    print(f"[execute] Inserted {blocks_inserted} reference blocks")
            else:
                block_errors.append(f"Reference DWG not found: {ref_dwg}")
                print(f"[execute] WARNING: Reference DWG not found at {ref_dwg}, skipping block insertion")

            drawing_dir = None
            try:
                drawing_path = str(doc.FullName)
                if drawing_path:
                    drawing_dir = os.path.dirname(drawing_path)
            except Exception:
                pass

            excel_path = ''
            try:
                excel_path = export_points_to_excel(points, precision, use_corners, drawing_dir)
                print(f"[execute] Excel exported to: {excel_path}")
                try:
                    os.startfile(excel_path)
                except Exception:
                    pass
            except Exception as e:
                block_errors.append(f"Excel export: {e}")
                print(f"[execute] Excel export error: {e}")

            return {
                'success': True,
                'points': points,
                'count': len(points),
                'layers': requested_layers,
                'excel_path': excel_path,
                'blocks_inserted': blocks_inserted,
                'block_errors': block_errors if block_errors else None,
                'error': None
            }

        except Exception as e:
            traceback.print_exc()
            return {
                'success': False,
                'points': [],
                'count': 0,
                'layers': [],
                'excel_path': '',
                'blocks_inserted': 0,
                'error': str(e)
            }
        finally:
            try:
                pythoncom.CoUninitialize()
            except:
                pass


# Initialize manager
def get_manager() -> AutoCADManager:
    global _manager
    if _manager is None:
        _manager = AutoCADManager()
    return _manager


# ========== API ENDPOINTS ==========

@app.route('/api/status', methods=['GET'])
@require_api_key
def api_status():
    """
    Health check endpoint - returns detailed AutoCAD connection status
    """
    manager = get_manager()
    status = manager.get_status()
    status['backend_id'] = 'coordinates-grabber-api'
    status['backend_version'] = '1.0.0'

    http_code = 200 if status['autocad_running'] else 503

    return jsonify(status), http_code


@app.route('/api/layers', methods=['GET'])
@require_api_key
def api_layers():
    """
    List available layers in the active AutoCAD drawing
    
    Response:
    {
        "success": bool,
        "layers": [str],  # Array of layer names
        "count": int,
        "error": str|null
    }
    """
    manager = get_manager()
    success, layers, error = manager.get_layers()
    
    response = {
        'success': success,
        'layers': layers,
        'count': len(layers),
        'error': error
    }
    
    return jsonify(response), 200 if success else 503


@app.route('/api/selection-count', methods=['GET'])
@require_api_key
def api_selection_count():
    """Get count of currently selected objects in AutoCAD (fresh COM)."""
    manager = get_manager()
    status = manager.get_status()
    
    if not status['drawing_open']:
        return jsonify({'success': False, 'count': 0, 'error': 'No drawing open'}), 503
    
    try:
        pythoncom.CoInitialize()
        acad = connect_autocad()
        if acad is None:
            return jsonify({'success': False, 'count': 0, 'error': 'Cannot connect to AutoCAD'}), 503
        
        doc = dyn(acad.ActiveDocument)
        
        # Try to delete an old temp selection set first
        try:
            old_ss = doc.SelectionSets.Item("TEMP_COUNT")
            old_ss.Delete()
        except Exception:
            pass
        
        ss = doc.SelectionSets.Add("TEMP_COUNT")
        ss.SelectOnScreen()
        count = ss.Count
        ss.Delete()
        
        return jsonify({'success': True, 'count': count, 'error': None})
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'count': 0, 'error': f'COM error: {str(e)}'}), 500
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


@app.route('/api/execute', methods=['POST'])
@require_api_key
def api_execute():
    """
    Execute coordinate extraction based on provided configuration.
    Uses late-bound COM via the manager's execute_layer_search method.
    """
    manager = get_manager()
    status = manager.get_status()
    
    if not status['drawing_open']:
        return jsonify({
            'success': False,
            'message': 'No drawing open in AutoCAD',
            'points_created': 0,
            'error_details': 'Please open a drawing before executing'
        }), 400
    
    try:
        raw_config = request.get_json()
        if not raw_config:
            raise ValueError('No configuration provided')
        
        # Validate and sanitize input
        config = validate_layer_config(raw_config)
        
        start_time = time.time()
        
        result = manager.execute_layer_search(config)
        
        duration = time.time() - start_time
        
        if result['success']:
            blocks_inserted = result.get('blocks_inserted', 0)
            block_errors = result.get('block_errors')
            layers = result.get('layers', [])
            if layers:
                msg = f'Extracted {result["count"]} points from {len(layers)} layer(s): {", ".join(layers)}'
            else:
                msg = f'Extracted {result["count"]} points'
            if blocks_inserted > 0:
                msg += f', inserted {blocks_inserted} reference blocks'
            if block_errors:
                msg += f' (warnings: {len(block_errors)})'
            return jsonify({
                'success': True,
                'message': msg,
                'points_created': result['count'],
                'blocks_inserted': blocks_inserted,
                'excel_path': result.get('excel_path', ''),
                'duration_seconds': round(duration, 2),
                'points': result['points'],
                'block_errors': block_errors,
                'error_details': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'message': result.get('error', 'No entities found'),
                'points_created': 0,
                'blocks_inserted': 0,
                'excel_path': '',
                'duration_seconds': round(duration, 2),
                'points': [],
                'error_details': result.get('error')
            }), 400
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Execution failed: {str(e)}',
            'points_created': 0,
            'error_details': str(e)
        }), 500


@app.route('/api/trigger-selection', methods=['POST'])
@require_api_key
def api_trigger_selection():
    """Bring AutoCAD to foreground (fresh COM)."""
    manager = get_manager()
    status = manager.get_status()
    
    if not status['drawing_open']:
        return jsonify({'success': False, 'message': 'No drawing open'}), 503
    
    try:
        pythoncom.CoInitialize()
        acad = connect_autocad()
        if acad is None:
            return jsonify({'success': False, 'message': 'Cannot connect to AutoCAD'}), 503
        
        acad.Visible = True
        acad.WindowState = 1  # Restore if minimized
        
        return jsonify({'success': True, 'message': 'AutoCAD activated'})
    
    except Exception as e:
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


@app.route('/health', methods=['GET'])
def health():
    """Simple health check for backend server"""
    return jsonify({
        'status': 'running',
        'server': 'Coordinates Grabber API',
        'backend_id': 'coordinates-grabber-api',
        'version': '1.0.0',
        'timestamp': time.time()
    })


# ========== MAIN ==========

if __name__ == '__main__':
    print("=" * 60)
    print("🚀 Coordinates Grabber API Server")
    print("=" * 60)
    print(f"Server starting on: http://localhost:5000")
    print(f"Health check: http://localhost:5000/health")
    print(f"Status endpoint: http://localhost:5000/api/status")
    print("")
    print("📋 Prerequisites:")
    print("  - AutoCAD must be running")
    print("  - A drawing must be open in AutoCAD")
    print("  - React frontend should connect to localhost:5000")
    print("")
    print("Press Ctrl+C to stop the server")
    print("=" * 60)
    
    # Initialize manager to show initial status
    manager = get_manager()
    initial_status = manager.get_status()
    
    if initial_status['autocad_running']:
        print(f"✅ AutoCAD detected: {initial_status['autocad_path']}")
        if initial_status['drawing_open']:
            print(f"✅ Drawing open: {initial_status['drawing_name']}")
        else:
            print("⚠️  No drawing is currently open")
    else:
        print("❌ AutoCAD not detected - waiting for it to start...")
    
    print("=" * 60)
    print("")
    
    # Run Flask server
    app.run(
        host='0.0.0.0',  # Listen on all interfaces
        port=5000,
        debug=False,  # Set to True for development
        threaded=True
    )
