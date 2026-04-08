from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from flask import Flask, g
from flask_limiter import Limiter
from openpyxl import load_workbook

from backend.route_groups.api_drawing_program import create_drawing_program_blueprint


class TestApiDrawingProgram(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.project_root = Path(self.temp_dir.name)
        (self.project_root / "Templates").mkdir(parents=True, exist_ok=True)
        (self.project_root / "Templates" / "sheet-template.dwg").write_text(
            "template", encoding="utf-8"
        )
        (self.project_root / "PROJ-00001-E3-0001 - Existing.dwg").write_text(
            "existing", encoding="utf-8"
        )

        app = Flask(__name__)
        app.config["TESTING"] = True
        limiter = Limiter(
            app=app,
            key_func=lambda: "test-client",
            default_limits=[],
            storage_uri="memory://",
            strategy="fixed-window",
        )

        def require_supabase_user(f):
            def wrapped(*args, **kwargs):
                g.supabase_user = {"id": "user-1", "email": "user@example.com"}
                return f(*args, **kwargs)

            wrapped.__name__ = getattr(f, "__name__", "wrapped")
            return wrapped

        app.register_blueprint(
            create_drawing_program_blueprint(
                limiter=limiter,
                logger=app.logger,
                require_supabase_user=require_supabase_user,
            )
        )
        self.client = app.test_client()

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_apply_plan_copies_templates_renames_drawings_and_writes_artifacts(self) -> None:
        response = self.client.post(
            "/api/drawing-program/apply-plan",
            json={
                "projectId": "project-1",
                "projectRootPath": str(self.project_root),
                "profile": {
                    "acadeLine1": "Client",
                    "acadeLine2": "Site",
                    "acadeLine4": "PROJ-00001",
                    "acadeProjectFilePath": str(self.project_root / "Test.wdp"),
                },
                "program": {"rows": []},
                "plan": {
                    "updatedProgram": {
                        "id": "program-1",
                        "projectId": "project-1",
                        "workbookMirror": {
                            "workbookRelativePath": "Drawing Index.xlsx",
                        },
                        "rows": [
                            {
                                "id": "row-new",
                                "drawingNumber": "PROJ-00001-E3-0002",
                                "title": "New Three Line",
                                "status": "planned",
                                "provisionState": "provisioned",
                                "dwgRelativePath": "PROJ-00001-E3-0002 - New Three Line.dwg",
                                "templateKey": "3LINE",
                                "sheetFamily": "Three-Line Diagram",
                                "discipline": "E",
                                "acadeSection": "SCHEMATIC",
                                "acadeGroup": "",
                            },
                            {
                                "id": "row-existing",
                                "drawingNumber": "PROJ-00001-E3-0003",
                                "title": "Existing",
                                "status": "planned",
                                "provisionState": "provisioned",
                                "dwgRelativePath": "PROJ-00001-E3-0003 - Existing.dwg",
                                "templateKey": "3LINE",
                                "sheetFamily": "Three-Line Diagram",
                                "discipline": "E",
                                "acadeSection": "SCHEMATIC",
                                "acadeGroup": "",
                            },
                        ],
                    },
                    "fileActions": [
                        {
                            "kind": "copy-template",
                            "rowId": "row-new",
                            "toRelativePath": "PROJ-00001-E3-0002 - New Three Line.dwg",
                            "templatePath": "Templates/sheet-template.dwg",
                            "blocked": False,
                        },
                        {
                            "kind": "rename-dwg",
                            "rowId": "row-existing",
                            "fromRelativePath": "PROJ-00001-E3-0001 - Existing.dwg",
                            "toRelativePath": "PROJ-00001-E3-0003 - Existing.dwg",
                            "blocked": False,
                        },
                    ],
                    "workbookRows": [
                        {
                            "suiteRowId": "row-new",
                            "sortOrder": 10,
                            "drawingNumber": "PROJ-00001-E3-0002",
                            "title": "New Three Line",
                            "status": "planned",
                            "discipline": "E",
                            "sheetFamily": "Three-Line Diagram",
                            "templateKey": "3LINE",
                            "provisionState": "provisioned",
                            "dwgRelativePath": "PROJ-00001-E3-0002 - New Three Line.dwg",
                            "acadeSection": "SCHEMATIC",
                            "acadeGroup": "",
                        },
                        {
                            "suiteRowId": "row-existing",
                            "sortOrder": 20,
                            "drawingNumber": "PROJ-00001-E3-0003",
                            "title": "Existing",
                            "status": "planned",
                            "discipline": "E",
                            "sheetFamily": "Three-Line Diagram",
                            "templateKey": "3LINE",
                            "provisionState": "provisioned",
                            "dwgRelativePath": "PROJ-00001-E3-0003 - Existing.dwg",
                            "acadeSection": "SCHEMATIC",
                            "acadeGroup": "",
                        },
                    ],
                    "warnings": [],
                },
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json() or {}
        self.assertTrue(payload.get("success"))
        self.assertTrue(
            (self.project_root / "PROJ-00001-E3-0002 - New Three Line.dwg").exists()
        )
        self.assertTrue(
            (self.project_root / "PROJ-00001-E3-0003 - Existing.dwg").exists()
        )
        self.assertFalse(
            (self.project_root / "PROJ-00001-E3-0001 - Existing.dwg").exists()
        )
        workbook = load_workbook(self.project_root / "Drawing Index.xlsx")
        self.assertEqual(workbook.sheetnames, ["Drawing Index"])
        sheet = workbook["Drawing Index"]
        self.assertEqual(sheet["A2"].value, "row-new")
        wdp_text = (self.project_root / "Test.wdp").read_text(encoding="utf-8")
        self.assertIn("PROJ-00001-E3-0002 - New Three Line.dwg", wdp_text)
        self.assertIn("PROJ-00001-E3-0003 - Existing.dwg", wdp_text)

    def test_sync_acade_writes_current_program_stack(self) -> None:
        response = self.client.post(
            "/api/drawing-program/sync-acade",
            json={
                "projectId": "project-1",
                "projectRootPath": str(self.project_root),
                "profile": {
                    "acadeLine1": "Client",
                    "acadeLine2": "Site",
                    "acadeLine4": "PROJ-00001",
                    "acadeProjectFilePath": str(self.project_root / "SyncOnly.wdp"),
                },
                "program": {
                    "id": "program-1",
                    "projectId": "project-1",
                    "workbookMirror": {
                        "workbookRelativePath": "Drawing Index.xlsx",
                    },
                    "rows": [
                        {
                            "id": "row-existing",
                            "drawingNumber": "PROJ-00001-E3-0001",
                            "title": "Existing",
                            "status": "planned",
                            "provisionState": "provisioned",
                            "dwgRelativePath": "PROJ-00001-E3-0001 - Existing.dwg",
                            "templateKey": "3LINE",
                            "sheetFamily": "Three-Line Diagram",
                            "discipline": "E",
                            "acadeSection": "SCHEMATIC",
                            "acadeGroup": "",
                        }
                    ],
                },
            },
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json() or {}
        self.assertTrue(payload.get("success"))
        wdp_text = (self.project_root / "SyncOnly.wdp").read_text(encoding="utf-8")
        self.assertIn("PROJ-00001-E3-0001 - Existing.dwg", wdp_text)

    def test_apply_plan_rejects_file_actions_outside_project_root(self) -> None:
        response = self.client.post(
            "/api/drawing-program/apply-plan",
            json={
                "projectId": "project-1",
                "projectRootPath": str(self.project_root),
                "profile": {
                    "acadeProjectFilePath": str(self.project_root / "Test.wdp"),
                },
                "program": {"rows": []},
                "plan": {
                    "updatedProgram": {
                        "id": "program-1",
                        "projectId": "project-1",
                        "workbookMirror": {
                            "workbookRelativePath": "Drawing Index.xlsx",
                        },
                        "rows": [],
                    },
                    "fileActions": [
                        {
                            "kind": "copy-template",
                            "rowId": "row-new",
                            "toRelativePath": "../escape.dwg",
                            "templatePath": "Templates/sheet-template.dwg",
                            "blocked": False,
                        }
                    ],
                    "workbookRows": [],
                    "warnings": [],
                },
            },
        )

        self.assertEqual(response.status_code, 400)
        payload = response.get_json() or {}
        self.assertFalse(payload.get("success"))
        self.assertEqual(payload.get("message"), "Invalid request parameters.")
        self.assertFalse((self.project_root.parent / "escape.dwg").exists())

    def test_sync_acade_rejects_project_file_outside_project_root(self) -> None:
        with tempfile.TemporaryDirectory() as other_dir:
            outside_wdp = Path(other_dir) / "Outside.wdp"
            outside_wdp.write_text("outside", encoding="utf-8")
            response = self.client.post(
                "/api/drawing-program/sync-acade",
                json={
                    "projectId": "project-1",
                    "projectRootPath": str(self.project_root),
                    "profile": {
                        "acadeProjectFilePath": str(outside_wdp),
                    },
                    "program": {
                        "id": "program-1",
                        "projectId": "project-1",
                        "workbookMirror": {
                            "workbookRelativePath": "Drawing Index.xlsx",
                        },
                        "rows": [],
                    },
                },
            )

        self.assertEqual(response.status_code, 400)
        payload = response.get_json() or {}
        self.assertFalse(payload.get("success"))
        self.assertEqual(payload.get("message"), "Invalid request parameters.")
