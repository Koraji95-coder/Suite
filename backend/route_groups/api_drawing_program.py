from __future__ import annotations

import os
import re
import shutil
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Sequence

from flask import Blueprint, jsonify, request
from ..response_helpers import make_error_response
from flask_limiter import Limiter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import safe_join

from backend.domains.project_setup import (
    DEFAULT_WDP_CONFIG_LINES,
    PANEL_DRAWING_TITLE_HINTS,
)
from backend.runtime_paths import (
    is_absolute_path_value,
    resolve_runtime_directory,
)


def create_drawing_program_blueprint(
    *,
    limiter: Limiter,
    logger: Any,
    require_supabase_user: Callable,
) -> Blueprint:
    bp = Blueprint("drawing_program_api", __name__, url_prefix="/api/drawing-program")

    def _normalize_text(value: Any) -> str:
        return str(value or "").strip()

    def _normalize_path(value: Any) -> str:
        return _normalize_text(value).replace("\\", "/")

    def _request_id() -> str:
        raw = (
            request.headers.get("X-Request-ID")
            or request.headers.get("X-Request-Id")
            or ""
        )
        normalized = re.sub(r"[^A-Za-z0-9._:-]+", "-", raw)
        if normalized:
            return normalized[:128]
        return f"drawing-program-{int(datetime.now().timestamp() * 1000)}"

    def _error(message: str, status_code: int, request_id: str):
        return make_error_response(message, status=status_code)

    def _parse_json() -> Dict[str, Any]:
        payload = request.get_json(silent=True)
        return payload if isinstance(payload, dict) else {}

    def _resolve_project_root(payload: Dict[str, Any]) -> Path:
        raw = _normalize_text(payload.get("projectRootPath"))
        if not raw:
            raise ValueError("projectRootPath is required.")
        if not is_absolute_path_value(raw):
            raise ValueError("projectRootPath must be an absolute path.")
        project_root = resolve_runtime_directory(raw)
        if project_root is None:
            raise ValueError(f"projectRootPath does not exist or is not a directory: {raw}")
        return project_root.resolve()

    def _resolve_profile(payload: Dict[str, Any], project_root: Path) -> Dict[str, str]:
        raw_profile = payload.get("profile")
        profile = raw_profile if isinstance(raw_profile, dict) else {}
        return {
            "acadeLine1": _normalize_text(profile.get("acadeLine1") or profile.get("acade_line1")),
            "acadeLine2": _normalize_text(profile.get("acadeLine2") or profile.get("acade_line2")),
            "acadeLine4": _normalize_text(profile.get("acadeLine4") or profile.get("acade_line4")),
            "acadeProjectFilePath": _normalize_text(
                profile.get("acadeProjectFilePath") or profile.get("acade_project_file_path")
            ),
            "projectRootPath": _normalize_text(profile.get("projectRootPath")) or str(project_root),
        }

    def _validate_program(payload: Dict[str, Any]) -> Dict[str, Any]:
        raw_program = payload.get("program")
        if not isinstance(raw_program, dict):
            raise ValueError("program is required.")
        rows = raw_program.get("rows")
        if not isinstance(rows, list):
            raise ValueError("program.rows must be an array.")
        return raw_program

    def _validate_plan(payload: Dict[str, Any]) -> Dict[str, Any]:
        raw_plan = payload.get("plan")
        if not isinstance(raw_plan, dict):
            raise ValueError("plan is required.")
        if not isinstance(raw_plan.get("fileActions"), list):
            raise ValueError("plan.fileActions must be an array.")
        return raw_plan

    def _project_root_realpath(project_root: Path) -> str:
        return os.path.realpath(str(project_root))

    def _ensure_under_project_root(project_root: Path, candidate_path: str, original_value: str, field_name: str) -> Path:
        project_root_real = _project_root_realpath(project_root)
        candidate_real = os.path.realpath(candidate_path)
        try:
            if os.path.commonpath([project_root_real, candidate_real]) != project_root_real:
                raise ValueError
        except ValueError as exc:
            raise ValueError(f"{field_name} resolves outside the project root: {original_value}") from exc
        return Path(candidate_real)

    def _resolve_relative_under_root(project_root: Path, relative_path: str) -> Path:
        normalized = _normalize_path(relative_path)
        if not normalized:
            raise ValueError("Relative path is required.")
        if is_absolute_path_value(normalized):
            raise ValueError("Relative path must not be absolute.")
        absolute = safe_join(_project_root_realpath(project_root), normalized)
        if absolute is None:
            raise ValueError(f"Relative path resolves outside the project root: {relative_path}")
        return _ensure_under_project_root(project_root, absolute, relative_path, "Relative path")

    def _resolve_template_path(project_root: Path, template_path: str) -> Path:
        raw = _normalize_text(template_path)
        if not raw:
            raise ValueError("Template path is required.")
        project_root_real = _project_root_realpath(project_root)
        if is_absolute_path_value(raw):
            runtime_candidate = resolve_runtime_path(raw)
            if runtime_candidate is None:
                raise ValueError(f"Template file does not exist: {template_path}")
            candidate_real = os.path.realpath(str(runtime_candidate))
        else:
            candidate = safe_join(project_root_real, raw.replace("\\", "/"))
            if candidate is None:
                raise ValueError(f"Template path resolves outside the project root: {template_path}")
            candidate_real = os.path.realpath(candidate)
        try:
            if os.path.commonpath([project_root_real, candidate_real]) != project_root_real:
                raise ValueError
        except ValueError as exc:
            raise ValueError(
                f"Template path resolves outside the project root: {template_path}"
            ) from exc
        if not os.path.isfile(candidate_real):
            raise ValueError(f"Template file does not exist: {candidate_real}")
        return Path(candidate_real)

    def _preflight_file_actions(project_root: Path, file_actions: Sequence[Dict[str, Any]]) -> None:
        for action in file_actions:
            if not isinstance(action, dict):
                continue
            if action.get("blocked"):
                raise ValueError("Plan still contains blocked file actions. Resolve them before apply.")
            kind = _normalize_text(action.get("kind"))
            if kind == "copy-template":
                target = _resolve_relative_under_root(
                    project_root,
                    _normalize_text(action.get("toRelativePath")),
                )
                source = _resolve_template_path(project_root, _normalize_text(action.get("templatePath")))
                if target.exists():
                    raise ValueError(f"Provision target already exists: {target}")
                if source.suffix.lower() != ".dwg":
                    raise ValueError(f"Template file must be a DWG: {source}")
            elif kind == "rename-dwg":
                source = _resolve_relative_under_root(
                    project_root,
                    _normalize_text(action.get("fromRelativePath")),
                )
                target = _resolve_relative_under_root(
                    project_root,
                    _normalize_text(action.get("toRelativePath")),
                )
                if not source.exists():
                    raise ValueError(f"Rename source does not exist: {source}")
                if target.exists() and target != source:
                    raise ValueError(f"Rename target already exists: {target}")

    def _write_workbook(workbook_path: Path, rows: Sequence[Dict[str, Any]]) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Drawing Index"
        header = [
            "Suite Row ID",
            "Sort Order",
            "Drawing Number",
            "Title",
            "Status",
            "Discipline",
            "Sheet Family",
            "Family Key",
            "Type Code",
            "Sequence Band",
            "Template Key",
            "Provision State",
            "DWG Path",
            "ACADE Section",
            "ACADE Group",
        ]
        worksheet.append(header)
        header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
        header_font = Font(bold=True)
        border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row in rows:
            worksheet.append(
                [
                    _normalize_text(row.get("suiteRowId")),
                    int(row.get("sortOrder") or 0),
                    _normalize_text(row.get("drawingNumber")),
                    _normalize_text(row.get("title")),
                    _normalize_text(row.get("status")),
                    _normalize_text(row.get("discipline")),
                    _normalize_text(row.get("sheetFamily")),
                    _normalize_text(row.get("familyKey")),
                    _normalize_text(row.get("typeCode")),
                    _normalize_text(row.get("sequenceBand")),
                    _normalize_text(row.get("templateKey")),
                    _normalize_text(row.get("provisionState")),
                    _normalize_path(row.get("dwgRelativePath")),
                    _normalize_text(row.get("acadeSection")),
                    _normalize_text(row.get("acadeGroup")),
                ]
            )
        for index, column_name in enumerate(header, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = min(
                max(len(column_name) + 2, 14),
                42,
            )
        workbook_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(workbook_path)

    def _build_wdp_text(existing_text: str | None, profile: Dict[str, str], rows: Sequence[Dict[str, Any]]) -> str:
        owner = _normalize_text(profile.get("acadeLine1")) or "Suite Project"
        desc = _normalize_text(profile.get("acadeLine2")) or owner
        project_number = _normalize_text(profile.get("acadeLine4"))
        config_lines: List[str] = []
        if existing_text:
            for line in existing_text.splitlines():
                stripped = line.strip()
                if stripped.startswith("+[") or stripped.startswith(";"):
                    config_lines.append(line)
        if not any(line.strip().startswith("+[") for line in config_lines):
            config_lines = list(DEFAULT_WDP_CONFIG_LINES)

        lines: List[str] = [f"*[1]{owner}"]
        if desc:
            lines.append(f"*[2]{desc}")
        if project_number:
            lines.append(f"*[4]{project_number}")
        lines.extend(config_lines)

        for row in rows:
            relative_path = _normalize_path(row.get("dwgRelativePath"))
            if not relative_path:
                continue
            title = _normalize_text(row.get("title")) or Path(relative_path).stem
            subtype_context = f"{relative_path} {title}".upper()
            subtype = _normalize_text(row.get("acadeSection")) or (
                "PANEL" if any(hint in subtype_context for hint in PANEL_DRAWING_TITLE_HINTS) else "SCHEMATIC"
            )
            lines.append(f"==={title}")
            lines.append(f"=====SUB={subtype}")
            lines.append(relative_path)
        return "\n".join(lines) + "\n"

    def _apply_program_files(project_root: Path, file_actions: Sequence[Dict[str, Any]]) -> Dict[str, List[str]]:
        project_root_real = _project_root_realpath(project_root)
        created_files: List[str] = []
        renamed_from: List[str] = []
        renamed_to: List[str] = []
        for action in file_actions:
            if not isinstance(action, dict):
                continue
            kind = _normalize_text(action.get("kind"))
            if kind == "copy-template":
                source = _resolve_template_path(project_root, _normalize_text(action.get("templatePath")))
                target = _resolve_relative_under_root(project_root, _normalize_text(action.get("toRelativePath")))
                source_real = os.path.realpath(str(source))
                target_real = os.path.realpath(str(target))
                try:
                    if os.path.commonpath([project_root_real, source_real]) != project_root_real:
                        raise ValueError
                    if os.path.commonpath([project_root_real, target_real]) != project_root_real:
                        raise ValueError
                except ValueError as exc:
                    raise ValueError(
                        "Drawing program file operation resolves outside the project root."
                    ) from exc
                os.makedirs(os.path.dirname(target_real), exist_ok=True)
                shutil.copy2(source_real, target_real)
                created_files.append(_normalize_path(Path(target_real).relative_to(project_root)))
            elif kind == "rename-dwg":
                source = _resolve_relative_under_root(project_root, _normalize_text(action.get("fromRelativePath")))
                target = _resolve_relative_under_root(project_root, _normalize_text(action.get("toRelativePath")))
                target.parent.mkdir(parents=True, exist_ok=True)
                source.rename(target)
                renamed_from.append(_normalize_path(source.relative_to(project_root)))
                renamed_to.append(_normalize_path(target.relative_to(project_root)))
        return {
            "createdFiles": created_files,
            "renamedFrom": renamed_from,
            "renamedTo": renamed_to,
        }

    def _finalize_program(program: Dict[str, Any], workbook_written: bool) -> Dict[str, Any]:
        now_iso = datetime.now(UTC).isoformat().replace("+00:00", "Z")
        rows = program.get("rows")
        if isinstance(rows, list):
            for row in rows:
                if not isinstance(row, dict):
                    continue
                if _normalize_text(row.get("status")).lower() == "inactive":
                    continue
                if workbook_written:
                    row["workbookSyncedAt"] = now_iso
                    row["workbookDriftDetectedAt"] = None
        workbook_mirror = program.get("workbookMirror")
        if isinstance(workbook_mirror, dict) and workbook_written:
            workbook_mirror["lastExportedAt"] = now_iso
            workbook_mirror["lastDriftEventAt"] = None
        program["lastAcadeSyncAt"] = now_iso
        program["acadeSyncPending"] = False
        program["updatedAt"] = now_iso
        return program

    @bp.route("/apply-plan", methods=["POST"])
    @require_supabase_user
    @limiter.limit("120 per hour")
    def api_drawing_program_apply_plan():
        request_id = _request_id()
        payload = _parse_json()
        try:
            project_root = _resolve_project_root(payload)
            profile = _resolve_profile(payload, project_root)
            _validate_program(payload)
            plan = _validate_plan(payload)
            file_actions = [entry for entry in (plan.get("fileActions") or []) if isinstance(entry, dict)]
            _preflight_file_actions(project_root, file_actions)

            updated_program = plan.get("updatedProgram")
            if not isinstance(updated_program, dict):
                raise ValueError("plan.updatedProgram is required.")
            file_result = _apply_program_files(project_root, file_actions)

            workbook_rows = [
                entry for entry in (plan.get("workbookRows") or []) if isinstance(entry, dict)
            ]
            workbook_relative_path = _normalize_path(
                ((updated_program.get("workbookMirror") or {}) if isinstance(updated_program.get("workbookMirror"), dict) else {}).get("workbookRelativePath")
            ) or "Drawing Index.xlsx"
            workbook_path = _resolve_relative_under_root(project_root, workbook_relative_path)
            _write_workbook(workbook_path, workbook_rows)

            active_rows = [
                row
                for row in (updated_program.get("rows") or [])
                if isinstance(row, dict) and _normalize_text(row.get("status")).lower() != "inactive"
            ]
            configured_wdp_path = _normalize_text(profile.get("acadeProjectFilePath"))
            raw_wdp_path = configured_wdp_path or f"{project_root.name}.wdp"
            project_root_real = _project_root_realpath(project_root)
            if is_absolute_path_value(raw_wdp_path):
                absolute_wdp_path = os.path.realpath(raw_wdp_path)
                try:
                    if os.path.commonpath([project_root_real, absolute_wdp_path]) != project_root_real:
                        raise ValueError
                except ValueError as exc:
                    raise ValueError(
                        "ACADE project file path resolves outside the project root."
                    ) from exc
                wdp_relative = os.path.relpath(absolute_wdp_path, project_root_real)
            else:
                wdp_relative = raw_wdp_path.replace("\\", "/")
            wdp_candidate = safe_join(project_root_real, wdp_relative.replace("\\", "/"))
            if wdp_candidate is None:
                raise ValueError(
                    "ACADE project file path resolves outside the project root."
                )
            wdp_real = os.path.realpath(str(wdp_candidate))
            try:
                if os.path.commonpath([project_root_real, wdp_real]) != project_root_real:
                    raise ValueError
            except ValueError as exc:
                raise ValueError(
                    "ACADE project file path resolves outside the project root."
                ) from exc
            wdp_path = Path(wdp_real)
            if os.path.exists(wdp_real):
                if not os.path.isfile(wdp_real):
                    raise ValueError(f"ACADE project file path does not exist: {wdp_real}")
                with open(wdp_real, "r", encoding="utf-8") as handle:
                    existing_text = handle.read()
            else:
                existing_text = None
            os.makedirs(os.path.dirname(wdp_real), exist_ok=True)
            with open(wdp_real, "w", encoding="utf-8", newline="\n") as handle:
                handle.write(_build_wdp_text(existing_text, profile, active_rows))

            finalized_program = _finalize_program(dict(updated_program), True)
            return (
                jsonify(
                    {
                        "success": True,
                        "requestId": request_id,
                        "message": "Drawing program apply completed.",
                        "data": {
                            "program": finalized_program,
                            "workbookPath": str(workbook_path),
                            "wdpPath": str(wdp_path),
                            "createdFiles": file_result["createdFiles"],
                            "renamedFiles": [
                                {
                                    "fromRelativePath": left,
                                    "toRelativePath": right,
                                }
                                for left, right in zip(file_result["renamedFrom"], file_result["renamedTo"])
                            ],
                        },
                        "warnings": plan.get("warnings") or [],
                    }
                ),
                200,
            )
        except ValueError as exc:
            return _error("Invalid request parameters.", 400, request_id)
        except Exception as exc:
            logger.exception("Drawing program apply failed: %s", exc)
            return _error("Unable to apply the drawing program.", 500, request_id)

    @bp.route("/sync-acade", methods=["POST"])
    @require_supabase_user
    @limiter.limit("120 per hour")
    def api_drawing_program_sync_acade():
        request_id = _request_id()
        payload = _parse_json()
        try:
            project_root = _resolve_project_root(payload)
            profile = _resolve_profile(payload, project_root)
            program = _validate_program(payload)
            workbook_rows = [
                {
                    "suiteRowId": _normalize_text(row.get("id")),
                    "sortOrder": row.get("sortOrder") or 0,
                    "drawingNumber": _normalize_text(row.get("drawingNumber")),
                    "title": _normalize_text(row.get("title")),
                    "status": _normalize_text(row.get("status")),
                    "discipline": _normalize_text(row.get("discipline")),
                    "sheetFamily": _normalize_text(row.get("sheetFamily")),
                    "familyKey": _normalize_text(row.get("familyKey")),
                    "typeCode": _normalize_text(row.get("typeCode")),
                    "sequenceBand": _normalize_text(row.get("sequenceBand")),
                    "templateKey": _normalize_text(row.get("templateKey")),
                    "provisionState": _normalize_text(row.get("provisionState")),
                    "dwgRelativePath": _normalize_text(row.get("dwgRelativePath")),
                    "acadeSection": _normalize_text(row.get("acadeSection")),
                    "acadeGroup": _normalize_text(row.get("acadeGroup")),
                }
                for row in (program.get("rows") or [])
                if isinstance(row, dict) and _normalize_text(row.get("status")).lower() != "inactive"
            ]
            workbook_relative_path = _normalize_path(
                ((program.get("workbookMirror") or {}) if isinstance(program.get("workbookMirror"), dict) else {}).get("workbookRelativePath")
            ) or "Drawing Index.xlsx"
            workbook_path = _resolve_relative_under_root(project_root, workbook_relative_path)
            _write_workbook(workbook_path, workbook_rows)

            active_rows = [
                row
                for row in (program.get("rows") or [])
                if isinstance(row, dict) and _normalize_text(row.get("status")).lower() != "inactive"
            ]
            configured_wdp_path = _normalize_text(profile.get("acadeProjectFilePath"))
            raw_wdp_path = configured_wdp_path or f"{project_root.name}.wdp"
            project_root_real = _project_root_realpath(project_root)
            if is_absolute_path_value(raw_wdp_path):
                absolute_wdp_path = os.path.realpath(raw_wdp_path)
                try:
                    if os.path.commonpath([project_root_real, absolute_wdp_path]) != project_root_real:
                        raise ValueError
                except ValueError as exc:
                    raise ValueError(
                        "ACADE project file path resolves outside the project root."
                    ) from exc
                wdp_relative = os.path.relpath(absolute_wdp_path, project_root_real)
            else:
                wdp_relative = raw_wdp_path.replace("\\", "/")
            wdp_candidate = safe_join(project_root_real, wdp_relative.replace("\\", "/"))
            if wdp_candidate is None:
                raise ValueError(
                    "ACADE project file path resolves outside the project root."
                )
            wdp_real = os.path.realpath(str(wdp_candidate))
            try:
                if os.path.commonpath([project_root_real, wdp_real]) != project_root_real:
                    raise ValueError
            except ValueError as exc:
                raise ValueError(
                    "ACADE project file path resolves outside the project root."
                ) from exc
            wdp_path = Path(wdp_real)
            if os.path.exists(wdp_real):
                if not os.path.isfile(wdp_real):
                    raise ValueError(f"ACADE project file path does not exist: {wdp_real}")
                with open(wdp_real, "r", encoding="utf-8") as handle:
                    existing_text = handle.read()
            else:
                existing_text = None
            os.makedirs(os.path.dirname(wdp_real), exist_ok=True)
            with open(wdp_real, "w", encoding="utf-8", newline="\n") as handle:
                handle.write(_build_wdp_text(existing_text, profile, active_rows))

            finalized_program = _finalize_program(dict(program), True)
            return (
                jsonify(
                    {
                        "success": True,
                        "requestId": request_id,
                        "message": "ACADE sync completed.",
                        "data": {
                            "program": finalized_program,
                            "workbookPath": str(workbook_path),
                            "wdpPath": str(wdp_path),
                        },
                    }
                ),
                200,
            )
        except ValueError as exc:
            return _error("Invalid request parameters.", 400, request_id)
        except Exception as exc:
            logger.exception("Drawing program ACADE sync failed: %s", exc)
            return _error("Unable to sync the ACADE project stack.", 500, request_id)

    return bp
