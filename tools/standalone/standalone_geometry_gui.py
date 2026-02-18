"""
Standalone Geometry GUI - Tkinter interface for coordinates extraction

Simple GUI wrapper for standalone_geometry.py that allows:
- File selection (DXF/DWG)
- Layer selection from file (or AutoCAD active drawing)
- Multiple layer processing
- Configuration options
- Progress tracking
- AutoCAD integration (if available)
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import sys
from pathlib import Path
from typing import List, Optional, Any

# Try to import AutoCAD COM support
try:
    import pythoncom
    import win32com.client
    AUTOCAD_AVAILABLE = True
except ImportError:
    AUTOCAD_AVAILABLE = False
    pythoncom = None  # type: ignore
    win32com = None  # type: ignore

# Import the standalone geometry module
try:
    from standalone_geometry import (
        StandaloneConfig,
        process_file,
        load_cad_file,
        PRESETS,
        apply_preset
    )
except ImportError:
    import standalone_geometry
    StandaloneConfig = standalone_geometry.StandaloneConfig
    process_file = standalone_geometry.process_file
    load_cad_file = standalone_geometry.load_cad_file
    PRESETS = standalone_geometry.PRESETS
    apply_preset = standalone_geometry.apply_preset


# -------------------------
# AutoCAD COM helpers (from coordtable.py)
# -------------------------
def dyn(obj: Any) -> Any:
    """Convert to dynamic dispatch object."""
    if not AUTOCAD_AVAILABLE:
        return obj
    try:
        if type(obj).__name__ == "CDispatch":
            return obj
    except Exception:
        pass
    try:
        ole = obj._oleobj_  # type: ignore[attr-defined]
    except Exception:
        ole = obj
    try:
        disp = ole.QueryInterface(pythoncom.IID_IDispatch)
        return win32com.client.dynamic.Dispatch(disp)
    except Exception:
        try:
            return win32com.client.dynamic.Dispatch(obj)
        except Exception:
            return obj


def connect_autocad() -> Optional[Any]:
    """Connect to running AutoCAD instance."""
    if not AUTOCAD_AVAILABLE:
        return None
    try:
        acad = win32com.client.dynamic.Dispatch("AutoCAD.Application")
        if acad is None:
            return None
        return dyn(acad)
    except Exception:
        return None


def get_or_create_selection_set(doc: Any, name: str) -> Any:
    """Get or create a named selection set in AutoCAD."""
    doc = dyn(doc)
    try:
        ss = doc.SelectionSets.Item(name)
        ss.Clear()
        return dyn(ss)
    except Exception:
        pass
    try:
        return dyn(doc.SelectionSets.Add(name))
    except Exception:
        try:
            ss = doc.SelectionSets.Item(name)
            ss.Delete()
        except Exception:
            pass
        return dyn(doc.SelectionSets.Add(name))


def list_layers_from_autocad(doc: Any) -> List[str]:
    """Extract layer names from AutoCAD document."""
    layers = set()
    try:
        doc = dyn(doc)
        # Method 1: From layer table
        try:
            for layer in doc.Layers:
                layer_name = str(layer.Name)
                if layer_name:
                    layers.add(layer_name)
        except Exception:
            pass

        # Method 2: Scan modelspace entities
        try:
            mspace = dyn(doc.ModelSpace)
            for entity in mspace:
                try:
                    layer_name = str(entity.Layer)
                    if layer_name:
                        layers.add(layer_name)
                except Exception:
                    pass
        except Exception:
            pass
    except Exception:
        pass

    return sorted(list(layers))


def extract_from_autocad_layer(doc, layer_name: str, extraction_mode: str, point_prefix: str, decimal_places: int, selected_entities=None):
    """
    Extract coordinates directly from AutoCAD ModelSpace for a specific layer.
    Similar to coordtable.py's _layer_entities_in_modelspace approach.

    Args:
        doc: AutoCAD document
        layer_name: Layer to extract from (or None to extract from all layers when using selection)
        extraction_mode: "center" or "corners"
        point_prefix: Prefix for point IDs
        decimal_places: Number of decimal places
        selected_entities: Optional list of pre-selected entities to process
    """
    from standalone_geometry import StandaloneRow

    doc = dyn(doc)
    rows = []
    counter = 1

    # If we have selected entities, process only those
    if selected_entities:
        entities_to_process = selected_entities
    else:
        # Otherwise, scan all ModelSpace entities on the target layer
        ms = dyn(doc.ModelSpace)
        target_layer_lower = layer_name.strip().lower() if layer_name else None
        entities_to_process = []

        try:
            count = int(ms.Count)
        except Exception:
            count = 0

        for i in range(count):
            try:
                ent = dyn(ms.Item(i))

                # Check if entity is on target layer (if layer filter specified)
                if target_layer_lower:
                    try:
                        ent_layer = str(ent.Layer).strip().lower()
                    except Exception:
                        continue

                    if ent_layer != target_layer_lower:
                        continue

                entities_to_process.append(ent)
            except Exception:
                continue

    # Process each entity
    for ent in entities_to_process:
        try:
            ent = dyn(ent)

            # Get layer name from entity
            try:
                entity_layer = str(ent.Layer)
            except Exception:
                entity_layer = layer_name if layer_name else "Unknown"

            # Get bounding box
            try:
                mn, mx = ent.GetBoundingBox()
                minx, miny = float(mn[0]), float(mn[1])
                maxx, maxy = float(mx[0]), float(mx[1])

                # Normalize
                if maxx < minx:
                    minx, maxx = maxx, minx
                if maxy < miny:
                    miny, maxy = maxy, miny

                if extraction_mode == "corners":
                    # Extract 4 corners
                    corners = [
                        ("SW", minx, miny),
                        ("SE", maxx, miny),
                        ("NE", maxx, maxy),
                        ("NW", minx, maxy),
                    ]
                    for corner_name, x, y in corners:
                        point_id = f"{point_prefix}{counter:03d}"
                        rows.append(StandaloneRow(
                            point_id=point_id,
                            x=round(x, decimal_places),
                            y=round(y, decimal_places),
                            layer=entity_layer,
                            corner=corner_name
                        ))
                        counter += 1
                else:
                    # Center mode
                    cx = (minx + maxx) / 2.0
                    cy = (miny + maxy) / 2.0
                    point_id = f"{point_prefix}{counter:03d}"
                    rows.append(StandaloneRow(
                        point_id=point_id,
                        x=round(cx, decimal_places),
                        y=round(cy, decimal_places),
                        layer=entity_layer,
                        corner=None
                    ))
                    counter += 1

            except Exception:
                # Entity doesn't have bounding box, skip
                continue

        except Exception:
            continue

    return rows


class CoordinatesGrabberGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Coordinates Grabber - DXF/DWG to Excel")

        # Set minimum size and make resizable
        self.root.minsize(850, 750)
        self.root.geometry("900x800")

        self.input_file = tk.StringVar()
        self.output_file = tk.StringVar()
        self.selected_layers = []
        self.available_layers = []

        # AutoCAD integration
        self.acad = None
        self.acad_doc = None
        self.acad_ss = None  # Selection set
        self.selected_entities = []  # List of selected entities
        self.autocad_mode = tk.BooleanVar(value=False)
        self.use_selection = tk.BooleanVar(value=False)  # Whether to use selected entities

        # Try to connect to AutoCAD
        if AUTOCAD_AVAILABLE:
            try:
                self.acad = connect_autocad()
                if self.acad:
                    self.acad_doc = dyn(self.acad.ActiveDocument)
                    self.acad_ss = get_or_create_selection_set(self.acad_doc, "COORDGRAB_SS")
                    self.autocad_mode.set(True)
            except Exception:
                pass

        # Set icon if available
        try:
            self.root.iconbitmap(default='')
        except:
            pass

        self.setup_ui()

        # Initialize mode state
        if self.acad:
            self.on_mode_change()
        
    def setup_ui(self):
        # Main container with padding
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configure grid weights for proper resizing
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=0)  # Labels column
        main_frame.columnconfigure(1, weight=1)  # Input fields column (expandable)
        main_frame.columnconfigure(2, weight=0)  # Buttons column
        
        row = 0

        # Title and instructions
        title = ttk.Label(main_frame, text="Coordinates Grabber", font=('Arial', 16, 'bold'))
        title.grid(row=row, column=0, columnspan=3, pady=(0, 5))
        row += 1

        subtitle = ttk.Label(main_frame, text="Extract coordinates from DXF/DWG files to Excel", font=('Arial', 9))
        subtitle.grid(row=row, column=0, columnspan=3, pady=(0, 15))
        row += 1

        # AutoCAD status indicator
        if self.acad:
            autocad_status = ttk.Label(
                main_frame,
                text="✓ AutoCAD Connected",
                font=('Arial', 9, 'bold'),
                foreground="green"
            )
            autocad_status.grid(row=row, column=0, columnspan=3, pady=(0, 5))
            row += 1

        # Step 1 header
        step1 = ttk.Label(main_frame, text="STEP 1: Select Input Source", font=('Arial', 10, 'bold'))
        step1.grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(5, 5))
        row += 1

        # Mode selection (AutoCAD vs File)
        if self.acad:
            mode_frame = ttk.Frame(main_frame)
            mode_frame.grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=5)

            ttk.Radiobutton(
                mode_frame,
                text="Use AutoCAD Active Drawing",
                variable=self.autocad_mode,
                value=True,
                command=self.on_mode_change
            ).pack(side=tk.LEFT, padx=5)

            ttk.Radiobutton(
                mode_frame,
                text="Use DXF/DWG File",
                variable=self.autocad_mode,
                value=False,
                command=self.on_mode_change
            ).pack(side=tk.LEFT, padx=5)
            row += 1

        # Input file section
        self.file_label = ttk.Label(main_frame, text="DXF/DWG File:")
        self.file_label.grid(row=row, column=0, sticky=tk.W, pady=5)
        self.file_entry = ttk.Entry(main_frame, textvariable=self.input_file, width=50)
        self.file_entry.grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5)
        self.file_button = ttk.Button(main_frame, text="Browse...", command=self.browse_input)
        self.file_button.grid(row=row, column=2, padx=(5, 0), pady=5)
        row += 1

        # Step 2 header
        step2 = ttk.Label(main_frame, text="STEP 2: Select Entities or Load Layers", font=('Arial', 10, 'bold'))
        step2.grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(15, 5))
        row += 1

        # Selection option (AutoCAD only)
        if self.acad:
            selection_frame = ttk.Frame(main_frame)
            selection_frame.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)

            ttk.Checkbutton(
                selection_frame,
                text="Use selected entities only (select polylines/entities in AutoCAD)",
                variable=self.use_selection
            ).pack(side=tk.LEFT)

            self.select_btn = ttk.Button(
                selection_frame,
                text="🎯 Select Entities in AutoCAD",
                command=self.select_entities_in_autocad
            )
            self.select_btn.pack(side=tk.LEFT, padx=(10, 0))

            self.selection_label = ttk.Label(selection_frame, text="(0 selected)", foreground="gray")
            self.selection_label.pack(side=tk.LEFT, padx=(5, 0))

            row += 1

        # Load layers button - make it more prominent
        load_btn = ttk.Button(main_frame, text="📂 Load Layers from File", command=self.load_layers)
        load_btn.grid(row=row, column=0, columnspan=3, pady=5, sticky=(tk.W, tk.E))
        row += 1
        
        # Layer selection - using dropdown/combobox
        ttk.Label(main_frame, text="Select Layer to Extract:").grid(row=row, column=0, sticky=tk.W, pady=5)
        row += 1

        # Combobox for layer selection
        layer_frame = ttk.Frame(main_frame)
        layer_frame.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=5)

        self.layer_combo = ttk.Combobox(layer_frame, state='readonly', width=60)
        self.layer_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

        # Add layer button
        self.add_layer_btn = ttk.Button(layer_frame, text="Add Layer", command=self.add_layer_to_list)
        self.add_layer_btn.pack(side=tk.LEFT)

        row += 1

        # Selected layers list - make it more prominent
        ttk.Label(main_frame, text="Layers to Extract:", font=('TkDefaultFont', 9, 'bold')).grid(row=row, column=0, sticky=tk.W, pady=(10, 5))
        row += 1

        # Listbox showing selected layers - increased height
        list_frame = ttk.Frame(main_frame)
        list_frame.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)

        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.selected_layers_listbox = tk.Listbox(
            list_frame,
            height=10,  # Increased from 6 to 10
            yscrollcommand=scrollbar.set,
            font=('Consolas', 9),
            selectmode=tk.EXTENDED
        )
        self.selected_layers_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.selected_layers_listbox.yview)

        main_frame.rowconfigure(row, weight=2)  # Give more weight for expansion
        row += 1

        # Layer management buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.grid(row=row, column=0, columnspan=3, pady=5)
        ttk.Button(btn_frame, text="Remove Selected", command=self.remove_layer_from_list).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="Clear All", command=self.clear_all_layers).pack(side=tk.LEFT, padx=2)
        row += 1
        
        # Step 3 header
        step3 = ttk.Label(main_frame, text="STEP 3: Configure Extraction", font=('Arial', 10, 'bold'))
        step3.grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(15, 5))
        row += 1
        
        # Extraction mode
        ttk.Label(main_frame, text="Extraction Mode:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.mode_var = tk.StringVar(value="center")
        mode_frame = ttk.Frame(main_frame)
        mode_frame.grid(row=row, column=1, sticky=tk.W, pady=5)
        ttk.Radiobutton(mode_frame, text="Center Point", variable=self.mode_var, value="center").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(mode_frame, text="4 Corners", variable=self.mode_var, value="corners").pack(side=tk.LEFT, padx=5)
        row += 1
        
        # Point prefix
        ttk.Label(main_frame, text="Point Prefix:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.prefix_var = tk.StringVar(value="P")
        ttk.Entry(main_frame, textvariable=self.prefix_var, width=10).grid(row=row, column=1, sticky=tk.W, pady=5)
        row += 1
        
        # Decimal places
        ttk.Label(main_frame, text="Decimal Places:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.decimals_var = tk.IntVar(value=3)
        ttk.Spinbox(main_frame, from_=0, to=6, textvariable=self.decimals_var, width=10).grid(row=row, column=1, sticky=tk.W, pady=5)
        row += 1
        
        # Output file
        ttk.Label(main_frame, text="Output Excel File:").grid(row=row, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.output_file, width=50).grid(row=row, column=1, sticky=(tk.W, tk.E), pady=5)
        ttk.Button(main_frame, text="Browse...", command=self.browse_output).grid(row=row, column=2, padx=(5, 0), pady=5)
        row += 1

        # Step 4 header
        step4 = ttk.Label(main_frame, text="STEP 4: Extract", font=('Arial', 10, 'bold'))
        step4.grid(row=row, column=0, columnspan=3, sticky=tk.W, pady=(15, 5))
        row += 1

        # Extract button - make it prominent
        extract_btn = ttk.Button(
            main_frame,
            text="🚀 Extract Coordinates to Excel",
            command=self.extract_coordinates
        )
        extract_btn.grid(row=row, column=0, columnspan=3, pady=5, sticky=(tk.W, tk.E), ipady=8)
        row += 1

        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='determinate', maximum=100)
        self.progress.grid(row=row, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        row += 1

        # Status label
        self.status_var = tk.StringVar(value="Ready - Select a DXF/DWG file to begin")
        status_label = ttk.Label(main_frame, textvariable=self.status_var, foreground="blue")
        status_label.grid(row=row, column=0, columnspan=3, pady=5)
        row += 1

    def on_mode_change(self):
        """Handle mode change between AutoCAD and File."""
        is_autocad = self.autocad_mode.get()

        # Enable/disable file selection widgets
        state = 'disabled' if is_autocad else 'normal'
        self.file_entry.config(state=state)
        self.file_button.config(state=state)

        if is_autocad:
            self.input_file.set("[AutoCAD Active Drawing]")
            self.status_var.set("AutoCAD mode - Click 'Load Layers' to read from active drawing")
        else:
            self.input_file.set("")
            self.status_var.set("File mode - Select a DXF/DWG file")

    def browse_input(self):
        """Browse for input DXF/DWG file."""
        filename = filedialog.askopenfilename(
            title="Select DXF/DWG File",
            filetypes=[
                ("CAD Files", "*.dxf *.dwg"),
                ("DXF Files", "*.dxf"),
                ("DWG Files", "*.dwg"),
                ("All Files", "*.*")
            ]
        )
        if filename:
            self.input_file.set(filename)
            # Auto-set output file
            if not self.output_file.get():
                base = os.path.splitext(filename)[0]
                self.output_file.set(f"{base}_extracted.xlsx")

    def browse_output(self):
        """Browse for output Excel file."""
        filename = filedialog.asksaveasfilename(
            title="Save Excel File As",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )
        if filename:
            self.output_file.set(filename)

    def load_layers(self):
        """Load layers from AutoCAD or DXF/DWG file."""
        # Check if using AutoCAD mode
        if self.autocad_mode.get() and self.acad_doc:
            self.load_layers_from_autocad()
            return

        # File mode
        input_path = self.input_file.get()
        if not input_path or input_path == "[AutoCAD Active Drawing]":
            messagebox.showwarning("No File", "Please select an input file first.")
            return

        if not os.path.exists(input_path):
            messagebox.showerror("File Not Found", f"File not found: {input_path}")
            return

        self.load_layers_from_file(input_path)

    def load_layers_from_autocad(self):
        """Load layers from AutoCAD active drawing."""
        try:
            self.status_var.set("Reading layers from AutoCAD...")
            self.root.update()

            if not self.acad_doc:
                messagebox.showerror("AutoCAD Error", "No active AutoCAD document found.")
                self.status_var.set("Error: No AutoCAD document")
                return

            layers = list_layers_from_autocad(self.acad_doc)

            if not layers:
                messagebox.showwarning(
                    "No Layers Found",
                    "No layers were found in the active AutoCAD drawing."
                )
                self.status_var.set("No layers found in AutoCAD")
                return

            self.available_layers = layers

            # Update dropdown
            self.layer_combo['values'] = self.available_layers
            if self.available_layers:
                self.layer_combo.current(0)

            self.status_var.set(f"✓ Loaded {len(self.available_layers)} layers from AutoCAD")

            # Show message
            messagebox.showinfo(
                "Layers Loaded from AutoCAD",
                f"Found {len(self.available_layers)} layers in active drawing.\n\n"
                f"Select layers from the dropdown and click 'Add Layer'.\n"
                f"Then click 'Extract Coordinates' when ready."
            )

        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(error_detail)
            messagebox.showerror("Error", f"Failed to load layers from AutoCAD:\n\n{str(e)}")
            self.status_var.set("Error loading AutoCAD layers")

    def load_layers_from_file(self, input_path: str):
        """Load layers from DXF/DWG file."""
        try:
            self.status_var.set("Loading file...")
            self.root.update()

            # Check if ezdxf is available
            try:
                import ezdxf
            except ImportError:
                messagebox.showerror(
                    "Missing Library",
                    "ezdxf library is required.\n\nInstall with:\npip install ezdxf"
                )
                self.status_var.set("Error: ezdxf not installed")
                return

            self.status_var.set("Reading CAD file...")
            self.root.update()

            doc = load_cad_file(input_path)
            if not doc:
                messagebox.showerror("Error", "Could not load CAD file. Check if file is valid DXF/DWG.")
                self.status_var.set("Error loading file")
                return

            self.status_var.set("Extracting layers...")
            self.root.update()

            # Extract layer names - comprehensive approach
            layers = set()

            # Method 1: From layer table
            try:
                for layer in doc.layers:
                    layer_name = layer.dxf.name
                    if layer_name and layer_name not in ['0']:  # Skip default layer 0 unless it has entities
                        layers.add(layer_name)
            except Exception as e:
                print(f"Warning: Could not read layer table: {e}")

            # Method 2: Scan all entities in modelspace
            try:
                mspace = doc.modelspace()
                entity_count = 0
                for entity in mspace:
                    entity_count += 1
                    try:
                        layer_name = entity.dxf.layer
                        if layer_name:
                            layers.add(layer_name)
                    except:
                        pass
                print(f"Scanned {entity_count} entities in modelspace")
            except Exception as e:
                print(f"Warning: Could not scan modelspace: {e}")

            # Method 3: Scan paperspace layouts
            try:
                for layout in doc.layouts:
                    for entity in layout:
                        try:
                            layer_name = entity.dxf.layer
                            if layer_name:
                                layers.add(layer_name)
                        except:
                            pass
            except Exception as e:
                print(f"Warning: Could not scan layouts: {e}")

            # Method 4: Scan block definitions
            try:
                for block in doc.blocks:
                    for entity in block:
                        try:
                            layer_name = entity.dxf.layer
                            if layer_name:
                                layers.add(layer_name)
                        except:
                            pass
            except Exception as e:
                print(f"Warning: Could not scan blocks: {e}")

            if not layers:
                messagebox.showwarning(
                    "No Layers Found",
                    "No layers were found in the file.\n\n"
                    "This could mean:\n"
                    "- The file is empty\n"
                    "- The file format is not supported\n"
                    "- All entities are on layer '0'"
                )
                self.status_var.set("No layers found")
                return

            self.available_layers = sorted(list(layers))

            # Update dropdown
            self.layer_combo['values'] = self.available_layers
            if self.available_layers:
                self.layer_combo.current(0)

            self.status_var.set(f"✓ Loaded {len(self.available_layers)} layers from file")

            # Show message
            messagebox.showinfo(
                "Layers Loaded from File",
                f"Found {len(self.available_layers)} layers.\n\n"
                f"Select layers from the dropdown and click 'Add Layer'.\n"
                f"Then click 'Extract Coordinates' when ready."
            )

        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            print(error_detail)
            messagebox.showerror("Error", f"Failed to load layers:\n\n{str(e)}\n\nSee console for details.")
            self.status_var.set("Error loading layers")

    def add_layer_to_list(self):
        """Add selected layer from dropdown to the extraction list."""
        selected_layer = self.layer_combo.get()
        if not selected_layer:
            messagebox.showwarning("No Layer Selected", "Please select a layer from the dropdown.")
            return

        # Check if already in list
        current_layers = self.selected_layers_listbox.get(0, tk.END)
        if selected_layer in current_layers:
            messagebox.showinfo("Already Added", f"Layer '{selected_layer}' is already in the extraction list.")
            return

        # Add to listbox
        self.selected_layers_listbox.insert(tk.END, selected_layer)
        self.status_var.set(f"Added layer: {selected_layer}")

    def remove_layer_from_list(self):
        """Remove selected layer from the extraction list."""
        selection = self.selected_layers_listbox.curselection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a layer to remove from the list.")
            return

        # Remove from listbox (in reverse to maintain indices)
        for index in reversed(selection):
            layer_name = self.selected_layers_listbox.get(index)
            self.selected_layers_listbox.delete(index)
            self.status_var.set(f"Removed layer: {layer_name}")

    def clear_all_layers(self):
        """Clear all layers from the extraction list."""
        self.selected_layers_listbox.delete(0, tk.END)
        self.status_var.set("Cleared all layers")

    def select_entities_in_autocad(self):
        """Allow user to select entities in AutoCAD."""
        if not self.acad_doc or not self.acad_ss:
            messagebox.showerror("AutoCAD Error", "No active AutoCAD connection.")
            return

        # Show instruction message
        result = messagebox.showinfo(
            "Select Entities in AutoCAD",
            "The GUI will minimize.\n\n"
            "1. Switch to AutoCAD\n"
            "2. Select polylines or entities\n"
            "3. Press ENTER when done\n\n"
            "Click OK to continue...",
            type=messagebox.OK
        )

        try:
            # Minimize the GUI window
            self.root.iconify()
            self.root.update()

            # Clear previous selection
            try:
                self.acad_ss.Clear()
            except Exception:
                pass

            # Prompt user to select entities
            self.status_var.set("Waiting for selection in AutoCAD...")

            # Send prompt to AutoCAD command line
            try:
                self.acad_doc.Utility.Prompt("\nSelect polylines/entities, then press ENTER...\n")
            except Exception:
                pass

            # This will pause and wait for user to select entities in AutoCAD
            try:
                self.acad_ss.SelectOnScreen()
            except Exception as e:
                # User cancelled or error occurred
                self.root.deiconify()
                if "rejected" in str(e).lower() or "cancelled" in str(e).lower():
                    self.status_var.set("Selection cancelled")
                    return
                raise

            # Restore the GUI window
            self.root.deiconify()
            self.root.lift()
            self.root.focus_force()

            # Collect selected entities
            self.selected_entities = []
            try:
                count = int(self.acad_ss.Count)
            except Exception:
                count = 0

            for i in range(count):
                try:
                    ent = dyn(self.acad_ss.Item(i))
                    self.selected_entities.append(ent)
                except Exception:
                    continue

            # Update UI
            if self.selected_entities:
                self.selection_label.config(text=f"({len(self.selected_entities)} selected)", foreground="green")
                self.use_selection.set(True)
                self.status_var.set(f"✓ Selected {len(self.selected_entities)} entities in AutoCAD")
                messagebox.showinfo(
                    "Selection Complete",
                    f"Selected {len(self.selected_entities)} entities.\n\n"
                    "These entities will be used for coordinate extraction."
                )
            else:
                self.selection_label.config(text="(0 selected)", foreground="gray")
                self.use_selection.set(False)
                self.status_var.set("No entities selected")

        except Exception as e:
            # Restore window if minimized
            try:
                self.root.deiconify()
            except Exception:
                pass

            import traceback
            error_detail = traceback.format_exc()
            print(error_detail)

            # Check if it's a cancellation
            if "rejected" in str(e).lower() or "cancelled" in str(e).lower():
                self.status_var.set("Selection cancelled")
            else:
                messagebox.showerror("Selection Error", f"Failed to select entities:\n\n{str(e)}")
                self.status_var.set("Selection failed")

    def update_progress(self, current, _total, message):
        """Update progress bar and status."""
        self.progress['value'] = current
        self.status_var.set(message)
        self.root.update()

    def extract_coordinates(self):
        """Run the coordinate extraction process."""
        # Check if using selection mode
        use_selection = self.use_selection.get() if hasattr(self, 'use_selection') else False

        if use_selection:
            # Selection mode - extract from selected entities
            if not self.selected_entities:
                messagebox.showwarning("No Selection", "Please select entities in AutoCAD first.\n\nClick 'Select Entities in AutoCAD' button.")
                return

            # In selection mode, we don't need layers
            selected_layers = ["[Selected Entities]"]
        else:
            # Layer mode - get selected layers from the listbox
            selected_layers = list(self.selected_layers_listbox.get(0, tk.END))
            if not selected_layers:
                messagebox.showwarning("No Layers", "Please add at least one layer to extract.\n\nSelect a layer from the dropdown and click 'Add Layer'.")
                return

        # Validate input source
        is_autocad = self.autocad_mode.get()
        if is_autocad:
            if not self.acad_doc:
                messagebox.showerror("AutoCAD Error", "No active AutoCAD document found.")
                return
            # Extract directly from AutoCAD - no file needed!
            input_path = None
        else:
            if use_selection:
                messagebox.showwarning("Selection Mode", "Selection mode requires AutoCAD connection.\n\nPlease use 'Use AutoCAD Active Drawing' mode.")
                return

            input_path = self.input_file.get()
            if not input_path or input_path == "[AutoCAD Active Drawing]":
                messagebox.showwarning("No Input", "Please select an input file.")
                return

            if not os.path.exists(input_path):
                messagebox.showerror("File Not Found", f"Input file not found: {input_path}")
                return

        output_path = self.output_file.get()
        if not output_path:
            messagebox.showwarning("No Output", "Please specify an output file.")
            return

        try:
            # Process each layer (or selected entities)
            all_rows = []
            total_layers = len(selected_layers)

            for idx, layer in enumerate(selected_layers):
                if use_selection:
                    self.status_var.set(f"Processing {len(self.selected_entities)} selected entities...")
                else:
                    self.status_var.set(f"Processing layer {idx+1}/{total_layers}: {layer}")
                self.progress['value'] = (idx / total_layers) * 90
                self.root.update()

                if is_autocad:
                    # Extract directly from AutoCAD ModelSpace
                    entities_to_use = self.selected_entities if use_selection else None
                    layer_to_use = None if use_selection else layer

                    rows = extract_from_autocad_layer(
                        self.acad_doc,
                        layer_to_use,
                        self.mode_var.get(),
                        self.prefix_var.get(),
                        self.decimals_var.get(),
                        selected_entities=entities_to_use
                    )
                    all_rows.extend(rows)
                else:
                    # Extract from DXF/DWG file
                    config = StandaloneConfig(
                        input_file=input_path,
                        output_excel=output_path,
                        target_layer=layer,
                        extraction_mode=self.mode_var.get(),
                        point_prefix=self.prefix_var.get(),
                        start_number=1,
                        decimal_places=self.decimals_var.get(),
                        use_blocks=True,
                        include_modelspace=True,
                        verbose=False
                    )

                    rows, _ = process_file(config, progress_callback=self.update_progress)
                    all_rows.extend(rows)

            # Export combined results using openpyxl (like coordtable.py)
            if all_rows:
                self.status_var.set("Exporting to Excel...")
                self.progress['value'] = 95
                self.root.update()

                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Alignment, Font, PatternFill
                    from openpyxl.utils import get_column_letter

                    # Create workbook
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Coordinates"

                    # Determine if we have corners mode
                    has_corners = self.mode_var.get() == "corners"
                    decimal_places = self.decimals_var.get()

                    # Create headers
                    if has_corners:
                        headers = ["Point ID", "Corner", "East (X)", "North (Y)", "Layer"]
                    else:
                        headers = ["Point ID", "East (X)", "North (Y)", "Layer"]

                    ws.append(headers)

                    # Style headers
                    header_font = Font(bold=True)
                    header_fill = PatternFill("solid", fgColor="D9E1F2")
                    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

                    for col_idx, h in enumerate(headers, start=1):
                        c = ws.cell(row=1, column=col_idx, value=h)
                        c.font = header_font
                        c.fill = header_fill
                        c.alignment = header_align

                    # Add data rows
                    for row in all_rows:
                        if has_corners:
                            ws.append([row.point_id, row.corner, row.x, row.y, row.layer])
                        else:
                            ws.append([row.point_id, row.x, row.y, row.layer])

                    # Freeze header row
                    ws.freeze_panes = "A2"

                    # Add auto-filter
                    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(all_rows) + 1}"

                    # Format numeric columns
                    num_fmt = f"0.{'0' * decimal_places}"
                    numeric_cols = {"East (X)", "North (Y)"}

                    for col_idx, h in enumerate(headers, start=1):
                        col_letter = get_column_letter(col_idx)

                        # Auto-size columns
                        width = len(h)
                        for row_idx in range(2, len(all_rows) + 2):
                            v = ws.cell(row=row_idx, column=col_idx).value
                            if v is not None:
                                width = max(width, len(str(v)))
                        ws.column_dimensions[col_letter].width = min(max(width + 2, 12), 70)

                        # Format numeric columns
                        if h in numeric_cols:
                            for row_idx in range(2, len(all_rows) + 2):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                if isinstance(cell.value, (int, float)):
                                    cell.number_format = num_fmt
                                    cell.alignment = Alignment(horizontal="right")

                    # Save workbook
                    wb.save(output_path)
                    final_path = output_path

                    self.progress['value'] = 100
                    self.status_var.set(f"Complete! Extracted {len(all_rows)} points")

                    messagebox.showinfo(
                        "Success",
                        f"Extraction complete!\n\n"
                        f"Points extracted: {len(all_rows)}\n"
                        f"Layers processed: {len(selected_layers)}\n"
                        f"Output: {final_path}"
                    )

                except Exception as e:
                    import traceback
                    error_detail = traceback.format_exc()
                    print(error_detail)
                    messagebox.showerror("Excel Export Error", f"Failed to export to Excel:\n\n{str(e)}")
                    self.status_var.set("Export failed")
            else:
                messagebox.showwarning("No Data", "No coordinates were extracted from the selected layers.")
                self.status_var.set("No data extracted")

        except Exception as e:
            messagebox.showerror("Error", f"Extraction failed:\n{str(e)}")
            self.status_var.set("Error during extraction")
            self.progress['value'] = 0


def main():
    """Launch the GUI application."""
    root = tk.Tk()
    app = CoordinatesGrabberGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()

